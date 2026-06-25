from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import close_old_connections, models
from django.db.models import Q
from urllib.parse import urlencode

from .models import TauxEvolution, Devise
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.hashers import make_password
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.views import PasswordChangeView
from django.contrib.messages.views import SuccessMessageMixin

from django.core.mail import send_mail, BadHeaderError

from .models import TauxEvolution_filiale
import json
import csv


import openpyxl
from django.template.loader import get_template
try:
    from xhtml2pdf import pisa
except ImportError:
    pisa = None
from kyc.models import (
    Notation, Historique, Kyc_pm, Kyc_pp, Anomalie, TauxEvolution, DATEREV,
    DataQualityRule, DataQualityRuleAudit, KycDocumentExtraction, KycExpiredDocumentScanMatch,
    KycDocumentMatchJob, KycDocumentMatchSettings, DOCUMENT_EXTRACTION_TYPE_CHOICES,
    KycFieldVisibilityConfig, KycDocumentType, Filiales, CLIENT_TYPE_CHOICES,
    DATA_QUALITY_FIELD_CHOICES,
)
from django.utils import timezone
from django.utils.decorators import method_decorator

from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, FileResponse, Http404
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode
from django.views.decorators.csrf import csrf_exempt
from oauthlib.oauth2.rfc6749.endpoints import token

from django.db.models import Max, Avg, Count, F, Q, Exists, OuterRef, Case, When, IntegerField, CharField
from django.db.models.functions import TruncDate, Length

# Register Length lookup to allow __length=1 in Q objects (faster than __regex)
CharField.register_lookup(Length)

from django.http import JsonResponse
from .models import Kyc_pp

from accounts.models import ProfileV, UserLoginHistory
from django.core.cache import cache
from kyc import forms
from kyc.forms import CustomUserCreationForm, LoginForm, ResetPasswordForm, UserEditForm, VoyageurProfileForm, \
    CambProfileForm, \
    ProfileModify, NotationForm, ProfileForm
from django.contrib.sessions.models import Session
from django.utils.timezone import now
from django.utils.dateparse import parse_date, parse_datetime

from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from .models import Kyc_pp
from django.conf import settings
from django.core.files.base import ContentFile
import os
import re
import sys
import subprocess
import hashlib
import math
import threading
import uuid
import zipfile
from .document_extraction import SUPPORTED_EXTENSIONS, extract_document_data, extract_pdf_grouped_documents

class CachedPaginator(Paginator):
    @property
    def count(self):
        query_str = str(self.object_list.query)
        query_hash = hashlib.md5(query_str.encode('utf-8')).hexdigest()
        cache_key = f"paginator_count_{query_hash}"
        cached_count = cache.get(cache_key)
        if cached_count is not None:
            return cached_count
        actual_count = super().count
        cache.set(cache_key, actual_count, 300)  # Cacher pendant 5 minutes
        return actual_count

def floor_one_decimal(value):
    return math.floor(value * 10) / 10

def compliance_rate_floor(ok_count, total, fail_count=0):
    if not total:
        return None

    rate = floor_one_decimal(ok_count / total * 100)
    if fail_count > 0 and rate >= 100:
        return 99.9
    return rate

def _pdf_link_callback(uri, rel):
    """Resolve static/media URIs for xhtml2pdf on local filesystem."""
    if uri.startswith(settings.MEDIA_URL):
        path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
    elif uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
        if not os.path.isfile(path):
            path = os.path.join(settings.BASE_DIR, "static", uri.replace(settings.STATIC_URL, ""))
    else:
        return uri

    if not os.path.isfile(path):
        raise Exception(f"Media URI must start with {settings.MEDIA_URL} or {settings.STATIC_URL}: {uri}")
    return path


def format_date_for_export(value, output_format="%d/%m/%Y", empty_value="-"):
    """Format date/datetime/string values safely for exports."""
    if value in (None, ""):
        return empty_value

    if hasattr(value, "strftime"):
        try:
            return value.strftime(output_format)
        except Exception:
            pass

    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return empty_value

        parsed_dt = parse_datetime(cleaned)
        if parsed_dt:
            return parsed_dt.strftime(output_format)

        parsed_d = parse_date(cleaned)
        if parsed_d:
            return parsed_d.strftime(output_format)

        for input_format in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(cleaned, input_format).strftime(output_format)
            except ValueError:
                continue

        return cleaned

    return str(value)

def get_data_quality_field_options():
    return {
        'PP': [
            ('FILIALE', 'FILIALE'),
            ('AGENCE', 'AGENCE'),
            ('LIB_AGENCE', 'LIB_AGENCE'),
            ('EXPL', 'EXPL'),
            ('CLIENT', 'CLIENT'),
            ('CODAPE', 'CODAPE'),
            ('IDP', 'IDP'),
            ('PAYNAIS', 'PAYNAIS'),
            ('PROFESSION', 'PROFESSION'),
            ('ADRESSE', 'ADRESSE'),
            ('PAYS_RESID', 'PAYS_RESID'),
            ('NUMID', 'NUMID'),
            ('SALAIRE', 'SALAIRE'),
            ('ORIGINE_REV', 'ORIGINE_REV'),
            ('DATVALID', 'DATVALID'),
            ('DATNAIS', 'DATNAIS'),
            ('TEL', 'TEL'),
            ('DATOUV', 'DATOUV'),
            ('PPE', 'PPE'),
            ('DEVISE', 'DEVISE'),
            ('RESID', 'RESID'),
            ('DATEREV', 'DATEREV'),
            ('RISQUE', 'RISQUE'),
            ('BOITE_POSTALE', 'BOITE_POSTALE'),
            ('CONSENT_BIC', 'CONSENT_BIC'),
            ('EMPLOYEUR', 'EMPLOYEUR'),
            ('INTITULE_COMPTE', 'INTITULE_COMPTE'),
            ('LIEU_DELIVRANCE_CIN', 'LIEU_DELIVRANCE_CIN'),
        ],
        'PM': [
            ('FILIALE', 'FILIALE'),
            ('AGENCE', 'AGENCE'),
            ('LIB_AGENCE', 'LIB_AGENCE'),
            ('EXPL', 'EXPL'),
            ('CLIENT', 'CLIENT'),
            ('AGEC', 'AGEC'),
            ('CODAPE', 'CODAPE'),
            ('IDM', 'IDM'),
            ('RCSNO', 'RCSNO'),
            ('CAPITAL', 'CAPITAL'),
            ('CA', 'CA'),
            ('RESULTAT', 'RESULTAT'),
            ('ORIGINE_REV', 'ORIGINE_REV'),
            ('DATOUV', 'DATOUV'),
            ('TEL', 'TEL'),
            ('DEVISE', 'DEVISE'),
            ('RESID', 'RESID'),
            ('DATEREV', 'DATEREV'),
            ('PPE', 'PPE'),
            ('RISQUE', 'RISQUE'),
            ('ACTIONNAIRE', 'ACTIONNAIRE'),
            ('ADRESSE_SOCIALE', 'ADRESSE_SOCIALE'),
            ('BOITE_POSTALE', 'BOITE_POSTALE'),
            ('CONSENT_BIC', 'CONSENT_BIC'),
            ('INTITULE_COMPTE', 'INTITULE_COMPTE'),
            ('MANDATAIRE', 'MANDATAIRE'),
            ('NUMERO_FISCAL', 'NUMERO_FISCAL'),
            ('PAYS_JUR', 'PAYS_JUR'),
        ],
    }


def evaluate_data_quality_scope(user):
    """Détermine le périmÃ¨tre de calcul qualité selon l'organe utilisateur."""
    organe = (getattr(user, 'organe', '') or '').strip()
    filiale = (getattr(user, 'filiale', '') or '').strip()
    agence = (getattr(user, 'agence', '') or '').strip()
    code_expl = (getattr(user, 'code_expl', '') or '').strip()

    if organe == 'Chargé Client':
        return {
            'filiale': filiale or None,
            'agence': agence or None,
            'expl': code_expl or None,
            'label': 'Mon portefeuille',
        }
    if organe == 'Directeur Agence':
        return {
            'filiale': filiale or None,
            'agence': agence or None,
            'expl': None,
            'label': f"Agence {agence}" if agence else 'Mon agence',
        }
    if organe == 'PASS' or 'Groupe' in organe:
        return {
            'filiale': None,
            'agence': None,
            'expl': None,
            'label': 'GROUPE (toutes filiales)',
        }

    return {
        'filiale': filiale or None,
        'agence': None,
        'expl': None,
        'label': filiale or 'Ma filiale',
    }


def _quality_cache_version():
    return cache.get('quality_control_rules_version', 1)

def _rule_eval_filiale(rule, user_filiale):
    from kyc.forms import DataQualityRuleForm
    parsed = DataQualityRuleForm._parse_filiales(rule.filiale)
    if not parsed:
        return user_filiale
    if user_filiale:
        return user_filiale if user_filiale in parsed else None
    if len(parsed) == 1:
        return parsed[0]
    return None

def _evaluate_data_quality_rule_scoped(rule, filiale=None, agence=None, expl=None):
    return evaluate_data_quality_rule(rule, filiale=filiale, agence=agence, expl=expl)

def get_incomplete_clients_queryset(queryset, client_type):
    from kyc.models import KycFieldVisibilityConfig
    from django.db.models import Q
    
    filiales = list(queryset.values_list('FILIALE', flat=True).distinct())
    if not filiales:
        return queryset.none()
        
    configs = list(KycFieldVisibilityConfig.objects.filter(client_type=client_type))
    
    combined_q = Q()
    for filiale in filiales:
        config = None
        if filiale:
            config = next((c for c in configs if filiale in (c.filiales or [])), None)
        if not config:
            config = next((c for c in configs if not c.filiales), None)
            
        if config and config.empty_check_fields:
            fields = config.empty_check_fields
        else:
            if client_type == 'pp':
                fields = ["NUMID", "DATNAIS", "ADRESSE", "TEL"]
            else:
                fields = ["NUMERO_FISCAL", "RCSNO", "ADRESSE_SOCIALE", "TEL"]
                
        field_q = Q()
        for f in fields:
            if f in ["CLIENT", "EXPL", "FILIALE", "AGENCE", "LIB_AGENCE"]:
                continue
            field_q |= Q(**{f"{f}__isnull": True}) | Q(**{f"{f}": ""})
            
        if filiale:
            combined_q |= Q(FILIALE=filiale) & field_q
        else:
            combined_q |= (Q(FILIALE__isnull=True) | Q(FILIALE="")) & field_q
            
    return queryset.filter(combined_q)

def evaluate_data_quality_rule(rule, filiale=None, agence=None, expl=None):
    model = Kyc_pp if rule.applicability == 'PP' else Kyc_pm
    field_names = [f.name for f in model._meta.get_fields() if not f.many_to_many and not f.one_to_many]
    if rule.control_type != 'composite' and rule.field_name not in field_names and rule.control_type not in ['expired_document', 'codape_agec_match']:
        return {'total': 0, 'fail_count': 0, 'ok_count': 0, 'clients': [], 'message': 'Champ de contrôle invalide'}

    from kyc.forms import DataQualityRuleForm
    parsed_filiales = DataQualityRuleForm._parse_filiales(rule.filiale)

    queryset = model.objects.all()
    if filiale and filiale != 'GROUPE':
        if parsed_filiales and filiale not in parsed_filiales:
            return {'total': 0, 'fail_count': 0, 'ok_count': 0, 'clients': [], 'message': 'Non applicable à cette filiale'}
        queryset = queryset.filter(FILIALE=filiale)
    elif parsed_filiales:
        queryset = queryset.filter(FILIALE__in=parsed_filiales)

    if agence:
        queryset = queryset.filter(AGENCE=agence)
    if expl:
        queryset = queryset.filter(EXPL=expl)
        
    total = queryset.count()
    if total == 0:
        return {'total': 0, 'fail_count': 0, 'ok_count': 0, 'clients': [], 'message': 'Aucune donnée disponible pour ce segment'}

    failures = []
    today = datetime.today().date()
    client_fields = ['CLIENT', 'EXPL', 'FILIALE', 'AGENCE']

    def safe_parse_date(value):
        if not value: return None
        if hasattr(value, 'date'): return value.date()
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y%m%d"):
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        return None

    def calculate_age(birth_date_str):
        parsed = safe_parse_date(birth_date_str)
        if not parsed: return None
        today_date = datetime.today().date()
        return today_date.year - parsed.year - ((today_date.month, today_date.day) < (parsed.month, parsed.day))

    def build_clients_from_values(rows, value_key):
        return [{
            'client': row.get('CLIENT', ''),
            'expl': row.get('EXPL', ''),
            'filiale': row.get('FILIALE', ''),
            'agence': row.get('AGENCE', ''),
            'field_value': str(row.get(value_key, '') or ''),
        } for row in rows]

    if rule.control_type == 'simple':
        param = (rule.parameter or '').strip().lower()
        
        # Determine if it's existence, length, or value match
        if not param or param == 'existence':
            # Existence check (is empty?)
            failures = queryset.filter(Q(**{f"{rule.field_name}__isnull": True}) | Q(**{f"{rule.field_name}": ""}))
            fail_count = failures.count()
            clients = build_clients_from_values(list(failures.values(*client_fields, rule.field_name)[:15]), rule.field_name)
        
        elif param.isdigit() or (param.startswith('len') or param.startswith('long')):
            # Length check
            try:
                # Extract number from param if it's like 'length:10'
                import re
                match = re.search(r'\d+', param)
                target_len = int(match.group()) if match else int(param)
                
                fail_count = 0
                clients = []
                for row in queryset.values(*client_fields, rule.field_name).iterator(chunk_size=2000):
                    val = str(row.get(rule.field_name) or '')
                    if len(val) != target_len:
                        fail_count += 1
                        if len(clients) < 15:
                            clients.append({
                                'client': row.get('CLIENT', ''),
                                'expl': row.get('EXPL', ''),
                                'filiale': row.get('FILIALE', ''),
                                'agence': row.get('AGENCE', ''),
                                'field_value': val,
                            })
            except:
                return {'total': total, 'fail_count': 0, 'ok_count': total, 'clients': [], 'message': 'ParamÃ¨tre de longueur invalide'}
        
        else:
            # Value match check
            failures = queryset.exclude(**{f"{rule.field_name}": rule.parameter})
            fail_count = failures.count()
            clients = build_clients_from_values(list(failures.values(*client_fields, rule.field_name)[:15]), rule.field_name)

        ok_count = total - fail_count
        return {
            'total': total,
            'fail_count': fail_count,
            'ok_count': ok_count,
            'clients': clients,
            'message': '',
        }

    elif rule.control_type == 'composite':
        conditions = rule.conditions.all()
        if not conditions.exists():
            return {'total': total, 'fail_count': 0, 'ok_count': total, 'clients': [], 'message': 'Pas de conditions'}
        
        fail_count = 0
        clients = []
        needed_fields = set(client_fields)
        for c in conditions:
            needed_fields.add(c.field_name)
        for row in queryset.values(*needed_fields).iterator(chunk_size=2000):
            all_match = True
            for cond in conditions:
                val = str(row.get(cond.field_name, '') or '').strip()
                target = (cond.value or '').strip()
                
                match = False
                op = cond.operator
                if op == '=': match = val == target
                elif op == '!=': match = val != target
                elif op == '>':
                    try: match = float(val.replace(',','.')) > float(target.replace(',','.'))
                    except: match = False
                elif op == '<':
                    try: match = float(val.replace(',','.')) < float(target.replace(',','.'))
                    except: match = False
                elif op == '>=':
                    try: match = float(val.replace(',','.')) >= float(target.replace(',','.'))
                    except: match = False
                elif op == '<=':
                    try: match = float(val.replace(',','.')) <= float(target.replace(',','.'))
                    except: match = False
                elif op == 'contains': match = target.lower() in val.lower()
                elif op == 'is_empty': match = not val
                elif op == 'is_not_empty': match = bool(val)
                elif op == 'expired':
                    p = safe_parse_date(val)
                    match = p and p < today
                elif op == 'age_gt':
                    age = calculate_age(val)
                    try: match = age is not None and age > int(target)
                    except: match = False
                elif op == 'age_lt':
                    age = calculate_age(val)
                    try: match = age is not None and age < int(target)
                    except: match = False
                elif op == 'min_length':
                    try: match = len(val) < int(target) # C'est un échec si la longueur est inférieure au min
                    except: match = False
                elif op == 'max_length':
                    try: match = len(val) > int(target) # C'est un échec si la longueur est supérieure au max
                    except: match = False
                
                if not match:
                    all_match = False
                    break
            
            if all_match:
                fail_count += 1
                if len(clients) < 15:
                    clients.append({
                        'client': row.get('CLIENT', ''),
                        'expl': row.get('EXPL', ''),
                        'filiale': row.get('FILIALE', ''),
                        'agence': row.get('AGENCE', ''),
                        'field_value': 'Multi-critÃ¨res',
                    })
        
        ok_count = total - fail_count
        return {
            'total': total,
            'fail_count': fail_count,
            'ok_count': ok_count,
            'clients': clients,
            'message': '',
        }

    return {
        'total': total,
        'fail_count': 0,
        'ok_count': total,
        'clients': [],
        'message': '',
    }


@login_required
def quality_control_view(request):
    user = request.user
    allowed_organs = ['Contrôle Permanent', 'Conformité', 'Qualité', 'DSI', 'Risques', 'DAI', 'PASS']
    user_organe = (getattr(user, 'organe', '') or '').strip()
    if user_organe not in allowed_organs:
        messages.error(request, "AccÃ¨s non autorisé au contrôle qualité.")
        return redirect('accueil')

    from kyc.forms import DataQualityRuleForm, DataQualityConditionFormSet

    # Vérification des droits de gestion
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    can_manage = user_organe in ['Conformité', 'Contrôle Permanent', 'PASS']
    user_filiale = getattr(request.user, 'filiale', '')
    
    if user_organe == 'PASS':
        from kyc.models import Filiales as ModelFiliales
        filiale_choices = [f[0] for f in ModelFiliales]
    else:
        filiale_choices = [user_filiale] if user_filiale else []
    
    if request.method == 'POST' and can_manage:
        form = DataQualityRuleForm(request.POST, filiale_choices=filiale_choices)
        formset = DataQualityConditionFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            rule = form.save(commit=False)
            rule.created_by = request.user
            rule.save()
            formset.instance = rule
            formset.save()
            
            # Audit creation
            DataQualityRuleAudit.objects.create(
                rule_name=rule.name,
                user=request.user,
                action='CREATION',
                details=f"Création de la rÃ¨gle '{rule.name}' ({rule.applicability})"
            )
            
            current_version = cache.get('quality_control_rules_version', 1)
            cache.set('quality_control_rules_version', current_version + 1, timeout=None)
            messages.success(request, 'RÃ¨gle de qualité enregistrée.')
            return redirect('kyc:quality_control')
    else:
        form = DataQualityRuleForm(filiale_choices=filiale_choices)
        formset = DataQualityConditionFormSet()

    rules = list(
        DataQualityRule.objects.all()
        .order_by('-active', '-created_at')
        .prefetch_related('conditions')
    )
    cache_ttl_seconds = 86400  # Cache journalier: donnees mises a jour 1 fois/jour
    rules_version = cache.get('quality_control_rules_version', 1)
    data_refresh_bucket = timezone.localdate().isoformat()
    stats = []
    
    # Portée de l'évaluation : vision groupe pour PASS et les organes Groupe
    group_organs = ['PASS', 'Conformité Groupe']
    eval_filiale = None if user_organe in group_organs else user_filiale

    for rule in rules:
        rule_signature = f"{rule.id}|{rule.name}|{rule.applicability}|{rule.field_name}|{rule.control_type}|{rule.parameter}|{rule.active}|{eval_filiale}"
        rule_hash = hashlib.md5(rule_signature.encode('utf-8')).hexdigest()
        rule_cache_key = f"quality_control:stat:v{rules_version}:d{data_refresh_bucket}:{rule_hash}"
        stat = cache.get(rule_cache_key)
        if stat is None:
            stat = evaluate_data_quality_rule(rule, filiale=eval_filiale)
            cache.set(rule_cache_key, stat, timeout=cache_ttl_seconds)
        stats.append(stat)
    for stat in stats:
        total = stat.get('total', 0)
        stat['compliance_rate'] = compliance_rate_floor(stat['ok_count'], total, stat.get('fail_count', 0))
    rules_with_stats = []
    for rule, stat in zip(rules, stats):
        grouped_conditions = {}
        for cond in rule.conditions.all():
            group_key = (cond.field_name, cond.operator)
            if group_key not in grouped_conditions:
                grouped_conditions[group_key] = {
                    'field_name': cond.field_name,
                    'operator_display': cond.get_operator_display(),
                    'values': [],
                }

            value = (cond.value or '').strip()
            if value and value not in grouped_conditions[group_key]['values']:
                grouped_conditions[group_key]['values'].append(value)

        from kyc.forms import DataQualityRuleForm
        parsed_filiales = DataQualityRuleForm._parse_filiales(rule.filiale)
        filiales_display = ", ".join(parsed_filiales) if parsed_filiales else "Toutes les filiales"

        rules_with_stats.append({
            'rule': rule,
            'stat': stat,
            'condition_groups': list(grouped_conditions.values()),
            'filiales_display': filiales_display,
        })

    field_options = get_data_quality_field_options()
    total_rules = len(rules)
    active_rules = sum(1 for rule in rules if rule.active)
    inactive_rules = total_rules - active_rules
    total_failures = sum(stat['fail_count'] for stat in stats)
    total_ok = sum(stat['ok_count'] for stat in stats)
    total_evaluated = sum(stat['total'] for stat in stats)
    global_compliance_rate = compliance_rate_floor(total_ok, total_evaluated, total_failures)

    return render(request, 'quality_control.html', {
        'form': form,
        'formset': formset,
        'rules': rules_with_stats,
        'field_options': field_options,
        'total_rules': total_rules,
        'active_rules': active_rules,
        'inactive_rules': inactive_rules,
        'total_failures': total_failures,
        'global_compliance_rate': global_compliance_rate,
        'form_has_errors': bool(form.errors),
        'can_manage': can_manage,
        'user_organe': user_organe,
    })

@login_required
def delete_quality_rule(request, pk):
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    if user_organe not in ['Conformité', 'Contrôle Permanent', 'PASS']:
        messages.error(request, "AccÃ¨s refusé.")
        return redirect('kyc:quality_control')
        
    rule = get_object_or_404(DataQualityRule, pk=pk)
    
    # Vérification filiale si pas PASS
    if user_organe != 'PASS' and rule.created_by and rule.created_by.filiale != request.user.filiale:
        messages.error(request, "Vous ne pouvez supprimer que les rÃ¨gles de votre filiale.")
        return redirect('kyc:quality_control')

    rule_name = rule.name
    DataQualityRuleAudit.objects.create(
        rule_name=rule_name,
        user=request.user,
        action='SUPPRESSION',
        details=f"Suppression de la rÃ¨gle '{rule_name}'"
    )
    
    rule.delete()
    current_version = cache.get('quality_control_rules_version', 1)
    cache.set('quality_control_rules_version', current_version + 1, timeout=None)
    messages.success(request, f"RÃ¨gle '{rule_name}' supprimée.")
    return redirect('kyc:quality_control')

@login_required
def edit_quality_rule(request, pk):
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    if user_organe not in ['Conformité', 'Contrôle Permanent', 'PASS']:
        messages.error(request, "AccÃ¨s refusé.")
        return redirect('kyc:quality_control')
        
    rule = get_object_or_404(DataQualityRule, pk=pk)
    
    # Vérification filiale si pas PASS
    if user_organe != 'PASS' and rule.created_by and rule.created_by.filiale != request.user.filiale:
        messages.error(request, "Vous ne pouvez modifier que les rÃ¨gles de votre filiale.")
        return redirect('kyc:quality_control')

    from kyc.forms import DataQualityRuleForm, DataQualityConditionFormSet
    
    if user_organe == 'PASS':
        from kyc.models import Filiales as ModelFiliales
        filiale_choices = [f[0] for f in ModelFiliales]
    else:
        filiale_choices = [request.user.filiale] if request.user.filiale else []
        
    if request.method == 'POST':
        form = DataQualityRuleForm(request.POST, instance=rule, filiale_choices=filiale_choices)
        formset = DataQualityConditionFormSet(request.POST, instance=rule)
        if form.is_valid() and formset.is_valid():
            changes = []
            if form.has_changed():
                for field in form.changed_data:
                    old = getattr(rule, field)
                    new = form.cleaned_data.get(field)
                    changes.append(f"{field}: {old} -> {new}")
            
            form.save()
            formset.save()
            
            DataQualityRuleAudit.objects.create(
                rule_name=rule.name,
                user=request.user,
                action='MODIFICATION',
                details="; ".join(changes) if changes else "Modification des conditions"
            )
            
            current_version = cache.get('quality_control_rules_version', 1)
            cache.set('quality_control_rules_version', current_version + 1, timeout=None)
            messages.success(request, "RÃ¨gle mise Ã  jour.")
            return redirect('kyc:quality_control')
    else:
        form = DataQualityRuleForm(instance=rule, filiale_choices=filiale_choices)
        formset = DataQualityConditionFormSet(instance=rule)
        
    return render(request, 'quality_rule_edit.html', {
        'form': form,
        'formset': formset,
        'rule': rule
    })

@login_required
def quality_control_audits(request):
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    if user_organe not in ['Conformité', 'Contrôle Permanent', 'PASS']:
        messages.error(request, "AccÃ¨s refusé.")
        return redirect('kyc:quality_control')
        
    if user_organe == 'PASS':
        audits = DataQualityRuleAudit.objects.all().order_by('-timestamp')
    else:
        # Inclure aussi les audits systeme (user null) pour eviter une page vide
        audits = DataQualityRuleAudit.objects.filter(
            Q(user__filiale=request.user.filiale) | Q(user__isnull=True)
        ).order_by('-timestamp')
        
    audits = list(audits)
    for audit in audits:
        raw_rule_name = (audit.rule_name or "")
        audit.rule_name_display = raw_rule_name.strip() or "N/A"
        audit.time_display = audit.timestamp.strftime("%H:%M:%S") if audit.timestamp else "--:--:--"
    return render(request, 'quality_control_audits.html', {'audits': audits})

@login_required
def export_audits_excel(request):
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    if user_organe not in ['Conformité', 'Contrôle Permanent', 'PASS']:
        return HttpResponseForbidden()
        
    if user_organe == 'PASS':
        audits = DataQualityRuleAudit.objects.all().order_by('-timestamp')
    else:
        audits = DataQualityRuleAudit.objects.filter(
            Q(user__filiale=request.user.filiale) | Q(user__isnull=True)
        ).order_by('-timestamp')
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit des Contrôles"
    ws.append(["Date & Heure", "Utilisateur", "RÃ¨gle", "Action", "Détails"])
    for audit in audits:
        ws.append([audit.timestamp.strftime("%d/%m/%Y %H:%M:%S"), audit.user.username if audit.user else "System", audit.rule_name, audit.action, audit.details])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=audit_controles.xlsx'
    wb.save(response)
    return response

@login_required
def export_audits_pdf(request):
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    if user_organe not in ['Conformité', 'Contrôle Permanent', 'PASS']:
        return HttpResponseForbidden()
        
    if user_organe == 'PASS':
        audits = DataQualityRuleAudit.objects.all().order_by('-timestamp')
    else:
        audits = DataQualityRuleAudit.objects.filter(
            Q(user__filiale=request.user.filiale) | Q(user__isnull=True)
        ).order_by('-timestamp')
        
    template_path = 'quality_control_audits_pdf.html'
    logo_rel_path = "images/boa.png"
    logo_full_path = os.path.join(settings.MEDIA_ROOT, logo_rel_path)
    context = {
        'audits': audits,
        'logo_path': logo_full_path if os.path.exists(logo_full_path) else None
    }
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="audit_controles.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    if not pisa:
        return HttpResponse("L'exportation PDF n'est pas disponible sur ce serveur (dépendances manquantes).")
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=_pdf_link_callback)
    if pisa_status.err: return HttpResponse('Erreur PDF')
    return response

@login_required
def export_rule_failures(request, rule_id):
    rule = get_object_or_404(DataQualityRule, pk=rule_id)
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    user_filiale = getattr(request.user, 'filiale', '')
    
    # Portée de l'évaluation
    group_organs = ['PASS', 'Conformité Groupe']
    eval_filiale = None if user_organe in group_organs else user_filiale
    
    # Re-évaluer pour obtenir TOUS les échecs (sans limite de 15)
    model = Kyc_pp if rule.applicability == 'PP' else Kyc_pm
    queryset = model.objects.all()
    if eval_filiale and eval_filiale != 'GROUPE':
        queryset = queryset.filter(FILIALE=eval_filiale)
        
    # Logic similar to evaluate_data_quality_rule but without chunking for building all failures
    # To avoid memory issues with huge datasets, we'll stream the response or just use values().iterator()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anomalies"
    
    # Headers dynamiques
    headers = ["CLIENT", "EXPL", "FILIALE", "AGENCE"]
    if rule.control_type == 'simple':
        headers.append(rule.field_name.upper())
    else:
        # Composite: tous les champs des conditions
        cond_fields = list(rule.conditions.values_list('field_name', flat=True))
        # Nettoyage des doublons tout en gardant l'ordre
        seen = set()
        unique_fields = [f for f in cond_fields if not (f in seen or seen.add(f))]
        for f in unique_fields:
            headers.append(f.upper())
            
    ws.append(headers)
    
    if rule.control_type == 'simple':
        param = (rule.parameter or '').strip().lower()
        if not param or param == 'existence':
            failures = queryset.filter(Q(**{f"{rule.field_name}__isnull": True}) | Q(**{f"{rule.field_name}": ""}))
            for row in failures.values("CLIENT", "EXPL", "FILIALE", "AGENCE", rule.field_name).iterator():
                ws.append([row['CLIENT'], row['EXPL'], row['FILIALE'], row['AGENCE'], str(row.get(rule.field_name) or '')])
        
        elif param.isdigit() or (param.startswith('len') or param.startswith('long')):
            import re
            match = re.search(r'\d+', param)
            target_len = int(match.group()) if match else int(param)
            for row in queryset.values("CLIENT", "EXPL", "FILIALE", "AGENCE", rule.field_name).iterator():
                val = str(row.get(rule.field_name) or '')
                if len(val) != target_len:
                    ws.append([row['CLIENT'], row['EXPL'], row['FILIALE'], row['AGENCE'], val])
        else:
            failures = queryset.exclude(**{f"{rule.field_name}": rule.parameter})
            for row in failures.values("CLIENT", "EXPL", "FILIALE", "AGENCE", rule.field_name).iterator():
                ws.append([row['CLIENT'], row['EXPL'], row['FILIALE'], row['AGENCE'], str(row.get(rule.field_name) or '')])
                
    elif rule.control_type == 'composite':
        conditions = rule.conditions.all()
        fields_to_fetch = ["CLIENT", "EXPL", "FILIALE", "AGENCE"]
        cond_fields = [c.field_name for c in conditions]
        unique_cond_fields = list(dict.fromkeys(cond_fields))
        fields_to_fetch.extend(unique_cond_fields)
        
        today = datetime.today().date()

        for row in queryset.values(*set(fields_to_fetch)).iterator():
            all_match = True
            for cond in conditions:
                val = str(row.get(cond.field_name, '') or '').strip()
                target = str(cond.value or '').strip()
                
                match = False
                op = cond.operator
                if op == '=': match = val == target
                elif op == '!=': match = val != target
                elif op == '>': 
                    try: match = float(val.replace(',','.')) > float(target.replace(',','.'))
                    except: match = False
                elif op == '<':
                    try: match = float(val.replace(',','.')) < float(target.replace(',','.'))
                    except: match = False
                elif op == '>=':
                    try: match = float(val.replace(',','.')) >= float(target.replace(',','.'))
                    except: match = False
                elif op == '<=':
                    try: match = float(val.replace(',','.')) <= float(target.replace(',','.'))
                    except: match = False
                elif op == 'contains': match = target.lower() in val.lower()
                elif op == 'is_empty': match = not val
                elif op == 'is_not_empty': match = bool(val)
                elif op == 'expired':
                    p = safe_parse_date(val)
                    match = p and p < today
                elif op == 'age_gt':
                    age = calculate_age(val)
                    try: match = age is not None and age > int(target)
                    except: match = False
                elif op == 'age_lt':
                    age = calculate_age(val)
                    try: match = age is not None and age < int(target)
                    except: match = False
                elif op == 'min_length':
                    try: match = len(val) < int(target)
                    except: match = False
                elif op == 'max_length':
                    try: match = len(val) > int(target)
                    except: match = False
                
                if not match:
                    all_match = False
                    break
            
            if all_match:
                line = [row['CLIENT'], row['EXPL'], row['FILIALE'], row['AGENCE']]
                for f in unique_cond_fields:
                    line.append(str(row.get(f) or ''))
                ws.append(line)

    from django.utils.text import slugify
    safe_name = slugify(rule.name).replace('-', '_') or f"anomalies_{rule.id}"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="{safe_name}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_rules_pdf(request):
    # Récupérer les rÃ¨gles avec la mÃªme logique que la vue principale
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    user_filiale = getattr(request.user, 'filiale', '')
    group_organs = ['PASS', 'Conformité Groupe']
    
    if user_organe == 'PASS':
        rules_qs = DataQualityRule.objects.all().order_by('-created_at')
    elif user_organe in ['Conformité Groupe']:
        rules_qs = DataQualityRule.objects.all().order_by('-created_at')
    else:
        # Filtrer par filiale
        rules_qs = DataQualityRule.objects.filter(
            Q(created_by__filiale=user_filiale) | Q(created_by__isnull=True)
        ).order_by('-created_at')

    # Ã‰valuation avec CACHE pour la rapidité
    import hashlib
    from django.core.cache import cache
    
    rules_with_stats = []
    eval_filiale = None if user_organe in group_organs else user_filiale
    rules_version = cache.get('quality_control_rules_version', 1)
    cache_ttl = 86400
    data_refresh_bucket = timezone.localdate().isoformat()

    for rule in rules_qs:
        # Signature identique Ã  la vue principale pour réutiliser le cache
        rule_signature = f"{rule.id}|{rule.name}|{rule.applicability}|{rule.field_name}|{rule.control_type}|{rule.parameter}|{rule.active}|{eval_filiale}"
        rule_hash = hashlib.md5(rule_signature.encode('utf-8')).hexdigest()
        rule_cache_key = f"quality_control:stat:v{rules_version}:d{data_refresh_bucket}:{rule_hash}"
        
        stat = cache.get(rule_cache_key)
        if stat is None:
            stat = evaluate_data_quality_rule(rule, filiale=eval_filiale)
            cache.set(rule_cache_key, stat, timeout=cache_ttl)
            
        # Calcul du taux
        total = stat.get('total', 0)
        stat['compliance_rate'] = compliance_rate_floor(stat['ok_count'], total, stat.get('fail_count', 0)) if total else 0
        
        from kyc.forms import DataQualityRuleForm
        parsed_filiales = DataQualityRuleForm._parse_filiales(rule.filiale)
        filiales_display = ", ".join(parsed_filiales) if parsed_filiales else "Toutes les filiales"

        rules_with_stats.append({
            'rule': rule,
            'stat': stat,
            'filiales_display': filiales_display,
        })

    template_path = 'quality_rules_pdf.html'
    logo_path = os.path.join(settings.MEDIA_ROOT, "images", "boa.png")
    context = {
        'rules': rules_with_stats,
        'user': request.user,
        'date': timezone.now(),
        'filiale': user_filiale if user_organe not in group_organs else "GROUPE",
        'logo_path': logo_path if os.path.exists(logo_path) else None,
    }
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="regles_qualite_kyc.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    if not pisa:
        return HttpResponse("L'exportation PDF n'est pas disponible sur ce serveur (dépendances manquantes).")
        
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=_pdf_link_callback)
    if pisa_status.err:
        return HttpResponse('Erreur lors de la génération du PDF')
        
    return response


@login_required
def accueil(request):
    user = request.user

    if user.is_authenticated:
        if user.organe == "Chargé Client":
            return redirect('non_rens')
        else:
            return redirect('agent')
    if not user.is_authenticated:
        return redirect('login_kyc')

    return render(request, 'accueil.html')


@login_required
def import_page(request):
    if not request.user.is_superuser:
        return redirect('accueil')
    log_dir = os.path.join(settings.BASE_DIR, "logs")
    os.makedirs(log_dir, exist_ok=True)
    history_path = os.path.join(log_dir, "import_history.log")
    run_dir = os.path.join(log_dir, "import_runs")
    os.makedirs(run_dir, exist_ok=True)

    def read_history(limit=50):
        if not os.path.exists(history_path):
            return []
        try:
            with open(history_path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            return [ln.rstrip("\n") for ln in lines[-limit:]]
        except Exception:
            return []

    def list_run_logs(limit=20):
        try:
            files = [f for f in os.listdir(run_dir) if f.endswith(".log")]
            files.sort(reverse=True)
            return files[:limit]
        except Exception:
            return []

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        data_dir = (request.POST.get("data_dir") or "").strip()
        filiales = (request.POST.get("filiales") or "").strip()
        only = (request.POST.get("only") or "").strip()
        bulk_size = (request.POST.get("bulk_size") or "").strip()
        taux_clear = request.POST.get("taux_clear") == "on"

        script = None
        if action == "run_kyc":
            script = "import_kyc.py"
        elif action == "run_premier":
            script = "import_premier.py"

        if not script:
            messages.error(request, "Action d'import inconnue.")
        else:
            env = os.environ.copy()
            if data_dir:
                env["KYC_DATA_DIR"] = data_dir
            if filiales:
                env["KYC_FILIALES"] = filiales
            if bulk_size:
                env["KYC_BULK_SIZE"] = bulk_size
            if only:
                env["KYC_ONLY"] = only
            if taux_clear:
                env["KYC_TAUX_CLEAR"] = "1"
            elif "KYC_TAUX_CLEAR" in env:
                env.pop("KYC_TAUX_CLEAR", None)

            cmd = [sys.executable, str(settings.BASE_DIR / script)]
            start_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
            detail_path = os.path.join(run_dir, f"{action}_{run_id}.log")
            try:
                result = subprocess.run(
                    cmd,
                    cwd=str(settings.BASE_DIR),
                    env=env,
                    capture_output=True,
                    text=True,
                    check=False,
                )
                status = "SUCCESS" if result.returncode == 0 else "FAILED"
                with open(detail_path, "w", encoding="utf-8") as df:
                    df.write(f"CMD: {cmd}\n")
                    df.write(f"START: {start_ts}\n")
                    df.write(f"RETURN: {result.returncode}\n")
                    df.write(f"DATA_DIR: {data_dir}\n")
                    df.write(f"FILIALES: {filiales}\n")
                    df.write(f"ONLY: {only}\n")
                    df.write(f"BULK_SIZE: {bulk_size}\n")
                    df.write(f"TAUX_CLEAR: {taux_clear}\n")
                    df.write("\n--- STDOUT ---\n")
                    df.write(result.stdout or "")
                    df.write("\n--- STDERR ---\n")
                    df.write(result.stderr or "")

                with open(history_path, "a", encoding="utf-8") as hf:
                    hf.write(f"{start_ts} | {action} | {status} | log={detail_path}\n")

                if status == "SUCCESS":
                    # Invalider le cache des rÃ¨gles de qualité aprÃ¨s un import réussi
                    current_v = cache.get('quality_control_rules_version', 1)
                    cache.set('quality_control_rules_version', current_v + 1, timeout=None)
                    messages.success(request, "Import terminé avec succÃ¨s.")
                else:
                    messages.error(request, f"Import échoué (code {result.returncode}).")

            except Exception as e:
                messages.error(request, f"Erreur d'exécution: {e}")

    context = {
        "history": read_history(),
        "run_logs": list_run_logs(),
        "history_log_name": "import_history.log",
    }
    return render(request, 'import.html', context)


DOCUMENT_EXTRACTION_FIELD_LABELS = [
    ("prenom", "Prenom"),
    ("nom", "Nom"),
    ("date_naissance", "Date de naissance"),
    ("lieu_naissance", "Lieu de naissance"),
    ("sexe", "Sexe"),
    ("pays_naissance", "Pays de naissance"),
    ("pays_delivrance", "Pays de delivrance"),
    ("date_expiration", "Date d'expiration"),
    ("adresse", "Adresse"),
    ("origine_revenu", "Origine du revenu"),
    ("numero_identification_nationale", "Numero identification nationale"),
    ("numero_document", "Numero document"),
    ("nationalite", "Nationalite"),
]

DOCUMENT_EXTRACTION_SEARCH_FIELDS = [
    ("all", "Tous les champs"),
    ("import_batch", "Lot d'import"),
    ("original_filename", "Nom du fichier"),
    ("source_filename", "Fichier source"),
    *DOCUMENT_EXTRACTION_FIELD_LABELS,
    ("extracted_text", "Texte extrait"),
]

KYC_PP_DOCUMENT_FIELD_MAP = [
    ("NUMID", "numero_identification_nationale", "NUMID"),
    ("NUMID", "numero_document", "NUMID"),
    ("DATNAIS", "date_naissance", "DATNAIS"),
    ("PAYNAIS", "pays_naissance", "PAYNAIS"),
    ("DATVALID", "date_expiration", "DATVALID"),
    ("ADRESSE", "adresse", "ADRESSE"),
    ("ORIGINE_REV", "origine_revenu", "ORIGINE_REV"),
]

KYC_PM_DOCUMENT_FIELD_MAP = [
    ("RCSNO", "numero_document", "RCSNO"),
    ("NUMERO_FISCAL", "numero_identification_nationale", "NUMERO_FISCAL"),
    ("ADRESSE_SOCIALE", "adresse", "ADRESSE_SOCIALE"),
    ("INTITULE_COMPTE", "nom", "INTITULE_COMPTE"),
]


def _normalize_match_value(value):
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


_field_visibility_configs_cache = None
_field_visibility_configs_cache_time = 0

def _get_cached_field_visibility_configs():
    global _field_visibility_configs_cache, _field_visibility_configs_cache_time
    import time
    now = time.time()
    if _field_visibility_configs_cache is None or now - _field_visibility_configs_cache_time > 10:
        _field_visibility_configs_cache = list(KycFieldVisibilityConfig.objects.all())
        _field_visibility_configs_cache_time = now
    return _field_visibility_configs_cache


def _get_field_sources(filiale, client_type_val):
    configs = _get_cached_field_visibility_configs()
    spec_config = next((c for c in configs if c.client_type == client_type_val and filiale in (c.filiales or [])), None)
    if spec_config:
        return spec_config.field_sources or {}
    global_config = next((c for c in configs if c.client_type == client_type_val and not c.filiales), None)
    if global_config:
        return global_config.field_sources or {}
    return {}



COUNTRY_ALIASES = {
    "SENEGAL": {"SENEGAL", "SEN", "SN", "BOASN"},
    "BENIN": {"BENIN", "BEN", "BJ", "BOABJ"},
    "COTE D IVOIRE": {"COTEDIVOIRE", "CIV", "CI", "IVOIRE", "BOACI"},
    "BURKINA FASO": {"BURKINAFASO", "BFA", "BF", "BOABF"},
    "MALI": {"MALI", "MLI", "ML", "BOAML"},
    "TOGO": {"TOGO", "TGO", "TG", "BOATG"},
    "NIGER": {"NIGER", "NER", "NE", "BOANE"},
}


def _country_key(value):
    normalized = _normalize_match_value(value)
    if not normalized:
        return ""
    for country, aliases in COUNTRY_ALIASES.items():
        if normalized in aliases or any(alias and alias in normalized for alias in aliases):
            return country
    return normalized


def _countries_are_compatible(left, right):
    left_key = _country_key(left)
    right_key = _country_key(right)
    return not left_key or not right_key or left_key == right_key


def _document_country_guard_passes(document, client):
    if document.pays_naissance and getattr(client, "PAYNAIS", "") and not _countries_are_compatible(document.pays_naissance, client.PAYNAIS):
        return False
    return True


def _is_empty_kyc_value(value):
    normalized = str(value or "").strip().lower()
    return normalized in {"", "-", "na", "n/a", "none", "null", "nan"}


def _document_identity_keys(document):
    return {
        key for key in [
            _normalize_match_value(document.numero_document),
            _normalize_match_value(document.numero_identification_nationale),
        ] if key
    }


def _values_match(left, right):
    left_value = _normalize_match_value(left)
    right_value = _normalize_match_value(right)
    return bool(left_value and right_value and left_value == right_value)


def _date_values_match(left, right):
    left_text = str(left or "").strip()
    right_text = str(right or "").strip()
    if not left_text or not right_text:
        return False

    left_formatted = format_date_for_export(left_text, empty_value="")
    right_formatted = format_date_for_export(right_text, empty_value="")
    if left_formatted and right_formatted and _normalize_match_value(left_formatted) == _normalize_match_value(right_formatted):
        return True

    return _normalize_match_value(left_text) == _normalize_match_value(right_text)


def _date_match_key(value):
    formatted = format_date_for_export(value, empty_value="")
    return _normalize_match_value(formatted or value)


def _nationality_match_key(value):
    return _country_key(value) or _normalize_match_value(value)


def _nationality_values_match(document_value, client_value):
    if not document_value or not client_value:
        return False
    if _countries_are_compatible(document_value, client_value):
        return True
    return _normalize_match_value(document_value) == _normalize_match_value(client_value)


DEFAULT_KYC_DOCUMENT_MATCH_WEIGHTS = {
    "birth_date_weight": 35,
    "document_validity_weight": 35,
    "birth_place_weight": 10,
    "nationality_weight": 30,
    "combination_threshold": 65,
}


def _get_kyc_document_match_weights():
    try:
        settings_obj = KycDocumentMatchSettings.get_active()
        return {
            "birth_date_weight": settings_obj.birth_date_weight,
            "document_validity_weight": settings_obj.document_validity_weight,
            "birth_place_weight": settings_obj.birth_place_weight,
            "nationality_weight": settings_obj.nationality_weight,
            "combination_threshold": settings_obj.combination_threshold,
        }
    except Exception:
        return DEFAULT_KYC_DOCUMENT_MATCH_WEIGHTS.copy()


def _document_client_identity_score(document, client, weights=None):
    weights = weights or DEFAULT_KYC_DOCUMENT_MATCH_WEIGHTS
    client_numid = getattr(client, "NUMID", "")
    if _values_match(document.numero_identification_nationale, client_numid) or _values_match(document.numero_document, client_numid):
        return 100

    score = 0
    checks = [
        (_date_values_match(document.date_naissance, getattr(client, "DATNAIS", "")), weights["birth_date_weight"]),
        (_date_values_match(document.date_expiration, getattr(client, "DATVALID", "")), weights["document_validity_weight"]),
        (_nationality_values_match(document.lieu_naissance, getattr(client, "PAYNAIS", "")), weights["birth_place_weight"]),
        (
            _nationality_values_match(document.nationalite, getattr(client, "PAYNAIS", ""))
            or _nationality_values_match(document.pays_naissance, getattr(client, "PAYNAIS", "")),
            weights["nationality_weight"],
        ),
    ]
    for matched, weight in checks:
        if matched:
            score += weight
    return score


def _document_client_haystack(document):
    return _normalize_match_value(
        " ".join([
            document.original_filename or "",
            document.source_filename or "",
            document.import_batch or "",
        ])
    )


def _document_client_tokens(document):
    raw_value = " ".join([
        document.original_filename or "",
        document.source_filename or "",
        document.import_batch or "",
    ]).upper()
    return {_normalize_match_value(token) for token in re.split(r"[^A-Z0-9]+", raw_value) if len(token) >= 4}


def _document_unique_key(document):
    identity_keys = sorted(_document_identity_keys(document))
    if identity_keys:
        country_parts = [
            _country_key(document.pays_delivrance),
            _country_key(document.pays_naissance),
        ]
        country_scope = "|".join([part for part in country_parts if part])
        if country_scope:
            return "identity:" + country_scope + ":" + "|".join(identity_keys)
        return "identity:" + "|".join(identity_keys)
    return "file:" + _normalize_match_value(
        "|".join([
            document.original_filename or "",
            document.source_filename or "",
            document.import_batch or "",
            document.page_range or "",
        ])
    )


def _client_dedup_key(client):
    normalized_idp = _normalize_match_value(getattr(client, "IDP", ""))
    if normalized_idp:
        return f"idp:{normalized_idp}"
    normalized_client = _normalize_match_value(getattr(client, "CLIENT", ""))
    if normalized_client:
        return f"client:{normalized_client}"
    return f"pk:{client.pk}"


def _build_kyc_pp_document_matches(document_queryset, limit=3000, result_limit=200, progress_callback=None):
    documents_for_match = list(document_queryset.order_by("-created_at")[:limit])
    if not documents_for_match:
        return [], {"documents_checked": 0, "documents_matched": 0, "clients_matched": 0, "suggestions_count": 0, "match_rate": 0}
    if progress_callback:
        progress_callback(0, len(documents_for_match), "Preparation du rapprochement")
    match_weights = _get_kyc_document_match_weights()

    document_keys = set()
    for document in documents_for_match:
        document_keys.update(_document_identity_keys(document))

    kyc_candidates = {}
    if document_keys:
        for client in Kyc_pp.objects.exclude(NUMID="").only(
            "id", "FILIALE", "AGENCE", "CLIENT", "IDP", "NUMID", "DATNAIS", "PAYNAIS", "DATVALID", "ADRESSE", "ORIGINE_REV"
        ):
            normalized_numid = _normalize_match_value(client.NUMID)
            if normalized_numid in document_keys:
                kyc_candidates.setdefault(normalized_numid, []).append(client)

    client_by_code = {}
    clients_by_birth_date = {}
    clients_by_validity_date = {}
    clients_by_nationality = {}
    clients_by_birth_place = {}
    for client in Kyc_pp.objects.only(
        "id", "FILIALE", "AGENCE", "CLIENT", "IDP", "NUMID", "DATNAIS", "PAYNAIS", "DATVALID", "ADRESSE", "ORIGINE_REV"
    )[:50000]:
        normalized_client = _normalize_match_value(client.CLIENT)
        if normalized_client:
            client_by_code.setdefault(normalized_client, []).append(client)
        birth_key = _date_match_key(client.DATNAIS)
        if birth_key:
            clients_by_birth_date.setdefault(birth_key, []).append(client)
        validity_key = _date_match_key(client.DATVALID)
        if validity_key:
            clients_by_validity_date.setdefault(validity_key, []).append(client)
        nationality_key = _nationality_match_key(client.PAYNAIS)
        if nationality_key:
            clients_by_nationality.setdefault(nationality_key, []).append(client)
            clients_by_birth_place.setdefault(nationality_key, []).append(client)

    matches = []
    client_match_index = {}
    matched_client_ids = set()
    matched_document_ids = set()
    for index, document in enumerate(documents_for_match, start=1):
        if progress_callback:
            progress_callback(index, len(documents_for_match), f"Analyse document {index}/{len(documents_for_match)}")
        candidate_clients = []
        for identity_key in _document_identity_keys(document):
            candidate_clients.extend(kyc_candidates.get(identity_key, []))

        for client_token in _document_client_tokens(document):
            candidate_clients.extend(client_by_code.get(client_token, []))

        birth_key = _date_match_key(document.date_naissance)
        validity_key = _date_match_key(document.date_expiration)
        nationality_key = _nationality_match_key(document.nationalite or document.pays_naissance)
        birth_place_key = _nationality_match_key(document.lieu_naissance)
        combination_pool = {}
        for client in clients_by_birth_date.get(birth_key, []):
            combination_pool[client.pk] = client
        for client in clients_by_validity_date.get(validity_key, []):
            combination_pool[client.pk] = client
        if birth_key or validity_key:
            for client in clients_by_birth_place.get(birth_place_key, []):
                combination_pool[client.pk] = client
        if birth_key or validity_key:
            for client in clients_by_nationality.get(nationality_key, []):
                combination_pool[client.pk] = client
        for client in combination_pool.values():
            if _document_client_identity_score(document, client, match_weights) >= match_weights["combination_threshold"]:
                candidate_clients.append(client)

        unique_clients = {}
        for client in candidate_clients:
            if not _document_country_guard_passes(document, client):
                continue
            unique_clients[client.pk] = client

        for client in unique_clients.values():
            suggestions = []
            used_kyc_fields = set()
            client_filiale = getattr(client, "FILIALE", "").strip()
            field_sources = _get_field_sources(client_filiale, "pp")
            for kyc_field, document_field, label in KYC_PP_DOCUMENT_FIELD_MAP:
                allowed_source = field_sources.get(kyc_field)
                if allowed_source and document.document_type != allowed_source:
                    continue
                document_value = getattr(document, document_field, "")
                if not document_value or not _is_empty_kyc_value(getattr(client, kyc_field, "")):
                    continue
                if kyc_field in used_kyc_fields:
                    continue
                used_kyc_fields.add(kyc_field)
                suggestions.append({
                    "field": kyc_field,
                    "label": label,
                    "document_value": document_value,
                })

            match_rate = _document_client_identity_score(document, client, match_weights)
            if match_rate < 30:
                continue

            candidate_match = {
                "client": client,
                "document": document,
                "suggestions": suggestions,
                "match_rate": match_rate,
            }
            if not _build_kyc_pp_match_action_items(candidate_match):
                continue

            client_dedup_key = _client_dedup_key(client)
            if client_dedup_key in client_match_index:
                existing_match = matches[client_match_index[client_dedup_key]]
                if _document_unique_key(existing_match["document"]) != _document_unique_key(document):
                    continue
                existing_fields = {suggestion["field"] for suggestion in existing_match["suggestions"]}
                for suggestion in suggestions:
                    if suggestion["field"] not in existing_fields:
                        existing_match["suggestions"].append(suggestion)
                        existing_fields.add(suggestion["field"])
                existing_extra_actions = existing_match.setdefault("extra_action_items", [])
                existing_action_keys = {
                    (action.get("kind"), (action.get("field") or "").strip().upper())
                    for action in _build_kyc_pp_match_action_items(existing_match)
                }
                for action in _build_kyc_pp_match_action_items(candidate_match):
                    action_key = (action.get("kind"), (action.get("field") or "").strip().upper())
                    if action_key not in existing_action_keys:
                        existing_extra_actions.append(action)
                        existing_action_keys.add(action_key)
                existing_match["match_rate"] = max(existing_match["match_rate"], match_rate)
                continue

            matched_client_ids.add(client.pk)
            matched_document_ids.add(document.pk)
            client_match_index[client_dedup_key] = len(matches)
            matches.append(candidate_match)

    suggestions_count = sum(len(match["suggestions"]) for match in matches)
    match_rate = round((len(matched_document_ids) / len(documents_for_match)) * 100, 1)

    summary = {
        "documents_checked": len(documents_for_match),
        "documents_matched": len(matched_document_ids),
        "clients_matched": len(client_match_index),
        "suggestions_count": suggestions_count,
        "match_rate": match_rate,
    }
    if result_limit:
        return matches[:result_limit], summary
    return matches, summary


def _build_kyc_pm_document_matches(document_queryset, limit=3000, result_limit=200, progress_callback=None):
    documents_for_match = list(document_queryset.order_by("-created_at")[:limit])
    if not documents_for_match:
        return [], {"documents_checked": 0, "documents_matched": 0, "clients_matched": 0, "suggestions_count": 0, "match_rate": 0}
    if progress_callback:
        progress_callback(0, len(documents_for_match), "Preparation du rapprochement")

    document_keys = set()
    for document in documents_for_match:
        if document.numero_document:
            document_keys.add(_normalize_match_value(document.numero_document))
        if document.numero_identification_nationale:
            document_keys.add(_normalize_match_value(document.numero_identification_nationale))

    kyc_candidates = {}
    if document_keys:
        for client in Kyc_pm.objects.exclude(RCSNO="").only(
            "id", "FILIALE", "AGENCE", "CLIENT", "IDM", "RCSNO", "NUMERO_FISCAL", "ADRESSE_SOCIALE", "INTITULE_COMPTE"
        ):
            norm_rcs = _normalize_match_value(client.RCSNO)
            if norm_rcs in document_keys:
                kyc_candidates.setdefault(norm_rcs, []).append(client)
        for client in Kyc_pm.objects.exclude(NUMERO_FISCAL="").only(
            "id", "FILIALE", "AGENCE", "CLIENT", "IDM", "RCSNO", "NUMERO_FISCAL", "ADRESSE_SOCIALE", "INTITULE_COMPTE"
        ):
            norm_nif = _normalize_match_value(client.NUMERO_FISCAL)
            if norm_nif in document_keys:
                kyc_candidates.setdefault(norm_nif, []).append(client)

    client_by_name = {}
    for client in Kyc_pm.objects.only(
        "id", "FILIALE", "AGENCE", "CLIENT", "IDM", "RCSNO", "NUMERO_FISCAL", "ADRESSE_SOCIALE", "INTITULE_COMPTE"
    )[:50000]:
        norm_name = _normalize_match_value(client.CLIENT)
        if norm_name:
            client_by_name.setdefault(norm_name, []).append(client)
        norm_intitule = _normalize_match_value(client.INTITULE_COMPTE)
        if norm_intitule and norm_intitule != norm_name:
            client_by_name.setdefault(norm_intitule, []).append(client)

    matches = []
    matched_client_ids = set()
    matched_document_ids = set()
    
    for index, document in enumerate(documents_for_match, start=1):
        if progress_callback:
            progress_callback(index, len(documents_for_match), f"Analyse document {index}/{len(documents_for_match)}")
        
        candidate_clients = []
        if document.numero_document:
            candidate_clients.extend(kyc_candidates.get(_normalize_match_value(document.numero_document), []))
        if document.numero_identification_nationale:
            candidate_clients.extend(kyc_candidates.get(_normalize_match_value(document.numero_identification_nationale), []))
            
        for client_token in _document_client_tokens(document):
            candidate_clients.extend(client_by_name.get(client_token, []))
            
        unique_clients = {}
        for client in candidate_clients:
            unique_clients[client.pk] = client
            
        for client in unique_clients.values():
            suggestions = []
            used_kyc_fields = set()
            client_filiale = getattr(client, "FILIALE", "").strip()
            field_sources = _get_field_sources(client_filiale, "pm")
            for kyc_field, document_field, label in KYC_PM_DOCUMENT_FIELD_MAP:
                allowed_source = field_sources.get(kyc_field)
                if allowed_source and document.document_type != allowed_source:
                    continue
                document_value = getattr(document, document_field, "")
                if not document_value or not _is_empty_kyc_value(getattr(client, kyc_field, "")):
                    continue
                if kyc_field in used_kyc_fields:
                    continue
                used_kyc_fields.add(kyc_field)
                suggestions.append({
                    "field": kyc_field,
                    "label": label,
                    "value": str(document_value),
                })
                
            match_rate = 0
            if (document.numero_document and _normalize_match_value(document.numero_document) == _normalize_match_value(client.RCSNO)) or \
               (document.numero_identification_nationale and _normalize_match_value(document.numero_identification_nationale) == _normalize_match_value(client.NUMERO_FISCAL)):
                match_rate = 100
            else:
                doc_name_norm = _normalize_match_value(document.nom or "")
                cli_name_norm = _normalize_match_value(client.CLIENT or "")
                if doc_name_norm and cli_name_norm:
                    if doc_name_norm == cli_name_norm:
                        match_rate = 90
                    elif doc_name_norm in cli_name_norm or cli_name_norm in doc_name_norm:
                        match_rate = 75
                    else:
                        match_rate = 50
                else:
                    match_rate = 40
            
            extra_action_items = []
            for kyc_field, document_field, label in KYC_PM_DOCUMENT_FIELD_MAP:
                allowed_source = field_sources.get(kyc_field)
                if allowed_source and document.document_type != allowed_source:
                    continue
                document_value = getattr(document, document_field, "")
                if document_value and _is_empty_kyc_value(getattr(client, kyc_field, "")):
                    extra_action_items.append({
                        "kind": "suggest",
                        "field": kyc_field,
                        "label": label,
                        "value": str(document_value),
                    })
                    
            matches.append({
                "client": client,
                "document": document,
                "suggestions": suggestions,
                "extra_action_items": extra_action_items,
                "match_rate": match_rate,
            })
            matched_client_ids.add(client.pk)
            matched_document_ids.add(document.pk)

    matches.sort(key=lambda m: m["match_rate"], reverse=True)
    summary = {
        "documents_checked": len(documents_for_match),
        "documents_matched": len(matched_document_ids),
        "clients_matched": len(matched_client_ids),
        "suggestions_count": sum(len(m["suggestions"]) for m in matches),
        "match_rate": int(len(matched_document_ids) / len(documents_for_match) * 100) if documents_for_match else 0
    }
    if result_limit:
        return matches[:result_limit], summary
    return matches, summary


LAST_KYC_PP_MATCH_SESSION_KEY = "document_extraction_last_kyc_pp_match_params"
LAST_KYC_PP_MATCH_RESULT_SESSION_KEY = "document_extraction_last_kyc_pp_match_result"
KYC_PP_MATCHED_BATCHES_SESSION_KEY = "document_extraction_kyc_pp_matched_batches"
LAST_UPLOADED_DOCUMENT_BATCH_SESSION_KEY = "document_extraction_last_uploaded_batch"
KYC_PP_MATCH_RESULT_VERSION = 3


def _filtered_document_extractions_from_params(params, user=None):
    documents = KycDocumentExtraction.objects.select_related("uploaded_by").all().order_by("-created_at", "-id")
    
    if user:
        users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
        is_group_user = (user.filiale in ["BOA Group", "BOA GROUP"]) or (user.organe in users_groupe) or (not user.filiale)
        if not is_group_user:
            documents = documents.filter(uploaded_by__filiale=user.filiale)

    client_type = params.get("client_type", "pp")
    documents = documents.filter(client_type=client_type)
    valid_document_types = dict(DOCUMENT_EXTRACTION_TYPE_CHOICES)
    selected_document_type = params.get("document_type", "")
    selected_import_batch = (params.get("import_batch") or "").strip()
    search_query = (params.get("q") or "").strip()
    search_field = params.get("field") or "all"
    allowed_search_fields = {field for field, _ in DOCUMENT_EXTRACTION_SEARCH_FIELDS}

    if selected_document_type in valid_document_types:
        documents = documents.filter(document_type=selected_document_type)

    if selected_import_batch:
        documents = documents.filter(import_batch=selected_import_batch)

    if search_field not in allowed_search_fields:
        search_field = "all"

    if search_query:
        if search_field == "all":
            search_filter = Q()
            for field_name, _ in DOCUMENT_EXTRACTION_SEARCH_FIELDS:
                if field_name == "all":
                    continue
                search_filter |= Q(**{f"{field_name}__icontains": search_query})
            documents = documents.filter(search_filter)
        else:
            documents = documents.filter(**{f"{search_field}__icontains": search_query})

    return documents


def _filtered_document_extractions_from_request(request):
    return _filtered_document_extractions_from_params(request.GET, user=request.user)


def _clean_document_match_params(params):
    return {
        key: value
        for key, value in params.items()
        if key not in {"page", "extraction_id", "match_kyc", "match_job", "show_match_results", "result_modal", "child_modal", "tab"} and value not in (None, "")
    }


def _document_match_scope_params(params):
    return {
        key: value
        for key, value in _clean_document_match_params(params).items()
        if key not in KYC_PP_MATCH_FILTER_FIELDS
    }


def _serialize_kyc_pp_matches(matches, summary, params):
    return {
        "version": KYC_PP_MATCH_RESULT_VERSION,
        "params": params,
        "summary": summary,
        "matches": [
            {
                "client_id": match["client"].pk,
                "document_id": match["document"].pk,
                "suggestions": match["suggestions"],
                "extra_action_items": match.get("extra_action_items") or [],
                "match_rate": match["match_rate"],
            }
            for match in matches
        ],
    }


def _hydrate_kyc_pp_match_result(result):
    if not result:
        return [], None, {}
    if result.get("version") != KYC_PP_MATCH_RESULT_VERSION:
        return [], None, result.get("params") or {}

    serialized_matches = result.get("matches") or []
    client_ids = [match.get("client_id") for match in serialized_matches if match.get("client_id")]
    document_ids = [match.get("document_id") for match in serialized_matches if match.get("document_id")]
    
    params = result.get("params") or {}
    client_type = params.get("client_type", "pp")
    
    if client_type == "pm":
        clients = Kyc_pm.objects.in_bulk(client_ids)
    else:
        clients = Kyc_pp.objects.in_bulk(client_ids)
        
    documents = KycDocumentExtraction.objects.in_bulk(document_ids)

    matches = []
    for serialized_match in serialized_matches:
        client = clients.get(serialized_match.get("client_id"))
        document = documents.get(serialized_match.get("document_id"))
        if not client or not document:
            continue
        matches.append({
            "client": client,
            "document": document,
            "suggestions": serialized_match.get("suggestions") or [],
            "extra_action_items": serialized_match.get("extra_action_items") or [],
            "match_rate": serialized_match.get("match_rate") or 0,
        })

    return matches, result.get("summary"), params


def _merge_kyc_pp_match_lists(match_lists):
    merged = []
    index_by_key = {}

    for matches in match_lists:
        for match in matches:
            client = match.get("client")
            client_id = getattr(client, "IDP", None) or getattr(client, "IDM", None)
            normalized_id = _normalize_match_value(client_id or "")
            client_pk = getattr(client, "pk", None)
            key = (
                ("idp", normalized_id)
                if normalized_id
                else ("client", client_pk, _document_unique_key(match.get("document")))
            )
            if key not in index_by_key:
                index_by_key[key] = len(merged)
                merged.append(match)
                continue

            existing_match = merged[index_by_key[key]]
            existing_fields = {
                (suggestion.get("field") or "").strip().upper()
                for suggestion in existing_match.get("suggestions", [])
            }
            for suggestion in match.get("suggestions", []):
                suggestion_field = (suggestion.get("field") or "").strip().upper()
                if suggestion_field and suggestion_field not in existing_fields:
                    existing_match.setdefault("suggestions", []).append(suggestion)
                    existing_fields.add(suggestion_field)

            existing_actions = existing_match.setdefault("extra_action_items", [])
            action_keys = {
                (action.get("kind"), (action.get("field") or "").strip().upper())
                for action in existing_actions
            }
            for action in match.get("extra_action_items", []):
                action_key = (action.get("kind"), (action.get("field") or "").strip().upper())
                if action_key not in action_keys:
                    existing_actions.append(action)
                    action_keys.add(action_key)
            existing_match["match_rate"] = max(existing_match.get("match_rate", 0), match.get("match_rate", 0))

    return merged


def _user_can_access_document_match_job(user, job):
    return bool(user.is_superuser or job.created_by_id == user.pk)


def _run_document_match_job(job_id):
    close_old_connections()
    try:
        job = KycDocumentMatchJob.objects.get(pk=job_id)
        scope_params = job.scope_params or {}
        job.status = "running"
        job.started_at = timezone.now()
        job.message = "Preparation du rapprochement"
        job.save(update_fields=["status", "started_at", "message", "updated_at"])

        last_saved_step = {"value": -1}

        def progress_callback(current, total, message):
            if current != total and current - last_saved_step["value"] < 5:
                return
            last_saved_step["value"] = current
            KycDocumentMatchJob.objects.filter(pk=job_id).update(
                progress_current=current,
                progress_total=total,
                message=message,
                updated_at=timezone.now(),
            )

        documents = _filtered_document_extractions_from_params(scope_params)
        client_type = scope_params.get("client_type", "pp")
        if client_type == "pm":
            matches, summary = _build_kyc_pm_document_matches(
                documents,
                progress_callback=progress_callback,
            )
        else:
            matches, summary = _build_kyc_pp_document_matches(
                documents,
                progress_callback=progress_callback,
            )
        result = _serialize_kyc_pp_matches(matches, summary, scope_params)
        KycDocumentMatchJob.objects.filter(pk=job_id).update(
            status="completed",
            progress_current=summary.get("documents_checked", 0),
            progress_total=summary.get("documents_checked", 0),
            message="Rapprochement termine",
            result=result,
            completed_at=timezone.now(),
            updated_at=timezone.now(),
        )
    except Exception as exc:
        KycDocumentMatchJob.objects.filter(pk=job_id).update(
            status="failed",
            message="Echec du rapprochement",
            error=str(exc),
            completed_at=timezone.now(),
            updated_at=timezone.now(),
        )
    finally:
        close_old_connections()


@login_required
def start_document_extraction_match_job(request):
    scope_params = _document_match_scope_params(request.GET)
    existing_job = (
        KycDocumentMatchJob.objects
        .filter(created_by=request.user, scope_params=scope_params, status="running")
        .order_by("-created_at")
        .first()
    )
    job = existing_job or KycDocumentMatchJob.objects.create(
        created_by=request.user,
        scope_params=scope_params,
        message="Rapprochement en attente",
    )
    if not existing_job:
        threading.Thread(target=_run_document_match_job, args=(job.pk,), daemon=True).start()

    redirect_params = dict(scope_params)
    redirect_params["match_job"] = job.pk
    return redirect(f"{reverse('document_extraction')}?{urlencode(redirect_params)}#suivi")


@login_required
def document_extraction_match_job_status(request, job_id):
    job = get_object_or_404(KycDocumentMatchJob, pk=job_id)
    if not _user_can_access_document_match_job(request.user, job):
        return JsonResponse({"error": "Acces non autorise"}, status=403)

    total = job.progress_total or 0
    current = job.progress_current or 0
    percent = min(100, int(current / total * 100)) if total else (100 if job.status == "completed" else 0)
    redirect_params = dict(job.scope_params or {})
    redirect_params["match_job"] = job.pk
    result_params = dict(redirect_params)
    result_params["show_match_results"] = "1"
    result_params["result_modal"] = "1"

    return JsonResponse({
        "id": job.pk,
        "status": job.status,
        "message": job.message,
        "error": job.error,
        "current": current,
        "total": total,
        "percent": percent,
        "redirect_url": f"{reverse('document_extraction')}?{urlencode(redirect_params)}#suivi",
        "result_url": f"{reverse('document_extraction')}?{urlencode(result_params)}#suivi",
    })


KYC_PP_MATCH_FILTER_FIELDS = {
    "match_client": "CLIENT",
    "match_idp": "IDP",
    "match_idm": "IDM",
    "match_filiale": "FILIALE",
    "match_agence": "AGENCE",
}


def _get_kyc_pp_match_filters(params):
    return {
        key: (params.get(key) or "").strip()
        for key in KYC_PP_MATCH_FILTER_FIELDS
    }


def _filter_kyc_pp_matches(matches, params):
    filters = _get_kyc_pp_match_filters(params)
    active_filters = {
        key: value.lower()
        for key, value in filters.items()
        if value
    }
    if not active_filters:
        return matches

    filtered_matches = []
    for match in matches:
        client = match["client"]
        keep_match = True
        for param_name, filter_value in active_filters.items():
            client_field = KYC_PP_MATCH_FILTER_FIELDS[param_name]
            client_value = str(getattr(client, client_field, "") or "").lower()
            if filter_value not in client_value:
                keep_match = False
                break
        if keep_match:
            filtered_matches.append(match)
    return filtered_matches


def _build_kyc_pp_match_action_items(match):
    document = match["document"]
    client = match["client"]
    actions = []
    action_keys = set()
    fields_with_actions = set()

    is_pm = isinstance(client, Kyc_pm)
    field_map = KYC_PM_DOCUMENT_FIELD_MAP if is_pm else KYC_PP_DOCUMENT_FIELD_MAP

    client_filiale = getattr(client, "FILIALE", "").strip()
    client_type_val = "pm" if is_pm else "pp"
    field_sources = _get_field_sources(client_filiale, client_type_val)

    def add_action(kind, field, text):
        normalized_field = (field or "").strip().upper()
        key = (kind, normalized_field) if normalized_field else (kind, "", _normalize_match_value(text))
        if not text or key in action_keys:
            return
        action_keys.add(key)
        fields_with_actions.add(normalized_field)
        actions.append({"kind": kind, "field": normalized_field, "text": text})

    for suggestion in match.get("suggestions", []):
        field = (suggestion.get("field") or "").strip().upper()
        document_value = suggestion.get("document_value", "") or suggestion.get("value", "")
        if not field or not document_value:
            continue
        add_action("complete", field, f"{field}: {document_value}")

    fields_in_order = []
    fields_seen = set()
    for kyc_field, _, _ in field_map:
        if kyc_field not in fields_seen:
            fields_seen.add(kyc_field)
            fields_in_order.append(kyc_field)

    for kyc_field in fields_in_order:
        mapped_fields = [
            (document_field, label)
            for field_name, document_field, label in field_map
            if field_name == kyc_field
        ]
        document_values = []
        for document_field, label in mapped_fields:
            allowed_source = field_sources.get(kyc_field)
            if allowed_source and document.document_type != allowed_source:
                continue
            val = getattr(document, document_field, "")
            if val:
                document_values.append((val, label))
                
        kyc_value = getattr(client, kyc_field, "")
        if not document_values or _is_empty_kyc_value(kyc_value):
            continue

        if not is_pm and kyc_field in {"DATNAIS", "DATVALID"}:
            values_match = any(_date_values_match(document_value, kyc_value) for document_value, _ in document_values)
        elif not is_pm and kyc_field == "PAYNAIS":
            values_match = any(_nationality_values_match(document_value, kyc_value) for document_value, _ in document_values)
        else:
            values_match = any(_values_match(document_value, kyc_value) for document_value, _ in document_values)

        if values_match:
            continue

        document_value, label = document_values[0]
        add_action("modify", kyc_field, f"{label} ({kyc_field}): {kyc_value or '-'} -> {document_value}")

    expired_match = match.get("expired_document_match")
    if expired_match and "DATVALID" not in fields_with_actions:
        allowed_source = field_sources.get("DATVALID")
        if not allowed_source or (expired_match.document and expired_match.document.document_type == allowed_source):
            add_action(
                "modify",
                "DATVALID",
                f"DATVALID: {expired_match.old_validity_date or '-'} -> {expired_match.document_validity_date or '-'}",
            )

    for action in match.get("extra_action_items") or []:
        add_action(action.get("kind") or "modify", action.get("field") or "", action.get("text") or "")

    return actions


@login_required
def export_document_extraction_matches(request):
    requested_scope_params = _document_match_scope_params(request.GET)
    matches = []

    last_match_result = request.session.get(LAST_KYC_PP_MATCH_RESULT_SESSION_KEY)
    stored_matches, _, stored_params = _hydrate_kyc_pp_match_result(last_match_result)
    client_type = requested_scope_params.get("client_type", "pp")

    if stored_matches and stored_params == requested_scope_params:
        matches = stored_matches
    else:
        documents = _filtered_document_extractions_from_params(requested_scope_params, user=request.user)
        if client_type == "pm":
            matches, _ = _build_kyc_pm_document_matches(documents, result_limit=None)
        else:
            matches, _ = _build_kyc_pp_document_matches(documents, result_limit=None)

    matches = _merge_kyc_pp_match_lists([matches])
    matches = _filter_kyc_pp_matches(matches, request.GET)

    selected_import_batch = (requested_scope_params.get("import_batch") or "").strip()
    expired_document_matches_qs = KycExpiredDocumentScanMatch.objects.select_related("client", "document").filter(
        status="a_valider"
    )
    if selected_import_batch:
        expired_document_matches_qs = expired_document_matches_qs.filter(document__import_batch=selected_import_batch)
    expired_match_by_client_id = {}
    for expired_match in expired_document_matches_qs.order_by("-match_rate", "-scan_date")[:500]:
        if expired_match.client_id in expired_match_by_client_id:
            continue
        expired_match_by_client_id[expired_match.client_id] = expired_match

    for match in matches:
        expired_match = expired_match_by_client_id.pop(match["client"].pk, None)
        if expired_match:
            match["expired_document_match"] = expired_match

    matched_idp_keys = {
        _normalize_match_value(getattr(match["client"], "IDM" if client_type == "pm" else "IDP", "")) or str(match["client"].pk)
        for match in matches
    }
    standalone_expired_matches = []
    if client_type == "pp":
        for expired_match in expired_match_by_client_id.values():
            expired_idp_key = _normalize_match_value(getattr(expired_match.client, "IDP", "") or expired_match.idp)
            expired_key = expired_idp_key or str(expired_match.client_id)
            if expired_key in matched_idp_keys:
                continue
            matched_idp_keys.add(expired_key)
            standalone_expired_matches.append(expired_match)

    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    filename = timezone.localtime(timezone.now()).strftime(f"correspondances_kyc_{client_type}_%Y%m%d_%H%M.csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")
    writer.writerow([
        "CLIENT",
        "IDM" if client_type == "pm" else "IDP",
        "FILIALE",
        "AGENCE",
        "TYPE DE DOCUMENT",
        "TAUX CORRESPONDANCE",
        "A completer / A modifier",
        "Numero document",
        "Numero d'identification nationale",
    ])

    for match in matches:
        document = match["document"]
        client = match["client"]
        action_items = _build_kyc_pp_match_action_items(match)
        client_id_val = getattr(client, "IDM", "") if client_type == "pm" else getattr(client, "IDP", "")

        writer.writerow([
            client.CLIENT,
            client_id_val,
            client.FILIALE,
            client.AGENCE,
            document.get_document_type_display(),
            match.get("match_rate", 0),
            " | ".join(action["text"] for action in action_items),
            document.numero_document,
            document.numero_identification_nationale,
        ])

    for expired_match in standalone_expired_matches:
        client = expired_match.client
        document = expired_match.document
        writer.writerow([
            getattr(client, "CLIENT", "") or expired_match.client_code,
            getattr(client, "IDP", "") or expired_match.idp,
            getattr(client, "FILIALE", "") or expired_match.filiale,
            getattr(client, "AGENCE", "") or expired_match.agence,
            document.get_document_type_display() if document else "",
            expired_match.match_rate or 0,
            f"DATVALID: {expired_match.old_validity_date or '-'} -> {expired_match.document_validity_date or '-'}",
            getattr(document, "numero_document", "") if document else "",
            getattr(document, "numero_identification_nationale", "") if document else "",
        ])
    return response


def _build_import_batch_name(request, uploaded_files):
    requested_name = (request.POST.get("batch_name") or "").strip()
    if requested_name:
        return requested_name[:120]

    first_file = uploaded_files[0].name if uploaded_files else "documents"
    timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d-%H%M%S")
    return f"LOT-{timestamp}-{os.path.splitext(os.path.basename(first_file))[0]}"[:120]


def _format_document_extraction_record(record):
    fields = {
        field_name: getattr(record, field_name, "")
        for field_name, _ in DOCUMENT_EXTRACTION_FIELD_LABELS
    }
    return {
        "id": record.pk,
        "filename": record.original_filename or os.path.basename(record.uploaded_file.name),
        "source_filename": record.source_filename,
        "file_url": record.uploaded_file.url if record.uploaded_file else "",
        "document_type": record.get_document_type_display(),
        "import_batch": record.import_batch,
        "page_number": record.page_number,
        "page_range": record.page_range,
        "text": record.extracted_text,
        "fields": fields,
        "warnings": [warning for warning in record.extraction_warnings.splitlines() if warning],
        "field_rows": [
            {"label": label, "value": fields.get(field_name)}
            for field_name, label in DOCUMENT_EXTRACTION_FIELD_LABELS
            if fields.get(field_name)
        ],
    }


def _fill_document_extraction_fields(record, extraction):
    extracted_fields = extraction.get("fields") or {}
    for field_name, _ in DOCUMENT_EXTRACTION_FIELD_LABELS:
        setattr(record, field_name, extracted_fields.get(field_name, ""))
    record.extracted_text = extraction.get("text") or ""
    record.extraction_warnings = "\n".join(extraction.get("warnings") or [])
    record.page_number = extraction.get("page_number") or record.page_number
    record.page_range = extraction.get("page_range") or record.page_range


DOCUMENT_EXTRACTION_TYPE_LABELS = dict(DOCUMENT_EXTRACTION_TYPE_CHOICES)


def _apply_detected_document_type(record, extraction, requested_type):
    detected_type = extraction.get("detected_document_type") or ""
    if detected_type not in DOCUMENT_EXTRACTION_TYPE_LABELS:
        return

    if detected_type != requested_type:
        warning = (
            "Type ajuste automatiquement: le document semble etre "
            f"{DOCUMENT_EXTRACTION_TYPE_LABELS[detected_type]} alors que "
            f"{DOCUMENT_EXTRACTION_TYPE_LABELS.get(requested_type, requested_type)} avait ete selectionne."
        )
        warnings = extraction.setdefault("warnings", [])
        if warning not in warnings:
            warnings.append(warning)
        record.extraction_warnings = "\n".join(warnings)

    record.document_type = detected_type


def _save_uploaded_document_record(uploaded_file, document_type, user, import_batch="", source_filename="", client_type="pp"):
    record = KycDocumentExtraction(
        document_type=document_type,
        original_filename=os.path.basename(uploaded_file.name),
        source_filename=source_filename or os.path.basename(uploaded_file.name),
        import_batch=import_batch,
        uploaded_by=user,
        client_type=client_type,
    )
    record.uploaded_file.save(uploaded_file.name, uploaded_file, save=False)
    extraction = extract_document_data(record.uploaded_file.path, uploaded_file.name)
    _fill_document_extraction_fields(record, extraction)
    _apply_detected_document_type(record, extraction, document_type)
    record.save()
    return record, extraction


def _save_zip_document_record(zip_file, member_name, document_type, user, import_batch, archive_name, client_type="pp"):
    safe_name = os.path.basename(member_name)
    _, extension = os.path.splitext(safe_name)
    if extension.lower() not in SUPPORTED_EXTENSIONS:
        return None, f"Format ignore dans le ZIP: {member_name}"

    with zip_file.open(member_name) as member:
        content = ContentFile(member.read(), name=safe_name)

    record = KycDocumentExtraction(
        document_type=document_type,
        original_filename=safe_name,
        source_filename=os.path.basename(archive_name or "archive.zip"),
        import_batch=import_batch,
        uploaded_by=user,
        client_type=client_type,
    )
    record.uploaded_file.save(safe_name, content, save=False)
    extraction = extract_document_data(record.uploaded_file.path, safe_name)
    _fill_document_extraction_fields(record, extraction)
    _apply_detected_document_type(record, extraction, document_type)
    record.save()
    return record, None


def _save_grouped_pdf_records(uploaded_file, document_type, user, import_batch, pages_per_document, client_type="pp"):
    if os.path.splitext(uploaded_file.name)[1].lower() != ".pdf":
        raise ValueError("Le mode document groupe accepte uniquement un fichier PDF.")

    shared_file_name = f"grouped_{uuid.uuid4().hex}_{os.path.basename(uploaded_file.name)}"
    base_record = KycDocumentExtraction(
        document_type=document_type,
        original_filename=os.path.basename(uploaded_file.name),
        source_filename=os.path.basename(uploaded_file.name),
        import_batch=import_batch,
        uploaded_by=user,
        client_type=client_type,
    )
    base_record.uploaded_file.save(shared_file_name, uploaded_file, save=False)

    grouped_extractions = extract_pdf_grouped_documents(
        base_record.uploaded_file.path,
        uploaded_file.name,
        pages_per_document=pages_per_document,
    )

    records = []
    for extraction in grouped_extractions:
        record = KycDocumentExtraction(
            document_type=document_type,
            uploaded_file=base_record.uploaded_file.name,
            original_filename=os.path.basename(uploaded_file.name),
            source_filename=os.path.basename(uploaded_file.name),
            import_batch=import_batch,
            uploaded_by=user,
            client_type=client_type,
        )
        _fill_document_extraction_fields(record, extraction)
        _apply_detected_document_type(record, extraction, document_type)
        record.save()
        records.append(record)
    return records


@login_required
def document_extraction(request):
    extraction = None
    # Ensure default types exist by calling detect_document_type with dummy input
    from kyc.document_extraction import detect_document_type
    detect_document_type("", "")
    
    valid_document_types = {dt.code: dt.label for dt in KycDocumentType.objects.all()}

    user = request.user
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    is_group_user = (user.filiale in ["BOA Group", "BOA GROUP"]) or (user.organe in users_groupe) or (not user.filiale)
    can_insert_batches = (user.organe == "DSI") or (user.organe == "PASS" and is_group_user)

    liste_filiales = [choice[0] for choice in Filiales]

    # Selected filiale for field sources (GET param 'filiale')
    if is_group_user:
        selected_filiale = request.GET.get("filiale", "").strip()
    else:
        selected_filiale = getattr(user, "filiale", "").strip()

    pp_fields = [
        ("CLIENT", "Nom & Prénom"),
        ("NUMID", "Numéro de document / NIN/NPI"),
        ("DATNAIS", "Date de naissance"),
        ("DATVALID", "Date d'expiration"),
        ("PAYNAIS", "Nationalité / Pays de naissance"),
        ("ADRESSE", "Adresse"),
        ("ORIGINE_REV", "Origine des revenus"),
    ]
    
    pm_fields = [
        ("CLIENT", "Raison sociale / Dénomination"),
        ("RCSNO", "Registre du commerce (RCS/RCCM)"),
        ("NUMERO_FISCAL", "Numéro fiscal (NIF)"),
        ("ADRESSE_SOCIALE", "Adresse sociale / Siège"),
    ]

    if request.method == "POST":
        if not can_insert_batches:
            messages.error(request, "Vous n'avez pas l'autorisation d'effectuer cette action.")
            return redirect("document_extraction")

        action = request.POST.get("action")
        if action == "save_document_field_sources":
            # Determine target filiale
            if is_group_user:
                target_filiale = request.GET.get("filiale", "").strip()
                apply_to_all = request.POST.get("apply_to_all_filiales") == "1"
            else:
                target_filiale = getattr(user, "filiale", "").strip()
                apply_to_all = False

            # Load the configs for the target filiale to filter
            if not target_filiale:
                pp_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if not c.filiales]
                pm_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if not c.filiales]
            else:
                pp_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if target_filiale in (c.filiales or [])]
                pm_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if target_filiale in (c.filiales or [])]
                if not pp_config_list:
                    pp_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if not c.filiales]
                if not pm_config_list:
                    pm_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if not c.filiales]

            pp_active_db_fields = set()
            if pp_config_list:
                pp_active_db_fields.update(pp_config_list[0].empty_check_fields or [])
            if not pp_active_db_fields:
                pp_active_db_fields = {"CLIENT", "NUMID", "DATNAIS", "DATVALID", "PAYNAIS", "ADRESSE", "ORIGINE_REV"}

            pm_active_db_fields = set()
            if pm_config_list:
                pm_active_db_fields.update(pm_config_list[0].empty_check_fields or [])
            if not pm_active_db_fields:
                pm_active_db_fields = {"CLIENT", "RCSNO", "NUMERO_FISCAL", "ADRESSE_SOCIALE"}

            filtered_pp_fields = [
                (f_name, label)
                for f_name, label in pp_fields
                if f_name in pp_active_db_fields
            ]
            filtered_pm_fields = [
                (f_name, label)
                for f_name, label in pm_fields
                if f_name in pm_active_db_fields
            ]

            for client_type_item in ["pp", "pm"]:
                fields_list = filtered_pp_fields if client_type_item == "pp" else filtered_pm_fields
                
                if apply_to_all:
                    # Save/Update the global config
                    global_configs = [c for c in KycFieldVisibilityConfig.objects.filter(client_type=client_type_item) if not c.filiales]
                    if global_configs:
                        config = global_configs[0]
                    else:
                        config = KycFieldVisibilityConfig(client_type=client_type_item, filiales=[])
                    
                    sources_dict = dict(config.field_sources or {})
                    for field_name, _ in fields_list:
                        val = request.POST.get(f"source_{client_type_item}_{field_name}", "").strip()
                        if val:
                            sources_dict[field_name] = val
                        else:
                            sources_dict.pop(field_name, None)
                    config.field_sources = sources_dict
                    config.save()

                    # Save to all other configs for this client_type
                    for oc in KycFieldVisibilityConfig.objects.filter(client_type=client_type_item):
                        oc.field_sources = sources_dict
                        oc.save()
                else:
                    # Find existing config or create a new one
                    if not target_filiale:
                        # Global config: filiales list is empty
                        global_configs = [c for c in KycFieldVisibilityConfig.objects.filter(client_type=client_type_item) if not c.filiales]
                        if global_configs:
                            config = global_configs[0]
                        else:
                            config = KycFieldVisibilityConfig(client_type=client_type_item, filiales=[])
                    else:
                        # Specific filiale config: filiales contains target_filiale
                        filiale_configs = [c for c in KycFieldVisibilityConfig.objects.filter(client_type=client_type_item) if target_filiale in (c.filiales or [])]
                        if filiale_configs:
                            config = filiale_configs[0]
                        else:
                            config = KycFieldVisibilityConfig(client_type=client_type_item, filiales=[target_filiale])
                    
                    sources_dict = dict(config.field_sources or {})
                    for field_name, _ in fields_list:
                        val = request.POST.get(f"source_{client_type_item}_{field_name}", "").strip()
                        if val:
                            sources_dict[field_name] = val
                        else:
                            sources_dict.pop(field_name, None)
                    config.field_sources = sources_dict
                    config.save()
            
            # Clear cache so updates are immediately visible
            global _field_visibility_configs_cache
            _field_visibility_configs_cache = None

            if apply_to_all:
                messages.success(request, "Configuration des sources documentaires enregistrée et appliquée à TOUTES les filiales.")
            else:
                messages.success(request, "Configuration des sources documentaires par champ enregistrée avec succès.")
            return_url = request.POST.get("return_url") or f"{reverse('document_extraction')}#sources"
            return redirect(return_url)
        uploaded_files = request.FILES.getlist("documents")
        if not uploaded_files and request.FILES.get("document"):
            uploaded_files = [request.FILES.get("document")]
        document_type = request.POST.get("document_type") or "piece_identite"
        import_mode = request.POST.get("import_mode") or "single"
        client_type = request.POST.get("client_type") or "pp"
        try:
            pages_per_document = max(int(request.POST.get("pages_per_document") or 1), 1)
        except ValueError:
            pages_per_document = 1

        if document_type not in valid_document_types:
            messages.error(request, "Veuillez choisir un type de document valide.")
            return redirect("document_extraction")

        if not uploaded_files:
            messages.error(request, "Veuillez selectionner au moins un document a analyser.")
        else:
            import_batch = _build_import_batch_name(request, uploaded_files)
            created_records = []
            errors = []

            if import_mode == "grouped_pdf":
                try:
                    created_records.extend(_save_grouped_pdf_records(
                        uploaded_files[0],
                        document_type,
                        request.user,
                        import_batch,
                        pages_per_document,
                        client_type=client_type,
                    ))
                except Exception as exc:
                    errors.append(str(exc))
            else:
                for uploaded_file in uploaded_files:
                    extension = os.path.splitext(uploaded_file.name)[1].lower()
                    if extension == ".zip":
                        try:
                            with zipfile.ZipFile(uploaded_file) as archive:
                                for member_name in archive.namelist():
                                    if member_name.endswith("/"):
                                        continue
                                    record, error = _save_zip_document_record(
                                        archive,
                                        member_name,
                                        document_type,
                                        request.user,
                                        import_batch,
                                        uploaded_file.name,
                                        client_type=client_type,
                                    )
                                    if record:
                                        created_records.append(record)
                                    if error:
                                        errors.append(error)
                        except zipfile.BadZipFile:
                            errors.append(f"Archive ZIP invalide: {uploaded_file.name}")
                    else:
                        try:
                            record, _ = _save_uploaded_document_record(
                                uploaded_file,
                                document_type,
                                request.user,
                                import_batch=import_batch,
                                client_type=client_type,
                            )
                            created_records.append(record)
                        except Exception as exc:
                            errors.append(f"{uploaded_file.name}: {exc}")

            if created_records:
                messages.success(
                    request,
                    f"{len(created_records)} document(s) charge(s), analyse(s) et enregistre(s) dans le lot {import_batch}.",
                )
                request.session[LAST_UPLOADED_DOCUMENT_BATCH_SESSION_KEY] = import_batch
                request.session.modified = True
            if errors:
                messages.warning(request, f"{len(errors)} element(s) non importe(s): " + " | ".join(errors[:5]))
            if created_records:
                return redirect(f"{reverse('document_extraction')}?{urlencode({'uploaded_batch': import_batch, 'client_type': client_type})}#charger")

    documents = _filtered_document_extractions_from_request(request)
    selected_document_type = request.GET.get("document_type", "")
    selected_import_batch = (request.GET.get("import_batch") or "").strip()
    uploaded_batch = (request.GET.get("uploaded_batch") or "").strip()
    if not uploaded_batch and not request.GET:
        uploaded_batch = (request.session.get(LAST_UPLOADED_DOCUMENT_BATCH_SESSION_KEY) or "").strip()
    search_query = (request.GET.get("q") or "").strip()
    search_field = request.GET.get("field") or "all"
    if selected_document_type not in valid_document_types:
        selected_document_type = ""
    if search_field not in {field for field, _ in DOCUMENT_EXTRACTION_SEARCH_FIELDS}:
        search_field = "all"

    selected_extraction_id = request.GET.get("extraction_id")
    if selected_extraction_id and selected_extraction_id.isdigit():
        selected_record = get_object_or_404(KycDocumentExtraction, pk=selected_extraction_id)
        extraction = _format_document_extraction_record(selected_record)

    uploaded_documents = KycDocumentExtraction.objects.none()
    uploaded_documents_count = 0
    uploaded_quality_alerts = []
    uploaded_quality_alerts_count = 0
    if uploaded_batch:
        uploaded_documents_queryset = KycDocumentExtraction.objects.filter(import_batch=uploaded_batch)
        uploaded_documents_count = uploaded_documents_queryset.count()
        uploaded_quality_alerts_count = uploaded_documents_queryset.exclude(extraction_warnings="").count()
        uploaded_documents = uploaded_documents_queryset.order_by("-created_at")[:50]
        for document in uploaded_documents:
            warnings = [warning for warning in document.extraction_warnings.splitlines() if warning]
            if warnings:
                uploaded_quality_alerts.append({
                    "filename": document.original_filename or os.path.basename(document.uploaded_file.name),
                    "document_type": document.get_document_type_display(),
                    "warnings": warnings[:3],
                })
            if len(uploaded_quality_alerts) >= 5:
                break

    requested_kyc_pp_matching = request.GET.get("match_kyc") == "1"
    client_type = request.GET.get("client_type", "pp")
    selected_match_job = None
    selected_match_job_id = request.GET.get("match_job")
    show_match_job_results = request.GET.get("show_match_results") == "1"
    show_match_result_modal = request.GET.get("result_modal") == "1"
    show_match_child_modal = request.GET.get("child_modal") == "1"
    wants_results_tab = request.GET.get("tab") == "results"
    if selected_match_job_id and selected_match_job_id.isdigit():
        selected_match_job = get_object_or_404(KycDocumentMatchJob, pk=selected_match_job_id)
        if not _user_can_access_document_match_job(request.user, selected_match_job):
            selected_match_job = None

    last_match_result = request.session.get(LAST_KYC_PP_MATCH_RESULT_SESSION_KEY)
    active_match_params = None
    is_global_consultation = not request.GET

    if requested_kyc_pp_matching:
        active_match_params = _document_match_scope_params(request.GET)
        request.session[LAST_KYC_PP_MATCH_SESSION_KEY] = active_match_params

    kyc_pp_matches = []
    kyc_pp_match_summary = {
        "documents_checked": 0,
        "documents_matched": 0,
        "clients_matched": 0,
        "suggestions_count": 0,
        "match_rate": 0,
    }
    if selected_match_job:
        active_match_params = selected_match_job.scope_params or {}
        if selected_match_job.status == "completed" and show_match_job_results:
            kyc_pp_matches, stored_summary, stored_params = _hydrate_kyc_pp_match_result(selected_match_job.result)
            if stored_summary is not None:
                kyc_pp_match_summary = stored_summary
                active_match_params = stored_params
                request.session[LAST_KYC_PP_MATCH_SESSION_KEY] = active_match_params
                request.session[LAST_KYC_PP_MATCH_RESULT_SESSION_KEY] = selected_match_job.result
                request.session.modified = True
        elif selected_match_job.status in {"pending", "running", "completed"}:
            kyc_pp_match_summary = {
                "documents_checked": selected_match_job.progress_total,
                "documents_matched": 0,
                "clients_matched": 0,
                "suggestions_count": 0,
                "match_rate": 0,
            }
    elif requested_kyc_pp_matching:
        match_documents = _filtered_document_extractions_from_params(active_match_params, user=request.user)
        if client_type == "pm":
            kyc_pp_matches, kyc_pp_match_summary = _build_kyc_pm_document_matches(match_documents)
        else:
            kyc_pp_matches, kyc_pp_match_summary = _build_kyc_pp_document_matches(match_documents)
        request.session[LAST_KYC_PP_MATCH_RESULT_SESSION_KEY] = _serialize_kyc_pp_matches(
            kyc_pp_matches,
            kyc_pp_match_summary,
            active_match_params,
        )
        matched_batch = (active_match_params.get("import_batch") or "").strip()
        if matched_batch:
            matched_batches = set(request.session.get(KYC_PP_MATCHED_BATCHES_SESSION_KEY) or [])
            matched_batches.add(matched_batch)
            request.session[KYC_PP_MATCHED_BATCHES_SESSION_KEY] = sorted(matched_batches)
        request.session.modified = True
    elif wants_results_tab or is_global_consultation:
        stored_matches, stored_summary, stored_params = _hydrate_kyc_pp_match_result(last_match_result)
        if not wants_results_tab and stored_summary is not None and stored_params == {}:
            kyc_pp_matches = stored_matches
            kyc_pp_match_summary = stored_summary
        else:
            if wants_results_tab:
                completed_jobs = KycDocumentMatchJob.objects.filter(
                    created_by=request.user,
                    status="completed",
                ).order_by("-completed_at", "-created_at")
                hydrated_match_lists = []
                documents_checked_total = 0
                for job in completed_jobs:
                    job_matches, job_summary, _ = _hydrate_kyc_pp_match_result(job.result)
                    if job_matches:
                        hydrated_match_lists.append(job_matches)
                    if job_summary:
                        documents_checked_total += job_summary.get("documents_checked", 0)

                kyc_pp_matches = _merge_kyc_pp_match_lists(hydrated_match_lists)
                if kyc_pp_matches:
                    matched_idp_keys = {
                        _normalize_match_value(getattr(match["client"], "IDM" if client_type == "pm" else "IDP", "")) or str(match["client"].pk)
                        for match in kyc_pp_matches
                    }
                    kyc_pp_match_summary = {
                        "documents_checked": documents_checked_total,
                        "documents_matched": len({match["document"].pk for match in kyc_pp_matches}),
                        "clients_matched": len(matched_idp_keys),
                        "suggestions_count": sum(len(_build_kyc_pp_match_action_items(match)) for match in kyc_pp_matches),
                        "match_rate": 0,
                    }
                    request.session[LAST_KYC_PP_MATCH_SESSION_KEY] = {}
                    request.session[LAST_KYC_PP_MATCH_RESULT_SESSION_KEY] = _serialize_kyc_pp_matches(
                        kyc_pp_matches,
                        kyc_pp_match_summary,
                        {},
                    )
                    request.session.modified = True
            else:
                global_job = (
                    KycDocumentMatchJob.objects
                    .filter(created_by=request.user, scope_params={}, status="completed")
                    .order_by("-completed_at", "-created_at")
                    .first()
                )
                if global_job:
                    kyc_pp_matches, stored_summary, stored_params = _hydrate_kyc_pp_match_result(global_job.result)
                    if stored_summary is not None:
                        kyc_pp_match_summary = stored_summary
                        request.session[LAST_KYC_PP_MATCH_SESSION_KEY] = {}
                        request.session[LAST_KYC_PP_MATCH_RESULT_SESSION_KEY] = global_job.result
                        request.session.modified = True
        if kyc_pp_matches:
            active_match_params = {}
    elif not is_global_consultation:
        kyc_pp_matches, stored_summary, stored_params = _hydrate_kyc_pp_match_result(last_match_result)
        stored_batch = (stored_params.get("import_batch") or "").strip()
        if selected_import_batch and stored_batch != selected_import_batch:
            kyc_pp_matches, stored_summary, stored_params = [], None, {}
        if stored_summary is not None:
            kyc_pp_match_summary = stored_summary
            active_match_params = stored_params

    kyc_pp_matches = _merge_kyc_pp_match_lists([kyc_pp_matches])
    run_kyc_pp_matching = active_match_params is not None
    kyc_pp_match_total_count = len(kyc_pp_matches)
    kyc_pp_match_filters = _get_kyc_pp_match_filters(request.GET)
    kyc_pp_matches = _filter_kyc_pp_matches(kyc_pp_matches, request.GET)
    expired_document_matches = KycExpiredDocumentScanMatch.objects.select_related("client", "document").filter(
        status="a_valider"
    )
    if selected_import_batch:
        expired_document_matches = expired_document_matches.filter(document__import_batch=selected_import_batch)
    unique_expired_document_matches = {}
    for expired_match in expired_document_matches.order_by("-match_rate", "-scan_date")[:500]:
        if expired_match.client_id in unique_expired_document_matches:
            continue
        unique_expired_document_matches[expired_match.client_id] = expired_match
        if len(unique_expired_document_matches) >= 100:
            break
    expired_match_by_client_id = unique_expired_document_matches
    for match in kyc_pp_matches:
        expired_match = expired_match_by_client_id.pop(match["client"].pk, None)
        if expired_match:
            match["expired_document_match"] = expired_match
        match["action_items"] = _build_kyc_pp_match_action_items(match)
        match["action_summary"] = " | ".join(action["text"] for action in match["action_items"])
    expired_document_matches = list(expired_match_by_client_id.values())
    matched_batches = set(request.session.get(KYC_PP_MATCHED_BATCHES_SESSION_KEY) or [])
    if is_group_user:
        match_jobs_qs = KycDocumentMatchJob.objects.all()
        extractions_qs = KycDocumentExtraction.objects.all()
    else:
        match_jobs_qs = KycDocumentMatchJob.objects.filter(created_by__filiale=user.filiale)
        extractions_qs = KycDocumentExtraction.objects.filter(uploaded_by__filiale=user.filiale)

    uploaded_batch_job_done = False
    uploaded_batch_running_job = None
    uploaded_batch_result_url = ""
    if uploaded_batch:
        uploaded_batch_running_job = (
            match_jobs_qs
            .filter(scope_params={"import_batch": uploaded_batch}, status__in=["pending", "running"])
            .order_by("-created_at")
            .first()
        )
        if uploaded_batch_running_job:
            follow_params = dict(uploaded_batch_running_job.scope_params or {})
            follow_params["match_job"] = uploaded_batch_running_job.pk
            uploaded_batch_running_job.follow_url = f"{reverse('document_extraction')}?{urlencode(follow_params)}#suivi"
        uploaded_batch_job_done = match_jobs_qs.filter(
            scope_params={"import_batch": uploaded_batch},
            status="completed",
        ).exists()
        uploaded_batch_completed_job = (
            match_jobs_qs
            .filter(scope_params={"import_batch": uploaded_batch}, status="completed")
            .order_by("-completed_at", "-created_at")
            .first()
        )
        if uploaded_batch_completed_job:
            result_params = dict(uploaded_batch_completed_job.scope_params or {})
            result_params["match_job"] = uploaded_batch_completed_job.pk
            result_params["show_match_results"] = "1"
            uploaded_batch_result_url = f"{reverse('document_extraction')}?{urlencode(result_params)}#consulter"
    uploaded_batch_matching_done = bool(uploaded_batch and (uploaded_batch in matched_batches or uploaded_batch_job_done))
    show_document_modal = (bool(extraction) and not show_match_child_modal) or request.GET.get("lot_view") == "1"
    recent_match_jobs = list(match_jobs_qs.order_by("-created_at")[:12])
    for job in recent_match_jobs:
        follow_params = dict(job.scope_params or {})
        follow_params["match_job"] = job.pk
        result_params = dict(follow_params)
        result_params["show_match_results"] = "1"
        result_params["result_modal"] = "1"
        job.follow_url = f"{reverse('document_extraction')}?{urlencode(follow_params)}#suivi"
        job.result_url = f"{reverse('document_extraction')}?{urlencode(result_params)}#suivi"
    upload_batch_queue_all = list(
        extractions_qs
        .exclude(import_batch="")
        .values("import_batch")
        .annotate(documents_count=Count("id"), latest_created_at=Max("created_at"))
        .order_by("-latest_created_at")[:12]
    )
    upload_batch_queue = []
    for batch in upload_batch_queue_all:
        batch_name = batch["import_batch"]
        batch["documents_url"] = f"{reverse('document_extraction')}?{urlencode({'import_batch': batch_name})}#base"
        latest_job = (
            match_jobs_qs
            .filter(scope_params={"import_batch": batch_name})
            .order_by("-created_at")
            .first()
        )
        batch["job"] = latest_job
        batch["status"] = "pending"
        batch["status_label"] = "En attente"
        batch["start_url"] = f"{reverse('start_document_extraction_match_job')}?{urlencode({'import_batch': batch_name})}"
        batch["follow_url"] = ""
        batch["result_url"] = ""
        if latest_job:
            follow_params = dict(latest_job.scope_params or {})
            follow_params["match_job"] = latest_job.pk
            result_params = dict(follow_params)
            result_params["show_match_results"] = "1"
            result_params["result_modal"] = "1"
            batch["follow_url"] = f"{reverse('document_extraction')}?{urlencode(follow_params)}#suivi"
            batch["result_url"] = f"{reverse('document_extraction')}?{urlencode(result_params)}#suivi"
            if latest_job.status == "completed":
                batch["status"] = "completed"
                batch["status_label"] = "Termine"
            elif latest_job.status in {"pending", "running"}:
                batch["status"] = "running"
                batch["status_label"] = "En cours"
            elif latest_job.status == "failed":
                batch["status"] = "failed"
                batch["status_label"] = "Echec"
        if batch["status"] in {"pending", "failed"}:
            upload_batch_queue.append(batch)

    paginator = Paginator(documents, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_params.pop("extraction_id", None)
    query_params.pop("match_kyc", None)

    match_query_params = request.GET.copy()
    match_query_params["match_kyc"] = "1"
    match_query_params.pop("page", None)
    match_query_params.pop("extraction_id", None)

    if active_match_params is not None:
        export_match_params = dict(active_match_params)
    else:
        export_match_params = _document_match_scope_params(request.GET)
    match_filter_hidden_params = {
        key: value
        for key, value in export_match_params.items()
        if key not in KYC_PP_MATCH_FILTER_FIELDS
        and key not in {"page", "extraction_id"}
        and value not in (None, "")
    }
    match_reset_params = dict(match_filter_hidden_params)
    if run_kyc_pp_matching:
        match_reset_params["match_kyc"] = "1"
    export_match_params.update({
        key: value
        for key, value in kyc_pp_match_filters.items()
        if value
    })
    export_match_querystring = urlencode(export_match_params)
    match_reset_querystring = urlencode(match_reset_params)
    selected_match_job_result_url = ""
    selected_match_job_follow_url = f"{reverse('document_extraction')}#suivi"
    selected_match_job_parent_modal_url = ""
    if selected_match_job:
        follow_params = dict(selected_match_job.scope_params or {})
        follow_params["match_job"] = selected_match_job.pk
        selected_match_job_follow_url = f"{reverse('document_extraction')}?{urlencode(follow_params)}#suivi"
        result_params = dict(selected_match_job.scope_params or {})
        result_params["match_job"] = selected_match_job.pk
        result_params["show_match_results"] = "1"
        result_params["result_modal"] = "1"
        selected_match_job_result_url = f"{reverse('document_extraction')}?{urlencode(result_params)}#suivi"
        selected_match_job_parent_modal_url = selected_match_job_result_url
    latest_global_match_job = (
        KycDocumentMatchJob.objects
        .filter(created_by=request.user, scope_params={}, status="completed")
        .order_by("-completed_at", "-created_at")
        .first()
    )
    all_match_results_url = f"{reverse('document_extraction')}#consulter"
    if latest_global_match_job:
        all_results_params = {
            "match_job": latest_global_match_job.pk,
            "show_match_results": "1",
            "result_modal": "1",
        }
        all_match_results_url = f"{reverse('document_extraction')}?{urlencode(all_results_params)}#suivi"

    if selected_match_job and show_match_result_modal:
        base_child_params = dict(selected_match_job.scope_params or {})
        base_child_params["match_job"] = selected_match_job.pk
        base_child_params["show_match_results"] = "1"
        base_child_params["result_modal"] = "1"
        base_child_params["child_modal"] = "1"
        for match in kyc_pp_matches:
            child_params = dict(base_child_params)
            child_params["extraction_id"] = match["document"].pk
            match["child_detail_url"] = f"{reverse('document_extraction')}?{urlencode(child_params)}#suivi"
    pp_sources = {}
    pm_sources = {}
    
    if not selected_filiale:
        pp_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if not c.filiales]
        pm_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if not c.filiales]
    else:
        pp_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if selected_filiale in (c.filiales or [])]
        pm_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if selected_filiale in (c.filiales or [])]
        
        if not pp_config:
            pp_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if not c.filiales]
        if not pm_config:
            pm_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if not c.filiales]
            
    pp_active_db_fields = set()
    if pp_config:
        pp_sources = getattr(pp_config[0], "field_sources", {}) or {}
        pp_active_db_fields.update(pp_config[0].empty_check_fields or [])
    if not pp_active_db_fields:
        pp_active_db_fields = {"CLIENT", "NUMID", "DATNAIS", "DATVALID", "PAYNAIS", "ADRESSE", "ORIGINE_REV"}

    pm_active_db_fields = set()
    if pm_config:
        pm_sources = getattr(pm_config[0], "field_sources", {}) or {}
        pm_active_db_fields.update(pm_config[0].empty_check_fields or [])
    if not pm_active_db_fields:
        pm_active_db_fields = {"CLIENT", "RCSNO", "NUMERO_FISCAL", "ADRESSE_SOCIALE"}

    filtered_pp_fields = [
        (f_name, label)
        for f_name, label in pp_fields
        if f_name in pp_active_db_fields
    ]
    filtered_pm_fields = [
        (f_name, label)
        for f_name, label in pm_fields
        if f_name in pm_active_db_fields
    ]

    document_field_source_sections = [
        {
            "title": "Particuliers (PP)",
            "client_type": "pp",
            "fields": filtered_pp_fields,
            "sources": pp_sources,
        },
        {
            "title": "Entreprises (PM)",
            "client_type": "pm",
            "fields": filtered_pm_fields,
            "sources": pm_sources,
        }
    ]

    context = {
        "extraction": extraction,
        "documents": page_obj,
        "documents_count": documents.count(),
        "uploaded_batch": uploaded_batch,
        "uploaded_documents": uploaded_documents,
        "uploaded_documents_count": uploaded_documents_count,
        "uploaded_quality_alerts": uploaded_quality_alerts,
        "uploaded_quality_alerts_count": uploaded_quality_alerts_count,
        "uploaded_batch_matching_done": uploaded_batch_matching_done,
        "uploaded_batch_running_job": uploaded_batch_running_job,
        "uploaded_batch_result_url": uploaded_batch_result_url,
        "upload_batch_queue": upload_batch_queue,
        "show_document_modal": show_document_modal,
        "kyc_pp_matches": kyc_pp_matches,
        "kyc_pp_match_summary": kyc_pp_match_summary,
        "kyc_pp_match_total_count": kyc_pp_match_total_count,
        "kyc_pp_match_filtered_count": len(kyc_pp_matches),
        "expired_document_matches": expired_document_matches,
        "kyc_pp_match_filters": kyc_pp_match_filters,
        "match_filter_hidden_params": match_filter_hidden_params,
        "run_kyc_pp_matching": run_kyc_pp_matching,
        "show_match_actions": run_kyc_pp_matching and not wants_results_tab and (not selected_match_job or show_match_job_results),
        "wants_results_tab": wants_results_tab,
        "selected_match_job": selected_match_job,
        "show_match_job_results": show_match_job_results,
        "show_match_result_modal": show_match_result_modal,
        "show_match_child_modal": show_match_child_modal,
        "selected_match_job_result_url": selected_match_job_result_url,
        "selected_match_job_follow_url": selected_match_job_follow_url,
        "selected_match_job_parent_modal_url": selected_match_job_parent_modal_url,
        "all_match_results_url": all_match_results_url,
        "recent_match_jobs": recent_match_jobs,
        "match_scope_import_batch": (active_match_params or {}).get("import_batch", ""),
        "match_querystring": match_query_params.urlencode(),
        "match_reset_querystring": match_reset_querystring,
        "export_match_querystring": export_match_querystring,
        "document_type_choices": [(dt.code, dt.label) for dt in KycDocumentType.objects.all()],
        "all_document_types": KycDocumentType.objects.all(),
        "client_type": client_type,
        "selected_document_type": selected_document_type,
        "selected_import_batch": selected_import_batch,
        "search_fields": DOCUMENT_EXTRACTION_SEARCH_FIELDS,
        "search_field": search_field,
        "search_query": search_query,
        "field_labels": DOCUMENT_EXTRACTION_FIELD_LABELS,
        "page_querystring": query_params.urlencode(),
        "is_group_user": is_group_user,
        "liste_filiales": liste_filiales,
        "selected_filiale": selected_filiale,
        "document_field_source_sections": document_field_source_sections,
        "document_field_source_return_url": request.get_full_path(),
        "can_insert_batches": can_insert_batches,
    }
    return render(request, 'document_extraction.html', context)


@login_required
def config_document_types(request):
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "save":
            type_id = request.POST.get("type_id")
            code = request.POST.get("code", "").strip().lower()
            label = request.POST.get("label", "").strip()
            filiale = request.POST.get("filiale", "").strip()
            client_type = request.POST.get("client_type", "pp").strip()
            min_score_str = request.POST.get("min_score", "2.0")
            keywords = request.POST.get("keywords", "").strip()
            
            try:
                min_score = float(min_score_str)
            except ValueError:
                min_score = 2.0
                
            if not code or not label:
                messages.error(request, "Le code technique et le libellé sont requis.")
            else:
                if type_id:
                    doc_type = get_object_or_404(KycDocumentType, pk=type_id)
                    if doc_type.code not in ['piece_identite', 'passeport']:
                        doc_type.code = code
                        doc_type.filiale = filiale
                    doc_type.label = label
                    doc_type.client_type = client_type
                    doc_type.min_score = min_score
                    doc_type.keywords = keywords
                    doc_type.save()
                    messages.success(request, f"Le type de document '{label}' a été mis à jour.")
                else:
                    if KycDocumentType.objects.filter(code=code, filiale=filiale, client_type=client_type).exists():
                        messages.error(request, f"Un type de document avec le code '{code}' existe déjà pour cette filiale et ce type de client.")
                    else:
                        KycDocumentType.objects.create(
                            code=code,
                            label=label,
                            filiale=filiale,
                            client_type=client_type,
                            min_score=min_score,
                            keywords=keywords
                        )
                        messages.success(request, f"Le type de document '{label}' a été créé avec succès.")
            return redirect("config_document_types")
            
        elif action == "delete":
            type_id = request.POST.get("type_id")
            if type_id:
                doc_type = get_object_or_404(KycDocumentType, pk=type_id)
                if doc_type.code in ['piece_identite', 'passeport']:
                    messages.error(request, "Les types de documents système ne peuvent pas être supprimés.")
                else:
                    label = doc_type.label
                    doc_type.delete()
                    messages.success(request, f"Le type de document '{label}' a été supprimé.")
            return redirect("config_document_types")
            
    document_types = KycDocumentType.objects.all()
    context = {
        "document_types": document_types,
        "filiale_choices": Filiales,
        "client_type_choices": CLIENT_TYPE_CHOICES,
    }
    return render(request, "config_document_types.html", context)


@login_required
def import_log_download(request, filename):
    if not request.user.is_superuser:
        return redirect('accueil')
    if not filename or ".." in filename or "/" in filename or "\\" in filename:
        raise Http404("Invalid file")

    log_dir = os.path.join(settings.BASE_DIR, "logs")
    run_dir = os.path.join(log_dir, "import_runs")
    allowed_paths = [
        os.path.join(run_dir, filename),
        os.path.join(log_dir, filename),
    ]
    file_path = next((p for p in allowed_paths if os.path.isfile(p)), None)
    if not file_path:
        raise Http404("File not found")

    return FileResponse(open(file_path, "rb"), as_attachment=True, filename=filename)


@login_required
def profile_update(request):
    if request.method == "POST":
        form = ProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect("profile")  # redirige aprÃ¨s update
    else:
        form = ProfileForm(instance=request.user)
    return render(request, "profile_update.html", {"form": form})


def rechercher_et_noter_agent(request):
    agent = None
    form = None
    message = None

    if request.method == "POST":
        expl = request.POST.get('expl', '')
        if expl:
            try:
                agent = ProfileV.objects.get(code_expl=expl)
            except ProfileV.DoesNotExist:
                message = "Agent introuvable."

            if agent:
                # Si l'utilisateur a le profil "Contrôle permanent"
                if request.user.is_authenticated and request.user.groups.filter(name='Contrôle permanent').exists():
                    form = NoteForm(request.POST or None)
                    if form.is_valid():
                        note = form.save(commit=False)
                        note.agent = agent
                        note.date_notation = timezone.now()  # Enregistrer la date de notation
                        note.save()
                        message = "Notation enregistrée avec succÃ¨s."
                        form = None  # Reset le formulaire aprÃ¨s enregistrement pour éviter la soumission répétée
                else:
                    message = "Vous n'avez pas la permission de noter cet agent."

    return render(request, 'accueil.html', {'agent': agent, 'form': form, 'message': message})


def password_reset_request(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data['email']
            associated_users = ProfileV.objects.filter(Q(username=data))
            if associated_users.exists():
                for user in associated_users:
                    subject = "Password Reset Requested"
                    email_template_name = "password_reset_email.txt"
                    c = {
                        "email": user.username,
                        'domain': request.get_host(),
                        'site_name': 'Plateforme KYC BOA',
                        "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                        'token': default_token_generator.make_token(user),
                        'protocol': 'https' if request.is_secure() else 'http',
                    }
                    email = render_to_string(email_template_name, c)
                    try:
                        send_mail(subject, email, 'mamadou@mamadou.sn', [user.username], fail_silently=False)
                    except BadHeaderError:
                        return HttpResponse('Invalid header found.')
                    return redirect('/password_resete/done')

                    messages.success(request, 'A message with reset password instructions has been sent to your inbox.')
                    return redirect("accueil")
            else:
                errors = 'Votre mail ne figure pas dans notre base.'
                return render(request=request, template_name="password_reset.html",
                              context={"password_reset_form": password_reset_form, "errors": errors})

    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="password_reset.html",
                  context={"password_reset_form": password_reset_form})


@login_required
def profil(request):
    roles_exclus = ["Chargé Client"]
    notes = Notation.objects.filter(flux_stock='Flux')
    user = request.user
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la derniÃ¨re note par agent
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)

    context = {'roles_exclus': roles_exclus,
               'notation': notation,
               }
    return render(request, 'profil.html', context)


def profile(request):
    if request.method == 'POST':
        user_form = ProfileModify(request.POST, instance=request.user)
        if user_form.is_valid():
            user_form.save()
            messages.success(request, 'Votre profil a été modifié avec succÃ¨s')
            return redirect('/perso/profil')

    else:
        user_form = ProfileModify(instance=request.user)
    return render(request, 'modify_profil.html', {'user_form': user_form})


@method_decorator(login_required, name='dispatch')
class ChangePasswordView(SuccessMessageMixin, PasswordChangeView):
    template_name = 'modify_pw.html'
    success_message = "Votre mot de passe a été changé avec succÃ¨s"
    success_url = reverse_lazy('profil')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ajout du contexte personnalisé
        context['roles_exclus'] = ["Chargé Client"]
        return context


def reset_user_password_b(request, user_id):
    roles_exclus = ["Chargé Client"]
    user = get_object_or_404(ProfileV, pk=user_id)

    if request.method == 'POST':
        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            # Enregistrer le nouveau mot de passe en le hachant
            new_password = form.cleaned_data['new_password']
            user.password = make_password(new_password)
            user.save()
            return redirect('profil')  # Rediriger vers la liste des utilisateurs aprÃ¨s modification
    else:
        form = ResetPasswordForm()

    return render(request, 'modify_pw.html', {'form': form, 'user': user, 'roles_exclus': roles_exclus})


@login_required
def perso(request):
    # Récupérer l'utilisateur connecté
    roles_exclus = ["Chargé Client"]
    user = request.user

    # Vérifier si l'utilisateur appartient Ã  "Contrôle", "Conformité" ou "Contrôle Groupe"
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        agents = ProfileV.objects.filter(filiale=user.filiale)
    else:
        agents = ProfileV.objects.all()
    # Gestion de la recherche par exploitant
    query = request.GET.get('q', '')
    if query:
        agents = agents.filter(code_expl__icontains=query)

    return render(request, 'mon_profile.html', {'agents': agents, 'query': query, 'roles_exclus': roles_exclus})


@login_required
def agent(request):
    roles_exclus = ["Chargé Client"]
    user = request.user

    # Vérifier si l'utilisateur appartient Ã  "Contrôle", "Conformité" ou "Contrôle Groupe"
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        notes = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Flux')
    elif user.organe in ['Directeur Agence']:
        notes = Notation.objects.filter(note_par__filiale=user.filiale, agent__agence=user.agence, flux_stock='Flux')
    else:
        notes = Notation.objects.filter(flux_stock='Flux')

    # Annoter chaque agent avec la derniÃ¨re date de notation
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la derniÃ¨re note par agent
    notes = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    notes = notes.order_by('-date_notation')
    # Gestion de la recherche par exploitant
    query = request.GET.get('q', '')
    if query:
        agents = ProfileV.objects.filter(code_expl__icontains=query)  # Utilisation de icontains pour une recherche partielle
        if agents.exists():
            notes = notes.all().filter(agent__in=agents)
        else:
            # Si aucun agent n'est trouvé, vider le queryset pour ne rien afficher
            notes = notes.none()

    return render(request, 'agent.html', {
        'notes': notes,
        'query': query,
        'roles_exclus': roles_exclus,
    })


def export_agents_excel(request):
    def strip_tz(value):
        if hasattr(value, 'tzinfo'):
            return value.replace(tzinfo=None)
        return value

    user = request.user

    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Flux')
    elif user.organe in ['Directeur Agence']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, agent__agence=user.agence, flux_stock='Flux')
    else:
        donnees = Notation.objects.filter(flux_stock='Flux')

    latest_notes = donnees.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la derniÃ¨re note par agent
    donnees = donnees.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    donnees = donnees.order_by('-date_notation')

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notation_Flux"

    # EntÃªtes
    headers = ["FILIALE", "EXPLOITANT", "NOM", "AGENCE", "EMAIL", "NOTE", "DerniÃ¨re notation", "Noté par le contrôleur",
               "Flux/Stock"]
    ws.append(headers)

    # Données
    for d in donnees:
        agent_nom = f"{d.agent.first_name} {d.agent.last_name}".strip()
        agent_email = d.agent.email or d.agent.username
        ws.append([
            d.agent.filiale, d.agent.code_expl, agent_nom, d.agent.agence, agent_email, d.note,
            strip_tz(d.date_notation), d.note_par.email, d.flux_stock

        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Notation_Flux_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_agents_excel_s(request):
    def strip_tz(value):
        if hasattr(value, 'tzinfo'):
            return value.replace(tzinfo=None)
        return value

    user = request.user

    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Stock')
    elif user.organe in ['Directeur Agence']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, agent__agence=user.agence, flux_stock='Stock')
    else:
        donnees = Notation.objects.filter(flux_stock='Stock')

    latest_notes = donnees.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la derniÃ¨re note par agent
    donnees = donnees.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    donnees = donnees.order_by('-date_notation')

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notation_Stock"

    # EntÃªtes
    headers = ["FILIALE", "EXPLOITANT", "NOM", "AGENCE", "EMAIL", "NOTE", "DerniÃ¨re notation", "Noté par le contrôleur",
               "Flux/Stock"]
    ws.append(headers)

    # Données
    for d in donnees:
        agent_nom = f"{d.agent.first_name} {d.agent.last_name}".strip()
        agent_email = d.agent.email or d.agent.username
        ws.append([
            d.agent.filiale, d.agent.code_expl, agent_nom, d.agent.agence, agent_email, d.note,
            strip_tz(d.date_notation), d.note_par.email, d.flux_stock

        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Notation_Stock_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def perso_stock(request):
    # Récupérer l'utilisateur connecté
    user = request.user

    # Vérifier si l'utilisateur appartient Ã  "Contrôle", "Conformité" ou "Contrôle Groupe"
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        notes = Notation.objects.filter(filiale=user.filiale, flux_stock='Stock')
    else:
        notes = Notation.objects.all().filter(flux_stock='Flux')
    # Gestion de la recherche par exploitant
    query = request.GET.get('q', '')
    if query:
        notes = notes.filter(agent__code_expl__icontains=query)

    return render(request, 'agent_stock.html', {'notes': notes, 'query': query})


@login_required
def agent_stock(request):
    user = request.user

    # Vérifier si l'utilisateur appartient Ã  "Contrôle", "Conformité" ou "Contrôle Groupe"
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        notes = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Stock')
    else:
        notes = Notation.objects.filter(flux_stock='Stock')

    # Annoter chaque agent avec la derniÃ¨re date de notation
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notes = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])

    # Filtrer pour obtenir uniquement la derniÃ¨re note par agent
    notes = notes.order_by('-date_notation')

    # Gestion de la recherche par exploitant
    query = request.GET.get('q', '')
    if query:
        agents = ProfileV.objects.filter(code_expl__icontains=query)  # Utilisation de icontains pour une recherche partielle
        if agents.exists():
            notes = notes.filter(agent__in=agents)
        else:
            # Si aucun agent n'est trouvé, vider le queryset pour ne rien afficher
            notes = notes.none()

    return render(request, 'agent_stock.html', {'notes': notes, 'query': query})


def notes(request):
    agent = None
    roles_exclus = ["Chargé Client", "Directeur Agence"]
    form = NotationForm()  # Initialisation du formulaire

    if request.method == 'POST':
        if 'search_agent' in request.POST:
            # Rechercher l'agent par son code exploitant
            code_exploitant = request.POST.get('code_exploitant')
            try:
                agent = ProfileV.objects.get(code_expl=code_exploitant, filiale=request.user.filiale)

                # Préremplir le formulaire avec l'agent trouvé
                form = NotationForm(initial={'agent': agent})
            except ProfileV.DoesNotExist:
                agent = None
                error_message = "L'agent avec ce code exploitant n'existe pas."
                return render(request, 'notation.html', {'form': form, 'error_message': error_message})

        else:
            # Soumission du formulaire de notation
            form = NotationForm(request.POST)
            if form.is_valid():
                user = request.user
                # Assigner l'agent Ã  partir du formulaire
                notation = form.save(commit=False)
                notation.filiale = request.user.filiale
                notation.note_par = request.user
                notation.date_notation = timezone.now()
                notation.save()
                messages.success(request, 'La notation a bien été sauvegardée.')

                return redirect('agent')
    else:
        form = NotationForm()  # Afficher un formulaire vide si la requÃªte n'est pas en POST

    return render(request, 'notation.html', {'form': form, 'agent': agent, 'roles_exclus': roles_exclus})


def agent_detail(request, agent_id):
    agent = get_object_or_404(ProfileV, id=agent_id)
    notations = agent.notations.all().order_by('-date_notation')
    return render(request, 'agent_detail.html', {'agent': agent, 'notations': notations})


@login_required
def historique(request):
    roles_exclus = ["Chargé Client", "Directeur Agence"]
    query = request.GET.get('q')

    if query:
        # Filtre les notations en fonction du code exploitant
        notations = Notation.objects.filter(note_par=request.user, agent__code_expl__icontains=query).order_by(
            "-date_notation")
    else:
        # RécupÃ¨re toutes les notations de l'utilisateur connecté
        notations = Notation.objects.filter(note_par=request.user).order_by("-date_notation")

    # Passe les notations au template
    context = {
        'notations': notations,
        'query': query,
    }  # Pour pré-remplir la barre de recherche
    return render(request, 'historique.html', {'notations': notations, 'roles_exclus': roles_exclus})


def test(request):
    return render(request, 'test.html')


@login_required
def register(request):
    roles_exclus = ["Chargé Client"]
    current_user = request.user

    # ðŸ”’ Vérification des droits d'accÃ¨s
    if current_user.organe not in ["PASS", "DSI"]:
        messages.error(request, "Vous nâ€™avez pas la permission de créer un compte utilisateur.")
        return redirect('user_list')

    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            new_user = form.save(commit=False)

            # Si l'utilisateur est DSI â†’ forcer la filiale du nouveau compte
            if current_user.organe == "DSI":
                new_user.filiale = current_user.filiale

            new_user.save()
            messages.success(request, "Utilisateur créé avec succÃ¨s.")
            return redirect('user_list')
    else:
        form = CustomUserCreationForm(current_user=current_user)  # ðŸ‘ˆ On passe lâ€™utilisateur connecté au formulaire

    return render(request, 'register.html', {'form': form, 'roles_exclus': roles_exclus})


# Fonction pour vérifier si l'utilisateur appartient Ã  l'organe "PASS"
def is_pass_user(user):
    return user.organe == 'PASS'


# Limiter l'accÃ¨s Ã  ceux de l'organe 'PASS'


@login_required
def user_list(request):
    user = request.user
    query = request.GET.get('q', '')
    page_number = request.GET.get('page')

    # 1. Filtrage par droits organe
    if user.organe == "PASS":
        users_base = ProfileV.objects.all().order_by('last_name')
    elif user.organe == "DSI":
        users_base = ProfileV.objects.filter(filiale=user.filiale).order_by('last_name')
    else:
        users_base = ProfileV.objects.none()

    # 2. Recherche multi-critÃ¨res
    if query:
        users_base = users_base.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(username__icontains=query)
        )

    # 3. Récupération des connectés (Sessions actives)
    active_sessions = Session.objects.filter(expire_date__gte=now())
    user_ids = []
    for s in active_sessions:
        data = s.get_decoded()
        if '_auth_user_id' in data:
            user_ids.append(data['_auth_user_id'])

    connected_users = users_base.filter(id__in=list(set(user_ids)))

    # 4. Pagination (10 par page)
    paginator = Paginator(users_base, 10)
    page_obj = paginator.get_page(page_number)

    context = {
        'total_users': users_base.count(),
        'connected_count': connected_users.count(),
        'connected_users': connected_users,
        'page_obj': page_obj,
        'query': query,
    }
    return render(request, 'user_list.html', context)



@login_required
def edit_user(request, user_id):
    current_user = request.user
    target_user = get_object_or_404(ProfileV, pk=user_id)

    # ðŸ”’ RÃ¨gles dâ€™accÃ¨s
    if current_user.organe != "PASS":
        if current_user.organe == "DSI":
            if current_user.filiale != target_user.filiale:
                messages.error(request, "Vous ne pouvez modifier que les utilisateurs de votre filiale.")
                return redirect('user_list')
        else:
            messages.error(request, "Vous nâ€™avez pas la permission de modifier cet utilisateur.")
            return redirect('user_list')

    if request.method == "POST":
        form = UserEditForm(request.POST, instance=target_user, current_user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Utilisateur modifié avec succÃ¨s.")
            return redirect('user_list')
    else:
        form = UserEditForm(instance=target_user, current_user=request.user)

    return render(request, 'edit_user.html', {'form': form, 'user': target_user})


@login_required
def change_user_password(request, user_id):
    current_user = request.user
    target_user = get_object_or_404(ProfileV, pk=user_id)

    # ðŸ”’ RÃ¨gles dâ€™accÃ¨s :
    # - PASS : peut changer le mot de passe de tous les utilisateurs
    # - DSI : peut changer le mot de passe uniquement des utilisateurs de sa filiale
    # - Autres : accÃ¨s refusé
    if current_user.organe != "PASS":
        if current_user.organe == "DSI":
            if current_user.filiale != target_user.filiale:
                messages.error(request, "Vous ne pouvez changer le mot de passe que des utilisateurs de votre filiale.")
                return redirect('user_list')
        else:
            messages.error(request, "Vous nâ€™avez pas la permission de changer ce mot de passe.")
            return redirect('user_list')

    # ðŸ§¾ Traitement du formulaire
    if request.method == 'POST':
        form = PasswordChangeForm(target_user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # éviter la déconnexion
            messages.success(request, "Le mot de passe a été modifié avec succÃ¨s.")
            return redirect('user_list')
    else:
        form = PasswordChangeForm(target_user)

    return render(request, 'change_user_password.html', {'form': form, 'user': target_user})


@login_required
def reset_user_password(request, user_id):
    current_user = request.user
    target_user = get_object_or_404(ProfileV, pk=user_id)

    # ðŸ”’ RÃ¨gles dâ€™accÃ¨s :
    if current_user.organe != "PASS":
        if current_user.organe == "DSI":
            if current_user.filiale != target_user.filiale:
                messages.error(request,
                               "Vous ne pouvez réinitialiser que les mots de passe des utilisateurs de votre filiale.")
                return redirect('user_list')
        else:
            messages.error(request, "Vous nâ€™avez pas la permission de réinitialiser ce mot de passe.")
            return redirect('user_list')

    # ðŸ§¾ Traitement du formulaire
    if request.method == 'POST':
        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            target_user.password = make_password(new_password)
            target_user.force_password_change = form.cleaned_data.get('force_password_change', False)
            target_user.save()
            messages.success(request, "Le mot de passe a été réinitialisé avec succÃ¨s.")
            return redirect('user_list')
    else:
        form = ResetPasswordForm(initial={
            'force_password_change': target_user.force_password_change
        })

    return render(request, 'reset_user_password.html', {'form': form, 'target_user': target_user})


@login_required
def user_statistics_view(request):
    roles_exclus = ["Chargé Client"]
    current_user = request.user  # utilisateur connecté

    # ðŸ”’ RÃ¨gles dâ€™accÃ¨s selon lâ€™organe
    if current_user.organe == "PASS":
        users = ProfileV.objects.all()

    elif current_user.organe == "DSI":
        users = ProfileV.objects.filter(filiale=current_user.filiale)

    else:
        messages.error(request, "Vous nâ€™avez pas la permission dâ€™accéder Ã  cette page.")
        return render(request, 'user_statistics.html', {
            'total_users': 0,
            'connected_count': 0,
            'connected_users': [],
            'users': [],
            'roles_exclus': roles_exclus,
            'connection_history_labels': json.dumps([]),
            'connection_history_values': json.dumps([]),
            'connection_history_rows': [],
            'history_days': 0,
        })

    # Nombre total d'utilisateurs visibles
    total_users = users.count()

    # ðŸ”„ Sessions actives
    active_sessions = Session.objects.filter(expire_date__gte=now())
    user_ids = []
    for session in active_sessions:
        data = session.get_decoded()
        if '_auth_user_id' in data:
            user_ids.append(data['_auth_user_id'])

    # Supprimer les doublons
    user_ids = list(set(user_ids))

    # ðŸ‘¥ Utilisateurs connectés visibles par le user connecté
    connected_users = users.filter(id__in=user_ids)
    connected_count = connected_users.count()

    visible_login_events = UserLoginHistory.objects.filter(user_id__in=users.values("id"))
    first_login_at = visible_login_events.order_by("login_at").values_list("login_at", flat=True).first()

    end_date = timezone.localdate()
    if first_login_at:
        if timezone.is_aware(first_login_at):
            start_date = timezone.localtime(first_login_at).date()
        else:
            start_date = first_login_at.date()
    else:
        start_date = end_date - timedelta(days=6)

    daily_connections = (
        visible_login_events
        .filter(login_at__date__range=(start_date, end_date))
        .annotate(day=TruncDate("login_at"))
        .values("day")
        .annotate(count=Count("user_id", distinct=True))
        .order_by("day")
    )
    daily_connections_map = {row["day"]: row["count"] for row in daily_connections}

    cursor = start_date
    chart_labels = []
    chart_values = []
    history_rows = []

    while cursor <= end_date:
        count = daily_connections_map.get(cursor, 0)
        chart_labels.append(cursor.strftime("%d/%m"))
        chart_values.append(count)
        history_rows.append(
            {
                "date": cursor.strftime("%d/%m/%Y"),
                "count": count,
            }
        )
        cursor += timedelta(days=1)

    # Contexte Ã  envoyer au template
    context = {
        'total_users': total_users,
        'connected_count': connected_count,
        'connected_users': connected_users,
        'users': users,
        'roles_exclus': roles_exclus,
        'connection_history_labels': json.dumps(chart_labels),
        'connection_history_values': json.dumps(chart_values),
        'connection_history_rows': list(reversed(history_rows)),
        'history_days': len(history_rows),
    }

    return render(request, 'user_statistics.html', context)


@login_required
def ppe(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')
    filiale_txt = (request.GET.get('col_filiale') or request.GET.get('filiale_txt', '')).strip()
    agence_txt = (request.GET.get('col_agence') or request.GET.get('agence_txt', '')).strip()
    lib_agence = (request.GET.get('col_lib_agence') or request.GET.get('lib_agence', '')).strip()
    expl_txt = (request.GET.get('col_expl') or request.GET.get('expl_txt', '')).strip()
    client_txt = (request.GET.get('col_client') or request.GET.get('client', '')).strip()
    risque_txt = (request.GET.get('col_risque') or request.GET.get('risque', '')).strip()

    donnees = Kyc_pp.objects.filter(PPE__icontains="O")

    # Filtrage automatique selon le rôle
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    if user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    if user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    if user.organe in users_groupe:
        donnees = donnees

    # Filtres manuels via GET
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    if filiale_txt:
        donnees = donnees.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees = donnees.filter(AGENCE__icontains=agence_txt)
    if expl_txt:
        donnees = donnees.filter(EXPL__icontains=expl_txt)
    if client_txt:
        donnees = donnees.filter(CLIENT__icontains=client_txt)
    if risque_txt:
        donnees = donnees.filter(RISQUE__icontains=risque_txt)
    if filiale_txt:
        donnees = donnees.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees = donnees.filter(AGENCE__icontains=agence_txt)
    if lib_agence:
        donnees = donnees.filter(LIB_AGENCE__icontains=lib_agence)
    if expl_txt:
        donnees = donnees.filter(EXPL__icontains=expl_txt)
    if client_txt:
        donnees = donnees.filter(CLIENT__icontains=client_txt)
    if risque_txt:
        donnees = donnees.filter(RISQUE__icontains=risque_txt)

    # === Valeurs du formulaire selon le rôle ===
    if user.organe == "Directeur Agence":
        exploitants = donnees.filter(AGENCE=user.agence).values_list('EXPL', flat=True).distinct()
        agences = donnees.filter(AGENCE=user.agence).values_list('AGENCE', flat=True).distinct()
    elif user.organe in users_filiale:
        agences = donnees.filter(FILIALE=user.filiale).values_list('AGENCE', flat=True).distinct()
        exploitants = donnees.filter(FILIALE=user.filiale).values_list('EXPL', flat=True).distinct()
    else:
        exploitants = donnees.values_list('EXPL', flat=True).distinct()
        agences = donnees.values_list('AGENCE', flat=True).distinct()

    filiales = donnees.values_list('FILIALE', flat=True).distinct()

    # PPE KPIs
    total_ppe = donnees.count()
    missing_ppe = get_incomplete_clients_queryset(donnees, 'pp').count()
    complete_ppe = max(0, total_ppe - missing_ppe)
    compliance_rate = round((complete_ppe / total_ppe) * 100, 1) if total_ppe > 0 else 100.0

    # Répartition par classe de risque
    risque_counts = list(
        donnees.values('RISQUE')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    risque_items = []
    for rc in risque_counts:
        label = rc['RISQUE'] or "Non renseigné"
        pct = round((rc['count'] / total_ppe) * 100, 1) if total_ppe > 0 else 0.0
        risque_items.append({
            'label': label,
            'value': rc['count'],
            'suffix': f'({pct}%)'
        })

    export_params = request.GET.copy()
    export_params['incompletes'] = '1'
    export_incomplets_url = f"{reverse('export_ppe')}?{export_params.urlencode()}"

    kpi_cards = [
        {
            'tone': 'emerald',
            'label': 'Répartition par Risque',
            'value': total_ppe,
            'subtitle': 'Clients PPE par classe de risque',
            'show_modal': True,
            'items': risque_items
        },
        {
            'tone': 'red',
            'label': 'PPE Incomplets',
            'value': missing_ppe,
            'subtitle': 'Dossiers avec données manquantes',
            'export_url': export_incomplets_url
        },
        {
            'tone': 'blue',
            'label': 'Taux de conformité',
            'value': compliance_rate,
            'suffix': '%',
            'subtitle': 'Dossiers complets / total',
        }
    ]

    context = {
        'donnees': donnees,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'kpi_cards': kpi_cards,
    }

    return render(request, 'ppe.html', context)



@login_required
def export_ppe(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # Récupérer les filtres GET
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')
    filiale_txt = request.GET.get('filiale_txt', '').strip()
    agence_txt = request.GET.get('agence_txt', '').strip()
    lib_agence = request.GET.get('lib_agence', '').strip()
    expl_txt = request.GET.get('expl_txt', '').strip()
    client_txt = request.GET.get('client', '').strip()
    risque_txt = request.GET.get('risque', '').strip()

    # Base queryset : PPE = "O"
    donnees = Kyc_pp.objects.filter(PPE__icontains="O")

    incompletes_only = request.GET.get('incompletes', '') == '1'
    if incompletes_only:
        donnees = get_incomplete_clients_queryset(donnees, 'pp')

    # Filtrage selon le rôle utilisateur
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    # else: si utilisateur avec rôle â€œgroupeâ€ ou autre --> pas de filtre rôle spécifique

    # Appliquer les filtres GET sâ€™ils sont fournis
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)

    # Création du classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Non rens PPE"

    headers = [
        "FILIALE", "AGENCE", "EXPL", "CLIENT", "CODAPE", "IDP",
        "PAYNAIS", "PROFESSION", "ADRESSE", "PAYS_RESID", "NUMID",
        "SALAIRE", "ORIGINE_REV", "DATVALID", "TEL"
    ]
    ws.append(headers)

    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.CODAPE, d.IDP,
            d.PAYNAIS, d.PROFESSION, d.ADRESSE, d.PAYS_RESID,
            d.NUMID, d.SALAIRE, d.ORIGINE_REV, d.DATVALID, d.TEL
        ])

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    if incompletes_only:
        filename = f"PPE_incomplets_{date_str}.xlsx"
    else:
        filename = f"PPE_non_rens_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def _apply_pp_header_filters(queryset, request, include_pays_resid=False, include_devise=False):
    field_map = {
        "lib_agence": "LIB_AGENCE",
        "client": "CLIENT",
        "idp": "IDP",
        "numid": "NUMID",
        "datnais": "DATNAIS",
        "paynais": "PAYNAIS",
        "adresse": "ADRESSE",
        "codape": "CODAPE",
        "profession": "PROFESSION",
        "salaire": "SALAIRE",
        "origine_rev": "ORIGINE_REV",
        "datvalid": "DATVALID",
        "tel": "TEL",
        "datouv": "DATOUV",
        "agence": "AGENCE",
        "expl": "EXPL",
    }
    if include_pays_resid:
        field_map["pays_resid"] = "PAYS_RESID"
    if include_devise:
        field_map["devise"] = "DEVISE"

    for param, field in field_map.items():
        value = (request.GET.get(f"col_{param}") or request.GET.get(param, "")).strip()
        if value:
            queryset = queryset.filter(**{f"{field}__icontains": value})
    return queryset


@login_required
def non_resid(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    # ðŸŒŸ CORRECTION / AMÃ‰LIORATION ðŸŒŸ
    # Utilisation de __exact pour filtrer strictement les non-résidents ('N')
    # Utilisez __icontains="N" si le champ peut contenir d'autres informations et que "N" suffit.
    donnees = Kyc_pp.objects.filter(RESID__icontains="N")
    # Si vous voulez l'ancienne logique avec moins de sensibilité Ã  la casse :
    # donnees = Kyc_pp.objects.filter(RESID__iexact="N")

    # === Filtrage automatique selon le rôle ===
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        # L'instruction "donnees = donnees" est inutile, on peut laisser le pass ou ne rien faire
        pass

    # === Filtres manuels via GET ===
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    donnees = _apply_pp_header_filters(donnees, request, include_pays_resid=True)

    # === Calcul des valeurs du formulaire (Listes de filtres) ===
    # On calcule les listes sur le QuerySet filtré par le rôle

    # Par défaut (pour les users_groupe ou si aucun rôle spécifique n'est atteint)
    exploitants = donnees.values_list('EXPL', flat=True).distinct()
    agences = donnees.values_list('AGENCE', flat=True).distinct()

    # Remplacer par des filtres plus stricts si l'utilisateur a un rôle restreint
    if user.organe == "Directeur Agence":
        # Les filtres par filiale et agence ont déjÃ  été appliqués ci-dessus, on peut utiliser donnees
        agences = donnees.values_list('AGENCE', flat=True).distinct()
        exploitants = donnees.values_list('EXPL', flat=True).distinct()
    elif user.organe in users_filiale:
        # Le filtre par filiale a déjÃ  été appliqué ci-dessus, on peut utiliser donnees
        agences = donnees.values_list('AGENCE', flat=True).distinct()
        exploitants = donnees.values_list('EXPL', flat=True).distinct()

    filiales = donnees.values_list('FILIALE', flat=True).distinct()
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

    # 2. Obtenir le QuerySet final pour la pagination
    # On utilise 'donnees' qui est le QuerySet filtré
    queryset = donnees.order_by('id')

    # Simulation de la variable
    ITEMS_PER_PAGE = 25

    # 3. Appliquer le Paginator
    paginator = Paginator(queryset, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')

    try:
        objets_page = paginator.page(page_number)
    except PageNotAnInteger:
        # Si 'page' n'est pas un entier, afficher la premiÃ¨re page
        objets_page = paginator.page(1)
    except EmptyPage:
        # Si la page est hors limites, afficher la derniÃ¨re page
        objets_page = paginator.page(paginator.num_pages)

    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

    # Non Résident PP KPIs
    total_non_resid = donnees.count()
    missing_non_resid = get_incomplete_clients_queryset(donnees, 'pp').count()
    complete_non_resid = max(0, total_non_resid - missing_non_resid)
    compliance_rate = round((complete_non_resid / total_non_resid) * 100, 1) if total_non_resid > 0 else 100.0

    # Répartition par pays
    country_counts = list(
        donnees.values('PAYS_RESID')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    country_items = []
    for cc in country_counts:
        country_code = cc['PAYS_RESID'] or "Non renseigné"
        pct = round((cc['count'] / total_non_resid) * 100, 1) if total_non_resid > 0 else 0.0
        country_items.append({
            'label': country_code,
            'value': cc['count'],
            'suffix': f'({pct}%)'
        })

    export_params = request.GET.copy()
    export_params['incompletes'] = '1'
    export_incomplets_url = f"{reverse('export_non_resid_pp')}?{export_params.urlencode()}"

    kpi_cards = [
        {
            'tone': 'emerald',
            'label': 'Répartition par Pays',
            'value': total_non_resid,
            'subtitle': 'Clients non résidents PP',
            'show_modal': True,
            'items': country_items
        },
        {
            'tone': 'red',
            'label': 'PP Incomplets',
            'value': missing_non_resid,
            'subtitle': 'Dossiers avec données manquantes',
            'export_url': export_incomplets_url,
        },
        {
            'tone': 'blue',
            'label': 'Taux de complétude',
            'value': compliance_rate,
            'suffix': '%',
            'subtitle': 'Dossiers complets / total',
        }
    ]

    context = {
        # 'donnees' est maintenant l'objet Page paginé
        "donnees": objets_page,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': query_params.urlencode(),
        'kpi_cards': kpi_cards,
    }
    return render(request, 'non_resid.html', context)

@login_required
def export_non_resid_pp(request):
        user = request.user

        users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
        users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                        "Contrôle Permanent Groupe", "PASS", "GUEST"]

        # Récupération des filtres GET pour la synchronisation
        filiale_param = request.GET.get('filiale', '')
        agence_param = request.GET.get('agence', '')
        expl_param = request.GET.get('expl', '')

        # Début du Queryset
        donnees = Kyc_pp.objects.filter(RESID__icontains="N")


        # === Filtrage automatique selon le rôle (identique Ã  devise) ===
        if user.organe == "Chargé Client":
            donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
        elif user.organe == "Directeur Agence":
            donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in users_filiale:
            donnees = donnees.filter(FILIALE=user.filiale)
        elif user.organe in users_groupe:
            pass  # on ne filtre pas davantage

        # === Filtres manuels via GET (synchronisation) ===
        if filiale_param:
            donnees = donnees.filter(FILIALE__icontains=filiale_param)
        if agence_param:
            donnees = donnees.filter(AGENCE__icontains=agence_param)
        if expl_param:
            donnees = donnees.filter(EXPL__icontains=expl_param)
        donnees = _apply_pp_header_filters(donnees, request, include_pays_resid=True)

        if request.GET.get('incompletes') == '1':
            donnees = get_incomplete_clients_queryset(donnees, 'pp')

        # Fin du Queryset filtré

        # Création du classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comptes Devise PP"  # J'ai renommé le titre

        # EntÃªtes
        headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "CODAPE", "IDP", "PAYNAIS", "PROFESSION", "ADRESSE", "PAYS_RESID",
                   "NUMID", "SALAIRE", "ORIGINE_REV", "DATVALID", "TEL"]
        ws.append(headers)

        # Données
        for d in donnees:
            ws.append([
                d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.CODAPE, d.IDP,
                d.PAYNAIS, d.PROFESSION, d.ADRESSE, d.PAYS_RESID, d.NUMID, d.SALAIRE, d.ORIGINE_REV, d.DATVALID, d.TEL
            ])

        # Ajustement largeur des colonnes (optionnel)
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15

        # Préparer la réponse HTTP
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        if request.GET.get('incompletes') == '1':
            filename = f"Comptes_non_resid_PP_incomplets_{date_str}.xlsx"
        else:
            filename = f"Comptes_non_resid_PP_{date_str}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@login_required

def non_resid_pm(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    # ðŸŒŸ CORRECTION / AMÃ‰LIORATION ðŸŒŸ
    # Utilisation de __exact pour filtrer strictement les non-résidents ('N')
    # Utilisez __icontains="N" si le champ peut contenir d'autres informations et que "N" suffit.
    donnees = Kyc_pm.objects.filter(RESID__exact="N")
    # Si vous voulez l'ancienne logique avec moins de sensibilité Ã  la casse :
    # donnees = Kyc_pp.objects.filter(RESID__iexact="N")

    # === Filtrage automatique selon le rôle ===
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        # L'instruction "donnees = donnees" est inutile, on peut laisser le pass ou ne rien faire
        pass

    # === Filtres manuels via GET ===
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)

    # === Calcul des valeurs du formulaire (Listes de filtres) ===
    # On calcule les listes sur le QuerySet filtré par le rôle

    # Par défaut (pour les users_groupe ou si aucun rôle spécifique n'est atteint)
    exploitants = donnees.values_list('EXPL', flat=True).distinct()
    agences = donnees.values_list('AGENCE', flat=True).distinct()

    # Remplacer par des filtres plus stricts si l'utilisateur a un rôle restreint
    if user.organe == "Directeur Agence":
        # Les filtres par filiale et agence ont déjÃ  été appliqués ci-dessus, on peut utiliser donnees
        agences = donnees.values_list('AGENCE', flat=True).distinct()
        exploitants = donnees.values_list('EXPL', flat=True).distinct()
    elif user.organe in users_filiale:
        # Le filtre par filiale a déjÃ  été appliqué ci-dessus, on peut utiliser donnees
        agences = donnees.values_list('AGENCE', flat=True).distinct()
        exploitants = donnees.values_list('EXPL', flat=True).distinct()

    filiales = donnees.values_list('FILIALE', flat=True).distinct()

    # 2. Obtenir le QuerySet final pour la pagination
    # On utilise 'donnees' qui est le QuerySet filtré
    queryset = donnees.order_by('id')

    # Simulation de la variable
    ITEMS_PER_PAGE = 25

    # 3. Appliquer le Paginator
    paginator = Paginator(queryset, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

    try:
        objets_page = paginator.page(page_number)
    except PageNotAnInteger:
        # Si 'page' n'est pas un entier, afficher la premiÃ¨re page
        objets_page = paginator.page(1)
    except EmptyPage:
        # Si la page est hors limites, afficher la derniÃ¨re page
        objets_page = paginator.page(paginator.num_pages)

    # Non Résident PM KPIs
    total_non_resid = donnees.count()
    missing_non_resid = get_incomplete_clients_queryset(donnees, 'pm').count()
    complete_non_resid = max(0, total_non_resid - missing_non_resid)
    compliance_rate = round((complete_non_resid / total_non_resid) * 100, 1) if total_non_resid > 0 else 100.0

    # Répartition par pays
    country_counts = list(
        donnees.values('PAYS_JUR')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    country_items = []
    for cc in country_counts:
        country_code = cc['PAYS_JUR'] or "Non renseigné"
        pct = round((cc['count'] / total_non_resid) * 100, 1) if total_non_resid > 0 else 0.0
        country_items.append({
            'label': country_code,
            'value': cc['count'],
            'suffix': f'({pct}%)'
        })

    export_params = request.GET.copy()
    export_params['incompletes'] = '1'
    export_incomplets_url = f"{reverse('export_non_resid_pm')}?{export_params.urlencode()}"

    kpi_cards = [
        {
            'tone': 'emerald',
            'label': 'Répartition par Pays',
            'value': total_non_resid,
            'subtitle': 'Clients non résidents PM',
            'show_modal': True,
            'items': country_items
        },
        {
            'tone': 'red',
            'label': 'PM Incomplets',
            'value': missing_non_resid,
            'subtitle': 'Dossiers avec données manquantes',
            'export_url': export_incomplets_url,
        },
        {
            'tone': 'blue',
            'label': 'Taux de complétude',
            'value': compliance_rate,
            'suffix': '%',
            'subtitle': 'Dossiers complets / total',
        }
    ]

    context = {
        # 'donnees' est maintenant l'objet Page paginé
        "donnees": objets_page,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': query_params.urlencode(),
        'kpi_cards': kpi_cards,
    }
    return render(request, 'non_resid_pm.html', context)


@login_required
def export_non_resid_pm(request):
    user = request.user

    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # Récupération des filtres GET pour la synchronisation
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    # Début du Queryset
    donnees = Kyc_pm.objects.filter(RESID__icontains="N")

    # === Filtrage automatique selon le rôle (identique Ã  devise) ===
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        pass  # on ne filtre pas davantage

    # === Filtres manuels via GET (synchronisation) ===
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)

    if request.GET.get('incompletes') == '1':
        donnees = get_incomplete_clients_queryset(donnees, 'pm')

    # Fin du Queryset filtré

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comptes Devise PP"  # J'ai renommé le titre

    # EntÃªtes
    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "AGEC", "CODAPE", "IDM", "RCSNO", "CAPITAL", "CA",
               "RESULTAT", "TEL"]
    ws.append(headers)

    # Données
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.AGEC, d.CODAPE, d.IDM,
            d.RCSNO, d.CAPITAL, d.CA, d.RESULTAT, d.TEL
        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    if request.GET.get('incompletes') == '1':
        filename = f"Comptes_non_resid_PM_incomplets_{date_str}.xlsx"
    else:
        filename = f"Comptes_non_resid_PM_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def scoring(request):
    # 1. Définition des rôles
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau", 'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]

    user = request.user
    today = date.today()

    # 2. Récupération des paramÃ¨tres
    periode_param = request.GET.get("periode", "")
    filiale_param = request.GET.get("filiale", "")
    agence_param = request.GET.get("agence", "")
    expl_param = request.GET.get("expl", "")
    filiale_txt = request.GET.get("filiale_txt", "").strip()
    agence_txt = request.GET.get("agence_txt", "").strip()
    lib_agence = request.GET.get("lib_agence", "").strip()
    expl_txt = request.GET.get("expl_txt", "").strip()
    client_txt = request.GET.get("client", "").strip()
    daterev_txt = request.GET.get("daterev", "").strip()

    # On commence par TOUT (on retire le filtre isnull pour tester)
    qs = DATEREV.objects.all()

    # 3. Filtrage par Rôle (Sécurité)
    organe = getattr(user, "organe", "")
    is_group_user = (organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
    if not is_group_user:
        if organe == "Chargé Client":
            qs = qs.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
        elif organe == "Directeur Agence":
            qs = qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif organe in users_filiale:
            qs = qs.filter(FILIALE=user.filiale)
    # Si users_groupe, on ne filtre pas initialement

    # 4. Filtrage par Période
    if periode_param == "today":
        qs = qs.filter(DATEREV__lte=today)
    elif periode_param == "3m":
        qs = qs.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=90))
    elif periode_param == "6m":
        qs = qs.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=180))
    elif periode_param == "1y":
        qs = qs.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=365))

    # 5. Filtres dynamiques de l'interface
    if filiale_param:
        qs = qs.filter(FILIALE=filiale_param)

    if agence_param:
        qs = qs.filter(AGENCE=agence_param)

    if expl_param:
        qs = qs.filter(EXPL=expl_param)
    if filiale_txt:
        qs = qs.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        qs = qs.filter(AGENCE__icontains=agence_txt)
    if lib_agence:
        qs = qs.filter(LIB_AGENCE__icontains=lib_agence)
    if expl_txt:
        qs = qs.filter(EXPL__icontains=expl_txt)
    if client_txt:
        qs = qs.filter(CLIENT__icontains=client_txt)
    if daterev_txt:
        daterev_txt = daterev_txt.strip()
        parsed = parse_date(daterev_txt)
        if not parsed:
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    parsed = datetime.strptime(daterev_txt, fmt).date()
                    break
                except (ValueError, TypeError):
                    continue
        if parsed:
            qs = qs.filter(DATEREV=parsed)

    # 6. Génération des options pour les menus déroulants (basé sur le QS filtré ou global)
    # Il est souvent préférable de baser les options sur le QS global ou par filiale
    filiales_opts = DATEREV.objects.values_list("FILIALE", flat=True).distinct().order_by("FILIALE")
    agences_opts = qs.values_list("AGENCE", flat=True).distinct().order_by("AGENCE")
    exploitants_opts = qs.values_list("EXPL", flat=True).distinct().order_by("EXPL")

    # 7. Tri et Pagination
    donnees_queryset = qs.order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")

    paginator = Paginator(donnees_queryset, 100)
    page = request.GET.get('page')
    try:
        donnees_page = paginator.page(page)
    except:
        donnees_page = paginator.page(1)

    context = {
        "donnees": donnees_page,
        "filiales": filiales_opts,
        "agences": agences_opts,
        "exploitants": exploitants_opts,
        "periode": periode_param,
        "filiale_param": filiale_param,
        "agence_param": agence_param,
        "expl_param": expl_param,
        "can_pick_filiale": is_group_user,
        "can_pick_agence": (is_group_user or organe in users_filiale or organe == "Directeur Agence"),
        "can_pick_expl": (is_group_user or organe in users_filiale or organe == "Directeur Agence"),
        "get_params": request.GET.urlencode(),
    }
    return render(request, "scoring.html", context)

from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
from .models import DATEREV

@login_required
def export_csv_scoring(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # Récupérer les filtres GET comme dans la vue scoring
    periode_param = request.GET.get("periode", "")
    filiale_param = request.GET.get("filiale", "")
    agence_param = request.GET.get("agence", "")
    expl_param = request.GET.get("expl", "")
    filiale_txt = request.GET.get("filiale_txt", "").strip()
    agence_txt = request.GET.get("agence_txt", "").strip()
    lib_agence = request.GET.get("lib_agence", "").strip()
    expl_txt = request.GET.get("expl_txt", "").strip()
    client_txt = request.GET.get("client", "").strip()
    daterev_txt = request.GET.get("daterev", "").strip()
    filiale_txt = request.GET.get("filiale_txt", "").strip()
    agence_txt = request.GET.get("agence_txt", "").strip()
    lib_agence = request.GET.get("lib_agence", "").strip()
    expl_txt = request.GET.get("expl_txt", "").strip()
    client_txt = request.GET.get("client", "").strip()
    daterev_txt = request.GET.get("daterev", "").strip()

    base_qs = DATEREV.objects.filter(DATEREV__isnull=False)

    if getattr(user, "organe", "") == "Chargé Client":
        base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        base_qs = base_qs.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        # pas de filtre organe â†’ laisse tout (selon ce que tu veux)
        pass
    else:
        # par sécurité, si organe non reconnu, on vide
        base_qs = DATEREV.objects.none()

    # Appliquer le filtre période si défini
    qs_period = base_qs
    today = date.today()
    if periode_param == "today":
        qs_period = qs_period.filter(DATEREV__lte=today)
    elif periode_param == "3m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=90))
    elif periode_param == "6m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=180))
    elif periode_param == "1y":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=365))

    # Filiale
    if filiale_param:
        qs_period = qs_period.filter(FILIALE=filiale_param)

    # Agence
    if agence_param:
        qs_period = qs_period.filter(AGENCE=agence_param)

    # Exploitant
    if expl_param:
        qs_period = qs_period.filter(EXPL=expl_param)
    if filiale_txt:
        qs_period = qs_period.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        qs_period = qs_period.filter(AGENCE__icontains=agence_txt)
    if lib_agence:
        qs_period = qs_period.filter(LIB_AGENCE__icontains=lib_agence)
    if expl_txt:
        qs_period = qs_period.filter(EXPL__icontains=expl_txt)
    if client_txt:
        qs_period = qs_period.filter(CLIENT__icontains=client_txt)
    if daterev_txt:
        parsed = parse_date(daterev_txt)
        if not parsed:
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    parsed = datetime.strptime(daterev_txt, fmt).date()
                    break
                except (ValueError, TypeError):
                    continue
        if parsed:
            qs_period = qs_period.filter(DATEREV=parsed)
    if filiale_txt:
        qs_period = qs_period.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        qs_period = qs_period.filter(AGENCE__icontains=agence_txt)
    if lib_agence:
        qs_period = qs_period.filter(LIB_AGENCE__icontains=lib_agence)
    if expl_txt:
        qs_period = qs_period.filter(EXPL__icontains=expl_txt)
    if client_txt:
        qs_period = qs_period.filter(CLIENT__icontains=client_txt)
    if daterev_txt:
        parsed = parse_date(daterev_txt)
        if not parsed:
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    parsed = datetime.strptime(daterev_txt, fmt).date()
                    break
                except (ValueError, TypeError):
                    continue
        if parsed:
            qs_period = qs_period.filter(DATEREV=parsed)

    donnees = qs_period

    # Création du classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Notation_Stock"

    # EntÃªtes (vérifie que les noms sont corrects)
    headers = ['FILIALE', 'AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'RISQUE']
    ws.append(headers)

    for d in donnees:
        daterev = d.DATEREV
        if hasattr(daterev, 'tzinfo'):
            daterev = daterev.replace(tzinfo=None)
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, daterev, d.PPE, d.RISQUE
        ])

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"scoring_export_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_csv_scoring_ppe(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # Récupérer les filtres GET comme dans la vue scoring
    periode_param = request.GET.get("periode", "")
    filiale_param = request.GET.get("filiale", "")
    agence_param = request.GET.get("agence", "")
    expl_param = request.GET.get("expl", "")

    if user.organe == 'Conformité':
        base_qs = DATEREV.objects.filter(FILIALE=user.filiale, PPE__icontains="O", DATEREV__isnull=False)
    elif user.organe == "Conformité Groupe":
        base_qs = DATEREV.objects.filter(PPE__icontains="O", DATEREV__isnull=False)

    # Appliquer le filtre période si défini
    qs_period = base_qs
    today = date.today()
    if periode_param == "today":
        qs_period = qs_period.filter(DATEREV__lte=today)
    elif periode_param == "3m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=90))
    elif periode_param == "6m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=180))
    elif periode_param == "1y":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=365))

    # Filiale
    if filiale_param:
        qs_period = qs_period.filter(FILIALE=filiale_param)

    # Agence
    if agence_param:
        qs_period = qs_period.filter(AGENCE=agence_param)

    # Exploitant
    if expl_param:
        qs_period = qs_period.filter(EXPL=expl_param)

    donnees = qs_period

    # Création du classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Notation_Stock"

    # EntÃªtes (vérifie que les noms sont corrects)
    headers = ['FILIALE', 'AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'RISQUE']
    ws.append(headers)

    for d in donnees:
        daterev = d.DATEREV
        if hasattr(daterev, 'tzinfo'):
            daterev = daterev.replace(tzinfo=None)
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, daterev, d.PPE, d.RISQUE
        ])

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"scoring_PPE_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def clients_scorer(request):
    from .models import Notation, Kyc_pm, Kyc_pp
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau", 'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]

    user = request.user
    notes = Notation.objects.filter(flux_stock='Flux')
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    if user.is_authenticated and getattr(user, 'filiale', ''):
        notation = notation.filter(agent__filiale=user.filiale)
    if user.is_authenticated and getattr(user, 'code_expl', ''):
        notation = notation.filter(agent__code_expl=user.code_expl)

    periode_param = request.GET.get("periode", "")
    filiale_param = request.GET.get("filiale", "")
    agence_param = request.GET.get("agence", "")
    expl_param = request.GET.get("expl", "")
    risque_param = request.GET.get("risque", "")
    client_type = request.GET.get('type_client', 'pp')

    col_agence = request.GET.get('col_agence', '')
    col_lib_agence = request.GET.get('col_lib_agence', '')
    col_expl = request.GET.get('col_expl', '')
    col_client = request.GET.get('col_client', '')
    col_daterev = request.GET.get('col_daterev', '')
    col_ppe = request.GET.get('col_ppe', '')
    col_risque = request.GET.get('col_risque', '')

    if client_type == 'pm':
        base_qs = DATEREV.objects.filter(Exists(Kyc_pm.objects.filter(CLIENT=OuterRef('CLIENT'))))
    else:
        base_qs = DATEREV.objects.filter(Exists(Kyc_pp.objects.filter(CLIENT=OuterRef('CLIENT'))))

    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
    if not is_group_user:
        if getattr(user, "organe", "") == "Chargé Client":
            base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
        elif user.organe == "Directeur Agence":
            base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in users_filiale:
            base_qs = base_qs.filter(FILIALE=user.filiale)

    today = date.today()
    qs_period = base_qs
    if periode_param == "today":
        qs_period = qs_period.filter(DATEREV__lte=today)
    elif periode_param == "3m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=90))
    elif periode_param == "6m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=180))
    elif periode_param == "1y":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=365))
    elif periode_param == "no_date":
        qs_period = qs_period.filter(DATEREV__isnull=True)

    can_pick_filiale = is_group_user
    selected_filiale = filiale_param if can_pick_filiale else getattr(user, "filiale", "")
    _user_scope = "groupe" if is_group_user else getattr(user, 'filiale', 'filiale')
    filiales_cache_key = f"filiales_opts_{client_type}_{periode_param}_{_user_scope}"
    filiales_opts = cache.get(filiales_cache_key)
    if filiales_opts is None:
        filiales_opts = list(qs_period.values_list("FILIALE", flat=True).distinct().order_by("FILIALE"))
        cache.set(filiales_cache_key, filiales_opts, 300)

    qs_filiale = qs_period
    if selected_filiale:
        qs_filiale = qs_filiale.filter(FILIALE=selected_filiale)

    can_pick_agence = (user.organe in users_groupe) or (user.organe in users_filiale) or (user.organe == "Directeur Agence")
    selected_agence = getattr(user, "agence", "") if user.organe == "Directeur Agence" else agence_param
    agences_cache_key = f"agences_opts_{client_type}_{periode_param}_{selected_filiale}_{_user_scope}"
    agences_opts = cache.get(agences_cache_key)
    if agences_opts is None:
        agences_opts = list(qs_filiale.values_list("AGENCE", flat=True).distinct().order_by("AGENCE"))
        cache.set(agences_cache_key, agences_opts, 300)

    qs_agence = qs_filiale
    if selected_agence:
        qs_agence = qs_agence.filter(AGENCE=selected_agence)

    can_pick_expl = getattr(user, "organe", "") != "Chargé Client"
    selected_expl = getattr(user, "code_expl", "") if getattr(user, "organe", "") == "Chargé Client" else expl_param
    exploitants_cache_key = f"expl_opts_{client_type}_{periode_param}_{selected_filiale}_{selected_agence}_{_user_scope}"
    exploitants_opts = cache.get(exploitants_cache_key)
    if exploitants_opts is None:
        exploitants_opts = list(qs_agence.values_list("EXPL", flat=True).distinct().order_by("EXPL"))
        cache.set(exploitants_cache_key, exploitants_opts, 300)

    donnees_queryset = qs_agence
    if selected_expl:
        donnees_queryset = donnees_queryset.filter(EXPL=selected_expl)

    if col_agence: donnees_queryset = donnees_queryset.filter(AGENCE__icontains=col_agence)
    if col_lib_agence: donnees_queryset = donnees_queryset.filter(LIB_AGENCE__icontains=col_lib_agence)
    if col_expl: donnees_queryset = donnees_queryset.filter(EXPL__icontains=col_expl)
    if col_client: donnees_queryset = donnees_queryset.filter(CLIENT__icontains=col_client)
    if col_daterev: donnees_queryset = donnees_queryset.filter(DATEREV__icontains=col_daterev)
    if col_ppe: donnees_queryset = donnees_queryset.filter(PPE__icontains=col_ppe)
    if col_risque: donnees_queryset = donnees_queryset.filter(RISQUE__icontains=col_risque)

    agg = donnees_queryset.aggregate(
        scored=Count(Case(When(~Q(RISQUE="") & ~Q(RISQUE__isnull=True), then=1), output_field=IntegerField())),
        unscored=Count(Case(When(Q(RISQUE="") | Q(RISQUE__isnull=True), then=1), output_field=IntegerField())),
        overdue_unscored=Count(Case(When((Q(RISQUE="") | Q(RISQUE__isnull=True)) & Q(DATEREV__lte=today), then=1), output_field=IntegerField()))
    )
    scorer_scored_count = agg['scored'] or 0
    scorer_unscored_count = agg['unscored'] or 0
    total_scorer = scorer_scored_count + scorer_unscored_count
    scoring_rate = (scorer_scored_count / total_scorer * 100) if total_scorer > 0 else 0.0
    if scorer_unscored_count > 0 and scoring_rate >= 99.9:
        scoring_rate = 99.9
    overdue_unscored_count = agg['overdue_unscored'] or 0

    risk_options = cache.get("risk_options_opts")
    if risk_options is None:
        risk_options = sorted(list(set(DATEREV.objects.exclude(RISQUE="").exclude(RISQUE__isnull=True).values_list('RISQUE', flat=True).distinct())))
        cache.set("risk_options_opts", risk_options, 300)

    if risque_param:
        if risque_param == "sans_classe":
            donnees_queryset = donnees_queryset.filter(Q(RISQUE="") | Q(RISQUE__isnull=True))
        else:
            donnees_queryset = donnees_queryset.filter(RISQUE=risque_param)

    donnees_queryset = donnees_queryset.values("FILIALE", "AGENCE", "LIB_AGENCE", "EXPL", "CLIENT", "DATEREV", "PPE", "RISQUE").order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")

    show_non_scored_modal = request.GET.get('show_non_scored_modal') == '1'
    is_overdue = request.GET.get('overdue') == '1'
    non_scored_page = None
    if show_non_scored_modal:
        non_scored_qs = qs_agence
        if selected_expl: non_scored_qs = non_scored_qs.filter(EXPL=selected_expl)
        if col_agence: non_scored_qs = non_scored_qs.filter(AGENCE__icontains=col_agence)
        if col_lib_agence: non_scored_qs = non_scored_qs.filter(LIB_AGENCE__icontains=col_lib_agence)
        if col_expl: non_scored_qs = non_scored_qs.filter(EXPL__icontains=col_expl)
        if col_client: non_scored_qs = non_scored_qs.filter(CLIENT__icontains=col_client)
        if col_daterev: non_scored_qs = non_scored_qs.filter(DATEREV__icontains=col_daterev)
        if col_ppe: non_scored_qs = non_scored_qs.filter(PPE__icontains=col_ppe)
        non_scored_qs = non_scored_qs.filter(Q(RISQUE="") | Q(RISQUE__isnull=True))
        if is_overdue:
            non_scored_qs = non_scored_qs.filter(DATEREV__lte=today)
        non_scored_qs = non_scored_qs.order_by("CLIENT")
        ns_paginator = CachedPaginator(non_scored_qs, 50)
        try: non_scored_page = ns_paginator.page(request.GET.get('non_scored_page', 1))
        except (PageNotAnInteger, EmptyPage): non_scored_page = ns_paginator.page(1)

    paginator = CachedPaginator(donnees_queryset, 100)
    try: donnees_page = paginator.page(request.GET.get('page'))
    except PageNotAnInteger: donnees_page = paginator.page(1)
    except EmptyPage: donnees_page = paginator.page(paginator.num_pages)

    current_get = request.GET.copy()
    current_get.pop('page', None)
    get_params = current_get.urlencode()
    
    close_get = current_get.copy()
    close_get.pop('show_non_scored_modal', None)
    close_get.pop('non_scored_page', None)
    non_scored_close_params = close_get.urlencode()
    
    modal_get = close_get.copy()
    modal_get['show_non_scored_modal'] = '1'
    non_scored_modal_params = modal_get.urlencode()
    
    overdue_modal_get = modal_get.copy()
    overdue_modal_get['overdue'] = '1'
    overdue_non_scored_modal_params = overdue_modal_get.urlencode()
    
    export_get = close_get.copy()
    export_get['export_unscored'] = '1'
    non_scored_export_params = export_get.urlencode()
    
    pp_nav = current_get.copy()
    pp_nav['type_client'] = 'pp'
    pp_nav_params = pp_nav.urlencode()
    
    pm_nav = current_get.copy()
    pm_nav['type_client'] = 'pm'
    pm_nav_params = pm_nav.urlencode()
    
    reset_params = f"type_client={client_type}"

    context = {
        "donnees": donnees_page, "filiales": filiales_opts, "agences": agences_opts, "exploitants": exploitants_opts,
        "notation": notation, "periode": periode_param, "filiale_param": selected_filiale, "agence_param": selected_agence,
        "expl_param": selected_expl, "risque_param": risque_param, "client_type": client_type, "risk_options": risk_options,
        "scorer_scored_count": scorer_scored_count, "scorer_unscored_count": scorer_unscored_count, "scoring_rate": scoring_rate,
        "overdue_unscored_count": overdue_unscored_count, "show_non_scored_modal": show_non_scored_modal,
        "non_scored_page": non_scored_page, "users_groupe": users_groupe, "users_filiale": users_filiale,
        "can_pick_filiale": can_pick_filiale, "can_pick_agence": can_pick_agence, "can_pick_expl": can_pick_expl,
        "get_params": get_params, "reset_params": reset_params, "non_scored_modal_params": non_scored_modal_params,
        "overdue_non_scored_modal_params": overdue_non_scored_modal_params, "is_overdue_modal": is_overdue,
        "non_scored_close_params": non_scored_close_params, "non_scored_export_params": non_scored_export_params,
        "pp_nav_params": pp_nav_params, "pm_nav_params": pm_nav_params
    }
    return render(request, "clients_scorer.html", context)

def export_csv_scoring_clients(request):
    from .models import Kyc_pm, Kyc_pp
    user = request.user
    periode_param, filiale_param, agence_param, expl_param = request.GET.get("periode", ""), request.GET.get("filiale", ""), request.GET.get("agence", ""), request.GET.get("expl", "")
    risque_param, client_type, export_unscored = request.GET.get("risque", ""), request.GET.get('type_client', 'pp'), request.GET.get('export_unscored') == '1'
    col_agence, col_lib_agence, col_expl, col_client, col_daterev, col_ppe, col_risque = request.GET.get('col_agence', ''), request.GET.get('col_lib_agence', ''), request.GET.get('col_expl', ''), request.GET.get('col_client', ''), request.GET.get('col_daterev', ''), request.GET.get('col_ppe', ''), request.GET.get('col_risque', '')

    if client_type == 'pm': base_qs = DATEREV.objects.filter(Exists(Kyc_pm.objects.filter(CLIENT=OuterRef('CLIENT'))))
    else: base_qs = DATEREV.objects.filter(Exists(Kyc_pp.objects.filter(CLIENT=OuterRef('CLIENT'))))

    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]
    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
    if not is_group_user:
        if getattr(user, "organe", "") == "Chargé Client": base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
        elif user.organe == "Directeur Agence": base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau", 'Risques', 'DAI', 'Qualité']: base_qs = base_qs.filter(FILIALE=user.filiale)

    today = date.today()
    if periode_param == "today": base_qs = base_qs.filter(DATEREV__lte=today)
    elif periode_param == "3m": base_qs = base_qs.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=90))
    elif periode_param == "6m": base_qs = base_qs.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=180))
    elif periode_param == "1y": base_qs = base_qs.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=365))
    elif periode_param == "no_date": base_qs = base_qs.filter(DATEREV__isnull=True)

    if filiale_param: base_qs = base_qs.filter(FILIALE=filiale_param)
    if agence_param: base_qs = base_qs.filter(AGENCE=agence_param)
    if expl_param: base_qs = base_qs.filter(EXPL=expl_param)

    if export_unscored: base_qs = base_qs.filter(Q(RISQUE="") | Q(RISQUE__isnull=True))
    elif risque_param:
        if risque_param == "sans_classe": base_qs = base_qs.filter(Q(RISQUE="") | Q(RISQUE__isnull=True))
        else: base_qs = base_qs.filter(RISQUE=risque_param)

    if col_agence: base_qs = base_qs.filter(AGENCE__icontains=col_agence)
    if col_lib_agence: base_qs = base_qs.filter(LIB_AGENCE__icontains=col_lib_agence)
    if col_expl: base_qs = base_qs.filter(EXPL__icontains=col_expl)
    if col_client: base_qs = base_qs.filter(CLIENT__icontains=col_client)
    if col_daterev: base_qs = base_qs.filter(DATEREV__icontains=col_daterev)
    if col_ppe: base_qs = base_qs.filter(PPE__icontains=col_ppe)
    if col_risque: base_qs = base_qs.filter(RISQUE__icontains=col_risque)

    donnees = base_qs.values("FILIALE", "AGENCE", "EXPL", "CLIENT", "DATEREV", "PPE", "RISQUE").distinct().order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['FILIALE', 'AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'RISQUE']
    ws.append(headers)
    for d in donnees:
        daterev_val = d["DATEREV"].strftime("%Y-%m-%d") if isinstance(d["DATEREV"], date) else str(d["DATEREV"] or "")
        ws.append([d["FILIALE"], d["AGENCE"], d["EXPL"], d["CLIENT"], daterev_val, d["PPE"], d["RISQUE"]])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Revue_scoring_{datetime.now().strftime("%Y-%m-%d_%H-%M")}.xlsx"'
    return response

@login_required
def sans_classe(request):
    user = request.user

    # 1. Rôles et paramÃ¨tres
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')
    filiale_txt = request.GET.get('filiale_txt', '').strip()
    agence_txt = request.GET.get('agence_txt', '').strip()
    lib_agence = request.GET.get('lib_agence', '').strip()
    expl_txt = request.GET.get('expl_txt', '').strip()
    client_txt = request.GET.get('client', '').strip()
    risque_txt = request.GET.get('risque', '').strip()

    # 2. Filtrage de base (Sécurité par rôle + Condition CLASSE vide)
    # On commence par le filtre "CLASSE vide ou nulle"
    donnees_queryset = DATEREV.objects.filter(Q(RISQUE__isnull=True) | Q(RISQUE=""))

    # Restriction du périmÃ¨tre selon l'organe de l'utilisateur
    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
    if not is_group_user:
        if user.organe == "Chargé Client":
            donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
        elif user.organe == "Directeur Agence":
            donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in users_filiale:
            donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale)

    # 3. Filtres manuels via le formulaire (GET)
    if filiale_param:
        donnees_queryset = donnees_queryset.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees_queryset = donnees_queryset.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees_queryset = donnees_queryset.filter(EXPL__icontains=expl_param)
    if filiale_txt:
        donnees_queryset = donnees_queryset.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees_queryset = donnees_queryset.filter(AGENCE__icontains=agence_txt)
    if lib_agence:
        donnees_queryset = donnees_queryset.filter(LIB_AGENCE__icontains=lib_agence)
    if expl_txt:
        donnees_queryset = donnees_queryset.filter(EXPL__icontains=expl_txt)
    if client_txt:
        donnees_queryset = donnees_queryset.filter(CLIENT__icontains=client_txt)
    if risque_txt:
        donnees_queryset = donnees_queryset.filter(RISQUE__icontains=risque_txt)

    # Tri cohérent pour la pagination
    donnees_queryset = donnees_queryset.order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")

    # 4. Options pour les menus déroulants (respectant le périmÃ¨tre)
    options_qs = DATEREV.objects.filter(Q(RISQUE__isnull=True) | Q(RISQUE=""))
    if not is_group_user:
        if user.organe == "Chargé Client":
            options_qs = options_qs.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
        elif user.organe == "Directeur Agence":
            options_qs = options_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in users_filiale:
            options_qs = options_qs.filter(FILIALE=user.filiale)

    filiales = options_qs.values_list('FILIALE', flat=True).distinct().order_by('FILIALE')
    agences = options_qs.values_list('AGENCE', flat=True).distinct().order_by('AGENCE')
    exploitants = options_qs.values_list('EXPL', flat=True).distinct().order_by('EXPL')

    # 5. Logique de pagination
    get_params = request.GET.copy()
    if 'page' in get_params:
        del get_params['page']

    paginator = Paginator(donnees_queryset, 100)
    page = request.GET.get('page')

    try:
        donnees_page = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        donnees_page = paginator.page(1)

    context = {
        'donnees': donnees_page,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': get_params.urlencode(),
        'filiale_param': filiale_param,
        'agence_param': agence_param,
        'expl_param': expl_param,
    }

    return render(request, 'sans_classe.html', context)


@login_required
def export_sans_classe(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # Récupérer les filtres GET
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    # Base queryset â€” uniquement ceux avec un RISQUE non nul
    donnees = DATEREV.objects.filter(Q(RISQUE__isnull=True) | Q(RISQUE=""))

    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
    # Filtrage selon le rôle
    if is_group_user:
        pass
    elif user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    else:
        # Si organe non reconnu ou pas autorisé — optionnel : renvoyer vide
        donnees = DATEREV.objects.none()

    # Appliquer les filtres GET si présents
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    if filiale_txt:
        donnees = donnees.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees = donnees.filter(AGENCE__icontains=agence_txt)
    if lib_agence:
        donnees = donnees.filter(LIB_AGENCE__icontains=lib_agence)
    if expl_txt:
        donnees = donnees.filter(EXPL__icontains=expl_txt)
    if client_txt:
        donnees = donnees.filter(CLIENT__icontains=client_txt)
    if risque_txt:
        donnees = donnees.filter(RISQUE__icontains=risque_txt)

    # Création du fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Sans_Classe_Export"

    headers = ['FILIALE', 'AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'RISQUE']
    ws.append(headers)

    for d in donnees:
        daterev = d.DATEREV
        if hasattr(daterev, 'tzinfo'):
            daterev = daterev.replace(tzinfo=None)
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, daterev, d.PPE, d.RISQUE
        ])

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Sans_Classe_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


from django.db.models import Q, Max


def sans_classe_s(request):
    user = request.user

    # 1. Rôles et paramÃ¨tres
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    # 2. Gestion des Notations (DerniÃ¨re note par agent selon périmÃ¨tre)
    notes = Notation.objects.filter(flux_stock='Flux')
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])

    # Sécurisation de l'affichage des notes
    if user.organe == "Chargé Client":
        notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)
    elif user.organe == "Directeur Agence":
        notation = notation.filter(agent__filiale=user.filiale, agent__agence=user.agence)
    elif user.organe in users_filiale:
        notation = notation.filter(agent__filiale=user.filiale)
    # Pour le Groupe, on voit toutes les notes (ou filtrer selon besoin)

    # 3. Filtrage du QuerySet Principal (Uniquement CLASSE non vide)
    # Correction : On exclut les vides ET les nuls
    donnees_queryset = DATEREV.objects.filter(Q(RISQUE__isnull=True) | Q(RISQUE=""))

    # Filtrage automatique selon le rôle (Sécurité)
    if user.organe == "Chargé Client":
        donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale)

    # Filtres manuels via le formulaire (GET)
    if filiale_param:
        donnees_queryset = donnees_queryset.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees_queryset = donnees_queryset.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees_queryset = donnees_queryset.filter(EXPL__icontains=expl_param)
    if filiale_txt:
        donnees_queryset = donnees_queryset.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees_queryset = donnees_queryset.filter(AGENCE__icontains=agence_txt)
    if expl_txt:
        donnees_queryset = donnees_queryset.filter(EXPL__icontains=expl_txt)
    if client_txt:
        donnees_queryset = donnees_queryset.filter(CLIENT__icontains=client_txt)
    if risque_txt:
        donnees_queryset = donnees_queryset.filter(RISQUE__icontains=risque_txt)

    donnees_queryset = donnees_queryset.order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")

    # 4. Valeurs du formulaire (Options des filtres)
    options_qs = DATEREV.objects.all()
    if user.organe == "Chargé Client":
        options_qs = options_qs.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        options_qs = options_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        options_qs = options_qs.filter(FILIALE=user.filiale)

    filiales = options_qs.values_list('FILIALE', flat=True).distinct().order_by('FILIALE')
    agences = options_qs.values_list('AGENCE', flat=True).distinct().order_by('AGENCE')
    exploitants = options_qs.values_list('EXPL', flat=True).distinct().order_by('EXPL')

    # 5. Pagination
    get_params = request.GET.copy()
    if 'page' in get_params:
        del get_params['page']

    paginator = Paginator(donnees_queryset, 100)
    page = request.GET.get('page')

    try:
        donnees_page = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        donnees_page = paginator.page(1)

    context = {
        'donnees': donnees_page,
        'filiales': filiales,
        'notation': notation,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': get_params.urlencode(),
        'filiale_param': filiale_param,
        'agence_param': agence_param,
        'expl_param': expl_param,
    }

    return render(request, 'sans_classe_s.html', context)


def export_sans_classe_s(request):
    def strip_tz(value):
        if hasattr(value, 'tzinfo'):
            return value.replace(tzinfo=None)
        return value

    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')
    filiale_txt = request.GET.get('filiale_txt', '').strip()
    agence_txt = request.GET.get('agence_txt', '').strip()
    expl_txt = request.GET.get('expl_txt', '').strip()
    client_txt = request.GET.get('client', '').strip()
    risque_txt = request.GET.get('risque', '').strip()

    donnees = DATEREV.objects.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)

    donnees = donnees.filter(Q(RISQUE__isnull=True) | Q(RISQUE=""))
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    if filiale_txt:
        donnees = donnees.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees = donnees.filter(AGENCE__icontains=agence_txt)
    if expl_txt:
        donnees = donnees.filter(EXPL__icontains=expl_txt)
    if client_txt:
        donnees = donnees.filter(CLIENT__icontains=client_txt)
    if risque_txt:
        donnees = donnees.filter(RISQUE__icontains=risque_txt)


    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients non classés"

    # EntÃªtes
    headers = ['AGENCE', 'AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'RISQUE']
    ws.append(headers)

    # Données
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.DATEREV, d.PPE, d.RISQUE

        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Clients sans classe de risque {date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response




ITEMS_PER_PAGE = 100  # Nombre d'éléments Ã  charger par page

from django.shortcuts import render
from django.db.models import Max
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Kyc_pm, Notation  # Assurez-vous que les imports correspondent Ã  vos fichiers

# --- FONCTION UTILITAIRE POUR FILTRE CHAMPS VIDES (KYC FIELD CONFIG) ---
from django.db.models import Q
from kyc.models import KycFieldVisibilityConfig

def apply_kyc_field_config_filter(queryset, client_type):
    """
    Filtre le queryset pour ne garder que les clients dont au moins un des 
    champs définis dans KycFieldVisibilityConfig (empty_check_fields) est vide.
    """
    # Si le queryset est vide, on s'arrête
    if not queryset.exists():
        return queryset

    filiales = list(queryset.order_by().values_list("FILIALE", flat=True).distinct())
    if not filiales:
        return queryset.none()

    combined_q = None
    
    # Préchauffer les configs pour éviter N requêtes
    configs = list(KycFieldVisibilityConfig.objects.filter(client_type=client_type))
    
    for filiale in filiales:
        # 1. Chercher la config spécifique à la filiale
        config = next((c for c in configs if filiale in (c.filiales or [])), None)
        
        # 2. Sinon, prendre la config globale (sans filiales)
        if not config:
            config = next((c for c in configs if not c.filiales), None)
            
        if config and config.empty_check_fields:
            missing_q = None
            for field_name in config.empty_check_fields:
                if field_name in ['SALAIRE', 'CAPITAL', 'CA', 'RESULTAT']:
                    field_q = (
                        Q(**{f"{field_name}__isnull": True}) | 
                        Q(**{f"{field_name}__exact": ""}) |
                        Q(**{f"{field_name}__iexact": "XX"}) |
                        Q(**{f"{field_name}__iexact": "RAS"}) |
                        Q(**{f"{field_name}__iexact": "R.A.S."}) |
                        Q(**{f"{field_name}__iexact": "R.A.S"}) |
                        Q(**{f"{field_name}__in": [".", "?", "-", "*"]})
                    )
                else:
                    field_q = (
                        Q(**{f"{field_name}__isnull": True}) | 
                        Q(**{f"{field_name}__exact": ""}) |
                        Q(**{f"{field_name}__iexact": "XX"}) |
                        Q(**{f"{field_name}__iexact": "RAS"}) |
                        Q(**{f"{field_name}__iexact": "R.A.S."}) |
                        Q(**{f"{field_name}__iexact": "R.A.S"}) |
                        Q(**{f"{field_name}__length": 1})
                    )
                missing_q = field_q if missing_q is None else missing_q | field_q
                
            if missing_q is not None:
                scoped_q = Q(FILIALE=filiale) & missing_q
                combined_q = scoped_q if combined_q is None else combined_q | scoped_q
        else:
            # S'il n'y a AUCUNE configuration pour cette filiale et aucune configuration globale,
            # ou que la configuration a un tableau empty_check_fields vide,
            # alors on suppose qu'aucun champ n'est requis d'être vide pour s'afficher.
            # Donc pour cette filiale, on ne retourne rien (car la condition 'au moins un champ vide' est impossible).
            pass

    if combined_q is None:
        return queryset.none()
        
    return queryset.filter(combined_q)

# --- 1. FONCTION DE SÉCURITÉ PM (Périmètre de données) ---
def get_filtered_queryset_pm(request):
    """Garantit que l'utilisateur ne voit que les entreprises (PM) de son périmÃ¨tre."""
    user = request.user
    queryset = Kyc_pm.objects.all()

    if user.organe == "Chargé Client":
        queryset = queryset.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)

    elif user.organe == "Directeur Agence":
        queryset = queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)

    elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        queryset = queryset.filter(FILIALE=user.filiale)

    elif user.organe in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                         "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]:
        # AccÃ¨s total pour le Groupe
        pass

    return queryset.order_by('id')

# --- 2. FONCTION DES LISTES DE FILTRES PM ---
def get_filter_lists_pm(user, request):
    """GénÃ¨re les options des menus déroulants PM selon les droits d'accÃ¨s."""
    filiale_list, agence_list, expl_list, datouv_list = [], [], [], []
    base_qs = Kyc_pm.objects.all()

    # Restriction de la base de données selon le rôle
    if user.organe == "Chargé Client":
        base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        base_qs = base_qs.filter(FILIALE=user.filiale)

    # 1. Liste des Filiales (Uniquement pour le Groupe)
    filiale_list = Kyc_pm.objects.values_list("FILIALE", flat=True).distinct().order_by("FILIALE")

    # 2. Logique dynamique des Agences et Exploitants
    f_filiale = request.GET.get("filiale")
    f_agence = request.GET.get("agence")

    # Agences
    if f_filiale:
        agence_list = Kyc_pm.objects.filter(FILIALE=f_filiale).values_list("AGENCE", flat=True).distinct()
    elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        agence_list = base_qs.values_list("AGENCE", flat=True).distinct()

    # Exploitants
    if f_agence:
        expl_list = Kyc_pm.objects.filter(AGENCE=f_agence).values_list("EXPL", flat=True).distinct()
    elif user.organe == "Directeur Agence":
        expl_list = base_qs.values_list("EXPL", flat=True).distinct()
    elif not f_agence and (f_filiale or user.organe in ["DSI", "Conformité", "Contrôle Permanent"]):
        expl_list = base_qs.values_list("EXPL", flat=True).distinct()

    # 3. Dates d'ouverture
    datouv_list = base_qs.exclude(DATOUV__isnull=True).values_list("DATOUV", flat=True).distinct().order_by('-DATOUV')

    return filiale_list, agence_list, expl_list, datouv_list

# --- 3. VUE PRINCIPALE PM ---
def non_rens_pm(request):
    user = request.user

    # A. Sécurité : Queryset restreint au rôle (PM)
    queryset = get_filtered_queryset_pm(request)
    
    # Application de la règle des champs vides KYC Field Config
    queryset = apply_kyc_field_config_filter(queryset, "pm")

    # B. Application des filtres du formulaire
    f_filiale = request.GET.get('filiale')
    f_agence = request.GET.get('agence')
    f_expl = request.GET.get('expl')
    f_datouv = request.GET.get('datouv')
    f_lib_agence = request.GET.get('col_lib_agence') or request.GET.get('lib_agence')
    f_client = request.GET.get('col_client') or request.GET.get('client')
    f_idm = request.GET.get('col_idm') or request.GET.get('idm')
    f_agec = request.GET.get('col_agec') or request.GET.get('agec')
    f_codape = request.GET.get('col_codape') or request.GET.get('codape')
    f_rcsno = request.GET.get('col_rcsno') or request.GET.get('rcsno')
    f_capital = request.GET.get('col_capital') or request.GET.get('capital')
    f_ca = request.GET.get('col_ca') or request.GET.get('ca')
    f_resultat = request.GET.get('col_resultat') or request.GET.get('resultat')

    col_agence = request.GET.get('col_agence')
    col_expl = request.GET.get('col_expl')
    col_datouv = request.GET.get('col_datouv')

    if f_filiale: queryset = queryset.filter(FILIALE=f_filiale)
    if f_agence: queryset = queryset.filter(AGENCE=f_agence)
    if f_expl: queryset = queryset.filter(EXPL=f_expl)
    if f_datouv: queryset = queryset.filter(DATOUV=f_datouv)

    if col_agence: queryset = queryset.filter(AGENCE__icontains=col_agence)
    if col_expl: queryset = queryset.filter(EXPL__icontains=col_expl)
    if col_datouv: queryset = queryset.filter(DATOUV__icontains=col_datouv)

    if f_lib_agence: queryset = queryset.filter(LIB_AGENCE__icontains=f_lib_agence)
    if f_client: queryset = queryset.filter(CLIENT__icontains=f_client)
    if f_idm: queryset = queryset.filter(IDM__icontains=f_idm)
    if f_agec: queryset = queryset.filter(AGEC__icontains=f_agec)
    if f_codape: queryset = queryset.filter(CODAPE__icontains=f_codape)
    if f_rcsno: queryset = queryset.filter(RCSNO__icontains=f_rcsno)
    if f_capital: queryset = queryset.filter(CAPITAL__icontains=f_capital)
    if f_ca: queryset = queryset.filter(CA__icontains=f_ca)
    if f_resultat: queryset = queryset.filter(RESULTAT__icontains=f_resultat)

    # C. Notations (MÃªme logique de sécurité que PP)
    notes = Notation.objects.filter(flux_stock='Flux')
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])

    if user.organe == "Chargé Client":
        notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)
    elif user.organe == "Directeur Agence":
        notation = notation.filter(agent__filiale=user.filiale, agent__agence=user.agence)
    elif user.organe not in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "GUEST"]:
        notation = notation.filter(agent__filiale=user.filiale)

    # D. Listes pour les menus déroulants
    filiale_list, agence_list, expl_list, datouv_list = get_filter_lists_pm(user, request)

    # E. Pagination et conservation des paramÃ¨tres
    query_params = request.GET.copy()
    if 'page' in query_params: del query_params['page']
    get_params = query_params.urlencode()

    paginator = Paginator(queryset.order_by('id'), 30)
    page_number = request.GET.get('page')
    try:
        objets_page = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        objets_page = paginator.page(1)

    context = {
        "donnees": objets_page,
        "get_params": get_params,
        "filiale_list": filiale_list,
        "agence_list": agence_list,
        "expl_list": expl_list,
        "datouv_list": datouv_list,
        "notation": notation,
        "users_filiale": ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité'],
        "users_groupe": ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                         "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"],
    }

    return render(request, "non_rens_pm.html", context)

from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from .models import Kyc_pm  # Vérifiez le nom de votre modÃ¨le


@login_required
def export_csv_pm(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # 1. Base de données initiale avec filtre des champs vides
    donnees = Kyc_pm.objects.all()
    donnees = apply_kyc_field_config_filter(donnees, "pm")

    # 2. Récupération des paramètres de filtrage depuis l'URL
    f_filiale = request.GET.get("filiale")
    f_agence = request.GET.get("agence")
    f_expl = request.GET.get("expl")
    f_datouv = request.GET.get("datouv")
    f_lib_agence = request.GET.get("lib_agence")
    f_client = request.GET.get("client")
    f_idm = request.GET.get("idm")
    f_agec = request.GET.get("agec")
    f_codape = request.GET.get("codape")
    f_rcsno = request.GET.get("rcsno")
    f_capital = request.GET.get("capital")
    f_ca = request.GET.get("ca")
    f_resultat = request.GET.get("resultat")

    # 3. Sécurité par rôle (PérimÃ¨tre de l'utilisateur)
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)

    # 4. Application des filtres de recherche choisis par l'utilisateur
    if f_filiale:
        donnees = donnees.filter(FILIALE=f_filiale)
    if f_agence:
        donnees = donnees.filter(AGENCE=f_agence)
    if f_expl:
        donnees = donnees.filter(EXPL=f_expl)
    if f_lib_agence:
        donnees = donnees.filter(LIB_AGENCE__icontains=f_lib_agence)
    if f_client:
        donnees = donnees.filter(CLIENT__icontains=f_client)
    if f_idm:
        donnees = donnees.filter(IDM__icontains=f_idm)
    if f_agec:
        donnees = donnees.filter(AGEC__icontains=f_agec)
    if f_codape:
        donnees = donnees.filter(CODAPE__icontains=f_codape)
    if f_rcsno:
        donnees = donnees.filter(RCSNO__icontains=f_rcsno)
    if f_capital:
        donnees = donnees.filter(CAPITAL__icontains=f_capital)
    if f_ca:
        donnees = donnees.filter(CA__icontains=f_ca)
    if f_resultat:
        donnees = donnees.filter(RESULTAT__icontains=f_resultat)

    # 5. Conversion et Filtrage par DATE (Crucial pour éviter l'export vide)
    if f_datouv and f_datouv.strip():
        try:
            # On tente de convertir "12/01/2026" en objet date Python
            date_objet = datetime.strptime(f_datouv.strip(), '%d/%m/%Y').date()
            donnees = donnees.filter(DATOUV=date_objet)
        except (ValueError, TypeError):
            # Si la date dans l'URL n'est pas au bon format, on ignore ce filtre
            pass

    # 6. Création du classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Export KYC PM"

    # En-tÃªtes (Headers) dynamiques basés sur la configuration
    from kyc.context_processors import kyc_display_fields_processor
    ctx = kyc_display_fields_processor(request)
    display_fields = ctx.get('kyc_pm_display_fields', [])
    headers = [label for field, label in display_fields]
    ws.append(headers)

    # Remplissage des lignes
    for d in donnees:
        row = []
        for field, label in display_fields:
            val = getattr(d, field, "")
            if field == "DATOUV":
                row.append(format_date_for_export(val))
            else:
                row.append(str(val) if val is not None else "")
        ws.append(row)

    # Mise en forme : Ajustement automatique de la largeur des colonnes
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 18

    # 7. Préparation de la réponse HTTP pour le téléchargement
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Export_KYC_PM_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

from django.shortcuts import render
from django.db.models import Max
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

# Assurez-vous que Kyc_pp et Notation sont importés
# from .models import Kyc_pp, Notation

# --- CONSTANTE DE TAILLE DE PAGE ---
ITEMS_PER_PAGE = 100  # Nombre d'éléments Ã  charger par page
from django.shortcuts import render
from django.db.models import Max
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Kyc_pp, Notation  # Vérifiez le nom de vos modÃ¨les


# --- 1. FONCTION DE SÃ‰CURITÃ‰ (PérimÃ¨tre de données) ---
def get_filtered_queryset(request):
    """Garantit que l'utilisateur ne voit que son périmÃ¨tre autorisé."""
    user = request.user
    queryset = Kyc_pp.objects.all()

    if user.organe == "Chargé Client":
        queryset = queryset.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)

    elif user.organe == "Directeur Agence":
        queryset = queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)

    elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        queryset = queryset.filter(FILIALE=user.filiale)

    elif user.organe in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                         "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]:
        # Le groupe voit tout par défaut, le filtrage se fera via le formulaire
        pass

    return queryset.order_by('id')


# --- 2. FONCTION DES LISTES DE FILTRES (Menus déroulants) ---
def get_filter_lists(user, request):
    """GénÃ¨re les options des menus déroulants selon les droits d'accÃ¨s."""
    filiale_list, agence_list, expl_list, datouv_list = [], [], [], []
    base_queryset = Kyc_pp.objects.all()

    if user.organe == "Chargé Client":
        base_queryset = base_queryset.filter(FILIALE=user.filiale, AGENCE= user.agence, EXPL=user.code_expl)
        expl_list = [user.code_expl]

    elif user.organe == "Directeur Agence":
        base_queryset = base_queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)
        expl_list = base_queryset.values_list("EXPL", flat=True).distinct()

    elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        base_queryset = base_queryset.filter(FILIALE=user.filiale)
        agence_list = base_queryset.values_list("AGENCE", flat=True).distinct()

        agence_filter = request.GET.get("agence")
        if agence_filter:
            expl_list = base_queryset.filter(AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()
        else:
            expl_list = base_queryset.values_list("EXPL", flat=True).distinct()

    elif user.organe in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                         "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]:
        filiale_list = Kyc_pp.objects.values_list("FILIALE", flat=True).distinct()

        filiale_filter = request.GET.get("filiale")
        agence_filter = request.GET.get("agence")

        if filiale_filter:
            base_queryset = base_queryset.filter(FILIALE=filiale_filter)
            agence_list = base_queryset.values_list("AGENCE", flat=True).distinct()
        if agence_filter:
            expl_list = base_queryset.filter(AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()

    datouv_list = base_queryset.exclude(DATOUV__isnull=True).values_list("DATOUV", flat=True).distinct().order_by(
        '-DATOUV')

    return filiale_list, agence_list, expl_list, datouv_list


# --- 3. VUE PRINCIPALE ---
@login_required
def non_rens(request):
    user = request.user

    # A. Sécurité de base : Queryset restreint au rôle
    queryset = get_filtered_queryset(request)
    
    # Application de la règle des champs vides KYC Field Config
    queryset = apply_kyc_field_config_filter(queryset, "pp")

    # B. Application des filtres du formulaire (Si renseignés)
    f_filiale = request.GET.get('filiale')
    f_agence = request.GET.get('agence')
    f_expl = request.GET.get('expl')
    f_datouv = request.GET.get('datouv')

    f_lib_agence = request.GET.get('col_lib_agence') or request.GET.get('lib_agence')
    f_client = request.GET.get('col_client') or request.GET.get('client')
    f_idp = request.GET.get('col_idp') or request.GET.get('idp')
    f_numid = request.GET.get('col_numid') or request.GET.get('numid')
    f_datnais = request.GET.get('col_datnais') or request.GET.get('datnais')
    f_paynais = request.GET.get('col_paynais') or request.GET.get('paynais')
    f_adresse = request.GET.get('col_adresse') or request.GET.get('adresse')
    f_codape = request.GET.get('col_codape') or request.GET.get('codape')
    f_profession = request.GET.get('col_profession') or request.GET.get('profession')
    f_salaire = request.GET.get('col_salaire') or request.GET.get('salaire')
    f_origine_rev = request.GET.get('col_origine_rev') or request.GET.get('origine_rev')
    f_datvalid = request.GET.get('col_datvalid') or request.GET.get('datvalid')
    f_tel = request.GET.get('col_tel') or request.GET.get('tel')

    col_agence = request.GET.get('col_agence')
    col_expl = request.GET.get('col_expl')
    col_datouv = request.GET.get('col_datouv')

    if f_filiale: queryset = queryset.filter(FILIALE=f_filiale)
    if f_agence: queryset = queryset.filter(AGENCE=f_agence)
    if f_expl: queryset = queryset.filter(EXPL=f_expl)
    if f_datouv: queryset = queryset.filter(DATOUV=f_datouv)

    if col_agence: queryset = queryset.filter(AGENCE__icontains=col_agence)
    if col_expl: queryset = queryset.filter(EXPL__icontains=col_expl)
    if col_datouv: queryset = queryset.filter(DATOUV__icontains=col_datouv)

    if f_lib_agence: queryset = queryset.filter(LIB_AGENCE__icontains=f_lib_agence)
    if f_client: queryset = queryset.filter(CLIENT__icontains=f_client)
    if f_idp: queryset = queryset.filter(IDP__icontains=f_idp)
    if f_numid: queryset = queryset.filter(NUMID__icontains=f_numid)
    if f_datnais: queryset = queryset.filter(DATNAIS__icontains=f_datnais)
    if f_paynais: queryset = queryset.filter(PAYNAIS__icontains=f_paynais)
    if f_adresse: queryset = queryset.filter(ADRESSE__icontains=f_adresse)
    if f_codape: queryset = queryset.filter(CODAPE__icontains=f_codape)
    if f_profession: queryset = queryset.filter(PROFESSION__icontains=f_profession)
    if f_salaire: queryset = queryset.filter(SALAIRE__icontains=f_salaire)
    if f_origine_rev: queryset = queryset.filter(ORIGINE_REV__icontains=f_origine_rev)
    if f_datvalid: queryset = queryset.filter(DATVALID__icontains=f_datvalid)
    if f_tel: queryset = queryset.filter(TEL__icontains=f_tel)

    # C. Données de notation (Flux) filtrées par périmÃ¨tre
    notes = Notation.objects.filter(flux_stock='Flux')
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])

    if user.organe == "Chargé Client":
        notation = notation.filter(agent__filiale=user.filiale, agent__agence=user.agence,agent__code_expl=user.code_expl)
    elif user.organe == "Directeur Agence":
        notation = notation.filter(agent__filiale=user.filiale, agent__agence=user.agence)
    elif user.organe not in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "GUEST"]:
        notation = notation.filter(agent__filiale=user.filiale)

    # D. Listes pour les menus déroulants
    filiale_list, agence_list, expl_list, datouv_list = get_filter_lists(user, request)

    # E. Pagination
    query_params = request.GET.copy()
    if 'page' in query_params: del query_params['page']
    get_params = query_params.urlencode()

    paginator = Paginator(queryset.order_by('id'), 30)
    page_number = request.GET.get('page')
    try:
        objets_page = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        objets_page = paginator.page(1)

    context = {
        "donnees": objets_page,
        "get_params": get_params,
        "filiale_list": filiale_list,
        "agence_list": agence_list,
        "expl_list": expl_list,
        "datouv_list": datouv_list,
        'users_groupe': ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                         "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"],
        'users_filiale': ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité'],
        'notation': notation,
    }

    return render(request, "non_rens.html", context)

@login_required
def export_csv_pp(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # Partir de tous les objets et appliquer le filtre des champs vides
    donnees = Kyc_pp.objects.all()
    donnees = apply_kyc_field_config_filter(donnees, "pp")

    # Appliquer les mêmes filtres que dans la vue de liste
    # selon lâ€™organe de lâ€™utilisateur + éventuellement les filtres GET
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)

    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
        expl_filter = request.GET.get("expl")
        if expl_filter:
            donnees = donnees.filter(EXPL=expl_filter)

    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
        agence_filter = request.GET.get("agence")
        expl_filter = request.GET.get("expl")
        if agence_filter:
            donnees = donnees.filter(AGENCE=agence_filter)
        if expl_filter:
            donnees = donnees.filter(EXPL=expl_filter)

    elif user.organe in users_groupe:
        filiale_filter = request.GET.get("filiale")
        agence_filter = request.GET.get("agence")
        expl_filter = request.GET.get("expl")

        if filiale_filter:
            donnees = donnees.filter(FILIALE=filiale_filter)
        if agence_filter:
            donnees = donnees.filter(AGENCE=agence_filter)
        if expl_filter:
            donnees = donnees.filter(EXPL=expl_filter)

    # Ensuite création du fichier Excel (ou CSV selon ton besoin)
    wb = Workbook()
    ws = wb.active
    ws.title = "Export KYC"

    from kyc.context_processors import kyc_display_fields_processor
    ctx = kyc_display_fields_processor(request)
    display_fields = ctx.get('kyc_pp_display_fields', [])
    headers = [label for field, label in display_fields]
    ws.append(headers)

    for d in donnees:
        row = []
        for field, label in display_fields:
            val = getattr(d, field, "")
            if field == "DATOUV" or field == "DATNAIS" or field == "DATVALID":
                row.append(format_date_for_export(val))
            else:
                row.append(str(val) if val is not None else "")
        ws.append(row)

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Champs_non_renseignés_PP_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_csv_anom(request):
    user = request.user

    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau", 'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # ????????????????????????????????????????????????
    # Logique de filtrage
    # ????????????????????????????????????????????????
    donnees = Anomalie.objects.all()

    filiale_filter = request.GET.get("filiale")
    agence_filter = request.GET.get("agence")
    expl_filter = request.GET.get("expl")

    filiale_txt = request.GET.get("filiale_txt")
    agence_txt = request.GET.get("agence_txt")
    expl_txt = request.GET.get("expl_txt")
    lib_agence = request.GET.get("lib_agence")
    client = request.GET.get("client")

    anom_age = request.GET.get("anom_age")
    anom_eer = request.GET.get("anom_eer")
    anom_cin = request.GET.get("anom_cin")

    if hasattr(user, "organe"):

        if user.organe == "Chargé Client":
            donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)

        elif user.organe == "Directeur Agence":
            donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
            if expl_filter:
                donnees = donnees.filter(EXPL=expl_filter)

        elif user.organe in users_filiale:
            donnees = donnees.filter(FILIALE=user.filiale)
            if agence_filter:
                donnees = donnees.filter(AGENCE=agence_filter)
            if expl_filter:
                donnees = donnees.filter(EXPL=expl_filter)

        elif user.organe in users_groupe:
            if filiale_filter:
                donnees = donnees.filter(FILIALE=filiale_filter)
            if agence_filter:
                donnees = donnees.filter(AGENCE=agence_filter)
            if expl_filter:
                donnees = donnees.filter(EXPL=expl_filter)

    else:
        if filiale_filter:
            donnees = donnees.filter(FILIALE=filiale_filter)
        if agence_filter:
            donnees = donnees.filter(AGENCE=agence_filter)
        if expl_filter:
            donnees = donnees.filter(EXPL=expl_filter)

    if filiale_txt:
        donnees = donnees.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees = donnees.filter(AGENCE__icontains=agence_txt)
    if expl_txt:
        donnees = donnees.filter(EXPL__icontains=expl_txt)
    if lib_agence:
        donnees = donnees.filter(LIB_AGENCE__icontains=lib_agence)
    if client:
        donnees = donnees.filter(CLIENT__icontains=client)

    if anom_age:
        donnees = donnees.filter(ANOMALIE_AGE=anom_age)
    if anom_eer:
        donnees = donnees.filter(ANOMALIE_DATE_EER=anom_eer)
    if anom_cin:
        donnees = donnees.filter(ANOMALIE_CIN=anom_cin)

    # ????????????????????????????????????????????????
    # Cr??ation du fichier Excel
    # ????????????????????????????????????????????????
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anomalies"

    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "ANOMALIE_AGE", "ANOMALIE_DATE_EER", "ANOMALIE_CIN"]
    ws.append(headers)

    for obj in donnees:
        ws.append([
            obj.FILIALE,
            obj.AGENCE,
            obj.EXPL,
            obj.CLIENT,
            obj.ANOMALIE_AGE,
            obj.ANOMALIE_DATE_EER,
            obj.ANOMALIE_CIN
        ])

    for idx, col_title in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Champs_en_anomalie_PP_{date_str}.xlsx"

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response

@login_required
def export_csv_anom_ppe(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # Récupérer les filtres GET envoyés par le template
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    # Base queryset : anomalies avec PPE = 'O'
    donnees = Anomalie.objects.filter(PPE='O')

    # Filtrer selon le rôle de lâ€™utilisateur
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    # Si user dans groupe ou autre, on laisse le queryset PPE='O'

    # Appliquer les filtres GET si fournis
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)

    # Création du classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "PPE en Anomalie"

    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT",
               "ANOMALIE_AGE", "ANOMALIE_DATE_EER", "ANOMALIE_CIN"]
    ws.append(headers)

    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT,
            d.ANOMALIE_AGE, d.ANOMALIE_DATE_EER, d.ANOMALIE_CIN
        ])

    for col_num, _ in enumerate(headers, 1):
        letter = get_column_letter(col_num)
        ws.column_dimensions[letter].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"PPE_anomalies_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


import json
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import TauxEvolution, TauxEvolution_filiale, Notation

def _dashboard_data_cache_version():
    """
    Versionne le cache avec les dates max des tables de taux.
    Quand une injection matinale met Ã  jour les données, la version change.
    """
    latest_filiale = TauxEvolution_filiale.objects.aggregate(last_date=Max('date')).get('last_date')
    latest_expl = TauxEvolution.objects.aggregate(last_date=Max('date')).get('last_date')
    rules_version = cache.get('quality_control_rules_version', 1)
    return f"{latest_filiale or 'none'}|{latest_expl or 'none'}|rules:{rules_version}"


def _build_dashboard_cache_key(prefix, user, request, extra=""):
    scope = "|".join([
        str(getattr(user, "id", "")),
        str(getattr(user, "organe", "")),
        str(getattr(user, "filiale", "")),
        str(getattr(user, "agence", "")),
        str(getattr(user, "code_expl", "")),
        request.GET.urlencode(),
        extra,
        _dashboard_data_cache_version(),
    ])
    scope_hash = hashlib.md5(scope.encode("utf-8")).hexdigest()
    return f"dashboard:{prefix}:{scope_hash}"


@login_required
def statistiques(request):
    context_cache_key = _build_dashboard_cache_key("statistiques", request.user, request)
    cached_context = cache.get(context_cache_key)
    if cached_context is not None:
        return render(request, 'statistiques.html', cached_context)

    user = request.user
    # Liste des roles ayant une vue globale (Groupe)
    user_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

    # 1. Mode + granularite
    mode = request.GET.get('mode', 'Flux')
    is_stock = (mode == 'Stock')
    code_flux_stock = "S" if is_stock else "F"
    periode = request.GET.get('periode', 'journalier')
    periodes_valides = {'journalier', 'hebdomadaire', 'mensuel', 'annuel'}
    if periode not in periodes_valides:
        periode = 'mensuel'

    # 2. Securisation de la filiale
    target_filiale = getattr(user, 'filiale', None)
    if user.organe in user_groupe:
        f_get = request.GET.get('filiale')
        if f_get:
            target_filiale = f_get
    else:
        target_filiale = user.filiale

    # 3. Exploitant / utilisateur
    selected_user_filter = request.GET.get('utilisateur', '').strip()
    if user.organe == "Chargé Client":
        selected_expl = user.code_expl
    elif user.organe == "Directeur Agence":
        agents_de_lagence = ProfileV.objects.filter(filiale=user.filiale, agence=user.agence).values_list('code_expl', flat=True)
        req_expl = request.GET.get('expl') or selected_user_filter
        if req_expl in agents_de_lagence:
            selected_expl = req_expl
        else:
            selected_expl = None
    else:
        selected_expl = request.GET.get('expl') or selected_user_filter

    # 4. Donnees moyennes filiale
    latest_filiale_data = TauxEvolution_filiale.objects.filter(filiale=target_filiale).order_by('-date').first()
    if is_stock:
        last_pp_fil = latest_filiale_data.stock_PP if latest_filiale_data else 0
        last_pm_fil = latest_filiale_data.stock_PM if latest_filiale_data else 0
    else:
        last_pp_fil = latest_filiale_data.flux_PP if latest_filiale_data else 0
        last_pm_fil = latest_filiale_data.flux_PM if latest_filiale_data else 0

    # 5. Donnees historiques exploitant selon granularite
    base_qs_expl = TauxEvolution.objects.filter(
        flux_stock=code_flux_stock,
        filiale=target_filiale,
        expl=selected_expl
    ).order_by('date')

    latest_taux_date = (
        TauxEvolution.objects
        .filter(flux_stock=code_flux_stock, filiale=target_filiale)
        .aggregate(last_date=Max('date'))
        .get('last_date')
    )

    def build_period_key(d):
        if periode == 'journalier':
            return d
        if periode == 'hebdomadaire':
            year, week, _ = d.isocalendar()
            return (year, week)
        if periode == 'annuel':
            return d.year
        return (d.year, d.month)

    def format_period_label(key):
        if periode == 'journalier':
            return key.strftime('%d/%m/%Y')
        if periode == 'hebdomadaire':
            return f"S{key[1]}-{key[0]}"
        if periode == 'annuel':
            return str(key)
        return f"{key[1]:02d}/{key[0]}"

    def aggregate_by_period(queryset):
        grouped = {}
        for obj in queryset:
            key = build_period_key(obj.date)
            bucket = grouped.setdefault(key, {"sum": 0.0, "count": 0})
            bucket["sum"] += float(obj.taux or 0)
            bucket["count"] += 1
        return {k: round(v["sum"] / v["count"], 2) for k, v in grouped.items() if v["count"] > 0}

    dict_expl_pp = aggregate_by_period(base_qs_expl.filter(pp_pm="P"))
    dict_expl_pm = aggregate_by_period(base_qs_expl.filter(pp_pm="M"))

    period_keys = sorted(set(dict_expl_pp.keys()) | set(dict_expl_pm.keys()))
    labels_chart = [format_period_label(k) for k in period_keys]
    labels_table = labels_chart[:]
    data_expl_pp = [float(dict_expl_pp.get(k, 0)) for k in period_keys]
    data_expl_pm = [float(dict_expl_pm.get(k, 0)) for k in period_keys]

    var_pp = round(data_expl_pp[-1] - data_expl_pp[-2], 2) if len(data_expl_pp) > 1 else 0
    var_pm = round(data_expl_pm[-1] - data_expl_pm[-2], 2) if len(data_expl_pm) > 1 else 0

    # 6. Liste dynamique des exploitants
    expl_queryset = TauxEvolution.objects.filter(filiale=target_filiale, flux_stock=code_flux_stock)
    if user.organe == "Directeur Agence":
        agents_de_lagence = ProfileV.objects.filter(filiale=user.filiale, agence=user.agence).values_list('code_expl', flat=True)
        expl_queryset = expl_queryset.filter(expl__in=agents_de_lagence)
    elif user.organe == "Chargé Client":
        expl_queryset = expl_queryset.filter(expl=user.code_expl)

    liste_expl = list(expl_queryset.values_list('expl', flat=True).distinct().order_by('expl'))
    profiles_by_expl = {
        p.code_expl: p
        for p in ProfileV.objects.filter(code_expl__in=liste_expl)
    }
    liste_expl_display = []
    for code in liste_expl:
        profile = profiles_by_expl.get(code)
        full_name = f"{getattr(profile, 'first_name', '')} {getattr(profile, 'last_name', '')}".strip() if profile else ""
        label = f"{code} - {full_name}" if full_name else code
        liste_expl_display.append({'code': code, 'label': label})

    # 7. Notation et identite
    agent_info = ProfileV.objects.filter(filiale=target_filiale, code_expl=selected_expl).first()
    notation_obj = Notation.objects.filter(agent__code_expl=selected_expl, flux_stock=mode).order_by('-date_notation').first()

    # 8. Liste des filiales
    if user.organe in user_groupe:
        liste_filiales = TauxEvolution_filiale.objects.values_list('filiale', flat=True).distinct().order_by('filiale')
    else:
        liste_filiales = [user.filiale] if user.filiale else []

    quality_scope = evaluate_data_quality_scope(user)
    if selected_expl:
        quality_scope = {
            'filiale': target_filiale,
            'agence': getattr(agent_info, 'agence', None) if agent_info else None,
            'expl': selected_expl,
            'label': f"Agent {selected_expl}",
        }
    rules_version = cache.get('quality_control_rules_version', 1)

    def compute_quality_rate_by_typology(applicability):
        scope_signature = (
            f"{quality_scope.get('filiale')}|{quality_scope.get('agence')}|"
            f"{quality_scope.get('expl')}|{user.organe}|{applicability}"
        )
        scope_hash = hashlib.md5(scope_signature.encode('utf-8')).hexdigest()
        cache_key = f"quality_control:dashboard_rate:v{rules_version}:{scope_hash}"
        cached_rate = cache.get(cache_key)
        if cached_rate is not None:
            return cached_rate

        rules = list(DataQualityRule.objects.filter(active=True, applicability=applicability))
        total_ok = 0
        total_evaluated = 0
        for rule in rules:
            stat = evaluate_data_quality_rule(
                rule,
                filiale=quality_scope.get('filiale'),
                agence=quality_scope.get('agence'),
                expl=quality_scope.get('expl'),
            )
            total_ok += stat.get('ok_count', 0)
            total_evaluated += stat.get('total', 0)

        rate = round(total_ok / total_evaluated * 100, 1) if total_evaluated else 0
        cache.set(cache_key, rate, timeout=3600)
        return rate

    quality_rate_pp = compute_quality_rate_by_typology('PP')
    quality_rate_pm = compute_quality_rate_by_typology('PM')

    context = {
        'mode': mode,
        'is_stock': is_stock,
        'periode': periode,
        'selected_filiale': target_filiale,
        'selected_expl': selected_expl,
        'selected_user_filter': selected_expl,
        'agent_nom': (
            (f"{agent_info.first_name} {agent_info.last_name}".strip() or agent_info.code_expl or agent_info.username)
            if agent_info else selected_expl
        ),
        'agent_note': notation_obj.note if notation_obj else "N/A",
        'labels_json': json.dumps(labels_chart),
        'data_expl_pp': json.dumps(data_expl_pp),
        'data_expl_pm': json.dumps(data_expl_pm),
        'last_pp_expl': data_expl_pp[-1] if data_expl_pp else 0,
        'last_pm_expl': data_expl_pm[-1] if data_expl_pm else 0,
        'last_pp_fil': last_pp_fil,
        'last_pm_fil': last_pm_fil,
        'var_pp': var_pp,
        'var_pm': var_pm,
        'historique': list(reversed(list(zip(labels_table, data_expl_pp, data_expl_pm)))),
        'liste_expl': liste_expl,
        'liste_expl_display': liste_expl_display,
        'user_groupe': user_groupe,
        'liste_filiales': liste_filiales,
        'quality_rate_pp': quality_rate_pp,
        'quality_rate_pm': quality_rate_pm,
        'quality_scope_label': quality_scope.get('label'),
        'latest_taux_date': latest_taux_date,
    }
    cache.set(context_cache_key, context, timeout=60 * 60 * 8)
    return render(request, 'statistiques.html', context)
def export_stats_pp(request):
    # mÃªme logique de filtrage que dans la vue principale
    user = request.user
    organe = user.organe
    filiale = user.filiale
    expl_user = user.expl

    if organe in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                  "Contrôle Permanent Groupe", "PASS", "GUEST"]:
        qs = TauxEvolution.objects.all()
    elif organe in ["Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        qs = TauxEvolution.objects.filter(filiale=filiale)
    elif organe == "Directeur Agence":
        qs = TauxEvolution.objects.filter(agence=user.agence)
    elif organe == "Chargé Client":
        qs = TauxEvolution.objects.filter(expl=expl_user)
    else:
        qs = TauxEvolution.objects.none()

    # filtrer éventuellement sur GET param
    selected_expl = request.GET.get('expl')
    if selected_expl:
        qs = qs.filter(expl=selected_expl)

    data = list(qs.values('filiale', 'agence', 'expl', 'date', 'taux'))
    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rapport_taux.xlsx"'
    df.to_excel(response, index=False)
    return response


@login_required
def daterev_ppe(request):
    # Rôles
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    user = request.user

    # Params GET
    periode_param = request.GET.get("periode", "")
    filiale_param = request.GET.get("filiale", "")
    agence_param = request.GET.get("agence", "")
    expl_param = request.GET.get("expl", "")

    base_qs = DATEREV.objects.all().filter(DATEREV__isnull=False, PPE='O')

    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
    if not is_group_user:
        if getattr(user, "organe", "") == "Chargé Client":
            base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
        elif user.organe == "Directeur Agence":
            base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in users_filiale:
            base_qs = base_qs.filter(FILIALE=user.filiale)

    today = date.today()
    qs_period = base_qs
    if periode_param == "today":
        qs_period = qs_period.filter(DATEREV__lte=today)
    elif periode_param == "3m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=90))
    elif periode_param == "6m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=180))
    elif periode_param == "1y":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=365))

    can_pick_filiale = is_group_user

    selected_filiale = filiale_param if can_pick_filiale else getattr(user, "filiale", "")

    filiales_opts = qs_period.values_list("FILIALE", flat=True).distinct().order_by("FILIALE")

    qs_filiale = qs_period
    if selected_filiale:
        qs_filiale = qs_filiale.filter(FILIALE=selected_filiale)

    can_pick_agence = (user.organe in users_groupe) or (user.organe in users_filiale) or (
            user.organe == "Directeur Agence")

    if user.organe == "Directeur Agence":
        selected_agence = getattr(user, "agence", "")
    else:
        selected_agence = agence_param

    agences_opts = qs_filiale.values_list("AGENCE", flat=True).distinct().order_by("AGENCE")

    qs_agence = qs_filiale
    if selected_agence:
        qs_agence = qs_agence.filter(AGENCE=selected_agence)

    can_pick_expl = (user.organe in users_groupe) or (user.organe in users_filiale) or (
            user.organe == "Directeur Agence")

    if getattr(user, "organe", "") == "Chargé Client":
        selected_expl = getattr(user, "code_expl", "")
    else:
        selected_expl = expl_param

    exploitants_opts = qs_agence.values_list("EXPL", flat=True).distinct().order_by("EXPL")

    donnees = qs_agence
    if selected_expl:
        donnees = donnees.filter(EXPL=selected_expl)
    count_risque_non_eleve = donnees.exclude(RISQUE="Risque eleve").count()

    context = {
        "donnees": donnees.order_by("FILIALE", "AGENCE", "EXPL", "CLIENT"),
        "total_count": donnees.count(), # Optionnel : le total général
        "count_risque_non_eleve": count_risque_non_eleve, # Le nouveau décompte
        # Options
        "filiales": filiales_opts,
        "agences": agences_opts,
        "exploitants": exploitants_opts,

        # Sélections courantes
        "periode": periode_param,
        "filiale_param": selected_filiale,
        "agence_param": selected_agence,
        "expl_param": selected_expl,

        # Rôles (si tu en as besoin dans le template)
        "users_groupe": users_groupe,
        "users_filiale": users_filiale,

        # Droits d'édition des selects
        "can_pick_filiale": can_pick_filiale,
        "can_pick_agence": can_pick_agence,
        "can_pick_expl": can_pick_expl,
    }
    return render(request, 'daterev_ppe.html', context)


@login_required
def non_anom_ppe(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    donnees = Anomalie.objects.filter(PPE="O")
    # Filtrage automatique selon le rôle
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    if user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    if user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    if user.organe in users_groupe:
        donnees = donnees

    # Filtres manuels via GET
    if filiale_param:
        donnees = donnees.filter(FILIALE=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL=expl_param)

    # === Valeurs du formulaire selon le rôle ===
    if user.organe == "Directeur Agence":
        exploitants = Anomalie.objects.filter(AGENCE=user.agence).values_list('EXPL', flat=True).distinct()
        agences = Anomalie.objects.filter(AGENCE=user.agence).values_list('AGENCE', flat=True).distinct()
    elif user.organe in users_filiale:
        agences = Anomalie.objects.filter(FILIALE=user.filiale).values_list('AGENCE', flat=True).distinct()
        exploitants = Anomalie.objects.filter(FILIALE=user.filiale).values_list('EXPL', flat=True).distinct()
    else:
        exploitants = Anomalie.objects.values_list('EXPL', flat=True).distinct()
        agences = Anomalie.objects.values_list('AGENCE', flat=True).distinct()

    filiales = Anomalie.objects.values_list('FILIALE', flat=True).distinct()

    context = {
        'donnees': donnees,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
    }

    return render(request, 'anom_ppe.html', context)

from django.shortcuts import render
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Max
from .models import Anomalie, Notation

@login_required
def non_anom(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau", 'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # ???????????????????????????????????????????????????????????????
    # 1. Notation (champ Flux)
    # ???????????????????????????????????????????????????????????????
    notes = Notation.objects.filter(flux_stock='Flux')
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])

    if hasattr(user, 'filiale') and hasattr(user, 'code_expl'):
        notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)

    # ???????????????????????????????????????????????????????????????
    # 2. Base Queryset des anomalies
    # ???????????????????????????????????????????????????????????????
    queryset = Anomalie.objects.all()

    # ???????????????????????????????????????????????????????????????
    # 3. Filtres GET
    # ???????????????????????????????????????????????????????????????
    filiale_filter = request.GET.get("filiale")
    agence_filter = request.GET.get("agence")
    expl_filter = request.GET.get("expl")

    filiale_txt = request.GET.get("filiale_txt")
    agence_txt = request.GET.get("agence_txt")
    expl_txt = request.GET.get("expl_txt")
    lib_agence = request.GET.get("lib_agence")
    client = request.GET.get("client")

    anom_age = request.GET.get("anom_age")
    anom_eer = request.GET.get("anom_eer")
    anom_cin = request.GET.get("anom_cin")

    filiale_list = []
    agence_list = []
    expl_list = []

    # ???????????????????????????????????????????????????????????????
    # 4. Filtrage selon r??le
    # ???????????????????????????????????????????????????????????????
    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)

    if is_group_user:
        from kyc.models import DataQualityRule
        from kyc.forms import DataQualityRuleForm
        anom_filiales = {f for f in Anomalie.objects.values_list("FILIALE", flat=True).distinct() if f}
        rule_filiales = set()
        for rule in DataQualityRule.objects.filter(active=True):
            parsed = DataQualityRuleForm._parse_filiales(rule.filiale)
            for f in parsed:
                if f:
                    rule_filiales.add(f)
        filiale_list = sorted(list(anom_filiales | rule_filiales))
        if filiale_filter:
            queryset = queryset.filter(FILIALE=filiale_filter)
            agence_list = Anomalie.objects.filter(FILIALE=filiale_filter).values_list("AGENCE", flat=True).distinct()
        if agence_filter:
            queryset = queryset.filter(AGENCE=agence_filter)
            expl_list = Anomalie.objects.filter(AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()
        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)

    elif user.organe == "Chargé Client":
        queryset = queryset.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)

    elif user.organe == "Directeur Agence":
        queryset = queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)
        agence_list = Anomalie.objects.filter(FILIALE=user.filiale).values_list("AGENCE", flat=True).distinct()
        expl_list = Anomalie.objects.filter(AGENCE=user.agence).values_list("EXPL", flat=True).distinct()
        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)

    elif user.organe in users_filiale:
        queryset = queryset.filter(FILIALE=user.filiale)
        filiale_list = [user.filiale]
        agence_list = Anomalie.objects.filter(FILIALE=user.filiale).values_list("AGENCE", flat=True).distinct()
        if agence_filter:
            queryset = queryset.filter(AGENCE=agence_filter)
            expl_list = Anomalie.objects.filter(AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()
        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)
            expl_list = Anomalie.objects.filter(AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()
        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)

    # Filtres texte (colonnes)
    if filiale_txt:
        queryset = queryset.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        queryset = queryset.filter(AGENCE__icontains=agence_txt)
    if expl_txt:
        queryset = queryset.filter(EXPL__icontains=expl_txt)
    if lib_agence:
        queryset = queryset.filter(LIB_AGENCE__icontains=lib_agence)
    if client:
        queryset = queryset.filter(CLIENT__icontains=client)

    # Options pour listes d??roulantes (avant filtre anomalie)
    opts_qs = queryset
    anom_age_opts = opts_qs.exclude(ANOMALIE_AGE__isnull=True).exclude(ANOMALIE_AGE="").values_list(
        "ANOMALIE_AGE", flat=True).distinct().order_by("ANOMALIE_AGE")
    anom_eer_opts = opts_qs.exclude(ANOMALIE_DATE_EER__isnull=True).exclude(ANOMALIE_DATE_EER="").values_list(
        "ANOMALIE_DATE_EER", flat=True).distinct().order_by("ANOMALIE_DATE_EER")
    anom_cin_opts = opts_qs.exclude(ANOMALIE_CIN__isnull=True).exclude(ANOMALIE_CIN="").values_list(
        "ANOMALIE_CIN", flat=True).distinct().order_by("ANOMALIE_CIN")

    # Filtres anomalies (via select)
    if anom_age:
        queryset = queryset.filter(ANOMALIE_AGE=anom_age)
    if anom_eer:
        queryset = queryset.filter(ANOMALIE_DATE_EER=anom_eer)
    if anom_cin:
        queryset = queryset.filter(ANOMALIE_CIN=anom_cin)

    # ???????????????????????????????????????????????????????????????
    # 5. Pagination
    # ???????????????????????????????????????????????????????????????
    queryset = queryset.order_by('CLIENT', 'id')

    ITEMS_PER_PAGE = 50
    paginator = Paginator(queryset, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')

    try:
        objets_page = paginator.page(page_number)
    except PageNotAnInteger:
        objets_page = paginator.page(1)
    except EmptyPage:
        objets_page = paginator.page(paginator.num_pages)

    import urllib.parse
    import hashlib
    from datetime import datetime
    from django.utils import timezone
    from django.core.cache import cache
    from kyc.models import DataQualityRule

    # Helper function to compute failures list in Python
    def get_rule_failures(rule, queryset_eval):
        client_fields = ['CLIENT', 'EXPL', 'FILIALE', 'AGENCE']
        failures = []
        today_date = timezone.localdate()

        def safe_parse_date(value):
            if not value: return None
            if hasattr(value, 'date'): return value.date()
            if isinstance(value, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y%m%d"):
                    try:
                        return datetime.strptime(value.strip(), fmt).date()
                    except ValueError:
                        continue
            return None

        def calculate_age(birth_date_str):
            parsed = safe_parse_date(birth_date_str)
            if not parsed: return None
            return today_date.year - parsed.year - ((today_date.month, today_date.day) < (parsed.month, parsed.day))

        if rule.control_type == 'simple':
            param = (rule.parameter or '').strip().lower()
            field_name = rule.field_name
            rows = queryset_eval.values(*client_fields, field_name).iterator(chunk_size=2000)
            
            if not param or param == 'existence':
                for row in rows:
                    val = row.get(field_name)
                    if val is None or str(val).strip() == "":
                        failures.append({
                            'client': row.get('CLIENT', ''),
                            'filiale': row.get('FILIALE', ''),
                            'agence': row.get('AGENCE', ''),
                            'expl': row.get('EXPL', ''),
                            'values': [str(val or '')]
                        })
            elif param.isdigit() or (param.startswith('len') or param.startswith('long')):
                import re
                match = re.search(r'\d+', param)
                target_len = int(match.group()) if match else int(param)
                for row in rows:
                    val = str(row.get(field_name) or '')
                    if len(val) != target_len:
                        failures.append({
                            'client': row.get('CLIENT', ''),
                            'filiale': row.get('FILIALE', ''),
                            'agence': row.get('AGENCE', ''),
                            'expl': row.get('EXPL', ''),
                            'values': [val]
                        })
            else:
                target_val = rule.parameter
                for row in rows:
                    val = row.get(field_name)
                    if str(val or '').strip() != str(target_val or '').strip():
                        failures.append({
                            'client': row.get('CLIENT', ''),
                            'filiale': row.get('FILIALE', ''),
                            'agence': row.get('AGENCE', ''),
                            'expl': row.get('EXPL', ''),
                            'values': [str(val or '')]
                        })
                        
        elif rule.control_type == 'composite':
            conditions = rule.conditions.all()
            cond_fields = [c.field_name for c in conditions]
            unique_cond_fields = list(dict.fromkeys(cond_fields))
            fields_to_fetch = list(client_fields) + unique_cond_fields
            rows = queryset_eval.values(*fields_to_fetch).iterator(chunk_size=2000)
            
            for row in rows:
                all_match = True
                for cond in conditions:
                    val = str(row.get(cond.field_name, '') or '').strip()
                    target = (cond.value or '').strip()
                    
                    match = False
                    op = cond.operator
                    if op == '=': match = val == target
                    elif op == '!=': match = val != target
                    elif op == '>':
                        try: match = float(val.replace(',','.')) > float(target.replace(',','.'))
                        except: match = False
                    elif op == '<':
                        try: match = float(val.replace(',','.')) < float(target.replace(',','.'))
                        except: match = False
                    elif op == '>=':
                        try: match = float(val.replace(',','.')) >= float(target.replace(',','.'))
                        except: match = False
                    elif op == '<=':
                        try: match = float(val.replace(',','.')) <= float(target.replace(',','.'))
                        except: match = False
                    elif op == 'contains': match = target.lower() in val.lower()
                    elif op == 'is_empty': match = not val
                    elif op == 'is_not_empty': match = bool(val)
                    elif op == 'expired':
                        p = safe_parse_date(val)
                        match = p and p < today_date
                    elif op == 'age_gt':
                        age = calculate_age(val)
                        try: match = age is not None and age > int(target)
                        except: match = False
                    elif op == 'age_lt':
                        age = calculate_age(val)
                        try: match = age is not None and age < int(target)
                        except: match = False
                    elif op == 'min_length':
                        try: match = len(val) < int(target)
                        except: match = False
                    elif op == 'max_length':
                        try: match = len(val) > int(target)
                        except: match = False
                    
                    if not match:
                        all_match = False
                        break
                
                if all_match:
                    failures.append({
                        'client': row.get('CLIENT', ''),
                        'filiale': row.get('FILIALE', ''),
                        'agence': row.get('AGENCE', ''),
                        'expl': row.get('EXPL', ''),
                        'values': [str(row.get(f, '') or '') for f in unique_cond_fields]
                    })
        return failures

    # 1. Fetch active rules
    rules_qs = DataQualityRule.objects.filter(active=True).prefetch_related('conditions')
    
    # 2. Filter by target filiale
    if is_group_user:
        target_filiale = filiale_filter
    else:
        target_filiale = user.filiale

    if target_filiale:
        rules_qs = rules_qs.filter(Q(filiale__icontains=f"|{target_filiale}|") | Q(filiale=""))

    # 3. Filter by search query q
    q = request.GET.get('q', '').strip()
    if q:
        rules_qs = rules_qs.filter(
            Q(name__icontains=q) |
            Q(field_name__icontains=q) |
            Q(parameter__icontains=q)
        )
    rules_qs = rules_qs.order_by('-created_at')

    # 4. Evaluation Scope
    if is_group_user:
        eval_filiale = filiale_filter
        eval_agence = agence_filter
        eval_expl = expl_filter
    elif user.organe == "Chargé Client":
        eval_filiale = user.filiale
        eval_agence = user.agence
        eval_expl = user.code_expl
    elif user.organe == "Directeur Agence":
        eval_filiale = user.filiale
        eval_agence = user.agence
        eval_expl = expl_filter
    elif user.organe in users_filiale:
        eval_filiale = user.filiale
        eval_agence = agence_filter
        eval_expl = expl_filter
    else:
        eval_filiale = filiale_filter
        eval_agence = agence_filter
        eval_expl = expl_filter

    rules_version = cache.get('quality_control_rules_version', 1)
    data_refresh_bucket = timezone.localdate().isoformat()
    cache_ttl_seconds = 86400

    rules_with_stats = []
    for rule in rules_qs:
        rule_eval_filiale = _rule_eval_filiale(rule, eval_filiale)
        non_anom_signature = (
            f"{rule.id}|{rule.name}|{rule.applicability}|{rule.field_name}|"
            f"{rule.control_type}|{rule.parameter}|{rule.filiale}|"
            f"{rule_eval_filiale}|{eval_agence}|{eval_expl}"
        )
        non_anom_key = f"quality_control:non_anom:v{rules_version}:d{data_refresh_bucket}:{hashlib.md5(non_anom_signature.encode('utf-8')).hexdigest()}"
        
        stat = cache.get(non_anom_key)
        if stat is None:
            stat = _evaluate_data_quality_rule_scoped(rule, filiale=rule_eval_filiale, agence=eval_agence, expl=eval_expl)
            cache.set(non_anom_key, stat, timeout=cache_ttl_seconds)

        total_eval = stat.get('total', 0)
        stat['compliance_rate'] = compliance_rate_floor(stat.get('ok_count', 0), total_eval, stat.get('fail_count', 0))

        from kyc.forms import DataQualityRuleForm
        parsed_filiales = DataQualityRuleForm._parse_filiales(rule.filiale)
        if user.organe not in users_groupe and user.filiale:
            if not parsed_filiales or user.filiale in parsed_filiales:
                parsed_filiales = [user.filiale]
            else:
                parsed_filiales = []

        if parsed_filiales:
            visible_filiales = parsed_filiales[:3]
            hidden_count = max(0, len(parsed_filiales) - 3)
            display_str = ", ".join(parsed_filiales)
        else:
            visible_filiales = ["Toutes les filiales"]
            hidden_count = 0
            display_str = "Toutes les filiales"

        filiales_summary = {
            'display': display_str,
            'visible': visible_filiales,
            'hidden_count': hidden_count,
        }

        rules_with_stats.append({
            'rule': rule,
            'stat': stat,
            'filiales_summary': filiales_summary,
        })

    # 5. Modal Logic
    selected_rule_id = request.GET.get("rule")
    selected_rule = None
    show_rule_modal = False
    failures_page = None
    selected_rule_conditions = []
    selected_rule_filiales_display = ""
    failure_columns = []
    
    failure_client_filter = request.GET.get("failure_client", "")
    failure_filiale_filter = request.GET.get("failure_filiale", "")
    failure_agence_filter = request.GET.get("failure_agence", "")
    failure_expl_filter = request.GET.get("failure_expl", "")
    failure_message = ""

    if selected_rule_id:
        try:
            selected_rule = DataQualityRule.objects.get(pk=selected_rule_id, active=True)
            show_rule_modal = True
            selected_rule_conditions = selected_rule.conditions.all()
            
            from kyc.forms import DataQualityRuleForm
            parsed_filiales = DataQualityRuleForm._parse_filiales(selected_rule.filiale)
            if user.organe not in users_groupe and user.filiale:
                if not parsed_filiales or user.filiale in parsed_filiales:
                    parsed_filiales = [user.filiale]
                else:
                    parsed_filiales = []
            selected_rule_filiales_display = ", ".join(parsed_filiales) if parsed_filiales else "Toutes les filiales"
            
            if selected_rule.control_type == 'simple':
                failure_columns = [{
                    'name': selected_rule.field_name.upper(),
                    'param': f'failure_{selected_rule.field_name}',
                    'filter_value': request.GET.get(f'failure_{selected_rule.field_name}', '')
                }]
            else:
                cond_fields = [c.field_name for c in selected_rule_conditions]
                unique_cond_fields = list(dict.fromkeys(cond_fields))
                failure_columns = [{
                    'name': f.upper(),
                    'param': f'failure_{f}',
                    'filter_value': request.GET.get(f'failure_{f}', '')
                } for f in unique_cond_fields]
                
            model = Kyc_pp if selected_rule.applicability == 'PP' else Kyc_pm
            queryset_eval = model.objects.all()
            
            rule_eval_filiale = _rule_eval_filiale(selected_rule, eval_filiale)
            if rule_eval_filiale and rule_eval_filiale != 'GROUPE':
                queryset_eval = queryset_eval.filter(FILIALE=rule_eval_filiale)
            elif parsed_filiales:
                queryset_eval = queryset_eval.filter(FILIALE__in=parsed_filiales)
                
            if eval_agence:
                queryset_eval = queryset_eval.filter(AGENCE=eval_agence)
            if eval_expl:
                queryset_eval = queryset_eval.filter(EXPL=eval_expl)
                
            failures = get_rule_failures(selected_rule, queryset_eval)
            
            if failure_client_filter:
                failures = [f for f in failures if failure_client_filter.lower() in str(f['client']).lower()]
            if failure_filiale_filter:
                failures = [f for f in failures if failure_filiale_filter.lower() in str(f['filiale']).lower()]
            if failure_agence_filter:
                failures = [f for f in failures if failure_agence_filter.lower() in str(f['agence']).lower()]
            if failure_expl_filter:
                failures = [f for f in failures if failure_expl_filter.lower() in str(f['expl']).lower()]

            for i, col in enumerate(failure_columns):
                val_filter = col['filter_value'].strip()
                if val_filter:
                    failures = [f for f in failures if val_filter.lower() in str(f['values'][i]).lower()]
            
            paginator = Paginator(failures, 15)
            page_number = request.GET.get('page')
            try:
                failures_page = paginator.page(page_number)
            except PageNotAnInteger:
                failures_page = paginator.page(1)
            except EmptyPage:
                failures_page = paginator.page(paginator.num_pages)
                
        except DataQualityRule.DoesNotExist:
            failure_message = "La règle spécifiée n'existe pas ou est inactive."

    # 6. Build query strings
    base_params = {}
    if q: base_params['q'] = q
    if filiale_filter: base_params['filiale'] = filiale_filter
    if agence_filter: base_params['agence'] = agence_filter
    if expl_filter: base_params['expl'] = expl_filter
    base_querystring = urllib.parse.urlencode(base_params)

    pagination_params = dict(request.GET.items())
    pagination_params.pop('page', None)
    pagination_querystring = urllib.parse.urlencode(pagination_params)

    export_querystring = pagination_querystring

    context = {
        "donnees": objets_page,
        "filiale_list": filiale_list,
        "agence_list": agence_list,
        "expl_list": expl_list,
        "users_groupe": users_groupe,
        "users_filiale": users_filiale,
        "notation": notation,
        "anom_age_opts": anom_age_opts,
        "anom_eer_opts": anom_eer_opts,
        "anom_cin_opts": anom_cin_opts,
        "anom_age": anom_age,
        "anom_eer": anom_eer,
        "anom_cin": anom_cin,
        
        "rules": rules_with_stats,
        "total_rules": len(rules_with_stats),
        "total_failures": sum(item['stat'].get('fail_count', 0) for item in rules_with_stats),
        "rule_search": q,
        "can_pick_filiale": is_group_user,
        "can_pick_agence": user.organe in users_filiale or is_group_user,
        "is_charge_client": user.organe == "Chargé Client",
        "is_group_user": is_group_user,
        "show_rule_modal": show_rule_modal,
        "selected_rule": selected_rule,
        "selected_rule_conditions": selected_rule_conditions,
        "selected_rule_filiales_display": selected_rule_filiales_display,
        "failure_columns": failure_columns,
        "failures_page": failures_page,
        "failure_client_filter": failure_client_filter,
        "failure_filiale_filter": failure_filiale_filter,
        "failure_agence_filter": failure_agence_filter,
        "failure_expl_filter": failure_expl_filter,
        "failure_message": failure_message,
        "base_querystring": base_querystring,
        "pagination_querystring": pagination_querystring,
        "export_querystring": export_querystring,
    }

    return render(request, "non_anom.html", context)

@login_required
def devise(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau", 'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    # 1. Récupérer LA devise de la filiale (on prend la premiÃ¨re trouvée)
    # On récupÃ¨re juste la valeur (ex: "XOF") pour la comparer aux données Kyc_pp
    devise_obj = Devise.objects.filter(filiale=user.filiale).first()
    devise_valeur = devise_obj.devise if devise_obj else None  # Remplacez 'nom_devise' par le nom réel de votre champ

    # 2. Filtrage de base : Exclure la devise de la filiale, les vides et "NA"
    donnees = Kyc_pp.objects.exclude(DEVISE=devise_valeur).exclude(DEVISE__isnull=True).exclude(DEVISE="").exclude(DEVISE="NA")

    # === Filtrage automatique selon le rôle ===
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    # Si dans users_groupe, on garde tout (déjÃ  géré par l'absence de filtre)

    # === Filtres manuels via GET ===
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    donnees = _apply_pp_header_filters(donnees, request, include_devise=True)

    # === Valeurs pour les menus déroulants du formulaire ===
    filiales = donnees.values_list('FILIALE', flat=True).distinct()
    agences = donnees.values_list('AGENCE', flat=True).distinct()
    exploitants = donnees.values_list('EXPL', flat=True).distinct()
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

    # === Paginator ===
    ITEMS_PER_PAGE = 25
    paginator = Paginator(donnees.order_by('id'), ITEMS_PER_PAGE)
    page_number = request.GET.get('page')

    try:
        objets_page = paginator.page(page_number)
    except PageNotAnInteger:
        objets_page = paginator.page(1)
    except EmptyPage:
        objets_page = paginator.page(paginator.num_pages)

    # Devise KPIs
    total_devise = donnees.count()
    missing_devise = get_incomplete_clients_queryset(donnees, 'pp').count()
    complete_devise = max(0, total_devise - missing_devise)
    compliance_rate = round((complete_devise / total_devise) * 100, 1) if total_devise > 0 else 100.0

    # Répartition par devise
    devise_counts = list(
        donnees.values('DEVISE')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    devise_items = []
    for dc in devise_counts:
        dev_code = dc['DEVISE'] or "Non renseignée"
        pct = round((dc['count'] / total_devise) * 100, 1) if total_devise > 0 else 0.0
        devise_items.append({
            'label': dev_code,
            'value': dc['count'],
            'suffix': f'({pct}%)'
        })

    export_params = request.GET.copy()
    export_params['incompletes'] = '1'
    export_incomplets_url = f"{reverse('export_devise_pp')}?{export_params.urlencode()}"

    kpi_cards = [
        {
            'tone': 'emerald',
            'label': 'Répartition par Devise',
            'value': total_devise,
            'subtitle': 'Comptes en devise étrangère',
            'show_modal': True,
            'items': devise_items
        },
        {
            'tone': 'red',
            'label': 'Comptes Incomplets',
            'value': missing_devise,
            'subtitle': 'Dossiers avec données manquantes',
            'export_url': export_incomplets_url,
        },
        {
            'tone': 'blue',
            'label': 'Taux de complétude',
            'value': compliance_rate,
            'suffix': '%',
            'subtitle': 'Dossiers complets / total',
        }
    ]

    context = {
        "donnees": objets_page,
        "devise_filiale": devise_valeur,  # Utile pour l'affichage dans le template
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': query_params.urlencode(),
        'kpi_cards': kpi_cards,
    }

    return render(request, 'devise.html', context)


@login_required
def export_devise_pp(request):
    user = request.user

    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # Récupération des filtres GET pour la synchronisation
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    devise_obj = Devise.objects.filter(filiale=user.filiale).first()
    devise_valeur = devise_obj.devise if devise_obj else None

    # Début du Queryset : Exclure la devise de la filiale, les vides et "NA"
    donnees = Kyc_pp.objects.exclude(DEVISE=devise_valeur).exclude(DEVISE__isnull=True).exclude(DEVISE="").exclude(DEVISE="NA")

    # === Filtrage automatique selon le rôle (identique Ã  devise) ===
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        pass  # on ne filtre pas davantage

    # === Filtres manuels via GET (synchronisation) ===
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    donnees = _apply_pp_header_filters(donnees, request, include_devise=True)

    if request.GET.get('incompletes') == '1':
        donnees = get_incomplete_clients_queryset(donnees, 'pp')

    # Fin du Queryset filtré

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comptes Devise PP"  # J'ai renommé le titre

    # EntÃªtes
    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "CODAPE", "IDP", "PAYNAIS", "PROFESSION", "ADRESSE", "PAYS_RESID",
               "NUMID", "SALAIRE", "ORIGINE_REV", "DATVALID", "TEL"]
    ws.append(headers)

    # Données
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.CODAPE, d.IDP,
            d.PAYNAIS, d.PROFESSION, d.ADRESSE, d.PAYS_RESID, d.NUMID, d.SALAIRE, d.ORIGINE_REV, d.DATVALID, d.TEL
        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    if request.GET.get('incompletes') == '1':
        filename = f"Comptes_en_devise_PP_incomplets_{date_str}.xlsx"
    else:
        filename = f"Comptes_en_devise_PP_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def devise_pm(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    # 1. Récupérer LA devise de la filiale
    devise_obj = Devise.objects.filter(filiale=user.filiale).first()
    devise_valeur = devise_obj.devise if devise_obj else None

    # 2. Filtrage de base : Exclure la devise de la filiale, les vides et "NA"
    donnees = Kyc_pm.objects.exclude(DEVISE=devise_valeur).exclude(DEVISE__isnull=True).exclude(DEVISE="").exclude(DEVISE="NA")

    # === Filtrage automatique selon le rôle ===
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        pass  # on ne filtre pas davantage

    # === Filtres manuels via GET ===
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)

    # === Valeurs du formulaire selon le rôle ===
    # Ceci doit Ãªtre recalculé aprÃ¨s l'application des filtres pour avoir les listes pertinentes

    # Simuler les listes pour le contexte
    filiales = donnees.values_list('FILIALE', flat=True).distinct()
    agences = donnees.values_list('AGENCE', flat=True).distinct()
    exploitants = donnees.values_list('EXPL', flat=True).distinct()

    # 2. Obtenir le QuerySet filtré
    # Utilisation de 'donnees' qui contient déjÃ  le QuerySet filtré.
    queryset = donnees.order_by('id')

    # 3. Appliquer le Paginator
    # Je vais simuler ITEMS_PER_PAGE pour l'exemple
    ITEMS_PER_PAGE = 25
    paginator = Paginator(queryset, ITEMS_PER_PAGE)

    page_number = request.GET.get('page')
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

    try:
        objets_page = paginator.page(page_number)
    except PageNotAnInteger:
        # Si 'page' n'est pas un entier, afficher la première page
        objets_page = paginator.page(1)
    except EmptyPage:
        # Si la page est hors limites, afficher la dernière page
        objets_page = paginator.page(paginator.num_pages)

    # Devise PM KPIs
    total_devise = donnees.count()
    missing_devise = get_incomplete_clients_queryset(donnees, 'pm').count()
    complete_devise = max(0, total_devise - missing_devise)
    compliance_rate = round((complete_devise / total_devise) * 100, 1) if total_devise > 0 else 100.0

    # Répartition par devise
    devise_counts = list(
        donnees.values('DEVISE')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    devise_items = []
    for dc in devise_counts:
        dev_code = dc['DEVISE'] or "Non renseignée"
        pct = round((dc['count'] / total_devise) * 100, 1) if total_devise > 0 else 0.0
        devise_items.append({
            'label': dev_code,
            'value': dc['count'],
            'suffix': f'({pct}%)'
        })

    export_params = request.GET.copy()
    export_params['incompletes'] = '1'
    export_incomplets_url = f"{reverse('export_devise_pm')}?{export_params.urlencode()}"

    kpi_cards = [
        {
            'tone': 'emerald',
            'label': 'Répartition par Devise',
            'value': total_devise,
            'subtitle': 'Comptes PM en devise',
            'show_modal': True,
            'items': devise_items
        },
        {
            'tone': 'red',
            'label': 'Comptes PM Incomplets',
            'value': missing_devise,
            'subtitle': 'Dossiers avec données manquantes',
            'export_url': export_incomplets_url,
        },
        {
            'tone': 'blue',
            'label': 'Taux de complétude',
            'value': compliance_rate,
            'suffix': '%',
            'subtitle': 'Dossiers complets / total',
        }
    ]

    context = {
        # 'donnees' est maintenant l'objet Page paginé
        "donnees": objets_page,
        "devise_filiale": devise_valeur,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': query_params.urlencode(),
        'kpi_cards': kpi_cards,
    }

    return render(request, 'devise_pm.html', context)


@login_required
def export_devise_pm(request):
    user = request.user

    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]

    # Récupération des filtres GET pour la synchronisation
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')


    devise_obj = Devise.objects.filter(filiale=user.filiale).first()
    devise_valeur = devise_obj.devise if devise_obj else None  # Remplacez 'nom_devise' par le nom réel de votre champ

    # 2. Filtrage de base : Exclure la devise de la filiale, les vides et "NA"
    donnees = Kyc_pm.objects.exclude(DEVISE=devise_valeur).exclude(DEVISE__isnull=True).exclude(DEVISE="").exclude(DEVISE="NA")

    # === Filtrage automatique selon le rôle (identique Ã  devise) ===
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        pass  # on ne filtre pas davantage

    # === Filtres manuels via GET (synchronisation) ===
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)

    if request.GET.get('incompletes') == '1':
        donnees = get_incomplete_clients_queryset(donnees, 'pm')

    # Fin du Queryset filtré

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comptes Devise PP"  # J'ai renommé le titre

    # EntÃªtes
    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "AGEC", "CODAPE", "IDM", "RCSNO", "CAPITAL", "CA",
               "RESULTAT", "TEL"]
    ws.append(headers)

    # Données
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.AGEC, d.CODAPE, d.IDM,
            d.RCSNO, d.CAPITAL, d.CA, d.RESULTAT, d.TEL
        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    if request.GET.get('incompletes') == '1':
        filename = f"Comptes_en_devise_PM_incomplets_{date_str}.xlsx"
    else:
        filename = f"Comptes_en_devise_PM_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def evolution_taux(request):
    user = request.user

    context = {}

    if user.organe == "Conformité Groupe":
        # Groupe : on récupÃ¨re toutes les filiales distinctes
        filiales = TauxEvolution_filiale.objects.values_list("filiale", flat=True).distinct()
        data_filiales = {}

        for filiale in filiales:
            qs = TauxEvolution_filiale.objects.filter(filiale=filiale).order_by("id")

            dates = [str(i.id) for i in qs]  # tu peux remplacer par une vraie date si tu en ajoutes une
            taux_pp = [round((t.flux_PP / t.stock_PP) * 100, 2) if t.stock_PP else 0 for t in qs]
            taux_pm = [round((t.flux_PM / t.stock_PM) * 100, 2) if t.stock_PM else 0 for t in qs]

            data_filiales[filiale] = {
                "dates": dates,
                "taux_pp": taux_pp,
                "taux_pm": taux_pm,
            }

        context["data_filiales"] = data_filiales

    elif user.organe == "Conformité":
        # Filiale : uniquement sa propre
        qs = TauxEvolution_filiale.objects.filter(filiale=user.filiale).order_by("id")

        dates = [str(i.id) for i in qs]  # pareil, si tu as un champ Date, utilise-le
        taux_pp = [round((t.flux_PP / t.stock_PP) * 100, 2) if t.stock_PP else 0 for t in qs]
        taux_pm = [round((t.flux_PM / t.stock_PM) * 100, 2) if t.stock_PM else 0 for t in qs]

        context.update({
            "filiale": user.filiale,
            "dates_pp": dates,
            "taux_pp": taux_pp,
            "dates_pm": dates,
            "taux_pm": taux_pm,
        })

    return render(request, "statistiques.html", context)



@login_required
def taux_evolution_view(request):
    user = request.user
    if user.organe in ["Chargé Client", "Directeur Agence"]:
        from django.shortcuts import redirect
        return redirect('statistiques')
    context_cache_key = _build_dashboard_cache_key("evolution_filiale", user, request)
    cached_context = cache.get(context_cache_key)
    if cached_context is not None:
        return render(request, 'evolution_par_filiale.html', cached_context)

    user_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

    filiale_sel = request.GET.get('filiale')

    if user.organe in user_groupe:
        filiales = TauxEvolution_filiale.objects.values_list('filiale', flat=True).distinct().order_by('filiale')
    else:
        filiales = TauxEvolution_filiale.objects.filter(filiale=user.filiale).values_list('filiale', flat=True).distinct().order_by('filiale')
        if not filiale_sel or filiale_sel != user.filiale:
            filiale_sel = user.filiale

    if not filiale_sel and filiales:
        filiale_sel = filiales[0]

    periode = request.GET.get('periode', 'journalier')
    periodes_valides = {'journalier', 'hebdomadaire', 'mensuel', 'annuel'}
    if periode not in periodes_valides:
        periode = 'journalier'

    rows = list(
        TauxEvolution_filiale.objects
        .filter(filiale=filiale_sel)
        .order_by('date')
        .values_list('date', 'flux_PM', 'flux_PP')
    )
    latest_taux_date = rows[-1][0] if rows else None

    def build_period_key(d):
        if periode == 'journalier':
            return d
        if periode == 'hebdomadaire':
            year, week, _ = d.isocalendar()
            return (year, week)
        if periode == 'annuel':
            return d.year
        return (d.year, d.month)

    def format_period_label(key):
        if periode == 'journalier':
            return key.strftime('%d/%m/%Y')
        if periode == 'hebdomadaire':
            return f"S{key[1]}-{key[0]}"
        if periode == 'annuel':
            return str(key)
        return f"{key[1]:02d}/{key[0]}"

    grouped = {}
    for d, pm, pp in rows:
        key = build_period_key(d)
        bucket = grouped.setdefault(
            key,
            {"sum_pm": 0.0, "sum_pp": 0.0, "count": 0, "latest_date": d},
        )
        bucket["sum_pm"] += float(pm or 0)
        bucket["sum_pp"] += float(pp or 0)
        bucket["count"] += 1
        if d > bucket["latest_date"]:
            bucket["latest_date"] = d

    period_keys = sorted(grouped.keys())
    labels = [format_period_label(k) for k in period_keys]
    data_pm = [round(grouped[k]["sum_pm"] / grouped[k]["count"], 2) for k in period_keys]
    data_pp = [round(grouped[k]["sum_pp"] / grouped[k]["count"], 2) for k in period_keys]
    history_rows = [
        (
            grouped[k]["latest_date"],
            round(grouped[k]["sum_pm"] / grouped[k]["count"], 2),
            round(grouped[k]["sum_pp"] / grouped[k]["count"], 2),
            labels[idx],
        )
        for idx, k in enumerate(period_keys)
    ]

    kpi_pm = {'last': 0, 'diff': 0, 'status': 'up'}
    kpi_pp = {'last': 0, 'diff': 0, 'status': 'up'}

    if len(data_pm) >= 1:
        kpi_pm['last'] = round(data_pm[-1], 2)
        if len(data_pm) >= 2:
            kpi_pm['diff'] = round(data_pm[-1] - data_pm[-2], 2)
            kpi_pm['status'] = 'up' if kpi_pm['diff'] >= 0 else 'down'

    if len(data_pp) >= 1:
        kpi_pp['last'] = round(data_pp[-1], 2)
        if len(data_pp) >= 2:
            kpi_pp['diff'] = round(data_pp[-1] - data_pp[-2], 2)
            kpi_pp['status'] = 'up' if kpi_pp['diff'] >= 0 else 'down'

    quality_scope = evaluate_data_quality_scope(user)
    rules_version = cache.get('quality_control_rules_version', 1)

    def compute_quality_rate_by_typology(applicability):
        scope_signature = (
            f"{quality_scope.get('filiale')}|{quality_scope.get('agence')}|"
            f"{quality_scope.get('expl')}|{user.organe}|{applicability}|evolution_filiale"
        )
        scope_hash = hashlib.md5(scope_signature.encode('utf-8')).hexdigest()
        cache_key = f"quality_control:evolution_rate:v{rules_version}:{scope_hash}"
        cached_rate = cache.get(cache_key)
        if cached_rate is not None:
            return cached_rate

        rules = list(DataQualityRule.objects.filter(active=True, applicability=applicability))
        total_ok = 0
        total_evaluated = 0
        for rule in rules:
            stat = evaluate_data_quality_rule(
                rule,
                filiale=quality_scope.get('filiale'),
                agence=quality_scope.get('agence'),
                expl=quality_scope.get('expl'),
            )
            total_ok += stat.get('ok_count', 0)
            total_evaluated += stat.get('total', 0)

        rate = round(total_ok / total_evaluated * 100, 1) if total_evaluated else 0
        cache.set(cache_key, rate, timeout=3600)
        return rate

    context = {
        'labels_json': json.dumps(labels),
        'data_pm_json': json.dumps(data_pm),
        'data_pp_json': json.dumps(data_pp),
        'filiales': list(filiales),
        'filiale_sel': filiale_sel,
        'periode': periode,
        'kpi_pm': kpi_pm,
        'kpi_pp': kpi_pp,
        'quality_rate_pp': compute_quality_rate_by_typology('PP'),
        'quality_rate_pm': compute_quality_rate_by_typology('PM'),
        'quality_scope_label': quality_scope.get('label'),
        'history_rows': list(reversed(history_rows[-10:])),
        'latest_taux_date': latest_taux_date,
        'is_filiale_user': user.organe not in user_groupe,
    }
    cache.set(context_cache_key, context, timeout=60 * 60 * 8)
    return render(request, 'evolution_par_filiale.html', context)


@login_required
def taux_evolution_view_stock(request):
    user = request.user
    if user.organe in ["Chargé Client", "Directeur Agence"]:
        from django.shortcuts import redirect
        return redirect('statistiques')
    context_cache_key = _build_dashboard_cache_key("evolution_filiale_stock", user, request)
    cached_context = cache.get(context_cache_key)
    if cached_context is not None:
        return render(request, 'evolution_par_filiale_stock.html', cached_context)

    user_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

    filiale_sel = request.GET.get('filiale')

    if user.organe in user_groupe:
        filiales = TauxEvolution_filiale.objects.values_list('filiale', flat=True).distinct().order_by('filiale')
    else:
        filiales = TauxEvolution_filiale.objects.filter(filiale=user.filiale).values_list('filiale', flat=True).distinct().order_by('filiale')
        if not filiale_sel or filiale_sel != user.filiale:
            filiale_sel = user.filiale

    if not filiale_sel and filiales:
        filiale_sel = filiales[0]

    periode = request.GET.get('periode', 'journalier')
    periodes_valides = {'journalier', 'hebdomadaire', 'mensuel', 'annuel'}
    if periode not in periodes_valides:
        periode = 'journalier'

    rows = list(
        TauxEvolution_filiale.objects
        .filter(filiale=filiale_sel)
        .order_by('date')
        .values_list('date', 'stock_PM', 'stock_PP')
    )
    latest_taux_date = rows[-1][0] if rows else None

    def build_period_key(d):
        if periode == 'journalier':
            return d
        if periode == 'hebdomadaire':
            year, week, _ = d.isocalendar()
            return (year, week)
        if periode == 'annuel':
            return d.year
        return (d.year, d.month)

    def format_period_label(key):
        if periode == 'journalier':
            return key.strftime('%d/%m/%Y')
        if periode == 'hebdomadaire':
            return f"S{key[1]}-{key[0]}"
        if periode == 'annuel':
            return str(key)
        return f"{key[1]:02d}/{key[0]}"

    grouped = {}
    for d, pm, pp in rows:
        key = build_period_key(d)
        bucket = grouped.setdefault(
            key,
            {"sum_pm": 0.0, "sum_pp": 0.0, "count": 0, "latest_date": d},
        )
        bucket["sum_pm"] += float(pm or 0)
        bucket["sum_pp"] += float(pp or 0)
        bucket["count"] += 1
        if d > bucket["latest_date"]:
            bucket["latest_date"] = d

    period_keys = sorted(grouped.keys())
    labels = [format_period_label(k) for k in period_keys]
    data_pm = [round(grouped[k]["sum_pm"] / grouped[k]["count"], 2) for k in period_keys]
    data_pp = [round(grouped[k]["sum_pp"] / grouped[k]["count"], 2) for k in period_keys]
    history_rows = [
        (
            grouped[k]["latest_date"],
            round(grouped[k]["sum_pm"] / grouped[k]["count"], 2),
            round(grouped[k]["sum_pp"] / grouped[k]["count"], 2),
            labels[idx],
        )
        for idx, k in enumerate(period_keys)
    ]

    kpi_pm = {'last': 0, 'diff': 0, 'status': 'up'}
    kpi_pp = {'last': 0, 'diff': 0, 'status': 'up'}

    if len(data_pm) >= 1:
        kpi_pm['last'] = round(data_pm[-1], 2)
        if len(data_pm) >= 2:
            kpi_pm['diff'] = round(data_pm[-1] - data_pm[-2], 2)
            kpi_pm['status'] = 'up' if kpi_pm['diff'] >= 0 else 'down'

    if len(data_pp) >= 1:
        kpi_pp['last'] = round(data_pp[-1], 2)
        if len(data_pp) >= 2:
            kpi_pp['diff'] = round(data_pp[-1] - data_pp[-2], 2)
            kpi_pp['status'] = 'up' if kpi_pp['diff'] >= 0 else 'down'

    quality_scope = evaluate_data_quality_scope(user)
    rules_version = cache.get('quality_control_rules_version', 1)

    def compute_quality_rate_by_typology(applicability):
        scope_signature = (
            f"{quality_scope.get('filiale')}|{quality_scope.get('agence')}|"
            f"{quality_scope.get('expl')}|{user.organe}|{applicability}|evolution_filiale_stock"
        )
        scope_hash = hashlib.md5(scope_signature.encode('utf-8')).hexdigest()
        cache_key = f"quality_control:evolution_stock_rate:v{rules_version}:{scope_hash}"
        cached_rate = cache.get(cache_key)
        if cached_rate is not None:
            return cached_rate

        rules = list(DataQualityRule.objects.filter(active=True, applicability=applicability))
        total_ok = 0
        total_evaluated = 0
        for rule in rules:
            stat = evaluate_data_quality_rule(
                rule,
                filiale=quality_scope.get('filiale'),
                agence=quality_scope.get('agence'),
                expl=quality_scope.get('expl'),
            )
            total_ok += stat.get('ok_count', 0)
            total_evaluated += stat.get('total', 0)

        rate = round(total_ok / total_evaluated * 100, 1) if total_evaluated else 0
        cache.set(cache_key, rate, timeout=3600)
        return rate

    context = {
        'labels_json': json.dumps(labels),
        'data_pm_json': json.dumps(data_pm),
        'data_pp_json': json.dumps(data_pp),
        'filiales': list(filiales),
        'filiale_sel': filiale_sel,
        'periode': periode,
        'kpi_pm': kpi_pm,
        'kpi_pp': kpi_pp,
        'quality_rate_pp': compute_quality_rate_by_typology('PP'),
        'quality_rate_pm': compute_quality_rate_by_typology('PM'),
        'quality_scope_label': quality_scope.get('label'),
        'history_rows': list(reversed(history_rows[-10:])),
        'latest_taux_date': latest_taux_date,
        'is_filiale_user': user.organe not in user_groupe,
    }
    cache.set(context_cache_key, context, timeout=60 * 60 * 8)
    return render(request, 'evolution_par_filiale_stock.html', context)


import csv
import io
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required

User = get_user_model()


@login_required
def bulk_user_upload(request):
    if request.method == "POST":
        csv_file = request.FILES.get('file')

        if not csv_file or not csv_file.name.endswith('.csv'):
            messages.error(request, "Veuillez sélectionner un fichier CSV valide.")
            return redirect('bulk_user_upload')

        try:
            # Lecture du fichier
            data_set = csv_file.read().decode('UTF-8')
            io_string = io.StringIO(data_set)
            next(io_string)  # Sauter l'en-tÃªte

            users_created = 0
            errors = 0

            for row in csv.reader(io_string, delimiter=',', quotechar='"'):
                # row[0]=username, row[1]=first_name, row[2]=last_name, etc.
                # Adaptez les index selon l'ordre de vos colonnes CSV
                try:
                    user, created = User.objects.get_or_create(
                        username=row[0],
                        defaults={
                            'first_name': row[1],
                            'last_name': row[2],
                            'organe': row[3],
                            'téléphone': row[4],
                            'agence': row[6],
                            'code_expl': row[7],
                        }
                    )
                    if created:
                        user.set_password(row[5])  # password1
                        user.save()
                        users_created += 1
                except Exception:
                    errors += 1
                    continue

            messages.success(request, f"{users_created} utilisateurs créés avec succÃ¨s. ({errors} erreurs)")

        except Exception as e:
            messages.error(request, f"Erreur lors du traitement : {e}")

    return render(request, 'bulk_upload.html')


from openpyxl import Workbook
from django.http import HttpResponse


@login_required
def download_excel_template(request):
    # Création d'un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Utilisateurs"

    # En-tÃªtes conformes Ã  votre script
    headers = ['username', 'first_name', 'last_name', 'organe', 'téléphone', 'password', 'agence', 'expl']
    ws.append(headers)

    # Exemple de données
    ws.append(['m.diop', 'Moussa', 'Diop', 'Conformité', '771234567', 'Boa2026!', 'Agence Dakar', 'EXPL001'])

    # Préparation de la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="template_kyc_bulk.xlsx"'

    wb.save(response)
    return response


@login_required
def kyc_field_config(request):
    KYC_PP_FIELD_LABELS = [
        ("CLIENT", "CLIENT"),
        ("EXPL", "EXPL"),
        ("FILIALE", "FILIALE"),
        ("AGENCE", "AGENCE"),
        ("LIB_AGENCE", "LIB_AGENCE"),
        ("IDP", "IDP"),
        ("PAYNAIS", "PAYNAIS"),
        ("PROFESSION", "PROFESSION"),
        ("SALAIRE", "SALAIRE"),
        ("NUMID", "NUMID"),
        ("CODAPE", "CODAPE"),
        ("TEL", "TEL"),
        ("DATNAIS", "DATNAIS"),
        ("ADRESSE", "ADRESSE"),
        ("DATVALID", "DATVALID"),
        ("ORIGINE_REV", "ORIGINE_REV"),
        ("INTITULE_COMPTE", "INTITULE_COMPTE"),
        ("EMPLOYEUR", "EMPLOYEUR"),
        ("PAYS_RESID", "PAYS_RESID"),
        ("LIEU_DELIVRANCE_CIN", "LIEU_DELIVRANCE_CIN"),
        ("BOITE_POSTALE", "BOITE_POSTALE"),
        ("CONSENT_BIC", "CONSENT_BIC"),
        ("DATOUV", "DATOUV"),
        ("PPE", "PPE"),
        ("DEVISE", "DEVISE"),
        ("RESID", "RESID"),
        ("DATEREV", "DATEREV"),
        ("RISQUE", "RISQUE"),
    ]

    KYC_PM_FIELD_LABELS = [
        ("CLIENT", "CLIENT"),
        ("EXPL", "EXPL"),
        ("FILIALE", "FILIALE"),
        ("AGENCE", "AGENCE"),
        ("LIB_AGENCE", "LIB_AGENCE"),
        ("IDM", "IDM"),
        ("CODAPE", "CODAPE"),
        ("AGEC", "AGEC"),
        ("CAPITAL", "CAPITAL"),
        ("CA", "CA"),
        ("RESULTAT", "RESULTAT"),
        ("RCSNO", "RCSNO"),
        ("ORIGINE_REV", "ORIGINE_REV"),
        ("TEL", "TEL"),
        ("INTITULE_COMPTE", "INTITULE_COMPTE"),
        ("ADRESSE_SOCIALE", "ADRESSE_SOCIALE"),
        ("NUMERO_FISCAL", "NUMERO_FISCAL"),
        ("PAYS_JUR", "PAYS_JUR"),
        ("ACTIONNAIRE", "ACTIONNAIRE"),
        ("MANDATAIRE", "MANDATAIRE"),
        ("BOITE_POSTALE", "BOITE_POSTALE"),
        ("CONSENT_BIC", "CONSENT_BIC"),
        ("DATOUV", "DATOUV"),
        ("DEVISE", "DEVISE"),
        ("RESID", "RESID"),
        ("DATEREV", "DATEREV"),
        ("PPE", "PPE"),
        ("RISQUE", "RISQUE"),
    ]

    # Retrieve all unique filiales
    filiale_choices = sorted(list(set(
        list(Kyc_pp.objects.exclude(FILIALE="").values_list('FILIALE', flat=True).distinct()) +
        list(Kyc_pm.objects.exclude(FILIALE="").values_list('FILIALE', flat=True).distinct())
    )))

    selected_filiale = request.GET.get('filiale_modal', '').strip()

    if request.method == "POST":
        action = request.POST.get('action', '')
        if action == "save_filiale_modal":
            # Save configs for the modal popup of a specific filiale
            sel_filiale = request.POST.get('selected_filiale', '').strip()
            if sel_filiale:
                for ct in ['pp', 'pm']:
                    config_id = request.POST.get(f'{ct}_config_id', '')
                    empty_fields = request.POST.getlist(f'{ct}_empty_fields')
                    display_fields = request.POST.getlist(f'{ct}_display_fields')
                    
                    if config_id:
                        config = KycFieldVisibilityConfig.objects.get(id=config_id)
                        config.empty_check_fields = empty_fields
                        config.display_fields = display_fields
                        config.save()
                    else:
                        # Create specific config for this filiale
                        KycFieldVisibilityConfig.objects.create(
                            client_type=ct,
                            filiales=[sel_filiale],
                            empty_check_fields=empty_fields,
                            display_fields=display_fields
                        )
                messages.success(request, f"Configurations spécifiques pour la filiale {sel_filiale} enregistrées.")
            return redirect('kyc_field_config')
            
        else:
            # Save standard global or filiale configurations
            client_type = request.POST.get('client_type', 'pp')
            config_id = request.POST.get('config_id', '')
            scope = request.POST.get('scope', 'global')
            filiales = request.POST.getlist('filiales') if scope == 'filiales' else []
            empty_fields = request.POST.getlist('empty_fields')
            display_fields = request.POST.getlist('display_fields')

            if action == "delete":
                if config_id:
                    KycFieldVisibilityConfig.objects.filter(id=config_id).delete()
                    messages.success(request, "Configuration supprimée.")
                return redirect('kyc_field_config')

            if config_id:
                config = KycFieldVisibilityConfig.objects.get(id=config_id)
                config.empty_check_fields = empty_fields
                config.display_fields = display_fields
                if not config.filiales or scope == 'filiales':
                    config.filiales = filiales
                config.save()
            else:
                KycFieldVisibilityConfig.objects.create(
                    client_type=client_type,
                    filiales=filiales,
                    empty_check_fields=empty_fields,
                    display_fields=display_fields
                )
            messages.success(request, "Configuration enregistrée.")
            return redirect('kyc_field_config')

    # Query all current configs
    configs_qs = KycFieldVisibilityConfig.objects.all()
    
    # Ensure default global configs exist
    for ct in ['pp', 'pm']:
        global_exists = any(not c.filiales for c in configs_qs if c.client_type == ct)
        if not global_exists:
            default_fields = [f[0] for f in (KYC_PP_FIELD_LABELS if ct == 'pp' else KYC_PM_FIELD_LABELS)]
            KycFieldVisibilityConfig.objects.create(
                client_type=ct,
                filiales=[],
                empty_check_fields=[],
                display_fields=default_fields
            )
            configs_qs = KycFieldVisibilityConfig.objects.all()

    configs_list = list(configs_qs)
    for c in configs_list:
        c.is_global = not c.filiales or len(c.filiales) == 0
        if c.is_global:
            c.scope_label = "Toutes les filiales"
        else:
            c.scope_label = f"Filiales : {', '.join(c.filiales)}"
        c.empty_fields = c.empty_check_fields
        c.display_field_names = c.display_fields

    sections = [
        {
            'client_type': 'pp',
            'title': 'Particuliers (PP)',
            'fields': KYC_PP_FIELD_LABELS,
            'configs': [c for c in configs_list if c.client_type == 'pp']
        },
        {
            'client_type': 'pm',
            'title': 'Entreprises (PM)',
            'fields': KYC_PM_FIELD_LABELS,
            'configs': [c for c in configs_list if c.client_type == 'pm']
        }
    ]

    selected_filiale_configs = []
    if selected_filiale:
        for ct in ['pp', 'pm']:
            spec_config = None
            for c in configs_list:
                if c.client_type == ct and not c.is_global and selected_filiale in (c.filiales or []):
                    spec_config = c
                    break
            
            if spec_config:
                is_specific = True
                config_id = spec_config.id
                empty_fields = spec_config.empty_check_fields
                display_fields = spec_config.display_fields
                scope_label = f"Règle spécifique pour {selected_filiale}"
            else:
                is_specific = False
                global_c = next((c for c in configs_list if c.client_type == ct and c.is_global), None)
                config_id = None
                empty_fields = global_c.empty_check_fields if global_c else []
                display_fields = global_c.display_fields if global_c else [f[0] for f in (KYC_PP_FIELD_LABELS if ct == 'pp' else KYC_PM_FIELD_LABELS)]
                scope_label = "Hérité du global (Toutes les filiales)"

            selected_filiale_configs.append({
                'client_type': ct,
                'title': 'Particuliers (PP)' if ct == 'pp' else 'Entreprises (PM)',
                'config_id': config_id,
                'is_specific': is_specific,
                'scope_label': scope_label,
                'empty_fields': empty_fields,
                'display_field_names': display_fields,
                'fields': KYC_PP_FIELD_LABELS if ct == 'pp' else KYC_PM_FIELD_LABELS,
                'filiales': [selected_filiale]
            })

    context = {
        'filia': getattr(request.user, 'filiale', ''),
        'filiale_choices': filiale_choices,
        'sections': sections,
        'selected_filiale': selected_filiale,
        'selected_filiale_configs': selected_filiale_configs,
    }
    return render(request, 'kyc_field_config.html', context)


def get_rate_color(rate, threshold):
    if rate < threshold:
        return "#ef4444"
    elif rate < threshold + 5:
        return "#f59e0b"
    else:
        return "#10b981"


def export_pilotage_excel(scope_data, summary, completeness_rows, quality_rows, notations_list, notation_kpis):
    wb = Workbook()
    
    # Sheet 1: Synthèse
    ws1 = wb.active
    ws1.title = "Synthèse"
    ws1.append(["RAPPORT DE PILOTAGE KYC - BOA GROUP"])
    ws1.append([f"Périmètre: {scope_data.get('selected_filiale') or 'GROUPE'}"])
    ws1.append([f"Date de génération: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"])
    ws1.append([f"Seuil d'analyse: {summary.get('threshold', 90.0)}%"])
    ws1.append([])
    
    ws1.append(["Indicateur", "Valeur", "Unité", "Statut"])
    ws1.append(["Taux de complétude global", summary.get("completeness_rate"), "%", "Sous seuil" if (summary.get("completeness_rate") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Taux de complétude PP", summary.get("completeness_rate_pp"), "%", "Sous seuil" if (summary.get("completeness_rate_pp") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Taux de complétude PM", summary.get("completeness_rate_pm"), "%", "Sous seuil" if (summary.get("completeness_rate_pm") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Taux de conformité qualité global", summary.get("quality_rate"), "%", "Sous seuil" if (summary.get("quality_rate") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Taux de conformité qualité PP", summary.get("quality_rate_pp"), "%", "Sous seuil" if (summary.get("quality_rate_pp") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Taux de conformité qualité PM", summary.get("quality_rate_pm"), "%", "Sous seuil" if (summary.get("quality_rate_pm") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Nombre de champs sous seuil", summary.get("low_completeness_count"), "", ""])
    ws1.append(["Nombre de règles sous seuil", summary.get("low_quality_count"), "", ""])
    
    # Sheet 2: Complétude
    ws2 = wb.create_sheet(title="Complétude")
    ws2.append(["Type", "Périmètre", "Champ (Code)", "Champ (Libellé)", "Total Clients", "Incomplets", "Taux", "Conformité"])
    for row in completeness_rows:
        status = "Sous seuil" if row.get("is_below_threshold") else "Conforme"
        ws2.append([
            row.get("type"),
            row.get("filiale"),
            row.get("field_name"),
            row.get("field_label"),
            row.get("total_clients"),
            row.get("missing_count"),
            row.get("rate"),
            status
        ])
        
    # Sheet 3: Qualité
    ws3 = wb.create_sheet(title="Qualité")
    ws3.append(["Type", "Périmètre", "Règle", "Champ", "Total Clients", "Anomalies", "Taux", "Conformité"])
    for row in quality_rows:
        status = "Sous seuil" if row.get("is_below_threshold") else "Conforme"
        ws3.append([
            row.get("type"),
            row.get("scope_label"),
            row.get("rule_name"),
            row.get("field_label"),
            row.get("total_clients"),
            row.get("fail_count"),
            row.get("rate"),
            status
        ])
        
    # Sheet 4: Notation
    ws4 = wb.create_sheet(title="Notation")
    ws4.append(["Agent Evalué", "Code Exploitant", "Filiale", "Note", "Flux / Stock", "Recommandations", "Evalué par", "Date évaluation"])
    for n in notations_list:
        ws4.append([
            n.agent.username,
            getattr(n.agent, "code_expl", "N/A"),
            n.agent.filiale,
            n.note,
            n.flux_stock,
            n.recommandation or "",
            n.note_par.username,
            n.date_notation.strftime("%d/%m/%Y %H:%M") if n.date_notation else ""
        ])
        
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    scope_label = "GROUPE" if scope_data.get("scope") == "groupe" else scope_data.get("selected_filiale", "FILIALE")
    scope_safe = re.sub(r"[^A-Za-z0-9_-]+", "_", scope_label)
    date_file = timezone.localtime().strftime("%Y%m%d")
    response['Content-Disposition'] = f'attachment; filename="rapport_pilotage_kyc_{scope_safe}_{date_file}.xlsx"'
    wb.save(response)
    return response


@login_required
def pilotage_kyc(request):
    user = request.user
    
    # 1. Access security: can the user access group scope?
    user_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]
    can_group = user.is_superuser or user.organe in user_groupe
    
    # Allowed filiales list
    allowed_filiales = sorted(list(set(
        list(Kyc_pp.objects.exclude(FILIALE="").values_list('FILIALE', flat=True).distinct()) +
        list(Kyc_pm.objects.exclude(FILIALE="").values_list('FILIALE', flat=True).distinct())
    )))
    if not can_group:
        user_filiale = getattr(user, 'filiale', '')
        allowed_filiales = [user_filiale] if user_filiale else []
        scope = "filiale"
        selected_filiale = user_filiale
    else:
        scope = request.GET.get('scope', 'filiale')
        if scope not in ['filiale', 'groupe']:
            scope = 'filiale'
        
        if scope == 'filiale':
            selected_filiale = request.GET.get('filiale', '')
            if not selected_filiale and allowed_filiales:
                selected_filiale = allowed_filiales[0]
        else:
            selected_filiale = ""

    # Threshold
    try:
        threshold = float(request.GET.get('threshold', '90.0').replace(',', '.'))
    except ValueError:
        threshold = 90.0

    KYC_PP_FIELD_LABELS = [
        ("CLIENT", "CLIENT"),
        ("EXPL", "EXPL"),
        ("FILIALE", "FILIALE"),
        ("AGENCE", "AGENCE"),
        ("LIB_AGENCE", "LIB_AGENCE"),
        ("IDP", "IDP"),
        ("PAYNAIS", "PAYNAIS"),
        ("PROFESSION", "PROFESSION"),
        ("SALAIRE", "SALAIRE"),
        ("NUMID", "NUMID"),
        ("CODAPE", "CODAPE"),
        ("TEL", "TEL"),
        ("DATNAIS", "DATNAIS"),
        ("ADRESSE", "ADRESSE"),
        ("DATVALID", "DATVALID"),
        ("ORIGINE_REV", "ORIGINE_REV"),
        ("INTITULE_COMPTE", "INTITULE_COMPTE"),
        ("EMPLOYEUR", "EMPLOYEUR"),
        ("PAYS_RESID", "PAYS_RESID"),
        ("LIEU_DELIVRANCE_CIN", "LIEU_DELIVRANCE_CIN"),
        ("BOITE_POSTALE", "BOITE_POSTALE"),
        ("CONSENT_BIC", "CONSENT_BIC"),
        ("DATOUV", "DATOUV"),
        ("PPE", "PPE"),
        ("DEVISE", "DEVISE"),
        ("RESID", "RESID"),
        ("DATEREV", "DATEREV"),
        ("RISQUE", "RISQUE"),
    ]

    KYC_PM_FIELD_LABELS = [
        ("CLIENT", "CLIENT"),
        ("EXPL", "EXPL"),
        ("FILIALE", "FILIALE"),
        ("AGENCE", "AGENCE"),
        ("LIB_AGENCE", "LIB_AGENCE"),
        ("IDM", "IDM"),
        ("CODAPE", "CODAPE"),
        ("AGEC", "AGEC"),
        ("CAPITAL", "CAPITAL"),
        ("CA", "CA"),
        ("RESULTAT", "RESULTAT"),
        ("RCSNO", "RCSNO"),
        ("ORIGINE_REV", "ORIGINE_REV"),
        ("TEL", "TEL"),
        ("INTITULE_COMPTE", "INTITULE_COMPTE"),
        ("ADRESSE_SOCIALE", "ADRESSE_SOCIALE"),
        ("NUMERO_FISCAL", "NUMERO_FISCAL"),
        ("PAYS_JUR", "PAYS_JUR"),
        ("ACTIONNAIRE", "ACTIONNAIRE"),
        ("MANDATAIRE", "MANDATAIRE"),
        ("BOITE_POSTALE", "BOITE_POSTALE"),
        ("CONSENT_BIC", "CONSENT_BIC"),
        ("DATOUV", "DATOUV"),
        ("DEVISE", "DEVISE"),
        ("RESID", "RESID"),
        ("DATEREV", "DATEREV"),
        ("PPE", "PPE"),
        ("RISQUE", "RISQUE"),
    ]

    # PP Active fields
    pp_config = None
    if scope == 'filiale' and selected_filiale:
        pp_config = next((c for c in KycFieldVisibilityConfig.objects.filter(client_type='pp') if selected_filiale in (c.filiales or [])), None)
    if not pp_config:
        pp_config = next((c for c in KycFieldVisibilityConfig.objects.filter(client_type='pp') if not c.filiales), None)
        
    if pp_config and pp_config.empty_check_fields:
        pp_fields_list = pp_config.empty_check_fields
    elif pp_config and pp_config.display_fields:
        pp_fields_list = pp_config.display_fields
    else:
        pp_fields_list = [f[0] for f in KYC_PP_FIELD_LABELS]
        
    pp_active_fields = [(f_name, dict(KYC_PP_FIELD_LABELS).get(f_name, f_name)) for f_name in pp_fields_list]

    # PM Active fields
    pm_config = None
    if scope == 'filiale' and selected_filiale:
        pm_config = next((c for c in KycFieldVisibilityConfig.objects.filter(client_type='pm') if selected_filiale in (c.filiales or [])), None)
    if not pm_config:
        pm_config = next((c for c in KycFieldVisibilityConfig.objects.filter(client_type='pm') if not c.filiales), None)
        
    if pm_config and pm_config.empty_check_fields:
        pm_fields_list = pm_config.empty_check_fields
    elif pm_config and pm_config.display_fields:
        pm_fields_list = pm_config.display_fields
    else:
        pm_fields_list = [f[0] for f in KYC_PM_FIELD_LABELS]
        
    pm_active_fields = [(f_name, dict(KYC_PM_FIELD_LABELS).get(f_name, f_name)) for f_name in pm_fields_list]

    # Filter by user selection for report
    selected_report_fields = request.GET.getlist('report_fields')
    if selected_report_fields:
        pp_fields_to_analyze = [f for f in pp_active_fields if f[0] in selected_report_fields]
        pm_fields_to_analyze = [f for f in pm_active_fields if f[0] in selected_report_fields]
    else:
        pp_fields_to_analyze = pp_active_fields
        pm_fields_to_analyze = pm_active_fields

    # 3. Base Querysets
    if scope == 'filiale':
        pp_queryset = Kyc_pp.objects.filter(FILIALE=selected_filiale)
        pm_queryset = Kyc_pm.objects.filter(FILIALE=selected_filiale)
    else:
        pp_queryset = Kyc_pp.objects.filter(FILIALE__in=allowed_filiales)
        pm_queryset = Kyc_pm.objects.filter(FILIALE__in=allowed_filiales)

    total_pp = pp_queryset.count()
    total_pm = pm_queryset.count()

    # 4. Completeness calculations
    completeness_rows_pp = []
    total_evaluated_pp = 0
    total_missing_pp = 0
    
    for f_name, f_label in pp_fields_to_analyze:
        missing_count = pp_queryset.filter(Q(**{f"{f_name}__isnull": True}) | Q(**{f_name: ""})).count()
        rate = compliance_rate_floor(total_pp - missing_count, total_pp)
        if rate is None:
            rate = 100.0
        completeness_rows_pp.append({
            'type': 'PP',
            'filiale': selected_filiale if scope == 'filiale' else 'GROUPE',
            'field_name': f_name,
            'field_label': f_label,
            'total_clients': total_pp,
            'missing_count': missing_count,
            'rate': rate,
            'is_below_threshold': rate < threshold
        })
        total_evaluated_pp += total_pp
        total_missing_pp += missing_count

    completeness_rows_pm = []
    total_evaluated_pm = 0
    total_missing_pm = 0
    
    for f_name, f_label in pm_fields_to_analyze:
        missing_count = pm_queryset.filter(Q(**{f"{f_name}__isnull": True}) | Q(**{f_name: ""})).count()
        rate = compliance_rate_floor(total_pm - missing_count, total_pm)
        if rate is None:
            rate = 100.0
        completeness_rows_pm.append({
            'type': 'PM',
            'filiale': selected_filiale if scope == 'filiale' else 'GROUPE',
            'field_name': f_name,
            'field_label': f_label,
            'total_clients': total_pm,
            'missing_count': missing_count,
            'rate': rate,
            'is_below_threshold': rate < threshold
        })
        total_evaluated_pm += total_pm
        total_missing_pm += missing_count

    completeness_rows = completeness_rows_pp + completeness_rows_pm
    low_completeness_rows = [r for r in completeness_rows if r['is_below_threshold']]
    low_completeness_rows.sort(key=lambda r: r['rate'])

    # Aggregate completeness KPIs
    total_pp_compliant = max(total_evaluated_pp - total_missing_pp, 0)
    completeness_rate_pp = round((total_pp_compliant / total_evaluated_pp) * 100, 1) if total_evaluated_pp > 0 else 100.0
    
    total_pm_compliant = max(total_evaluated_pm - total_missing_pm, 0)
    completeness_rate_pm = round((total_pm_compliant / total_evaluated_pm) * 100, 1) if total_evaluated_pm > 0 else 100.0
    
    total_evaluated_global = total_evaluated_pp + total_evaluated_pm
    total_compliant_global = total_pp_compliant + total_pm_compliant
    completeness_rate = round((total_compliant_global / total_evaluated_global) * 100, 1) if total_evaluated_global > 0 else 100.0

    low_completeness_count_pp = sum(1 for r in completeness_rows_pp if r['is_below_threshold'])
    low_completeness_count_pm = sum(1 for r in completeness_rows_pm if r['is_below_threshold'])
    low_completeness_count = low_completeness_count_pp + low_completeness_count_pm

    # 5. Quality calculations
    all_quality_rules = DataQualityRule.objects.filter(active=True)
    selected_report_rules = request.GET.getlist('report_rules')
    if selected_report_rules:
        rules_to_evaluate = all_quality_rules.filter(id__in=selected_report_rules)
    else:
        rules_to_evaluate = all_quality_rules

    quality_rows_pp = []
    quality_rows_pm = []
    total_ok_pp = 0
    total_eval_rules_pp = 0
    total_ok_pm = 0
    total_eval_rules_pm = 0

    for rule in rules_to_evaluate:
        eval_fil = selected_filiale if scope == 'filiale' else None
        stat = evaluate_data_quality_rule(rule, filiale=eval_fil)
        total = stat.get('total', 0)
        fail_count = stat.get('fail_count', 0)
        ok_count = stat.get('ok_count', 0)
        rate = compliance_rate_floor(ok_count, total, fail_count)
        if rate is None:
            rate = 100.0

        row = {
            'id': rule.id,
            'type': rule.applicability,
            'scope_label': selected_filiale if scope == 'filiale' else 'GROUPE',
            'rule_name': rule.name,
            'field_label': dict(DATA_QUALITY_FIELD_CHOICES).get(rule.field_name, rule.field_name),
            'total_clients': total,
            'fail_count': fail_count,
            'rate': rate,
            'is_below_threshold': rate < threshold,
            'export_url': reverse('kyc:export_rule_failures', kwargs={'rule_id': rule.id})
        }

        if rule.applicability == 'PP':
            quality_rows_pp.append(row)
            total_ok_pp += ok_count
            total_eval_rules_pp += total
        else:
            quality_rows_pm.append(row)
            total_ok_pm += ok_count
            total_eval_rules_pm += total

    quality_rows = quality_rows_pp + quality_rows_pm
    low_quality_rows = [r for r in quality_rows if r['is_below_threshold']]
    low_quality_rows.sort(key=lambda r: r['rate'])

    # Aggregate quality KPIs
    quality_rate_pp = round((total_ok_pp / total_eval_rules_pp) * 100, 1) if total_eval_rules_pp > 0 else 100.0
    quality_rate_pm = round((total_ok_pm / total_eval_rules_pm) * 100, 1) if total_eval_rules_pm > 0 else 100.0
    
    total_ok_global = total_ok_pp + total_ok_pm
    total_eval_global = total_eval_rules_pp + total_eval_rules_pm
    quality_rate = round((total_ok_global / total_eval_global) * 100, 1) if total_eval_global > 0 else 100.0

    low_quality_count_pp = sum(1 for r in quality_rows_pp if r['is_below_threshold'])
    low_quality_count_pm = sum(1 for r in quality_rows_pm if r['is_below_threshold'])
    low_quality_count = low_quality_count_pp + low_quality_count_pm

    summary_dict = {
        'threshold': threshold,
        'completeness_rate': completeness_rate,
        'completeness_rate_pp': completeness_rate_pp,
        'completeness_rate_pm': completeness_rate_pm,
        'low_completeness_count': low_completeness_count,
        'low_completeness_count_pp': low_completeness_count_pp,
        'low_completeness_count_pm': low_completeness_count_pm,
        'completeness_total': total_pp + total_pm,
        'quality_rate': quality_rate,
        'quality_rate_pp': quality_rate_pp,
        'quality_rate_pm': quality_rate_pm,
        'low_quality_count': low_quality_count,
        'low_quality_count_pp': low_quality_count_pp,
        'low_quality_count_pm': low_quality_count_pm,
        'quality_total': total_eval_global
    }

    # 6. Notation and KPIs
    if scope == 'filiale':
        notations = Notation.objects.filter(agent__filiale=selected_filiale).select_related('agent', 'note_par')
    else:
        notations = Notation.objects.filter(agent__filiale__in=allowed_filiales).select_related('agent', 'note_par')

    total_notations = notations.count()
    total_agents = notations.values('agent').distinct().count()
    excellence_count = notations.filter(note__in=['Très Bien', 'Bien']).count()
    excellence_rate = round((excellence_count / total_notations) * 100, 1) if total_notations > 0 else 0.0

    notation_kpis = {
        'total_agents': total_agents,
        'total_notations': total_notations,
        'excellence_rate': excellence_rate
    }
    notations_list = list(notations.order_by('-date_notation'))

    # Exports
    export_format = request.GET.get('export')
    if export_format:
        scope_data = {
            'scope': scope,
            'selected_filiale': selected_filiale
        }
        if export_format == 'pdf':
            from kyc.pilotage_exports import export_pilotage_pdf
            return export_pilotage_pdf(scope_data, summary_dict, completeness_rows, quality_rows)
        elif export_format == 'pptx':
            from kyc.pilotage_exports import export_pilotage_pptx
            return export_pilotage_pptx(scope_data, summary_dict, completeness_rows, quality_rows)
        elif export_format == 'excel':
            return export_pilotage_excel(scope_data, summary_dict, completeness_rows, quality_rows, notations_list, notation_kpis)

    # 7. Chart preparation
    chart_comp_pp_labels = [r['field_label'] for r in completeness_rows_pp]
    chart_comp_pp_values = [r['rate'] for r in completeness_rows_pp]
    chart_comp_pp_colors = [get_rate_color(r['rate'], threshold) for r in completeness_rows_pp]

    chart_comp_pm_labels = [r['field_label'] for r in completeness_rows_pm]
    chart_comp_pm_values = [r['rate'] for r in completeness_rows_pm]
    chart_comp_pm_colors = [get_rate_color(r['rate'], threshold) for r in completeness_rows_pm]

    chart_qual_pp_labels = [r['rule_name'] for r in quality_rows_pp]
    chart_qual_pp_values = [r['rate'] for r in quality_rows_pp]
    chart_qual_pp_colors = [get_rate_color(r['rate'], threshold) for r in quality_rows_pp]

    chart_qual_pm_labels = [r['rule_name'] for r in quality_rows_pm]
    chart_qual_pm_values = [r['rate'] for r in quality_rows_pm]
    chart_qual_pm_colors = [get_rate_color(r['rate'], threshold) for r in quality_rows_pm]

    chart_notation_overall_labels = ['Très Bien', 'Bien', 'Passable', 'Insuffisant']
    chart_notation_overall_values = [
        notations.filter(note='Très Bien').count(),
        notations.filter(note='Bien').count(),
        notations.filter(note='Passable').count(),
        notations.filter(note='Insuffisant').count()
    ]
    chart_notation_overall_colors = ['#10b981', '#3b82f6', '#f59e0b', '#ef4444']

    chart_notation_filiales = sorted(list(set(notations.values_list('agent__filiale', flat=True).distinct())))
    chart_notation_filiales = [f for f in chart_notation_filiales if f]
    
    chart_notation_by_filiale_datasets = []
    notes_order = ['Très Bien', 'Bien', 'Passable', 'Insuffisant']
    colors_map = {
        'Très Bien': '#10b981',
        'Bien': '#3b82f6',
        'Passable': '#f59e0b',
        'Insuffisant': '#ef4444'
    }
    for note in notes_order:
        data = []
        for fil in chart_notation_filiales:
            count = notations.filter(agent__filiale=fil, note=note).count()
            data.append(count)
        chart_notation_by_filiale_datasets.append({
            'label': note,
            'data': data,
            'backgroundColor': colors_map[note]
        })

    context = {
        'scope': scope,
        'can_group': can_group,
        'allowed_filiales': allowed_filiales,
        'selected_filiale': selected_filiale,
        'threshold': threshold,
        'summary': summary_dict,
        
        'pp_active_fields': pp_active_fields,
        'pm_active_fields': pm_active_fields,
        'selected_report_fields': selected_report_fields,
        'selected_report_rules': [int(rid) for rid in selected_report_rules if rid.isdigit()],
        'all_quality_rules': all_quality_rules,
        
        'low_completeness_rows': low_completeness_rows,
        'low_quality_rows': low_quality_rows,
        
        'notations_list': notations_list,
        'notation_kpis': notation_kpis,
        
        # Charts
        'chart_comp_pp_labels': chart_comp_pp_labels,
        'chart_comp_pp_values': chart_comp_pp_values,
        'chart_comp_pp_colors': chart_comp_pp_colors,
        'chart_comp_pm_labels': chart_comp_pm_labels,
        'chart_comp_pm_values': chart_comp_pm_values,
        'chart_comp_pm_colors': chart_comp_pm_colors,
        
        'chart_qual_pp_labels': chart_qual_pp_labels,
        'chart_qual_pp_values': chart_qual_pp_values,
        'chart_qual_pp_colors': chart_qual_pp_colors,
        'chart_qual_pm_labels': chart_qual_pm_labels,
        'chart_qual_pm_values': chart_qual_pm_values,
        'chart_qual_pm_colors': chart_qual_pm_colors,
        
        'chart_notation_overall_labels': chart_notation_overall_labels,
        'chart_notation_overall_values': chart_notation_overall_values,
        'chart_notation_overall_colors': chart_notation_overall_colors,
        
        'chart_notation_filiales': chart_notation_filiales,
        'chart_notation_by_filiale_datasets': chart_notation_by_filiale_datasets,
        
        'pp_fields_json': json.dumps(pp_active_fields),
        'pm_fields_json': json.dumps(pm_active_fields),
    }

    return render(request, 'pilotage_kyc.html', context)








