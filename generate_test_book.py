import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

organes = [
    'Directeur Agence', 'Directeur de Zone', 'Chargé Client', 'Contrôle Permanent', 
    'Directeur Réseau', 'Contrôle Permanent Groupe', 'Conformité', 'Conformité Groupe', 
    'PASS', 'DSI', 'GUEST', 'Qualité', 'DAI', 'Risques'
]

onglets_base = [
    "Tableau de bord", "Agents notés", "Champs non-renseignés", "Clients en anomalie", 
    "Scoring Clients", "Screening KYC ID", "Nouvelle Notation", "Historique (Notations)", 
    "PPE (Conformité)", "Comptes spécifiques (Conformité)", "Paramètres Système", 
    "Régles de Qualité", "Champs KYC", "Types de documents", "Importation", "Pilotage"
]

def a_acces(organe, onglet):
    # Tableau de bord : tout le monde
    if onglet == "Tableau de bord":
        return "Oui"
    
    # Agents notés : tout le monde sauf Directeur Agence et Chargé Client
    if onglet == "Agents notés":
        return "Non" if organe in ["Directeur Agence", "Chargé Client"] else "Oui"
    
    # Menus de base : tout le monde
    if onglet in ["Champs non-renseignés", "Clients en anomalie", "Scoring Clients", "Screening KYC ID"]:
        return "Oui"
    
    # Notations
    if onglet in ["Nouvelle Notation", "Historique (Notations)"]:
        return "Oui" if organe in ["Contrôle Permanent", "PASS"] else "Non"
        
    # Conformité
    if onglet in ["PPE (Conformité)", "Comptes spécifiques (Conformité)"]:
        return "Oui" if organe in ["Conformité", "Conformité Groupe", "PASS"] else "Non"
        
    # Administration DSI / PASS
    if onglet in ["Paramètres Système", "Régles de Qualité"]:
        return "Oui" if organe in ["PASS", "DSI"] else "Non"
        
    # Administration PASS seul
    if onglet in ["Champs KYC", "Types de documents", "Importation", "Pilotage"]:
        return "Oui" if organe == "PASS" else "Non"
        
    return "Non"

def get_specific_tests(organe, onglet):
    tests = []
    
    if onglet == "Tableau de bord":
        tests.append("Allez sur le Tableau de bord. Regardez si les gros chiffres en haut et les graphiques s'affichent correctement à l'écran.")
        if organe in ["Directeur Agence", "Chargé Client"]:
            tests.append("Vérifiez que vous ne voyez que les informations de votre propre agence ou de vos propres clients, et non celles des autres agences.")
        else:
            tests.append("Choisissez une agence ou un exploitant dans les petites cases en haut (les filtres). Vérifiez que les chiffres et les graphiques changent pour correspondre à votre choix.")
            
    elif onglet == "Nouvelle Notation":
        tests.append("Allez dans 'Nouvelle Notation'. Tapez un vrai code d'exploitant dans la case prévue.")
        tests.append("Vérifiez que le nom et le prénom du Chargé Client s'affichent tout seuls une fois le code tapé.")
        tests.append("Essayez de valider sans remplir toutes les cases obligatoires : l'application doit vous en empêcher avec un message rouge.")
        tests.append("Remplissez tout correctement, cliquez sur le bouton de validation, et vérifiez qu'un petit message de succès vert apparaît.")
        
    elif onglet == "Historique (Notations)":
        tests.append("Allez dans 'Historique'. Cherchez dans la liste pour voir si vous retrouvez la notation que vous venez juste de créer.")
        tests.append("Tapez un nom dans la barre de recherche pour vérifier que ça trouve bien la bonne personne.")
        tests.append("Cliquez sur les numéros en bas de page (1, 2, 3...) pour vérifier qu'on peut bien passer d'une page à l'autre.")
        
    elif onglet in ["PPE (Conformité)", "Comptes spécifiques (Conformité)"]:
        tests.append("Allez dans cet onglet. Vérifiez que la liste des clients s'affiche bien au milieu de l'écran.")
        tests.append("Essayez de chercher un nom précis dans la barre de recherche et vérifiez que le tableau se met à jour.")
        
    elif onglet in ["Champs non-renseignés", "Clients en anomalie"]:
        tests.append("Allez dans cet onglet. Vérifiez qu'un grand tableau avec les clients à problème s'affiche bien.")
        tests.append("S'il y a un bouton 'Télécharger' ou 'Export Excel', cliquez dessus et vérifiez que le fichier arrive bien sur votre ordinateur.")
        
    elif onglet == "Paramètres Système":
        tests.append("Allez dans les Paramètres Système. Cliquez sur le bouton pour créer un nouvel utilisateur (inventez un nom) et enregistrez-le.")
        tests.append("Retrouvez cet utilisateur dans la liste, cliquez pour le modifier (changez son prénom par exemple) et enregistrez.")
        tests.append("Essayez de supprimer ou bloquer cet utilisateur, puis vérifiez qu'il disparaît ou apparaît comme bloqué.")
        
    elif onglet == "Agents notés":
        tests.append("Allez dans 'Agents notés'. Vérifiez que vous voyez bien une liste de personnes avec leurs notes ou scores.")
        tests.append("Cliquez sur une des personnes dans la liste et vérifiez qu'on vous montre bien le détail de ses notes.")
        
    if not tests:
        tests.append("Cliquez sur cet onglet. Vérifiez simplement que la page s'ouvre normalement, qu'elle n'est pas toute blanche et qu'il n'y a pas de message d'erreur incompréhensible à l'écran.")
    
    return tests

# Create the excel file
wb = Workbook()

# Styling tokens
header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Slate Premium
header_font = Font(name="Segoe UI", size=11, color="FFFFFF", bold=True)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

cell_font = Font(name="Segoe UI", size=10, color="334155")
cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

border_thin = Border(
    left=Side(style='thin', color="E2E8F0"), 
    right=Side(style='thin', color="E2E8F0"), 
    top=Side(style='thin', color="E2E8F0"), 
    bottom=Side(style='thin', color="E2E8F0")
)

alt_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
alt_fill_2 = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

headers = [
    "Onglet à Tester", 
    "Accès Prévu", 
    "Tests Spécifiques à Réaliser", 
    "Résultat attendu", 
    "Résultat obtenu", 
    "Date test", 
    "Motif KO/Commentaire", 
    "Copie écran associée"
]

def apply_header_style(ws, headers_list):
    ws.append(headers_list)
    ws.row_dimensions[1].height = 35
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_thin

# First sheet: Recap
ws_recap = wb.active
ws_recap.title = "Matrice des Accès"
apply_header_style(ws_recap, ["Organe", "Onglet", "Accès Prévu", "Statut du Test (OK/KO)", "Commentaires"])

row_num = 2
for organe in organes:
    for onglet in onglets_base:
        acces = a_acces(organe, onglet)
        ws_recap.append([organe, onglet, acces, "", ""])
        for cell in ws_recap[row_num]:
            cell.font = cell_font
            cell.border = border_thin
            cell.alignment = center_alignment
            cell.fill = alt_fill_1 if row_num % 2 == 0 else alt_fill_2
        row_num += 1

# Adjust columns for Recap
for col in ws_recap.columns:
    ws_recap.column_dimensions[col[0].column_letter].width = 25

# Add individual sheets for each Organe
for organe in organes:
    ws = wb.create_sheet(title=organe[:31])
    apply_header_style(ws, headers)
    
    current_row = 2
    
    for idx_onglet, onglet in enumerate(onglets_base):
        acces = a_acces(organe, onglet)
        attendu = "Fonctionnement sans erreur" if acces == "Oui" else "Accès refusé / invisible"
        
        start_row = current_row
        
        if acces == "Oui":
            tests_specifiques = get_specific_tests(organe, onglet)
            for test in tests_specifiques:
                ws.append([onglet, acces, test, attendu, "", "", "", ""])
                current_row += 1
        else:
            ws.append([onglet, acces, "N/A", attendu, "", "", "", ""])
            current_row += 1
            
        # Merge cells for "Onglet", "Accès Prévu", "Résultat attendu"
        if current_row - 1 > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row-1, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row-1, end_column=2)
            ws.merge_cells(start_row=start_row, start_column=4, end_row=current_row-1, end_column=4)
            
        # Apply styles
        group_fill = alt_fill_1 if idx_onglet % 2 == 0 else alt_fill_2
        
        for row in range(start_row, current_row):
            ws.row_dimensions[row].height = 45 # Make rows taller for better readability
            for col_idx, cell in enumerate(ws[row], 1):
                cell.font = cell_font
                cell.border = border_thin
                
                # Alignments
                if col_idx in [1, 2, 4]:
                    cell.alignment = center_alignment
                else:
                    cell.alignment = cell_alignment
                
                # Highlights for "Accès Prévu"
                if col_idx == 2:
                    if acces == "Oui":
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Emerald-100
                        cell.font = Font(name="Segoe UI", size=10, color="065F46", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Red-100
                        cell.font = Font(name="Segoe UI", size=10, color="991B1B", bold=True)
                else:
                    cell.fill = group_fill

    # Set column widths
    ws.column_dimensions['A'].width = 25 # Onglet
    ws.column_dimensions['B'].width = 15 # Accès
    ws.column_dimensions['C'].width = 50 # Tests
    ws.column_dimensions['D'].width = 22 # Résultat attendu
    ws.column_dimensions['E'].width = 20 # Obtenu
    ws.column_dimensions['F'].width = 15 # Date
    ws.column_dimensions['G'].width = 35 # Motif
    ws.column_dimensions['H'].width = 25 # Ecran

wb.save("Cahier_de_Tests_KYC_v5.xlsx")
print("Cahier de tests généré avec succès : Cahier_de_Tests_KYC_v5.xlsx")
