"""Onglet Administration > Audit : piste d'audit unifiee de la plateforme KYC.

Agrege dans une seule chronologie :
  - les evenements du journal d'audit (AuditEvent) : connexions, deconnexions,
    echecs d'authentification, habilitations, actions sensibles ;
  - l'historique de connexion anterieur au journal (UserLoginHistory) ;
  - les audits de regles de qualite (DataQualityRuleAudit) ;
  - les traitements Screening KYC ID (jobs OCR, rapprochements, decisions) ;
  - les notations des agents (Notation).

Perimetre de consultation : PASS voit l'ensemble du Groupe ; au niveau
filiale, les organes DSI, Conformite, Qualite et Controle Permanent ne
consultent que les evenements de leur propre filiale.
"""

from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date

from accounts.audit import log_audit
from accounts.models import AuditEvent, UserLoginHistory
from kyc.models import (
    DataQualityRuleAudit, KycDocumentMatchJob, KycDocumentOcrJob, KycMatchDecision,
    Notation,
)

                                                                            
                                                                            
                                                                                
AUDIT_ORGANES = (
    "PASS", "Conformité Groupe", "Contrôle Permanent Groupe",
    "DSI", "Conformité", "Qualité", "Contrôle Permanent",
)

                                                                  
AUDIT_GROUP_ORGANES = ("PASS", "Conformité Groupe", "Contrôle Permanent Groupe")

                                                                        
SOURCE_LIMIT = 3000

                                                                              
                                       
EXCLUDED_CATEGORIES = (AuditEvent.CAT_IMPORT,)

VISIBLE_CATEGORIES = tuple(
    (code, label) for code, label in AuditEvent.CATEGORIES if code not in EXCLUDED_CATEGORIES
)

CATEGORY_LABELS = dict(AuditEvent.CATEGORIES)

CATEGORY_STYLES = {
    AuditEvent.CAT_CONNEXION: ("bg-emerald-50", "text-emerald-600", "border-emerald-100"),
    AuditEvent.CAT_HABILITATION: ("bg-indigo-50", "text-indigo-600", "border-indigo-100"),
    AuditEvent.CAT_IMPORT: ("bg-amber-50", "text-amber-600", "border-amber-100"),
    AuditEvent.CAT_DONNEES: ("bg-blue-50", "text-blue-600", "border-blue-100"),
    AuditEvent.CAT_CONFIG: ("bg-violet-50", "text-violet-600", "border-violet-100"),
    AuditEvent.CAT_EXPORT: ("bg-teal-50", "text-teal-600", "border-teal-100"),
    AuditEvent.CAT_SCREENING: ("bg-cyan-50", "text-cyan-600", "border-cyan-100"),
    AuditEvent.CAT_NOTATION: ("bg-orange-50", "text-orange-600", "border-orange-100"),
    AuditEvent.CAT_SECURITE: ("bg-red-50", "text-red-600", "border-red-100"),
}


def _can_view_audit(user):
    return bool(
        user.is_authenticated
        and (user.is_superuser or (getattr(user, "organe", "") or "").strip() in AUDIT_ORGANES)
    )


def _user_display(user):
    if user is None:
        return "Systeme", "", ""
    full = (user.get_full_name() or "").strip()
    return (full or user.username), (user.filiale or ""), (user.organe or "")


def _row(dt, category, action, *, user=None, user_label=None, filiale="", organe="",
         target="", details="", ip="", success=True, source=""):
    if user is not None:
        label, fil, org = _user_display(user)
        user_label = user_label or label
        filiale = filiale or fil
        organe = organe or org
    cat_bg, cat_text, cat_border = CATEGORY_STYLES.get(category, ("bg-slate-50", "text-slate-600", "border-slate-100"))
    label = user_label or "Systeme"
    return {
        "dt": dt,
        "category": category,
        "category_label": CATEGORY_LABELS.get(category, category),
        "cat_bg": cat_bg,
        "cat_text": cat_text,
        "cat_border": cat_border,
        "action": action,
        "user_label": label,
        "initials": (label[:2] or "SY").upper(),
        "filiale": filiale or "-",
        "organe": organe or "-",
        "target": target or "-",
        "details": details or "",
        "ip": ip or "-",
        "success": success,
        "source": source,
    }


def _scope_filiale(user):
    """Filiale imposee a l'utilisateur : None = perimetre Groupe.

    Les organes filiale (DSI, Conformite, Qualite, Controle Permanent) ne
    consultent que les evenements de leur propre filiale.
    """
    if user.is_superuser:
        return None
    organe = (getattr(user, "organe", "") or "").strip()
    if organe in AUDIT_GROUP_ORGANES:
        return None
    return (getattr(user, "filiale", "") or "").strip() or None


def _collect_rows(scope_filiale=None):
    """Construit la chronologie unifiee a partir de toutes les sources.

    `scope_filiale` restreint la lecture aux evenements d'une filiale (DSI).
    """
    rows = []

                              
    events = AuditEvent.objects.exclude(category__in=EXCLUDED_CATEGORIES)
    if scope_filiale:
        events = events.filter(Q(filiale=scope_filiale) | Q(user__filiale=scope_filiale))
    for ev in events.select_related("user").order_by("-timestamp")[:SOURCE_LIMIT]:
        label = ev.username or (_user_display(ev.user)[0] if ev.user else "Systeme")
        rows.append(_row(
            ev.timestamp, ev.category, ev.action,
            user_label=label,
            filiale=ev.filiale, organe=ev.organe,
            target=ev.target, details=ev.details, ip=ev.ip_address or "",
            success=ev.success, source="AuditEvent",
        ))

                                                                          
                                                              
    first_login_event = (AuditEvent.objects
                         .filter(category=AuditEvent.CAT_CONNEXION)
                         .order_by("timestamp")
                         .values_list("timestamp", flat=True)
                         .first())
    legacy = UserLoginHistory.objects.select_related("user")
    if scope_filiale:
        legacy = legacy.filter(user__filiale=scope_filiale)
    if first_login_event:
                                                                                
                                                       
        legacy = legacy.filter(login_at__lt=first_login_event - timedelta(seconds=5))
    for ev in legacy.order_by("-login_at")[:SOURCE_LIMIT]:
        rows.append(_row(
            ev.login_at, AuditEvent.CAT_CONNEXION, "Connexion",
            user=ev.user,
            target=getattr(ev.user, "username", ""),
            details="Connexion enregistree dans l'historique des acces.",
            source="UserLoginHistory",
        ))

                                      
    rule_audits = DataQualityRuleAudit.objects.all()
    if scope_filiale:
        rule_audits = rule_audits.filter(user__filiale=scope_filiale)
    for a in rule_audits.select_related("user").order_by("-timestamp")[:SOURCE_LIMIT]:
        rows.append(_row(
            a.timestamp, AuditEvent.CAT_CONFIG,
            f"Regle qualite - {(a.action or '').capitalize()}",
            user=a.user,
            target=(a.rule_name or "").strip() or "Regle",
            details=a.details or "",
            source="DataQualityRuleAudit",
        ))

                                                          
    match_jobs = KycDocumentMatchJob.objects.all()
    if scope_filiale:
        match_jobs = match_jobs.filter(created_by__filiale=scope_filiale)
    for j in match_jobs.select_related("created_by").order_by("-created_at")[:SOURCE_LIMIT]:
        rows.append(_row(
            j.created_at, AuditEvent.CAT_SCREENING, "Rapprochement documents",
            user=j.created_by,
            target=f"Job #{j.pk}",
            details=f"Statut : {j.get_status_display()} | Progression : {j.progress_current}/{j.progress_total}"
                    + (f" | {j.message}" if j.message else "")
                    + (f" | Erreur : {j.error[:200]}" if j.error else ""),
            success=(j.status != "failed"),
            source="KycDocumentMatchJob",
        ))

    ocr_jobs = KycDocumentOcrJob.objects.all()
    if scope_filiale:
        ocr_jobs = ocr_jobs.filter(created_by__filiale=scope_filiale)
    for j in ocr_jobs.select_related("created_by").order_by("-created_at")[:SOURCE_LIMIT]:
        rows.append(_row(
            j.created_at, AuditEvent.CAT_SCREENING, "Chargement / OCR de lot",
            user=j.created_by,
            target=f"Lot {j.import_batch}",
            details=f"Statut : {j.get_status_display()} | Traites : {j.done_count} | Echecs : {j.failed_count}"
                    + (f" | {j.message}" if j.message else ""),
            success=(j.status != "failed"),
            source="KycDocumentOcrJob",
        ))

    decisions = KycMatchDecision.objects.exclude(decided_at__isnull=True)
    if scope_filiale:
        decisions = decisions.filter(
            Q(filiale=scope_filiale) | Q(decided_by__filiale=scope_filiale)
        )
    for d in (decisions.select_related("decided_by")
              .order_by("-decided_at")[:SOURCE_LIMIT]):
        rows.append(_row(
            d.decided_at, AuditEvent.CAT_SCREENING,
            f"Correspondance - {d.get_status_display()}",
            user=d.decided_by,
            filiale=d.filiale or "",
            target=d.client_code or f"Client #{d.client_id}",
            details=f"Taux de correspondance : {d.match_rate}% | Agence : {d.agence or '-'}"
                    + (f" | {d.notes[:200]}" if d.notes else ""),
            success=(d.status != "rejected"),
            source="KycMatchDecision",
        ))

                             
    notations = Notation.objects.all()
    if scope_filiale:
        notations = notations.filter(
            Q(note_par__filiale=scope_filiale) | Q(agent__filiale=scope_filiale)
        )
    for n in (notations.select_related("agent", "note_par")
              .order_by("-date_notation")[:SOURCE_LIMIT]):
        agent_label, agent_filiale, _ = _user_display(n.agent)
        code = (getattr(n.agent, "code_expl", "") or "").strip()
        rows.append(_row(
            n.date_notation, AuditEvent.CAT_NOTATION,
            f"Notation - {n.note}",
            user=n.note_par,
            filiale=agent_filiale,
            target=f"{agent_label}" + (f" ({code})" if code else ""),
            details=(f"Flux/Stock : {n.flux_stock or '-'}"
                     + (f" | Recommandation : {n.recommandation[:200]}" if n.recommandation else "")),
            source="Notation",
        ))

    rows.sort(key=lambda r: r["dt"] or timezone.now(), reverse=True)
    return rows


def _apply_filters(rows, *, category, filiale, user_label, date_from, date_to, query, status):
    def keep(r):
        if category and r["category"] != category:
            return False
        if filiale and r["filiale"] != filiale:
            return False
        if user_label and r["user_label"] != user_label:
            return False
        if status == "ko" and r["success"]:
            return False
        if status == "ok" and not r["success"]:
            return False
        dt = r["dt"]
        if dt is not None:
            local = timezone.localtime(dt) if timezone.is_aware(dt) else dt
            if date_from and local.date() < date_from:
                return False
            if date_to and local.date() > date_to:
                return False
        if query:
            haystack = " ".join([
                r["user_label"], r["action"], r["target"], r["details"],
                r["filiale"], r["organe"], r["ip"], r["category_label"],
            ]).lower()
            if query not in haystack:
                return False
        return True

    return [r for r in rows if keep(r)]


def _read_filters(request):
    return {
        "category": (request.GET.get("category") or "").strip(),
        "filiale": (request.GET.get("filiale") or "").strip(),
        "user_label": (request.GET.get("acteur") or "").strip(),
        "date_from": parse_date((request.GET.get("date_from") or "").strip() or "1900-01-01"),
        "date_to": parse_date((request.GET.get("date_to") or "").strip() or "2999-12-31"),
        "query": (request.GET.get("q") or "").strip().lower(),
        "status": (request.GET.get("status") or "").strip(),
    }


@login_required
def audit_view(request):
    if not _can_view_audit(request.user):
        messages.error(request, "Acces refuse : la piste d'audit est reservee a l'administration.")
        return redirect("accueil")

    scope = _scope_filiale(request.user)
    filters = _read_filters(request)
    if scope:
                                                                                  
                                                 
        filters["filiale"] = ""
    all_rows = _collect_rows(scope)
    rows = _apply_filters(all_rows, **filters)

    now = timezone.localtime(timezone.now())
    today = now.date()
    stats = {
        "total": len(rows),
        "connexions": sum(1 for r in rows if r["category"] == AuditEvent.CAT_CONNEXION),
        "incidents": sum(1 for r in rows if not r["success"]),
        "aujourdhui": sum(
            1 for r in rows
            if r["dt"] and (timezone.localtime(r["dt"]) if timezone.is_aware(r["dt"]) else r["dt"]).date() == today
        ),
    }

    paginator = Paginator(rows, 50)
    page = paginator.get_page(request.GET.get("page"))

    querystring = request.GET.copy()
    querystring.pop("page", None)

    context = {
        "page_obj": page,
        "rows": page.object_list,
        "stats": stats,
        "categories": VISIBLE_CATEGORIES,
        "scope_filiale": scope,
        "filiales": sorted({r["filiale"] for r in all_rows if r["filiale"] and r["filiale"] != "-"}),
        "acteurs": sorted({r["user_label"] for r in all_rows if r["user_label"]}),
        "f_category": filters["category"],
        "f_filiale": filters["filiale"],
        "f_acteur": filters["user_label"],
        "f_status": filters["status"],
        "f_q": request.GET.get("q", ""),
        "f_date_from": request.GET.get("date_from", ""),
        "f_date_to": request.GET.get("date_to", ""),
        "querystring": querystring.urlencode(),
        "current_name": "audit",
    }
    return render(request, "audit.html", context)


@login_required
def export_audit_excel(request):
    if not _can_view_audit(request.user):
        messages.error(request, "Acces refuse : la piste d'audit est reservee a l'administration.")
        return redirect("accueil")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    scope = _scope_filiale(request.user)
    filters = _read_filters(request)
    if scope:
        filters["filiale"] = ""
    rows = _apply_filters(_collect_rows(scope), **filters)

    log_audit(
        request,
        category=AuditEvent.CAT_EXPORT,
        action="Export de la piste d'audit",
        target="piste_audit_kyc.xlsx",
        details=f"{len(rows)} evenement(s) exporte(s) | Filtres : {request.GET.urlencode() or 'aucun'}",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Piste d'audit"
    headers = ["Date & heure", "Categorie", "Action", "Acteur", "Organe", "Filiale",
               "Cible", "Details", "Adresse IP", "Statut", "Source"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="123B6D")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for r in rows:
        dt = r["dt"]
        if dt is not None and timezone.is_aware(dt):
            dt = timezone.localtime(dt)
        ws.append([
            dt.strftime("%d/%m/%Y %H:%M:%S") if dt else "",
            r["category_label"], r["action"], r["user_label"], r["organe"], r["filiale"],
            r["target"], r["details"], r["ip"],
            "Succes" if r["success"] else "Echec", r["source"],
        ])

    for col, width in zip("ABCDEFGHIJK", (20, 22, 34, 26, 22, 14, 32, 70, 16, 12, 22)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    stamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M")
    response["Content-Disposition"] = f'attachment; filename=piste_audit_kyc_{stamp}.xlsx'
    wb.save(response)
    return response
