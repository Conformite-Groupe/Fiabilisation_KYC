
import datetime
from django.db.models import Q
import labels
from django.db.models import OuterRef, Subquery, F, DateField, FloatField

import json

from django.db.models import Max

import json
from datetime import timedelta
from django.db.models import OuterRef, Subquery
from django.db.models import Q
import json
from .models import TauxEvolution, TAUX_FILIALE
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.hashers import make_password
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.views import PasswordChangeView
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.sites import requests
from django.core.mail import send_mail, BadHeaderError
from django.db.models import Q
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO


from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.views import generic

from kyc.models import Notation, Historique, Kyc_pm, Kyc_pp, Anomalie, TauxEvolution, DATEREV
from django.utils import timezone
from django.utils.decorators import method_decorator

from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode
from django.views.decorators.csrf import csrf_exempt
from oauthlib.oauth2.rfc6749.endpoints import token

from django.http import JsonResponse
from .models import Kyc_pp

from accounts.models import ProfileV
from kyc import forms
from kyc.forms import CustomUserCreationForm, LoginForm, ResetPasswordForm, UserEditForm, VoyageurProfileForm, \
    CambProfileForm, \
    ProfileModify, NotationForm, ProfileForm
from django.contrib.sessions.models import Session
from django.utils.timezone import now



@login_required
def accueil(request):
    return render(request, 'accueil.html')

def profile_update(request):
    if request.method == "POST":
        form = ProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect("profile")  # redirige après update
    else:
        form = ProfileForm(instance=request.user)
    return render(request, "profile_update.html", {"form": form})


@csrf_exempt
def rechercher_et_noter_agent(request):
    agent = None
    form = None
    message = None

    if request.method == "POST":
        expl = request.POST.get('expl', '')
        if expl:
            try:
                agent = ProfileV.objects.get(code_expl=expl)
            except ProfileV.DoesNotExist:
                message = "Agent introuvable."

            if agent:
                # Si l'utilisateur a le profil "Contrôle permanent"
                if request.user.is_authenticated and request.user.groups.filter(name='Contrôle permanent').exists():
                    form = NoteForm(request.POST or None)
                    if form.is_valid():
                        note = form.save(commit=False)
                        note.agent = agent
                        note.date_notation = timezone.now()  # Enregistrer la date de notation
                        note.save()
                        message = "Notation enregistrée avec succès."
                        form = None  # Reset le formulaire après enregistrement pour éviter la soumission répétée
                else:
                    message = "Vous n'avez pas la permission de noter cet agent."

    return render(request, 'accueil.html', {'agent': agent, 'form': form, 'message': message})


@csrf_exempt
def password_reset_request(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data['email']
            associated_users = ProfileV.objects.filter(Q(username=data))
            if associated_users.exists():
                for user in associated_users:
                    subject = "Password Reset Requested"
                    email_template_name = "password_reset_email.txt"
                    c = {
                        "email": user.username,
                        'domain': '127.0.0.1:8000',
                        'site_name': 'BuurChange',
                        "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                        'token': default_token_generator.make_token(user),
                        'protocol': 'http',
                    }
                    email = render_to_string(email_template_name, c)
                    try:
                        send_mail(subject, email, 'mamadou@mamadou.sn', [user.username], fail_silently=False)
                    except BadHeaderError:
                        return HttpResponse('Invalid header found.')
                    return redirect('/password_resete/done')

                    messages.success(request, 'A message with reset password instructions has been sent to your inbox.')
                    return redirect("accueil")
            else:
                errors = 'Votre mail ne figure pas dans notre base.'
                return render(request=request, template_name="password_reset.html",
                              context={"password_reset_form": password_reset_form, "errors": errors})

    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="password_reset.html",
                  context={"password_reset_form": password_reset_form})


@login_required
@csrf_exempt
def profil(request):
    roles_exclus = ["Chargé Client"]
    notes = Notation.objects.filter(flux_stock='Flux')
    user = request.user
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la dernière note par agent
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)

    context= {'roles_exclus':roles_exclus,
              'notation':notation,
              }
    return render(request, 'profil.html', context)


@csrf_exempt
def profile(request):
    if request.method == 'POST':
        user_form = ProfileModify(request.POST, instance=request.user)
        if user_form.is_valid():
            user_form.save()
            messages.success(request, 'Votre profil a été modifié avec succès')
            return redirect('/perso/profil')

    else:
        user_form = ProfileModify(instance=request.user)
    return render(request, 'modify_profil.html', {'user_form': user_form})


@method_decorator(csrf_exempt, name='dispatch')
class ChangePasswordView(SuccessMessageMixin, PasswordChangeView):
    template_name = 'modify_pw.html'
    success_message = "Votre mot de passe a été changé avec succès"
    success_url = reverse_lazy('profil')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ajout du contexte personnalisé
        context['roles_exclus'] = ["Chargé Client"]
        return context

@csrf_exempt
def reset_user_password_b(request, user_id):
    roles_exclus = ["Chargé Client"]
    user = get_object_or_404(ProfileV, pk=user_id)

    if request.method == 'POST':
        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            # Enregistrer le nouveau mot de passe en le hachant
            new_password = form.cleaned_data['new_password']
            user.password = make_password(new_password)
            user.save()
            return redirect('profil')  # Rediriger vers la liste des utilisateurs après modification
    else:
        form = ResetPasswordForm()

    return render(request, 'modify_pw.html', {'form': form, 'user': user,'roles_exclus':roles_exclus})


@login_required
def perso(request):
    # Récupérer l'utilisateur connecté
    roles_exclus=["Chargé Client"]
    user = request.user

    # Vérifier si l'utilisateur appartient à "Contrôle", "Conformité" ou "Contrôle Groupe"
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Qualité']:
        agents = ProfileV.objects.filter(filiale=user.filiale)
    else:
        agents = ProfileV.objects.all()
    # Gestion de la recherche par exploitant
    query = request.GET.get('q', '')
    if query:
        agents = agents.filter(code_expl__icontains=query)

    return render(request, 'mon_profile.html', {'agents': agents, 'query': query, 'roles_exclus':roles_exclus})


@login_required
@csrf_exempt
def agent(request):
    roles_exclus = ["Chargé Client"]
    user = request.user

    # Vérifier si l'utilisateur appartient à "Contrôle", "Conformité" ou "Contrôle Groupe"
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Qualité']:
        notes = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Flux')
    elif user.organe in ['Directeur Agence']:
        notes = Notation.objects.filter(note_par__filiale=user.filiale, agent__agence=user.agence, flux_stock='Flux')
    else:
        notes = Notation.objects.filter(flux_stock='Flux')

    # Annoter chaque agent avec la dernière date de notation
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la dernière note par agent
    notes = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    notes = notes.order_by('-date_notation')
    # Gestion de la recherche par exploitant
    query = request.GET.get('q', '')
    if query:
        agents = ProfileV.objects.filter(code_expl__icontains=query)  # Utilisation de icontains pour une recherche partielle
        if agents.exists():
            notes = notes.all().filter(agent__in=agents)
        else:
            # Si aucun agent n'est trouvé, vider le queryset pour ne rien afficher
            notes = notes.none()

    return render(request, 'agent.html', {'notes': notes, 'query': query, 'roles_exclus':roles_exclus})


def export_agents_excel(request):
    def strip_tz(value):
        if hasattr(value, 'tzinfo'):
            return value.replace(tzinfo=None)
        return value

    user = request.user

    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Qualité']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Flux')
    elif user.organe in ['Directeur Agence']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, agent__agence=user.agence, flux_stock='Flux')
    else:
        donnees = Notation.objects.filter(flux_stock='Flux')

    latest_notes = donnees.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la dernière note par agent
    donnees = donnees.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    donnees = donnees.order_by('-date_notation')


    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notation_Flux"

    # Entêtes
    headers = ["FILIALE", "EXPLOITANT", "NOM", "AGENCE", "EMAIL", "NOTE", "Dernière notation", "Noté par le contrôleur", "Flux/Stock"]
    ws.append(headers)

    # Données
    for d in donnees:
        agent_nom = f"{d.agent.first_name} {d.agent.last_name}".strip()
        agent_email = d.agent.email or d.agent.username
        ws.append([
            d.note_par.filiale, d.agent.code_expl, agent_nom, d.agent.agence, agent_email, d.note,   strip_tz(d.date_notation), d.note_par.email, d.flux_stock

        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Notation_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_agents_excel_s(request):
    def strip_tz(value):
        if hasattr(value, 'tzinfo'):
            return value.replace(tzinfo=None)
        return value

    user = request.user

    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Qualité']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Stock')
    elif user.organe in ['Directeur Agence']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, agent__agence=user.agence, flux_stock='Stock')
    else:
        donnees = Notation.objects.filter(flux_stock='Stock')

    latest_notes = donnees.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la dernière note par agent
    donnees = donnees.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    donnees = donnees.order_by('-date_notation')


    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notation_Stock"

    # Entêtes
    headers = ["FILIALE", "EXPLOITANT", "NOM", "AGENCE", "EMAIL", "NOTE", "Dernière notation", "Noté par le contrôleur", "Flux/Stock"]
    ws.append(headers)

    # Données
    for d in donnees:
        agent_nom = f"{d.agent.first_name} {d.agent.last_name}".strip()
        agent_email = d.agent.email or d.agent.username
        ws.append([
            d.note_par.filiale, d.agent.code_expl, agent_nom, d.agent.agence, agent_email, d.note,   strip_tz(d.date_notation), d.note_par.email, d.flux_stock

        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Notation_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@csrf_exempt
def perso_stock(request):
    # Récupérer l'utilisateur connecté
    user = request.user

    # Vérifier si l'utilisateur appartient à "Contrôle", "Conformité" ou "Contrôle Groupe"
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Qualité']:
        notes = Notation.objects.filter(filiale=user.filiale, flux_stock='Stock')
    else:
        notes = Notation.objects.all().filter(flux_stock='Flux')
    # Gestion de la recherche par exploitant
    query = request.GET.get('q', '')
    if query:
        notes = notes.filter(agent__code_expl__icontains=query)

    return render(request, 'agent_stock.html', {'notes': notes, 'query': query})


@login_required
@csrf_exempt
def agent_stock(request):
    user = request.user

    # Vérifier si l'utilisateur appartient à "Contrôle", "Conformité" ou "Contrôle Groupe"
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Qualité']:
        notes = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Stock')
    else:
        notes = Notation.objects.filter(flux_stock='Stock')

    # Annoter chaque agent avec la dernière date de notation
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notes = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])

    # Filtrer pour obtenir uniquement la dernière note par agent
    notes = notes.order_by('-date_notation')

    # Gestion de la recherche par exploitant
    query = request.GET.get('q', '')
    if query:
        agents = ProfileV.objects.filter(code_expl__icontains=query)  # Utilisation de icontains pour une recherche partielle
        if agents.exists():
            notes = notes.filter(agent__in=agents)
        else:
            # Si aucun agent n'est trouvé, vider le queryset pour ne rien afficher
            notes = notes.none()

    return render(request, 'agent_stock.html', {'notes': notes, 'query': query})


@csrf_exempt
def notes(request):
    agent = None
    roles_exclus = ["Chargé Client", "Directeur Agence"]
    form = NotationForm()  # Initialisation du formulaire

    if request.method == 'POST':
        if 'search_agent' in request.POST:
            # Rechercher l'agent par son code exploitant
            code_exploitant = request.POST.get('code_exploitant')
            try:
                agent = ProfileV.objects.get(code_expl=code_exploitant, filiale=request.user.filiale)

                # Préremplir le formulaire avec l'agent trouvé
                form = NotationForm(initial={'agent': agent})
            except ProfileV.DoesNotExist:
                agent = None
                error_message = "L'agent avec ce code exploitant n'existe pas."
                return render(request, 'notation.html', {'form': form, 'error_message': error_message})

        else:
            # Soumission du formulaire de notation
            form = NotationForm(request.POST)
            if form.is_valid():
                user = request.user
                # Assigner l'agent à partir du formulaire
                notation = form.save(commit=False)
                notation.filiale =request.user.filiale
                notation.note_par = request.user
                notation.date_notation = timezone.now()
                notation.save()
                messages.success(request, 'La notation a bien été sauvegardée.')

                return redirect('agent')
    else:
        form = NotationForm()  # Afficher un formulaire vide si la requête n'est pas en POST

    return render(request, 'notation.html', {'form': form, 'agent': agent,'roles_exclus':roles_exclus})


def agent_detail(request, agent_id):
    agent = get_object_or_404(ProfileV, id=agent_id)
    notations = agent.notations.all().order_by('-date_notation')
    return render(request, 'agent_detail.html', {'agent': agent, 'notations': notations})


@login_required
@csrf_exempt
def historique(request):
    roles_exclus = ["Chargé Client", "Directeur Agence"]
    query = request.GET.get('q')

    if query:
        # Filtre les notations en fonction du code exploitant
        notations = Notation.objects.filter(note_par=request.user, agent__code_expl__icontains=query).order_by(
            "-date_notation")
    else:
        # Récupère toutes les notations de l'utilisateur connecté
        notations = Notation.objects.filter(note_par=request.user).order_by("-date_notation")

    # Passe les notations au template
    context = {
        'notations': notations,
        'query': query,
    }  # Pour pré-remplir la barre de recherche
    return render(request, 'historique.html', {'notations': notations, 'roles_exclus':roles_exclus})


def test(request):
    return render(request, 'test.html')


def register(request):
    roles_exclus = ["Chargé Client"]
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            return redirect('user_list')  # Rediriger vers la page d'accueil après l'inscription
    else:
        form = CustomUserCreationForm()
    return render(request, 'register.html', {'form': form,'roles_exclus':roles_exclus})


# Fonction pour vérifier si l'utilisateur appartient à l'organe "PASS"
def is_pass_user(user):
    return user.organe == 'PASS'


# Limiter l'accès à ceux de l'organe 'PASS'
def user_list(request):
    roles_exclus = ["Chargé Client"]
    users = ProfileV.objects.all()  # Récupérer tous les utilisateurs
    total_users = ProfileV.objects.count()

    # Sessions actives
    active_sessions = Session.objects.filter(expire_date__gte=now())
    user_ids = []
    for session in active_sessions:
        data = session.get_decoded()
        if '_auth_user_id' in data:
            user_ids.append(data['_auth_user_id'])

    # Supprimer les doublons
    user_ids = list(set(user_ids))

    # Liste des utilisateurs connectés
    connected_users = ProfileV.objects.filter(id__in=user_ids)
    connected_count = connected_users.count()

    # Envoyer les données au template
    context = {
        'total_users': total_users,
        'connected_count': connected_count,
        'connected_users': connected_users,
        'users': users,
        'roles_exclus':roles_exclus
    }
    return render(request, 'user_list.html', context)


@user_passes_test(is_pass_user)
def edit_user(request, user_id):
    user = get_object_or_404(ProfileV, pk=user_id)

    if request.method == "POST":
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('user_list')  # Rediriger vers la liste des utilisateurs après modification
    else:
        form = UserEditForm(instance=user)

    return render(request, 'edit_user.html', {'form': form, 'user': user})


@user_passes_test(is_pass_user)
def change_user_password(request, user_id):
    user = get_object_or_404(ProfileV, pk=user_id)

    if request.method == 'POST':
        form = PasswordChangeForm(user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important pour éviter de déconnecter l'utilisateur
            return redirect('user_list')  # Rediriger vers la liste des utilisateurs après modification
    else:
        form = PasswordChangeForm(user)

    return render(request, 'change_user_password.html', {'form': form, 'user': user})


@user_passes_test(is_pass_user)
def reset_user_password(request, user_id):
    target_user = get_object_or_404(ProfileV, pk=user_id)

    if request.method == 'POST':
        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            # Enregistrer le nouveau mot de passe en le hachant
            new_password = form.cleaned_data['new_password']
            target_user.password = make_password(new_password)
            target_user.save()
            return redirect('user_list')  # Rediriger vers la liste des utilisateurs après modification
    else:
        form = ResetPasswordForm()

    return render(request, 'reset_user_password.html', {'form': form, 'target_user': target_user})


def user_statistics_view(request):
    # Nombre total d'utilisateurs
    roles_exclus = ["Chargé Client"]
    users = ProfileV.objects.all()  # Récupérer tous les utilisateurs
    total_users = ProfileV.objects.count()

    # Sessions actives
    active_sessions = Session.objects.filter(expire_date__gte=now())
    user_ids = []
    for session in active_sessions:
        data = session.get_decoded()
        if '_auth_user_id' in data:
            user_ids.append(data['_auth_user_id'])

    # Supprimer les doublons
    user_ids = list(set(user_ids))

    # Liste des utilisateurs connectés
    connected_users = ProfileV.objects.filter(id__in=user_ids)
    connected_count = connected_users.count()

    # Envoyer les données au template
    context = {
        'total_users': total_users,
        'connected_count': connected_count,
        'connected_users': connected_users,
        'users': users,
        'roles_exclus':roles_exclus

    }
    return render(request, 'user_statistics.html', context)


@login_required
@csrf_exempt
def ppe(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    donnees = Kyc_pp.objects.filter(PPE__icontains="O")

    # Filtrage automatique selon le rôle
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    if user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    if user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    if user.organe in users_groupe:
        donnees = donnees

    # Filtres manuels via GET
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(code_expl__icontains=expl_param)

    # === Valeurs du formulaire selon le rôle ===
    if user.organe == "Directeur Agence":
        exploitants = Kyc_pp.objects.filter(AGENCE=user.agence).values_list('EXPL', flat=True).distinct()
        agences = Kyc_pp.objects.filter(AGENCE=user.agence).values_list('AGENCE', flat=True).distinct()
    elif user.organe in users_filiale:
        agences = Kyc_pp.objects.filter(FILIALE=user.filiale).values_list('AGENCE', flat=True).distinct()
        exploitants = Kyc_pp.objects.filter(FILIALE=user.filiale).values_list('EXPL', flat=True).distinct()
    else:
        exploitants = Kyc_pp.objects.values_list('EXPL', flat=True).distinct()
        agences = Kyc_pp.objects.values_list('AGENCE', flat=True).distinct()

    filiales = Kyc_pp.objects.values_list('FILIALE', flat=True).distinct()

    context = {
        'donnees': donnees,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
    }

    return render(request, 'ppe.html', context)

def export_ppe(request):
    user = request.user

    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    donnees = Kyc_pp.objects.filter(PPE__icontains="O")

    # Filtrage par organe
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    else:
        donnees = donnees

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Non rens PPE"

    # Entêtes
    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "CODAPE", "IDP", "PAYNAIS", "PROFESSION", "ADRESSE", "PAYS_RESID",
               "NUMID", "SALAIRE", "ORIGINE_REV", "DATVALID", "TEL"]
    ws.append(headers)

    # Données
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.CODAPE, d.IDP,
            d.PAYNAIS, d.PROFESSION, d.ADRESSE, d.PAYS_RESID, d.NUMID, d.SALAIRE, d.ORIGINE_REV, d.DATVALID, d.TEL
        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Champs_non_renseignés_PPE_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@csrf_exempt
def non_resid(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')
    donnees = Kyc_pp.objects.filter(~Q(RESID=""), RESID__isnull=False)

    # Filtrage automatique selon le rôle
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    if user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    if user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    if user.organe in users_groupe:
        donnees = donnees

    # Filtres manuels via GET
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(code_expl__icontains=expl_param)

    # === Valeurs du formulaire selon le rôle ===
    if user.organe == "Directeur Agence":
        exploitants = donnees.objects.filter(AGENCE=user.agence).values_list('EXPL', flat=True).distinct()
        agences = donnees.objects.filter(AGENCE=user.agence).values_list('AGENCE', flat=True).distinct()
    elif user.organe in users_filiale:
        agences = Kyc_pp.objects.filter(FILIALE=user.filiale).values_list('AGENCE', flat=True).distinct()
        exploitants = Kyc_pp.objects.filter(FILIALE=user.filiale).values_list('EXPL', flat=True).distinct()
    else:
        exploitants = Kyc_pp.objects.values_list('EXPL', flat=True).distinct()
        agences = Kyc_pp.objects.values_list('AGENCE', flat=True).distinct()

    filiales = Kyc_pp.objects.values_list('FILIALE', flat=True).distinct()

    context = {
        'donnees': donnees,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
    }
    return render(request, 'non_resid.html', context)

@login_required
@csrf_exempt
def non_resid_pm(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')
    donnees = Kyc_pm.objects.filter(~Q(RESID=""), RESID__isnull=False)

    # Filtrage automatique selon le rôle
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    if user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    if user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    if user.organe in users_groupe:
        donnees = donnees

    # Filtres manuels via GET
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(code_expl__icontains=expl_param)

    # === Valeurs du formulaire selon le rôle ===
    if user.organe == "Directeur Agence":
        exploitants = donnees.objects.filter(AGENCE=user.agence).values_list('EXPL', flat=True).distinct()
        agences = donnees.objects.filter(AGENCE=user.agence).values_list('AGENCE', flat=True).distinct()
    elif user.organe in users_filiale:
        agences = Kyc_pm.objects.filter(FILIALE=user.filiale).values_list('AGENCE', flat=True).distinct()
        exploitants = Kyc_pm.objects.filter(FILIALE=user.filiale).values_list('EXPL', flat=True).distinct()
    else:
        exploitants = Kyc_pm.objects.values_list('EXPL', flat=True).distinct()
        agences = Kyc_pm.objects.values_list('AGENCE', flat=True).distinct()

    filiales = Kyc_pm.objects.values_list('FILIALE', flat=True).distinct()

    context = {
        'donnees': donnees,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
    }
    return render(request, 'non_resid_pm.html', context)

from datetime import date, timedelta
@login_required
@csrf_exempt

def scoring(request):
    # Rôles
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe  = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    user = request.user

    # Params GET
    periode_param = request.GET.get("periode", "")
    filiale_param = request.GET.get("filiale", "")
    agence_param  = request.GET.get("agence", "")
    expl_param    = request.GET.get("expl", "")

    base_qs = DATEREV.objects.all().filter(DATEREV__isnull=False)

    if getattr(user, "organe", "") == "Chargé Client":
        base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        base_qs = base_qs.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        pass

    today = date.today()
    qs_period = base_qs
    if periode_param == "today":
        qs_period = qs_period.filter(DATEREV__lte=today)
    elif periode_param == "3m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=90))
    elif periode_param == "6m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=180))
    elif periode_param == "1y":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=365))


    can_pick_filiale = user.organe in users_groupe


    selected_filiale = filiale_param if can_pick_filiale else getattr(user, "filiale", "")


    filiales_opts = qs_period.values_list("FILIALE", flat=True).distinct().order_by("FILIALE")


    qs_filiale = qs_period
    if selected_filiale:
        qs_filiale = qs_filiale.filter(FILIALE=selected_filiale)


    can_pick_agence = (user.organe in users_groupe) or (user.organe in users_filiale) or (user.organe == "Directeur Agence")


    if user.organe == "Directeur Agence":
        selected_agence = getattr(user, "agence", "")
    else:
        selected_agence = agence_param


    agences_opts = qs_filiale.values_list("AGENCE", flat=True).distinct().order_by("AGENCE")


    qs_agence = qs_filiale
    if selected_agence:
        qs_agence = qs_agence.filter(AGENCE=selected_agence)


    can_pick_expl = (user.organe in users_groupe) or (user.organe in users_filiale) or (user.organe == "Directeur Agence")


    if getattr(user, "organe", "") == "Chargé Client":
        selected_expl = getattr(user, "code_expl", "")
    else:
        selected_expl = expl_param


    exploitants_opts = qs_agence.values_list("EXPL", flat=True).distinct().order_by("EXPL")


    donnees = qs_agence
    if selected_expl:
        donnees = donnees.filter(EXPL=selected_expl)

    context = {
        "donnees": donnees.order_by("FILIALE", "AGENCE", "EXPL", "CLIENT"),

        # Options
        "filiales": filiales_opts,
        "agences": agences_opts,
        "exploitants": exploitants_opts,

        # Sélections courantes
        "periode": periode_param,
        "filiale_param": selected_filiale,
        "agence_param": selected_agence,
        "expl_param": selected_expl,

        # Rôles (si tu en as besoin dans le template)
        "users_groupe": users_groupe,
        "users_filiale": users_filiale,

        # Droits d'édition des selects
        "can_pick_filiale": can_pick_filiale,
        "can_pick_agence": can_pick_agence,
        "can_pick_expl": can_pick_expl,
    }
    return render(request, "scoring.html", context)


def export_csv_scoring(request):
    def strip_tz(value):
        if hasattr(value, 'tzinfo'):
            return value.replace(tzinfo=None)
        return value

    user = request.user

    if user.organe == 'Conformité':
        donnees = DATEREV.objects.filter(FILIALE=user.filiale)
    elif user.organe == "Conformité Groupe":
        donnees = DATEREV.objects.all()

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notation_Stock"

    # Entêtes
    headers = ['AGENCE','AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'CLASSE']
    ws.append(headers)

    # Données
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.DATEREV, d.PPE, d.CLASSE

        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Revue_scoring_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



@login_required
@csrf_exempt
def sans_classe(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    donnees = DATEREV.objects.filter(CLASSE="N")

    # Filtrage automatique selon le rôle
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    if user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    if user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    if user.organe in users_groupe:
        donnees = DATEREV.objects.all()

    # Filtres manuels via GET
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(code_expl__icontains=expl_param)

    # === Valeurs du formulaire selon le rôle ===
    if user.organe == "Directeur Agence":
        exploitants = DATEREV.objects.filter(AGENCE=user.agence).values_list('EXPL', flat=True).distinct()
        agences = DATEREV.objects.filter(AGENCE=user.agence).values_list('AGENCE', flat=True).distinct()
    elif user.organe in users_filiale:
        agences = DATEREV.objects.filter(FILIALE=user.filiale).values_list('AGENCE', flat=True).distinct()
        exploitants = DATEREV.objects.filter(FILIALE=user.filiale).values_list('EXPL', flat=True).distinct()
    else:
        exploitants = DATEREV.objects.values_list('EXPL', flat=True).distinct()
        agences = DATEREV.objects.values_list('AGENCE', flat=True).distinct()

    filiales = DATEREV.objects.values_list('FILIALE', flat=True).distinct()

    context = {
        'donnees': donnees,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
    }

    return render(request, 'sans_classe.html', context)

def export_sans_classe(request):
    def strip_tz(value):
        if hasattr(value, 'tzinfo'):
            return value.replace(tzinfo=None)
        return value

    user = request.user

    if user.organe == 'Conformité':
        donnees = DATEREV.objects.filter(FILIALE=user.filiale, CLASSE='N')
    elif user.organe == "Conformité Filiale":
        donnees = DATEREV.objects.filter(CLASSE='N')

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients non classés"

    # Entêtes
    headers = ['AGENCE','AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'CLASSE']
    ws.append(headers)

    # Données
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.DATEREV, d.PPE, d.CLASSE

        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Revue_scoring_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



@login_required
@csrf_exempt

def non_rens_pm(request):
    user = request.user
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    notes = Notation.objects.filter(flux_stock='Flux')

    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la dernière note par agent
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)

    queryset = Kyc_pm.objects.all()

    # Listes de filtres
    filiale_list, agence_list, expl_list = [], [], []

    # Cas Chargé Client
    if user.organe == "Chargé Client":
        queryset = queryset.filter(FILIALE=user.filiale, EXPL=user.code_expl)

    # Cas Directeur Agence
    elif user.organe == "Directeur Agence":
        queryset = queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)
        expl_filter = request.GET.get("expl")
        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)
        expl_list = Kyc_pm.objects.filter(AGENCE=user.agence).values_list("EXPL", flat=True).distinct()

    # Cas Conformité (filiale)
    elif user.organe in users_filiale:
        queryset = queryset.filter(FILIALE=user.filiale)
        agence_filter = request.GET.get("agence")
        expl_filter = request.GET.get("expl")

        # Liste agences de la filiale
        agence_list = Kyc_pm.objects.filter(FILIALE=user.filiale).values_list("AGENCE", flat=True).distinct()

        if agence_filter:
            queryset = queryset.filter(AGENCE=agence_filter)
            expl_list = Kyc_pm.objects.filter(FILIALE=user.filiale, AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()

        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)

    # Cas Conformité Groupe
    elif user.organe in users_groupe:
        filiale_filter = request.GET.get("filiale")
        agence_filter = request.GET.get("agence")
        expl_filter = request.GET.get("expl")

        filiale_list = Kyc_pm.objects.values_list("FILIALE", flat=True).distinct()

        if filiale_filter:
            queryset = queryset.filter(FILIALE=filiale_filter)
            agence_list = Kyc_pm.objects.filter(FILIALE=filiale_filter).values_list("AGENCE", flat=True).distinct()

        if agence_filter:
            queryset = queryset.filter(AGENCE=agence_filter)
            expl_list = Kyc_pm.objects.filter(AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()

        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)

    context = {
        "donnees": queryset,
        "filiale_list": filiale_list,
        "agence_list": agence_list,
        "expl_list": expl_list,
        'users_groupe':users_groupe,
        'users_filiale':users_filiale,
        'notation':notation,
    }
    return render(request, "non_rens_pm.html", context)



def non_rens(request):
    user = request.user
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    notes = Notation.objects.filter(flux_stock='Flux')

    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la dernière note par agent
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)

    queryset = Kyc_pp.objects.all()

    # Listes de filtres
    filiale_list, agence_list, expl_list = [], [], []

    # Cas Chargé Client
    if user.organe == "Chargé Client":
        queryset = queryset.filter(FILIALE=user.filiale, EXPL=user.code_expl)

    # Cas Directeur Agence
    elif user.organe == "Directeur Agence":
        queryset = queryset.filter(FILIALE=user.filiale,AGENCE=user.agence)
        expl_filter = request.GET.get("expl")
        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)
        expl_list = Kyc_pp.objects.filter(AGENCE=user.agence).values_list("EXPL", flat=True).distinct()

    # Cas Conformité (filiale)
    elif user.organe in users_filiale:
        queryset = queryset.filter(FILIALE=user.filiale)
        agence_filter = request.GET.get("agence")
        expl_filter = request.GET.get("expl")

        # Liste agences de la filiale
        agence_list = Kyc_pp.objects.filter(FILIALE=user.filiale).values_list("AGENCE", flat=True).distinct()

        if agence_filter:
            queryset = queryset.filter(AGENCE=agence_filter)
            expl_list = Kyc_pp.objects.filter(FILIALE=user.filiale, AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()

        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)

    # Cas Conformité Groupe
    elif user.organe in users_groupe:
        filiale_filter = request.GET.get("filiale")
        agence_filter = request.GET.get("agence")
        expl_filter = request.GET.get("expl")

        filiale_list = Kyc_pp.objects.values_list("FILIALE", flat=True).distinct()

        if filiale_filter:
            queryset = queryset.filter(FILIALE=filiale_filter)
            agence_list = Kyc_pp.objects.filter(FILIALE=filiale_filter).values_list("AGENCE", flat=True).distinct()

        if agence_filter:
            queryset = queryset.filter(AGENCE=agence_filter)
            expl_list = Kyc_pp.objects.filter(AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()

        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)

    context = {
        "donnees": queryset,
        "filiale_list": filiale_list,
        "agence_list": agence_list,
        "expl_list": expl_list,
        'users_groupe':users_groupe,
        'users_filiale':users_filiale,
        'notation':notation,
    }
    return render(request, "non_rens.html", context)


from django.db.models import Max
from django.shortcuts import render
from .models import Anomalie, Notation

def non_anom(request):
    user = request.user
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    notes = Notation.objects.filter(flux_stock='Flux')

    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la dernière note par agent
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)

    queryset = Anomalie.objects.all()

    # Listes de filtres
    filiale_list, agence_list, expl_list = [], [], []

    # Cas Chargé Client
    if user.organe == "Chargé Client":
        queryset = queryset.filter(FILIALE=user.filiale, EXPL=user.code_expl)

    # Cas Directeur Agence
    elif user.organe == "Directeur Agence":
        queryset = queryset.filter(FILIALE=user.filiale,AGENCE=user.agence)
        expl_filter = request.GET.get("expl")
        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)
        expl_list = Anomalie.objects.filter(AGENCE=user.agence).values_list("EXPL", flat=True).distinct()

    # Cas Conformité (filiale)
    elif user.organe in users_filiale:
        queryset = queryset.filter(FILIALE=user.filiale)
        agence_filter = request.GET.get("agence")
        expl_filter = request.GET.get("expl")

        # Liste agences de la filiale
        agence_list = Anomalie.objects.filter(FILIALE=user.filiale).values_list("AGENCE", flat=True).distinct()

        if agence_filter:
            queryset = queryset.filter(AGENCE=agence_filter)
            expl_list = Anomalie.objects.filter(FILIALE=user.filiale, AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()

        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)

    # Cas Conformité Groupe
    elif user.organe in users_groupe:
        filiale_filter = request.GET.get("filiale")
        agence_filter = request.GET.get("agence")
        expl_filter = request.GET.get("expl")

        filiale_list = Anomalie.objects.values_list("FILIALE", flat=True).distinct()

        if filiale_filter:
            queryset = queryset.filter(FILIALE=filiale_filter)
            agence_list = Anomalie.objects.filter(FILIALE=filiale_filter).values_list("AGENCE", flat=True).distinct()

        if agence_filter:
            queryset = queryset.filter(AGENCE=agence_filter)
            expl_list = Anomalie.objects.filter(AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()

        if expl_filter:
            queryset = queryset.filter(EXPL=expl_filter)

    context = {
        "donnees": queryset,
        "filiale_list": filiale_list,
        "agence_list": agence_list,
        "expl_list": expl_list,
        'users_groupe':users_groupe,
        'users_filiale':users_filiale,
        'notation':notation,
    }
    return render(request, "non_anom.html", context)


def export_csv_anom(request):
    user = request.user

    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]


    donnees = Anomalie.objects.all()

    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    else:
        donnees = donnees.objects.filter()

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anomalie"

    # Entêtes
    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "ANOMALIE_AGE", "ANOMALIE_DATE_EER", "ANOMALIE_CIN"]
    ws.append(headers)

    # Données
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.ANOMALIE_AGE, d.ANOMALIE_DATE_EER,
            d.ANOMALIE_CIN
        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Champs_en_anomalie_PP_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



def export_csv_anom_ppe(request):
    user = request.user

    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]


    donnees = Anomalie.objects.filter(PPE='O')

    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    else:
        donnees = donnees.objects.filter()

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PPE en Anomalie"

    # Entêtes
    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "ANOMALIE_AGE", "ANOMALIE_DATE_EER", "ANOMALIE_CIN"]
    ws.append(headers)

    # Données
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.ANOMALIE_AGE, d.ANOMALIE_DATE_EER,
            d.ANOMALIE_CIN
        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Champs_en_anomalie_PPE_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_csv(request):
    user = request.user
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    donnees = Kyc_pm.objects.all()

    # Filtrage par organe
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    else:
        donnees = donnees.objects.filter()


    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export KYC"

    # Entêtes
    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "AGEC", "CODAPE", "IDM", "RCSNO", "CAPITAL", "CA", "RESULTAT",
               "ORIGINE_REVENU", "TEL"]
    ws.append(headers)

    # Données
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.AGEC, d.CODAPE, d.IDM,
            d.RCSNO, d.CAPITAL, d.CA, d.RESULTAT, d.ORIGINE_REVENU, d.TEL
        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"champs_non_renseignés_PM_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_csv_pp(request):
    user = request.user

    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]


    donnees = Kyc_pp.objects.all()
    # Filtrage par organe
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    else:
        donnees = donnees.objects.filter()

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export KYC"

    # Entêtes
    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "CODAPE", "IDP", "PAYNAIS", "PROFESSION", "ADRESSE", "PAYS_RESID",
               "NUMID", "SALAIRE", "ORIGINE_REV", "DATVALID", "TEL"]
    ws.append(headers)

    # Données
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.CODAPE, d.IDP,
            d.PAYNAIS, d.PROFESSION, d.ADRESSE, d.PAYS_RESID, d.NUMID, d.SALAIRE, d.ORIGINE_REV, d.DATVALID, d.TEL
        ])

    # Ajustement largeur des colonnes (optionnel)
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Préparer la réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"champs_non_renseignés_PP_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response






from django.contrib.auth.decorators import login_required

@login_required
def statistiques(request):
    user = request.user
    # Rôles "filiale" (accès restreint à leur propre filiale)
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]


    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    notes = Notation.objects.filter(flux_stock='Flux')

    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la dernière note par agent
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)

    # Paramètres GET
    selected_expl = request.GET.get('expl')
    # Seul le Groupe" peut choisir librement une filiale dans le GET
    selected_filiale = request.GET.get('filiale') if user.organe in users_groupe else None

    # Querysets vides par défaut
    qs_pm = TauxEvolution.objects.none()
    qs_pp = TauxEvolution.objects.none()

    liste_expl = []
    liste_filiales = []

    base_qs = TauxEvolution.objects.filter(flux_stock="F")

    # ---------- RÔLES ----------
    if user.organe == "Chargé Client":
        # Pas de dropdown : son propre exploitant
        liste_expl = [user.code_expl]
        selected_expl = user.code_expl
        qs_pm = base_qs.filter(filiale=user.filiale, expl=user.code_expl, pp_pm="M").order_by('date')
        qs_pp = base_qs.filter(filiale=user.filiale, expl=user.code_expl, pp_pm="P").order_by('date')

    elif user.organe == "Directeur Agence":
        # Exploitants de l’agence + de la filiale de l’utilisateur
        liste_expl = (base_qs
                      .filter(filiale=user.filiale, agence=user.agence)
                      .values_list('expl', flat=True).distinct().order_by('expl'))
        if selected_expl:
            qs_pm = base_qs.filter(filiale=user.filiale, expl=selected_expl, pp_pm="M").order_by('date')
            qs_pp = base_qs.filter(filiale=user.filiale, expl=selected_expl, pp_pm="P").order_by('date')

    elif user.organe in users_filiale:
        # Exploitants de la filiale de l’utilisateur
        liste_expl = (base_qs
                      .filter(filiale=user.filiale)
                      .values_list('expl', flat=True).distinct().order_by('expl'))
        if selected_expl:
            qs_pm = base_qs.filter(filiale=user.filiale, expl=selected_expl, pp_pm="M").order_by('date')
            qs_pp = base_qs.filter(filiale=user.filiale, expl=selected_expl, pp_pm="P").order_by('date')

    elif user.organe in users_groupe:
        # Peut voir TOUTES les filiales + dropdown "filiale"
        liste_filiales = (base_qs.values_list('filiale', flat=True)
                          .distinct().order_by('filiale'))

        # ➜ Les exploitants sont CEUX de la filiale sélectionnée
        if selected_filiale:
            liste_expl = (base_qs
                          .filter(filiale=selected_filiale)
                          .values_list('expl', flat=True).distinct().order_by('expl'))
            if selected_expl:
                qs_pm = base_qs.filter(filiale=selected_filiale, expl=selected_expl, pp_pm="M").order_by('date')
                qs_pp = base_qs.filter(filiale=selected_filiale, expl=selected_expl, pp_pm="P").order_by('date')
        else:
            # Pas de filiale choisie : on attend le choix de filiale (liste_expl vide)
            liste_expl = []

    else:
        # Cas générique : tous les exploitants (toutes filiales)
        liste_expl = (base_qs.values_list('expl', flat=True)
                      .distinct().order_by('expl'))
        if selected_expl:
            qs_pm = base_qs.filter(pp_pm="M").order_by('date')
            qs_pp = base_qs.filter(pp_pm="P").order_by('date')

    # ---------- Données pour les charts ----------
    dates_pm = [obj.date.strftime('%d-%m-%Y') for obj in qs_pm]
    taux_pm = [obj.taux for obj in qs_pm]
    dates_pp = [obj.date.strftime('%d-%m-%Y') for obj in qs_pp]
    taux_pp = [obj.taux for obj in qs_pp]

    context = {
        'liste_filiales': liste_filiales,      # pour le dropdown filiale
        'selected_filiale': selected_filiale,  # valeur retenue
        'liste_expl': liste_expl,              # dépend désormais de la filiale choisie
        'selected_expl': selected_expl,
        'dates_pm': json.dumps(dates_pm),
        'taux_pm': json.dumps(taux_pm),
        'dates_pp': json.dumps(dates_pp),
        'taux_pp': json.dumps(taux_pp),
        'notation': notation,
    }
    return render(request, 'statistiques.html', context)


@login_required
def statistiques_stock(request):
    user = request.user
    # Rôles "filiale" (accès restreint à leur propre filiale)
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]


    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    notes = Notation.objects.filter(flux_stock='Flux')

    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))

    # Filtrer pour obtenir uniquement la dernière note par agent
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)

    # Paramètres GET
    selected_expl = request.GET.get('expl')
    # Seul "Conformité Groupe" peut choisir librement une filiale dans le GET
    selected_filiale = request.GET.get('filiale') if user.organe in users_groupe else None

    # Querysets vides par défaut
    qs_pm = TauxEvolution.objects.none()
    qs_pp = TauxEvolution.objects.none()

    liste_expl = []
    liste_filiales = []

    base_qs = TauxEvolution.objects.filter(flux_stock="S")

    # ---------- RÔLES ----------
    if user.organe == "Chargé Client":
        # Pas de dropdown : son propre exploitant
        liste_expl = [user.code_expl]
        selected_expl = user.code_expl
        qs_pm = base_qs.filter(filiale=user.filiale, expl=user.code_expl, pp_pm="M").order_by('date')
        qs_pp = base_qs.filter(filiale=user.filiale, expl=user.code_expl, pp_pm="P").order_by('date')

    elif user.organe == "Directeur Agence":
        # Exploitants de l’agence + de la filiale de l’utilisateur
        liste_expl = (base_qs
                      .filter(filiale=user.filiale, agence=user.agence)
                      .values_list('expl', flat=True).distinct().order_by('expl'))
        if selected_expl:
            qs_pm = base_qs.filter(filiale=user.filiale, expl=selected_expl, pp_pm="M").order_by('date')
            qs_pp = base_qs.filter(filiale=user.filiale, expl=selected_expl, pp_pm="P").order_by('date')

    elif user.organe in users_filiale:
        # Exploitants de la filiale de l’utilisateur
        liste_expl = (base_qs
                      .filter(filiale=user.filiale)
                      .values_list('expl', flat=True).distinct().order_by('expl'))
        if selected_expl:
            qs_pm = base_qs.filter(filiale=user.filiale, expl=selected_expl, pp_pm="M").order_by('date')
            qs_pp = base_qs.filter(filiale=user.filiale, expl=selected_expl, pp_pm="P").order_by('date')

    elif user.organe in users_groupe:
        # Peut voir TOUTES les filiales + dropdown "filiale"
        liste_filiales = (base_qs.values_list('filiale', flat=True)
                          .distinct().order_by('filiale'))

        # ➜ Les exploitants sont CEUX de la filiale sélectionnée
        if selected_filiale:
            liste_expl = (base_qs
                          .filter(filiale=selected_filiale)
                          .values_list('expl', flat=True).distinct().order_by('expl'))
            if selected_expl:
                qs_pm = base_qs.filter(filiale=selected_filiale, expl=selected_expl, pp_pm="M").order_by('date')
                qs_pp = base_qs.filter(filiale=selected_filiale, expl=selected_expl, pp_pm="P").order_by('date')
        else:
            # Pas de filiale choisie : on attend le choix de filiale (liste_expl vide)
            liste_expl = []

    else:
        # Cas générique : tous les exploitants (toutes filiales)
        liste_expl = (base_qs.values_list('expl', flat=True)
                      .distinct().order_by('expl'))
        if selected_expl:
            qs_pm = base_qs.filter(pp_pm="M").order_by('date')
            qs_pp = base_qs.filter(pp_pm="P").order_by('date')

    # ---------- Données pour les charts ----------
    dates_pm = [obj.date.strftime('%d-%m-%Y') for obj in qs_pm]
    taux_pm = [obj.taux for obj in qs_pm]
    dates_pp = [obj.date.strftime('%d-%m-%Y') for obj in qs_pp]
    taux_pp = [obj.taux for obj in qs_pp]

    context = {
        'liste_filiales': liste_filiales,      # pour le dropdown filiale
        'selected_filiale': selected_filiale,  # valeur retenue
        'liste_expl': liste_expl,              # dépend désormais de la filiale choisie
        'selected_expl': selected_expl,
        'dates_pm': json.dumps(dates_pm),
        'taux_pm': json.dumps(taux_pm),
        'dates_pp': json.dumps(dates_pp),
        'taux_pp': json.dumps(taux_pp),
        'notation': notation,
    }
    return render(request, 'statistiques_stock.html', context)

def export_stats_pp(request):
    # même logique de filtrage que dans la vue principale
    user = request.user
    organe = user.organe
    filiale = user.filiale
    expl_user = user.expl

    if organe in ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]:
        qs = TauxEvolution.objects.all()
    elif organe in ["Conformité", "Contrôle Permanent", "Directeur Réseau"]:
        qs = TauxEvolution.objects.filter(filiale=filiale)
    elif organe == "Directeur Agence":
        qs = TauxEvolution.objects.filter(agence=user.agence)
    elif organe == "Chargé de client":
        qs = TauxEvolution.objects.filter(expl=expl_user)
    else:
        qs = TauxEvolution.objects.none()

    # filtrer éventuellement sur GET param
    selected_expl = request.GET.get('expl')
    if selected_expl:
        qs = qs.filter(expl=selected_expl)

    data = list(qs.values('filiale', 'agence', 'expl', 'date', 'taux'))
    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rapport_taux.xlsx"'
    df.to_excel(response, index=False)
    return response


def daterev_ppe(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    donnees = DATEREV.objects.filter(PPE="O")

    # Filtrage automatique selon le rôle
    if user.organe =="Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale,AGENCE=user.agence, EXPL=user.code_expl)
    if user.organe =="Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    if user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    if user.organe in users_groupe:
        donnees = DATEREV.objects.all()

    # Filtres manuels via GET
    if filiale_param:
        donnees = donnees.filter(FILIALE=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL=expl_param)

    # === Valeurs du formulaire selon le rôle ===
    if user.organe == "Directeur Agence":
        exploitants = DATEREV.objects.filter(AGENCE=user.agence).values_list('EXPL', flat=True).distinct()
        agences = DATEREV.objects.filter(AGENCE=user.agence).values_list('AGENCE', flat=True).distinct()
    elif user.organe in users_filiale:
        agences = DATEREV.objects.filter(FILIALE=user.filiale).values_list('AGENCE', flat=True).distinct()
        exploitants = DATEREV.objects.filter(FILIALE=user.filiale).values_list('EXPL', flat=True).distinct()
    else:
        exploitants = DATEREV.objects.values_list('EXPL', flat=True).distinct()
        agences = DATEREV.objects.values_list('AGENCE', flat=True).distinct()

    filiales = DATEREV.objects.values_list('FILIALE', flat=True).distinct()

    context = {
        'donnees': donnees,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale':users_filiale,
    }

    return render(request, 'daterev_ppe.html', context)


def non_anom_ppe(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    donnees = Anomalie.objects.filter(PPE="O")
    # Filtrage automatique selon le rôle
    if user.organe =="Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale,AGENCE=user.agence, EXPL=user.code_expl)
    if user.organe =="Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    if user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    if user.organe in users_groupe:
        donnees = donnees

    # Filtres manuels via GET
    if filiale_param:
        donnees = donnees.filter(FILIALE=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL=expl_param)

    # === Valeurs du formulaire selon le rôle ===
    if user.organe == "Directeur Agence":
        exploitants = Anomalie.objects.filter(AGENCE=user.agence).values_list('EXPL', flat=True).distinct()
        agences = Anomalie.objects.filter(AGENCE=user.agence).values_list('AGENCE', flat=True).distinct()
    elif user.organe in users_filiale:
        agences = Anomalie.objects.filter(FILIALE=user.filiale).values_list('AGENCE', flat=True).distinct()
        exploitants = Anomalie.objects.filter(FILIALE=user.filiale).values_list('EXPL', flat=True).distinct()
    else:
        exploitants = Anomalie.objects.values_list('EXPL', flat=True).distinct()
        agences = Anomalie.objects.values_list('AGENCE', flat=True).distinct()

    filiales = Anomalie.objects.values_list('FILIALE', flat=True).distinct()

    context = {
        'donnees': donnees,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale':users_filiale,
    }

    return render(request, 'anom_ppe.html', context)

@login_required
@csrf_exempt

def devise(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    donnees = Kyc_pp.objects.filter(~Q(DEVISE=""), DEVISE__isnull=False)

    # Filtrage automatique selon le rôle
    if user.organe =="Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale,AGENCE=user.agence, EXPL=user.code_expl)
    if user.organe =="Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    if user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    if user.organe in users_groupe:
        donnees = Kyc_pp.objects.all()

    # Filtres manuels via GET
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(code_expl__icontains=expl_param)

    # === Valeurs du formulaire selon le rôle ===
    if user.organe == "Directeur Agence":
        exploitants = Kyc_pp.objects.filter(AGENCE=user.agence).values_list('EXPL', flat=True).distinct()
        agences = Kyc_pp.objects.filter(AGENCE=user.agence).values_list('AGENCE', flat=True).distinct()
    elif user.organe in users_filiale:
        agences = Kyc_pp.objects.filter(FILIALE=user.filiale).values_list('AGENCE', flat=True).distinct()
        exploitants = Kyc_pp.objects.filter(FILIALE=user.filiale).values_list('EXPL', flat=True).distinct()
    else:
        exploitants = Kyc_pp.objects.values_list('EXPL', flat=True).distinct()
        agences = Kyc_pp.objects.values_list('AGENCE', flat=True).distinct()

    filiales = Kyc_pp.objects.values_list('FILIALE', flat=True).distinct()

    context = {
        'donnees': donnees,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale':users_filiale,
    }

    return render(request, 'devise.html', context)

def devise_pm(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["Conformité", "Contrôle Permanent", "Directeur Réseau"]
    users_groupe = ["Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    donnees = Kyc_pm.objects.filter(~Q(DEVISE=""), DEVISE__isnull=False)

    # Filtrage automatique selon le rôle
    if user.organe =="Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale,AGENCE=user.agence, EXPL=user.code_expl)
    if user.organe =="Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    if user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    if user.organe in users_groupe:
        donnees = donnees

    # Filtres manuels via GET
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(code_expl__icontains=expl_param)

    # === Valeurs du formulaire selon le rôle ===
    if user.organe == "Directeur Agence":
        exploitants = donnees.objects.filter(AGENCE=user.agence).values_list('EXPL', flat=True).distinct()
        agences = donnees.objects.filter(AGENCE=user.agence).values_list('AGENCE', flat=True).distinct()
    elif user.organe in users_filiale:
        agences = Kyc_pm.objects.filter(FILIALE=user.filiale).values_list('AGENCE', flat=True).distinct()
        exploitants = Kyc_pm.objects.filter(FILIALE=user.filiale).values_list('EXPL', flat=True).distinct()
    else:
        exploitants = Kyc_pm.objects.values_list('EXPL', flat=True).distinct()
        agences = Kyc_pm.objects.values_list('AGENCE', flat=True).distinct()

    filiales = Kyc_pm.objects.values_list('FILIALE', flat=True).distinct()

    context = {
        'donnees': donnees,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale':users_filiale,
    }

    return render(request, 'devise_pm.html', context)


@login_required
def evolution_taux(request):
    user = request.user

    context = {}

    if user.organe == "Conformité Groupe":
        # Groupe : on récupère toutes les filiales distinctes
        filiales = TAUX_FILIALE.objects.values_list("FILIALE", flat=True).distinct()
        data_filiales = {}

        for filiale in filiales:
            qs = TAUX_FILIALE.objects.filter(FILIALE=filiale).order_by("id")

            dates = [str(i.id) for i in qs]  # tu peux remplacer par une vraie date si tu en ajoutes une
            taux_pp = [round((t.FLUX_PP / t.STOCK_PP) * 100, 2) if t.STOCK_PP else 0 for t in qs]
            taux_pm = [round((t.FLUX_PM / t.STOCK_PM) * 100, 2) if t.STOCK_PM else 0 for t in qs]

            data_filiales[filiale] = {
                "dates": dates,
                "taux_pp": taux_pp,
                "taux_pm": taux_pm,
            }

        context["data_filiales"] = data_filiales

    elif user.organe == "Conformité":
        # Filiale : uniquement sa propre
        qs = TAUX_FILIALE.objects.filter(FILIALE=user.filiale).order_by("id")

        dates = [str(i.id) for i in qs]  # pareil, si tu as un champ Date, utilise-le
        taux_pp = [round((t.FLUX_PP / t.STOCK_PP) * 100, 2) if t.STOCK_PP else 0 for t in qs]
        taux_pm = [round((t.FLUX_PM / t.STOCK_PM) * 100, 2) if t.STOCK_PM else 0 for t in qs]

        context.update({
            "filiale": user.filiale,
            "dates_pp": dates,
            "taux_pp": taux_pp,
            "dates_pm": dates,
            "taux_pm": taux_pm,
        })

    return render(request, "statistiques.html", context)
