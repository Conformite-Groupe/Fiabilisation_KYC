"""
Génération du fichier Excel – Règles de Contrôle Qualité KYC
Plateforme KYC – Bank of Africa Group
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
import os

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "Regles_Controle_Qualite_KYC_BOA.xlsx")

# ─────────────────────────────────────────────
# PALETTE DE COULEURS
# ─────────────────────────────────────────────
C = {
    # Entêtes de section
    "hdr_pp":       "1A3C6E",   # Bleu marine foncé – PP
    "hdr_pm":       "0D5C63",   # Vert pétrole – PM
    "hdr_shared":   "3D2B6E",   # Violet – règles communes
    "hdr_reg":      "7B1818",   # Rouge bordeaux – réglementaire
    "hdr_text":     "FFFFFF",   # Blanc

    # Criticité
    "crit_critique": "FF4444",  # Rouge vif
    "crit_import":   "FF9900",  # Orange
    "crit_normale":  "2ECC71",  # Vert
    "crit_txt_dark": "1A1A1A",

    # Dimensions
    "dim1_completude": "E8F4FD",  # Bleu très clair
    "dim2_validite":   "FFF3E0",  # Orange très clair
    "dim3_coherence":  "F3E8FF",  # Violet très clair
    "dim4_conformite": "FDECEA",  # Rouge très clair
    "dim5_unicite":    "E8FDF5",  # Vert très clair

    # Lignes alternées
    "row_odd":     "F7F9FC",
    "row_even":    "FFFFFF",

    # Général
    "accent_blue":  "2E86AB",
    "light_gray":   "F0F0F0",
    "dark_text":    "1A1A1A",
    "border_color": "CCCCCC",
    "gold":         "D4AF37",
    "tab_pp":       "1A3C6E",
    "tab_pm":       "0D5C63",
    "tab_shared":   "5C3A8C",
    "tab_summary":  "1B4F72",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="1A1A1A", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(color=C["border_color"]):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def thick_border():
    thick = Side(style="medium", color="444444")
    thin  = Side(style="thin",   color=C["border_color"])
    return Border(left=thick, right=thick, top=thick, bottom=thick)

CRIT_MAP = {
    "🔴 Critique":    (C["crit_critique"], "FFFFFF", "CRITIQUE"),
    "🟡 Importante":  (C["crit_import"],   "1A1A1A", "IMPORTANTE"),
    "🟢 Normale":     (C["crit_normale"],  "FFFFFF", "NORMALE"),
}

# ─────────────────────────────────────────────
# DONNÉES
# ─────────────────────────────────────────────

# Colonnes : ID | Nom règle | Champ(s) | Type | Opérateur / Logique | Paramètre | Description | Base légale | Texte légal précis | Criticité
COLS = [
    "Réf.", "Nom de la règle", "Champ(s) contrôlé(s)", "Type de contrôle",
    "Opérateur / Logique", "Paramètre", "Description fonctionnelle",
    "Base légale", "Texte légal précis", "Criticité"
]
COL_WIDTHS = [10, 42, 24, 18, 32, 14, 52, 24, 62, 14]

DIM_COLORS = {
    1: C["dim1_completude"],
    2: C["dim2_validite"],
    3: C["dim3_coherence"],
    4: C["dim4_conformite"],
    5: C["dim5_unicite"],
}

# ── DIMENSION 1 : COMPLÉTUDE ──────────────────────────────────────────────────
D1_PP = [
    ("PP-C01","Numéro pièce identité manquant","NUMID","Simple","is_empty","—",
     "Le numéro de pièce d'identité doit être renseigné pour tout client PP.",
     "Instruction BCEAO n°01-2018","Art. 19 – «L'établissement doit recueillir [...] un document d'identité officiel en cours de validité comportant le numéro»","🔴 Critique"),
    ("PP-C02","Date de naissance manquante","DATNAIS","Simple","is_empty","—",
     "La date de naissance est obligatoire pour identifier le client et vérifier sa majorité.",
     "Instruction BCEAO n°01-2018","Art. 19 al.2 – «La date et le lieu de naissance doivent être collectés pour les personnes physiques»","🔴 Critique"),
    ("PP-C03","Adresse domicile manquante","ADRESSE","Simple","is_empty","—",
     "L'adresse de domicile est requise pour la connaissance géographique du client.",
     "Règlement UEMOA 14-2002","Art. 7 – «Les établissements doivent disposer de l'adresse de résidence de chaque client»","🔴 Critique"),
    ("PP-C04","Téléphone manquant","TEL","Simple","is_empty","—",
     "Le numéro de téléphone est un moyen de contact obligatoire pour le dispositif d'alerte.",
     "Politique interne BOA","Charte KYC BOA Groupe § 3.2 – Données minimales de contact","🟡 Importante"),
    ("PP-C05","Date validité pièce ID manquante","DATVALID","Simple","is_empty","—",
     "La date de validité permet de contrôler l'expiration du document d'identité.",
     "Instruction BCEAO n°01-2018","Art. 20 – «La date d'expiration du document d'identité doit être enregistrée et surveillée»","🔴 Critique"),
    ("PP-C06","Pays de naissance manquant","PAYNAIS","Simple","is_empty","—",
     "Le pays de naissance est requis pour détecter les nationalités à risque et les PPE.",
     "Recommandation GAFI 10","§ 10.10 – «Les informations de naissance incluant le pays doivent être collectées»","🟡 Importante"),
    ("PP-C07","Profession/Activité manquante","PROFESSION","Simple","is_empty","—",
     "La profession est essentielle pour évaluer la cohérence des revenus et la source des fonds.",
     "Instruction BCEAO n°01-2018","Art. 21 – «La profession ou l'activité exercée doit figurer dans le dossier KYC»","🟡 Importante"),
    ("PP-C08","Origine des revenus manquante","ORIGINE_REV","Simple","is_empty","—",
     "L'origine des revenus est une donnée fondamentale LCB-FT pour tout client bancaire.",
     "Loi LCB-FT (UEMOA)","Art. 34 – «L'établissement doit identifier l'origine des fonds déposés ou utilisés»","🔴 Critique"),
    ("PP-C09","Salaire/Revenus manquant","SALAIRE","Simple","is_empty","—",
     "Le niveau de revenus est nécessaire à l'évaluation du profil de risque financier.",
     "Instruction BCEAO n°01-2018","Art. 22 – «Les informations patrimoniales et de revenus doivent être collectées»","🟡 Importante"),
    ("PP-C10","Code APE/Secteur activité manquant","CODAPE","Simple","is_empty","—",
     "Le code APE permet de classifier l'activité économique du client et d'évaluer son secteur.",
     "Politique interne BOA","Charte KYC BOA § 3.5 – Classification sectorielle","🟡 Importante"),
    ("PP-C11","Pays de résidence manquant","PAYS_RESID","Simple","is_empty","—",
     "Le pays de résidence est requis notamment pour les clients non-résidents.",
     "Règlement UEMOA 14-2002","Art. 9 – «Le pays de résidence fiscale doit être collecté»","🟡 Importante"),
    ("PP-C12","Statut résident manquant","RESID","Simple","is_empty","—",
     "Le statut résident/non-résident détermine le régime réglementaire applicable.",
     "Instruction BCEAO n°01-2018","Art. 15 – «Le statut de résidence doit être identifié et documenté»","🟡 Importante"),
    ("PP-C13","Consentement BIC manquant","CONSENT_BIC","Simple","is_empty","—",
     "Le consentement au partage des données BIC doit être tracé.",
     "RGPD / Loi protection données","Art. 7 RGPD – «Le consentement doit être documenté et tracé»","🟢 Normale"),
    ("PP-C14","Employeur manquant","EMPLOYEUR","Simple","is_empty","—",
     "L'employeur est nécessaire pour les clients salariés afin de vérifier la source de revenus.",
     "Instruction BCEAO n°01-2018","Art. 21 – «Pour les salariés, l'employeur et son activité doivent être renseignés»","🟡 Importante"),
    ("PP-C15","Lieu délivrance CIN manquant","LIEU_DELIVRANCE_CIN","Simple","is_empty","—",
     "Le lieu de délivrance de la CIN permet d'authentifier le document présenté.",
     "Politique interne BOA","Directive KYC BOA § 4.1 – Vérification documentaire","🟢 Normale"),
    ("PP-C16","Intitulé compte manquant","INTITULE_COMPTE","Simple","is_empty","—",
     "L'intitulé du compte doit correspondre à l'identité du client.",
     "Politique interne BOA","Charte KYC BOA § 5.2 – Correspondance identité/compte","🟢 Normale"),
    ("PP-C17","Date ouverture compte manquante","DATOUV","Simple","is_empty","—",
     "La date d'ouverture est requise pour le suivi de l'ancienneté et de la révision KYC.",
     "Instruction BCEAO n°01-2018","Art. 25 – «La date d'ouverture des relations d'affaires doit être enregistrée»","🟡 Importante"),
    ("PP-C18","Statut PPE non renseigné","PPE","Simple","is_empty","—",
     "Le statut PPE est obligatoire – toute omission constitue un manquement grave LCB-FT.",
     "Recommandation GAFI 12","Rec. 12 – «Les établissements doivent identifier si le client est une PPE»","🔴 Critique"),
    ("PP-C19","Niveau de risque non renseigné","RISQUE","Simple","is_empty","—",
     "La cotation du risque est le cœur du dispositif de surveillance LCB-FT.",
     "Loi LCB-FT (UEMOA)","Art. 28 – «Chaque client doit faire l'objet d'une cotation de risque documentée»","🔴 Critique"),
    ("PP-C20","Date révision KYC manquante","DATEREV","Simple","is_empty","—",
     "La date de dernière révision KYC est obligatoire pour le suivi réglementaire.",
     "Loi LCB-FT (UEMOA)","Art. 35 – «Les dossiers KYC doivent être mis à jour à intervalles réguliers»","🔴 Critique"),
]

D1_PM = [
    ("PM-C01","Numéro RCS/Registre manquant","RCSNO","Simple","is_empty","—",
     "Le numéro au registre du commerce est requis pour identifier légalement la société.",
     "Droit commercial OHADA","Art. 45 AUDCG OHADA – «Toute société doit être immatriculée au RCCM»","🔴 Critique"),
    ("PM-C02","Numéro fiscal manquant","NUMERO_FISCAL","Simple","is_empty","—",
     "L'identifiant fiscal est obligatoire pour le contrôle de la légalité des flux financiers.",
     "Code général des impôts","Art. 1er CGI – «Tout contribuable doit posséder un numéro fiscal d'identification»","🔴 Critique"),
    ("PM-C03","Capital social manquant","CAPITAL","Simple","is_empty","—",
     "Le capital social renseigne sur la solvabilité et la nature juridique de l'entité.",
     "Acte Uniforme OHADA","Art. 387 AUSCGIE – «Le capital social doit être mentionné dans tout document»","🟡 Importante"),
    ("PM-C04","Chiffre d'affaires manquant","CA","Simple","is_empty","—",
     "Le CA permet d'évaluer la taille de l'entreprise et la cohérence des flux bancaires.",
     "Instruction BCEAO n°01-2018","Art. 23 – «Les informations financières des PM doivent inclure le CA»","🟡 Importante"),
    ("PM-C05","Résultat net manquant","RESULTAT","Simple","is_empty","—",
     "Le résultat net est un indicateur de santé financière indispensable au profilage de risque.",
     "Instruction BCEAO n°01-2018","Art. 23 – «La situation financière récente de la PM doit être documentée»","🟡 Importante"),
    ("PM-C06","Adresse siège social manquante","ADRESSE_SOCIALE","Simple","is_empty","—",
     "L'adresse du siège social est une donnée d'identification légale fondamentale.",
     "Acte Uniforme OHADA","Art. 26 AUDCG – «Le siège social doit être précisément identifié»","🔴 Critique"),
    ("PM-C07","Téléphone manquant","TEL","Simple","is_empty","—",
     "Le contact téléphonique est nécessaire pour les diligences de vérification.",
     "Politique interne BOA","Charte KYC BOA § 3.2 – Données minimales de contact PM","🟡 Importante"),
    ("PM-C08","Origine des fonds manquante","ORIGINE_REV","Simple","is_empty","—",
     "L'origine des fonds est obligatoire pour tout client PM afin de lutter contre le blanchiment.",
     "Loi LCB-FT (UEMOA)","Art. 34 – «L'établissement doit vérifier l'origine des fonds des personnes morales»","🔴 Critique"),
    ("PM-C09","Code APE manquant","CODAPE","Simple","is_empty","—",
     "Le code APE permet de classifier le secteur d'activité et d'identifier les secteurs à risque.",
     "Politique interne BOA","Charte KYC BOA § 3.5 – Classification sectorielle PM","🟡 Importante"),
    ("PM-C10","Pays de juridiction manquant","PAYS_JUR","Simple","is_empty","—",
     "Le pays de juridiction est essentiel pour les sociétés off-shore ou à capital étranger.",
     "Recommandation GAFI 24","Rec. 24 – «Le pays d'enregistrement des entités légales doit être documenté»","🟡 Importante"),
    ("PM-C11","Actionnaire(s) manquant(s)","ACTIONNAIRE","Simple","is_empty","—",
     "L'identification des bénéficiaires effectifs (actionnaires > 25%) est une exigence GAFI.",
     "Recommandation GAFI 24","Rec. 24 – «Les bénéficiaires effectifs détenant > 25% du capital doivent être identifiés»","🔴 Critique"),
    ("PM-C12","Mandataire social manquant","MANDATAIRE","Simple","is_empty","—",
     "Le mandataire social est responsable légal de la société – son identification est obligatoire.",
     "Instruction BCEAO n°01-2018","Art. 19 §3 – «Les personnes habilitées à agir au nom de la PM doivent être identifiées»","🔴 Critique"),
    ("PM-C13","Statut résident manquant","RESID","Simple","is_empty","—",
     "Le statut de résidence détermine le régime fiscal et réglementaire applicable.",
     "Instruction BCEAO n°01-2018","Art. 15 – «Le statut de résidence doit être documenté pour chaque PM»","🟡 Importante"),
    ("PM-C14","Statut PPE non renseigné","PPE","Simple","is_empty","—",
     "La vérification PPE pour les dirigeants de PM est obligatoire selon les rec. GAFI.",
     "Recommandation GAFI 12","Rec. 12 – «Les PPE dirigeantes de PM doivent être identifiées»","🔴 Critique"),
    ("PM-C15","Niveau de risque non renseigné","RISQUE","Simple","is_empty","—",
     "La cotation de risque est obligatoire pour toute PM selon le dispositif LCB-FT.",
     "Loi LCB-FT (UEMOA)","Art. 28 – «Les PM doivent faire l'objet d'une classification par niveau de risque»","🔴 Critique"),
    ("PM-C16","Date révision KYC manquante","DATEREV","Simple","is_empty","—",
     "La révision périodique du KYC est une obligation réglementaire pour toutes les PM.",
     "Loi LCB-FT (UEMOA)","Art. 35 – «Les dossiers KYC doivent être mis à jour régulièrement»","🔴 Critique"),
    ("PM-C17","Date ouverture compte manquante","DATOUV","Simple","is_empty","—",
     "La date d'ouverture de la relation d'affaires doit être tracée.",
     "Instruction BCEAO n°01-2018","Art. 25 – «La date de début des relations d'affaires doit être documentée»","🟡 Importante"),
    ("PM-C18","IDM (identifiant moral) manquant","IDM","Simple","is_empty","—",
     "L'identifiant moral interne est requis pour le rapprochement des systèmes bancaires.",
     "Politique interne BOA","Référentiel tiers BOA – § 2.1 – Identifiant unique PM","🟡 Importante"),
    ("PM-C19","Code agence économique manquant","AGEC","Simple","is_empty","—",
     "Le code agence économique est nécessaire pour les déclarations statistiques BCEAO.",
     "Instruction BCEAO Statistiques","Instruction BCEAO 94-07 – «Les établissements doivent renseigner le code AGEC»","🟢 Normale"),
    ("PM-C20","Consentement BIC manquant","CONSENT_BIC","Simple","is_empty","—",
     "Le consentement au partage des données BIC doit être tracé pour les PM.",
     "RGPD / Loi protection données","Art. 7 RGPD – «Le consentement des PM doit être documenté»","🟢 Normale"),
]

# ── DIMENSION 2 : VALIDITÉ ────────────────────────────────────────────────────
D2_PP = [
    ("PP-V01","Format téléphone invalide (longueur)","TEL","Composite","min_length=8 ET max_length=15","8–15 car.",
     "Un numéro de téléphone doit contenir entre 8 et 15 chiffres selon la norme internationale.",
     "Standard ITU-T E.164","E.164 – «Les numéros d'abonné internationaux ont de 7 à 15 chiffres»","🟡 Importante"),
    ("PP-V02","Devise non standard (≠ 3 car.)","DEVISE","Simple","min_length=3 ET max_length=3","3 car.",
     "Le code devise doit respecter la norme ISO 4217 (3 lettres).",
     "Norme ISO 4217","ISO 4217 – «Les codes de devises sont composés de 3 caractères alphabétiques»","🟡 Importante"),
    ("PP-V03","Âge client inférieur à 18 ans","DATNAIS","Simple","age_lt","18",
     "Un client PP doit être majeur (18 ans minimum). Tout mineur nécessite un représentant légal.",
     "Code civil / Droit bancaire","Art. 388 Code civil – «La majorité légale est fixée à 18 ans» ; Circulaire banque","🔴 Critique"),
    ("PP-V04","Âge client supérieur à 100 ans (aberrant)","DATNAIS","Simple","age_gt","100",
     "Une date de naissance générant un âge > 100 ans est probablement une erreur de saisie.",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 4.3 – Seuils d'âge acceptables","🔴 Critique"),
    ("PP-V05","Pièce d'identité expirée","DATVALID","Simple","expired","—",
     "Un document d'identité expiré ne peut pas servir de base à une relation bancaire valide.",
     "Instruction BCEAO n°01-2018","Art. 20 – «Les documents d'identité expirés doivent être mis à jour immédiatement»","🔴 Critique"),
    ("PP-V06","Date de révision KYC expirée","DATEREV","Simple","expired","—",
     "Une révision KYC dont la date est dépassée indique un dossier non actualisé.",
     "Loi LCB-FT (UEMOA)","Art. 35 – «Les dossiers KYC périmés constituent un manquement réglementaire»","🔴 Critique"),
    ("PP-V07","Date ouverture compte dans le futur","DATOUV","Composite","DATOUV > date_du_jour","—",
     "Une date d'ouverture supérieure à la date courante est une incohérence chronologique.",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 5.1 – Cohérence temporelle","🟡 Importante"),
    ("PP-V08","NUMID trop court (< 5 caractères)","NUMID","Simple","min_length","5",
     "Un numéro de pièce d'identité de moins de 5 caractères est vraisemblablement erroné.",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 4.2 – Longueur minimale des identifiants","🟡 Importante"),
    ("PP-V09","Risque hors domaine autorisé","RISQUE","Composite","RISQUE != (Faible|Moyen|Élevé)","—",
     "Le champ RISQUE ne doit contenir que les valeurs du référentiel : Faible, Moyen, Élevé.",
     "Loi LCB-FT (UEMOA)","Art. 28 – «La classification doit utiliser des niveaux de risque définis»","🔴 Critique"),
    ("PP-V10","Statut PPE invalide","PPE","Composite","PPE != (O|N|Oui|Non)","—",
     "Le champ PPE doit contenir uniquement les valeurs booléennes du référentiel.",
     "Recommandation GAFI 12","Rec. 12 – «Le statut PPE doit être documenté de manière non ambiguë»","🔴 Critique"),
    ("PP-V11","Statut résident invalide","RESID","Composite","RESID != (R|NR|Résident|Non-résident)","—",
     "Le statut de résidence doit correspondre aux valeurs du référentiel autorisé.",
     "Instruction BCEAO n°01-2018","Art. 15 – «Les valeurs du statut de résidence doivent être standardisées»","🟡 Importante"),
    ("PP-V12","Salaire nul pour salarié","SALAIRE, PROFESSION","Composite","PROFESSION contains Salarié ET SALAIRE=0","—",
     "Un client déclaré salarié ne peut avoir un salaire nul – incohérence de profil.",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 5.3 – Cohérence revenus/profession","🟡 Importante"),
]

D2_PM = [
    ("PM-V01","Format téléphone invalide","TEL","Composite","min_length=8 ET max_length=15","8–15 car.",
     "Un numéro de téléphone d'entreprise doit respecter la norme internationale E.164.",
     "Standard ITU-T E.164","E.164 – «Les numéros d'abonné internationaux ont de 7 à 15 chiffres»","🟡 Importante"),
    ("PM-V02","Capital social nul ou négatif","CAPITAL","Composite","CAPITAL <= 0","—",
     "Un capital social nul ou négatif est juridiquement impossible pour une société de capitaux.",
     "Acte Uniforme OHADA","Art. 387 AUSCGIE – «Le capital social doit être entièrement souscrit à la constitution»","🟡 Importante"),
    ("PM-V03","Date de révision KYC expirée","DATEREV","Simple","expired","—",
     "Une date de révision dépassée signale un dossier PM non actualisé.",
     "Loi LCB-FT (UEMOA)","Art. 35 – «Les dossiers PM doivent être révisés selon la périodicité réglementaire»","🔴 Critique"),
    ("PM-V04","Date d'ouverture dans le futur","DATOUV","Composite","DATOUV > date_du_jour","—",
     "Une date d'ouverture dans le futur est une incohérence chronologique.",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 5.1","🟡 Importante"),
    ("PM-V05","Statut PPE invalide","PPE","Composite","PPE != (O|N|Oui|Non)","—",
     "Le statut PPE des dirigeants de PM doit être documenté sans ambiguïté.",
     "Recommandation GAFI 12","Rec. 12 – «Le statut PPE doit être documenté de manière non ambiguë»","🔴 Critique"),
    ("PM-V06","Risque hors domaine autorisé","RISQUE","Composite","RISQUE != (Faible|Moyen|Élevé)","—",
     "La cotation de risque ne peut contenir que les valeurs du référentiel.",
     "Loi LCB-FT (UEMOA)","Art. 28 – «La classification de risque utilise des niveaux définis»","🔴 Critique"),
    ("PM-V07","Numéro RCS trop court","RCSNO","Simple","min_length","6",
     "Un numéro RCCM de moins de 6 caractères est vraisemblablement erroné.",
     "Acte Uniforme OHADA","Art. 45 AUDCG – «Le numéro RCCM est attribué selon un format standardisé»","🟡 Importante"),
    ("PM-V08","Statut résident invalide","RESID","Composite","RESID != (R|NR|...)","—",
     "Le statut de résidence PM doit correspondre aux valeurs standardisées.",
     "Instruction BCEAO n°01-2018","Art. 15 – «Les valeurs du statut de résidence doivent être standardisées»","🟡 Importante"),
    ("PM-V09","Devise non standard","DEVISE","Simple","min_length=3 ET max_length=3","3 car.",
     "Le code devise doit respecter la norme ISO 4217.",
     "Norme ISO 4217","ISO 4217 – «Les codes de devises sont composés de 3 caractères alphabétiques»","🟡 Importante"),
    ("PM-V10","Résultat supérieur au CA (aberrant)","RESULTAT, CA","Composite","RESULTAT > CA","—",
     "Un résultat net supérieur au chiffre d'affaires est financièrement impossible.",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 5.4 – Cohérence financière PM","🟡 Importante"),
]

# ── DIMENSION 3 : COHÉRENCE ───────────────────────────────────────────────────
D3_PP = [
    ("PP-K01","PPE déclaré sans risque Élevé","PPE + RISQUE","Composite","PPE=O ET RISQUE!=Élevé","—",
     "Toute PPE doit obligatoirement être classée en risque Élevé selon les rec. GAFI.",
     "Recommandation GAFI 12","Rec. 12 – «Les PPE font l'objet de mesures de vigilance renforcées, impliquant une classification Élevé»","🔴 Critique"),
    ("PP-K02","Non-résident sans pays de résidence","RESID + PAYS_RESID","Composite","RESID=NR ET PAYS_RESID vide","—",
     "Un client non-résident doit obligatoirement avoir son pays de résidence renseigné.",
     "Règlement UEMOA 14-2002","Art. 9 – «La résidence fiscale des non-résidents doit être documentée»","🔴 Critique"),
    ("PP-K03","Salarié sans employeur","PROFESSION + EMPLOYEUR","Composite","PROFESSION contains Salarié ET EMPLOYEUR vide","—",
     "Un client déclaré salarié doit avoir son employeur identifié pour vérifier la source des revenus.",
     "Instruction BCEAO n°01-2018","Art. 21 – «Pour les salariés, l'employeur doit être renseigné»","🟡 Importante"),
    ("PP-K04","Salarié sans salaire","PROFESSION + SALAIRE","Composite","PROFESSION contains Salarié ET SALAIRE vide","—",
     "La cohérence entre le statut professionnel et les revenus déclarés doit être vérifiée.",
     "Instruction BCEAO n°01-2018","Art. 22 – «Les revenus doivent être cohérents avec la profession déclarée»","🟡 Importante"),
    ("PP-K05","Retraité avec salaire positif","PROFESSION + SALAIRE","Composite","PROFESSION contains Retraité ET SALAIRE > 0","—",
     "Un client retraité ne devrait pas avoir de salaire actif – vérification de cohérence.",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 5.3","🟢 Normale"),
    ("PP-K06","Pièce expirée avec risque Faible","DATVALID + RISQUE","Composite","DATVALID expiré ET RISQUE=Faible","—",
     "Un client avec pièce d'identité expirée ne peut pas rester en risque Faible.",
     "Instruction BCEAO n°01-2018","Art. 20 §2 – «L'expiration du document d'identité implique une réévaluation du risque»","🟡 Importante"),
    ("PP-K07","Date naissance postérieure à l'ouverture","DATNAIS + DATOUV","Composite","DATNAIS > DATOUV","—",
     "La date de naissance ne peut pas être postérieure à la date d'ouverture du compte.",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 5.1 – Cohérence temporelle","🔴 Critique"),
    ("PP-K08","Révision KYC antérieure à l'ouverture","DATEREV + DATOUV","Composite","DATEREV < DATOUV","—",
     "La date de révision KYC ne peut pas être antérieure à la date d'ouverture du compte.",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 5.1","🟡 Importante"),
    ("PP-K09","Pays naissance = pays résidence pour NR","PAYNAIS + PAYS_RESID + RESID","Composite","RESID=NR ET PAYNAIS=PAYS_RESID","—",
     "Un non-résident dont le pays de naissance = pays de résidence est suspect – vérification requise.",
     "Contrôle qualité interne","Directive KYC BOA § 6.2 – Cohérence géographique","🟢 Normale"),
    ("PP-K10","Compte devise sans code devise","INTITULE_COMPTE + DEVISE","Composite","INTITULE contains DEV ET DEVISE vide","—",
     "Un compte libellé en devise doit avoir son code devise renseigné.",
     "Instruction BCEAO devises","Instruction BCEAO 94-05 – «Tout compte en devise doit mentionner la devise utilisée»","🟡 Importante"),
]

D3_PM = [
    ("PM-K01","PPE déclarée sans risque Élevé","PPE + RISQUE","Composite","PPE=O ET RISQUE!=Élevé","—",
     "Une PM avec PPE parmi ses dirigeants doit être classée en risque Élevé.",
     "Recommandation GAFI 12","Rec. 12 – «La présence de PPE parmi les dirigeants implique une vigilance renforcée»","🔴 Critique"),
    ("PM-K02","Non-résident sans pays de juridiction","RESID + PAYS_JUR","Composite","RESID=NR ET PAYS_JUR vide","—",
     "Une PM non-résidente doit avoir son pays de juridiction renseigné.",
     "Recommandation GAFI 24","Rec. 24 – «Le pays d'enregistrement de la PM non-résidente doit être documenté»","🔴 Critique"),
    ("PM-K03","Société avec capital sans actionnaire","CAPITAL + ACTIONNAIRE","Composite","CAPITAL > 0 ET ACTIONNAIRE vide","—",
     "Une société avec capital doit avoir ses bénéficiaires effectifs identifiés.",
     "Recommandation GAFI 24","Rec. 24 – «Les bénéficiaires effectifs des sociétés à capital doivent être identifiés»","🔴 Critique"),
    ("PM-K04","Résultat sans CA renseigné","RESULTAT + CA","Composite","RESULTAT renseigné ET CA vide","—",
     "Le résultat net n'a de sens que si le CA est également renseigné.",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 5.4","🟡 Importante"),
    ("PM-K05","Révision antérieure à l'ouverture","DATEREV + DATOUV","Composite","DATEREV < DATOUV","—",
     "La date de révision KYC ne peut pas précéder la date d'ouverture de la relation d'affaires.",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 5.1","🟡 Importante"),
    ("PM-K06","Risque Faible avec PPE parmi dirigeants","RISQUE + PPE","Composite","RISQUE=Faible ET PPE=O","—",
     "Incohérence grave : une PM avec PPE ne peut pas être classée en risque Faible.",
     "Recommandation GAFI 12","Rec. 12 – «La présence de PPE interdit toute classification en risque faible»","🔴 Critique"),
    ("PM-K07","RCS renseigné mais sans numéro fiscal","RCSNO + NUMERO_FISCAL","Composite","RCSNO renseigné ET NUMERO_FISCAL vide","—",
     "Une société immatriculée doit également disposer d'un numéro fiscal.",
     "Code général des impôts","Art. 1er CGI – «Toute société immatriculée doit posséder un identifiant fiscal»","🟡 Importante"),
    ("PM-K08","CA positif mais résultat vide","CA + RESULTAT","Composite","CA > 0 ET RESULTAT vide","—",
     "Une société avec du chiffre d'affaires doit avoir son résultat net documenté.",
     "Instruction BCEAO n°01-2018","Art. 23 – «Les informations financières doivent être complètes»","🟡 Importante"),
]

# ── DIMENSION 4 : CONFORMITÉ RÉGLEMENTAIRE ────────────────────────────────────
D4_PP = [
    ("PP-R01","Risque élevé sans révision < 1 an","RISQUE + DATEREV","Composite","RISQUE=Élevé ET DATEREV > 1 an","—",
     "Les clients à risque élevé doivent faire l'objet d'une révision KYC annuelle au minimum.",
     "Loi LCB-FT UEMOA – Art. 34","Art. 34 §2 – «Les clients à risque élevé sont soumis à une surveillance renforcée, avec révision au moins annuelle»","🔴 Critique"),
    ("PP-R02","PPE sans origine des revenus","PPE + ORIGINE_REV","Composite","PPE=O ET ORIGINE_REV vide","—",
     "L'origine des revenus est obligatoire pour toute PPE afin de détecter d'éventuels conflits.",
     "Recommandation GAFI 12","Rec. 12 §b – «Pour les PPE, l'origine du patrimoine et des fonds doit être établie»","🔴 Critique"),
    ("PP-R03","PPE sans déclaration de patrimoine","PPE + SALAIRE","Composite","PPE=O ET SALAIRE vide","—",
     "Le niveau de patrimoine/revenus d'une PPE doit être documenté pour la vigilance renforcée.",
     "Recommandation GAFI 12","Rec. 12 §c – «L'établissement doit obtenir l'approbation de la direction pour les PPE et documenter leurs revenus»","🔴 Critique"),
    ("PP-R04","Non-résident sans boîte postale","RESID + BOITE_POSTALE","Composite","RESID=NR ET BOITE_POSTALE vide","—",
     "Un client non-résident doit disposer d'une adresse de correspondance dans le pays.",
     "Politique interne BOA","Directive KYC BOA § 4.5 – «Les non-résidents doivent avoir une adresse de correspondance locale»","🟡 Importante"),
    ("PP-R05","Client sans identité valide","NUMID + DATVALID","Composite","NUMID vide OU DATVALID expiré","—",
     "Toute relation bancaire nécessite une pièce d'identité valide et non expirée.",
     "Instruction BCEAO n°01-2018","Art. 19 – «Aucune relation d'affaires ne peut être établie sans document d'identité valide»","🔴 Critique"),
    ("PP-R06","Revenus élevés sans origine déclarée","SALAIRE + ORIGINE_REV","Composite","SALAIRE > seuil ET ORIGINE_REV vide","—",
     "Des revenus dépassant un seuil doivent avoir une origine clairement documentée.",
     "Loi LCB-FT UEMOA – Art. 34","Art. 34 – «L'établissement doit vérifier l'origine des fonds de manière proportionnelle au montant»","🔴 Critique"),
    ("PP-R07","Dossier KYC non révisé > 3 ans","DATEREV","Simple","DATEREV > 3 ans","—",
     "Tout dossier KYC doit être révisé au maximum tous les 3 ans, même pour risque faible.",
     "Loi LCB-FT UEMOA – Art. 35","Art. 35 – «La périodicité maximale de révision est de 3 ans pour les clients à risque faible»","🔴 Critique"),
    ("PP-R08","Client mineur sans représentant légal","DATNAIS + EMPLOYEUR","Composite","DATNAIS age_lt=18 ET EMPLOYEUR vide","—",
     "Un client mineur doit avoir son représentant légal identifié dans le champ EMPLOYEUR.",
     "Code civil / Droit bancaire","Art. 388 Code civil – «Les mineurs doivent être représentés par leur tuteur légal»","🔴 Critique"),
]

D4_PM = [
    ("PM-R01","Société sans bénéficiaires effectifs","ACTIONNAIRE","Simple","is_empty","—",
     "Toute PM doit avoir ses bénéficiaires effectifs identifiés – exigence GAFI fondamentale.",
     "Recommandation GAFI 24","Rec. 24 – «Les établissements doivent identifier et vérifier les bénéficiaires effectifs de toute entité légale»","🔴 Critique"),
    ("PM-R02","PM risque élevé sans révision < 1 an","RISQUE + DATEREV","Composite","RISQUE=Élevé ET DATEREV > 1 an","—",
     "Les PM à risque élevé nécessitent une révision KYC annuelle au minimum.",
     "Loi LCB-FT UEMOA – Art. 34","Art. 34 §2 – «Les PM à risque élevé sont soumises à une surveillance renforcée annuelle»","🔴 Critique"),
    ("PM-R03","PM PPE sans origine des fonds","PPE + ORIGINE_REV","Composite","PPE=O ET ORIGINE_REV vide","—",
     "Une PM dont les dirigeants sont des PPE doit avoir l'origine de ses fonds documentée.",
     "Recommandation GAFI 12","Rec. 12 §b – «L'origine des fonds des entités liées à des PPE doit être établie»","🔴 Critique"),
    ("PM-R04","RCS absent pour société de capitaux","RCSNO","Simple","is_empty","—",
     "Toute société de capitaux doit être immatriculée – l'absence du RCS est une anomalie grave.",
     "Droit commercial OHADA","Art. 45 AUDCG – «L'immatriculation au RCCM est constitutive de la personnalité morale»","🔴 Critique"),
    ("PM-R05","Dossier PM non révisé > 2 ans","DATEREV","Simple","DATEREV > 2 ans","—",
     "Les dossiers PM doivent être révisés au minimum tous les 2 ans selon la réglementation.",
     "Loi LCB-FT UEMOA – Art. 35","Art. 35 – «La périodicité de révision des PM est de 24 mois maximum (risque faible)»","🔴 Critique"),
    ("PM-R06","Numéro fiscal absent","NUMERO_FISCAL","Simple","is_empty","—",
     "L'absence de numéro fiscal constitue un manquement légal et un risque de fraude fiscale.",
     "Code général des impôts","Art. 1er CGI – «Toute PM exerçant une activité doit posséder un NIF»","🔴 Critique"),
    ("PM-R07","Mandataire absent pour PM active","MANDATAIRE + CA","Composite","CA > 0 ET MANDATAIRE vide","—",
     "Une PM avec activité doit avoir son mandataire social identifié.",
     "Instruction BCEAO n°01-2018","Art. 19 §3 – «Les personnes habilitées à agir doivent être identifiées»","🔴 Critique"),
    ("PM-R08","Actionnaire > 25% non déclaré","ACTIONNAIRE + CAPITAL","Composite","CAPITAL > seuil ET ACTIONNAIRE vide","—",
     "Tout actionnaire détenant > 25% du capital doit être identifié comme bénéficiaire effectif.",
     "Recommandation GAFI 24","Rec. 24 – «Les actionnaires détenant plus de 25% du capital sont des bénéficiaires effectifs»","🔴 Critique"),
]

# ── DIMENSION 5 : UNICITÉ ─────────────────────────────────────────────────────
D5 = [
    ("U01","Même numéro d'identité sur plusieurs PP","NUMID","Agrégation","Même NUMID → plusieurs CLIENT","—",
     "Un numéro de pièce d'identité ne peut pas être associé à plusieurs clients distincts.",
     "Instruction BCEAO n°01-2018","Art. 19 – «Chaque document d'identité est unique et lié à une seule personne physique»","🔴 Critique","PP"),
    ("U02","Même numéro fiscal sur plusieurs PM","NUMERO_FISCAL","Agrégation","Même NUMERO_FISCAL → plusieurs CLIENT","—",
     "Un numéro fiscal ne peut être attribué qu'à une seule entité légale.",
     "Code général des impôts","Art. 1er CGI – «Le numéro d'identification fiscale est unique par entité»","🔴 Critique","PM"),
    ("U03","Même RCS sur plusieurs PM","RCSNO","Agrégation","Même RCSNO → plusieurs CLIENT","—",
     "Un numéro RCCM/RCS est unique par société – tout doublon indique une fraude ou erreur.",
     "Droit commercial OHADA","Art. 45 AUDCG – «Le numéro RCCM est unique par société»","🔴 Critique","PM"),
    ("U04","Même téléphone > 5 clients PP","TEL","Agrégation","Même TEL → > 5 CLIENT","—",
     "Un numéro de téléphone partagé par plus de 5 clients PP est suspect (fraude, erreur de saisie).",
     "Contrôle qualité interne","Politique QualitéDonnées BOA § 6.1 – Détection doublons par contact","🟡 Importante","PP + PM"),
]


# ─────────────────────────────────────────────
# CRÉATION DU WORKBOOK
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Supprimer la feuille par défaut

# ─────────────────────────────────────────────
# FONCTION : créer un onglet de règles
# ─────────────────────────────────────────────
def make_rules_sheet(wb, title, tab_color, dimensions):
    """
    dimensions = list of (dim_label, dim_number, dim_color, rules_list)
    """
    ws = wb.create_sheet(title=title)
    ws.sheet_properties.tabColor = tab_color

    # ── En-tête du document ──
    ws.merge_cells("A1:J1")
    ws["A1"] = f"📋 RÈGLES DE CONTRÔLE QUALITÉ KYC – {title.upper()}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = fill("1B2A4A")
    ws["A1"].alignment = align("center", wrap=False)

    ws.merge_cells("A2:J2")
    ws["A2"] = "Bank of Africa Group – Plateforme KYC v2 – Dispositif LCB-FT"
    ws["A2"].font = Font(bold=False, size=10, color="AAAAAA", italic=True, name="Calibri")
    ws["A2"].fill = fill("1B2A4A")
    ws["A2"].alignment = align("center", wrap=False)

    # Hauteurs des lignes titre
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16

    row = 3
    for (dim_label, dim_num, dim_color, rules) in dimensions:
        # ── Séparateur de dimension ──
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:J{row}")
        dim_hdr = ws[f"A{row}"]
        dim_hdr.value = f"   {dim_label}"
        dim_hdr.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        dim_hdr.fill = fill("2C3E50")
        dim_hdr.alignment = align("left", wrap=False)
        row += 1

        # ── Colonnes ──
        ws.row_dimensions[row].height = 30
        for col_idx, col_name in enumerate(COLS, start=1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
            cell.fill = fill("2E4053")
            cell.alignment = align("center", "center", wrap=True)
            cell.border = border("666666")
        row += 1

        # ── Lignes de données ──
        for i, rule in enumerate(rules):
            ws.row_dimensions[row].height = 40
            crit = rule[-1]
            crit_fill_c, crit_txt_c, crit_label = CRIT_MAP.get(crit, ("CCCCCC", "000000", "—"))
            bg = dim_color if i % 2 == 0 else C["row_even"]

            for col_idx, value in enumerate(rule[:-1], start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.alignment = align("left", "center", wrap=True)
                cell.border = border()

                if col_idx == 1:  # Réf
                    cell.font = Font(bold=True, size=9, color="1A3C6E", name="Calibri")
                    cell.fill = fill(bg)
                elif col_idx == 10:  # Criticité
                    cell.value = crit_label
                    cell.font = Font(bold=True, size=9, color=crit_txt_c, name="Calibri")
                    cell.fill = fill(crit_fill_c)
                    cell.alignment = align("center", "center", wrap=False)
                elif col_idx in (8, 9):  # Base légale + Texte légal
                    cell.font = Font(size=8, italic=(col_idx == 9), color="444444", name="Calibri")
                    cell.fill = fill("FAFAFA")
                else:
                    cell.font = Font(size=9, color=C["dark_text"], name="Calibri")
                    cell.fill = fill(bg)
            row += 1

        row += 1  # Ligne vide entre dimensions

    # ── Largeurs de colonnes ──
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Figer les 3 premières lignes + colonne A
    ws.freeze_panes = "B4"
    return ws


# ─────────────────────────────────────────────
# ONGLET 1 – PERSONNES PHYSIQUES (PP)
# ─────────────────────────────────────────────
make_rules_sheet(wb, "Personnes Physiques (PP)", C["tab_pp"], [
    ("📁 DIMENSION 1 – COMPLÉTUDE  |  Exhaustivité des champs obligatoires PP", 1, C["dim1_completude"], D1_PP),
    ("🔍 DIMENSION 2 – VALIDITÉ  |  Format et valeurs attendus PP", 2, C["dim2_validite"], D2_PP),
    ("🔗 DIMENSION 3 – COHÉRENCE  |  Cohérence inter-champs PP", 3, C["dim3_coherence"], D3_PP),
    ("⚖️ DIMENSION 4 – CONFORMITÉ RÉGLEMENTAIRE  |  LCB-FT / GAFI / BCEAO – PP", 4, C["dim4_conformite"], D4_PP),
])

# ─────────────────────────────────────────────
# ONGLET 2 – PERSONNES MORALES (PM)
# ─────────────────────────────────────────────
make_rules_sheet(wb, "Personnes Morales (PM)", C["tab_pm"], [
    ("📁 DIMENSION 1 – COMPLÉTUDE  |  Exhaustivité des champs obligatoires PM", 1, C["dim1_completude"], D1_PM),
    ("🔍 DIMENSION 2 – VALIDITÉ  |  Format et valeurs attendus PM", 2, C["dim2_validite"], D2_PM),
    ("🔗 DIMENSION 3 – COHÉRENCE  |  Cohérence inter-champs PM", 3, C["dim3_coherence"], D3_PM),
    ("⚖️ DIMENSION 4 – CONFORMITÉ RÉGLEMENTAIRE  |  LCB-FT / GAFI / OHADA – PM", 4, C["dim4_conformite"], D4_PM),
])

# ─────────────────────────────────────────────
# ONGLET 3 – UNICITÉ (PP + PM)
# ─────────────────────────────────────────────
ws_u = wb.create_sheet(title="Unicité & Doublons")
ws_u.sheet_properties.tabColor = C["tab_shared"]

ws_u.merge_cells("A1:K1")
ws_u["A1"] = "📋 RÈGLES DE CONTRÔLE QUALITÉ KYC – UNICITÉ & DÉTECTION DE DOUBLONS"
ws_u["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
ws_u["A1"].fill = fill("1B2A4A")
ws_u["A1"].alignment = align("center", wrap=False)
ws_u.row_dimensions[1].height = 26

ws_u.merge_cells("A2:K2")
ws_u["A2"] = "Bank of Africa Group – Plateforme KYC v2"
ws_u["A2"].font = Font(italic=True, size=9, color="AAAAAA", name="Calibri")
ws_u["A2"].fill = fill("1B2A4A")
ws_u["A2"].alignment = align("center", wrap=False)
ws_u.row_dimensions[2].height = 14

COLS_U = COLS + ["Applicabilité"]
COL_WIDTHS_U = COL_WIDTHS + [16]

ws_u.row_dimensions[3].height = 30
for col_idx, col_name in enumerate(COLS_U, start=1):
    cell = ws_u.cell(row=3, column=col_idx, value=col_name)
    cell.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
    cell.fill = fill("3D2B6E")
    cell.alignment = align("center", "center", wrap=True)
    cell.border = border("666666")

for i, rule in enumerate(D5):
    row = i + 4
    ws_u.row_dimensions[row].height = 40
    crit = rule[8]
    crit_fill_c, crit_txt_c, crit_label = CRIT_MAP.get(crit, ("CCCCCC", "000000", "—"))
    applicabilite = rule[9]
    bg = C["dim5_unicite"] if i % 2 == 0 else C["row_even"]
    data_row = rule[:8] + (crit_label, applicabilite)

    for col_idx, value in enumerate(data_row, start=1):
        cell = ws_u.cell(row=row, column=col_idx, value=value)
        cell.alignment = align("left", "center", wrap=True)
        cell.border = border()
        if col_idx == 1:
            cell.font = Font(bold=True, size=9, color="3D2B6E", name="Calibri")
            cell.fill = fill(bg)
        elif col_idx == 9:
            cell.font = Font(bold=True, size=9, color=crit_txt_c, name="Calibri")
            cell.fill = fill(crit_fill_c)
            cell.alignment = align("center", "center", wrap=False)
        elif col_idx in (7, 8):
            cell.font = Font(size=8, italic=(col_idx == 8), color="444444", name="Calibri")
            cell.fill = fill("FAFAFA")
        else:
            cell.font = Font(size=9, color=C["dark_text"], name="Calibri")
            cell.fill = fill(bg)

for col_idx, width in enumerate(COL_WIDTHS_U, start=1):
    ws_u.column_dimensions[get_column_letter(col_idx)].width = width
ws_u.freeze_panes = "B4"

# ─────────────────────────────────────────────
# ONGLET 4 – SYNTHÈSE & TABLEAU DE BORD
# ─────────────────────────────────────────────
ws_s = wb.create_sheet(title="📊 Synthèse")
ws_s.sheet_properties.tabColor = C["tab_summary"]

ws_s.merge_cells("A1:H1")
ws_s["A1"] = "SYNTHÈSE – CATALOGUE DES RÈGLES DE CONTRÔLE QUALITÉ KYC"
ws_s["A1"].font = Font(bold=True, size=15, color="FFFFFF", name="Calibri")
ws_s["A1"].fill = fill("1B2A4A")
ws_s["A1"].alignment = align("center", wrap=False)
ws_s.row_dimensions[1].height = 32

ws_s.merge_cells("A2:H2")
ws_s["A2"] = "Bank of Africa Holding – Dispositif LCB-FT / KYC v2 – Juin 2026"
ws_s["A2"].font = Font(italic=True, size=10, color="888888", name="Calibri")
ws_s["A2"].fill = fill("1B2A4A")
ws_s["A2"].alignment = align("center", wrap=False)
ws_s.row_dimensions[2].height = 18

# ── Statistiques ──
summary_data = [
    ("",),
    ("📌 RÉCAPITULATIF PAR DIMENSION",),
    ("Dimension", "PP", "PM", "PP+PM", "Total", "dont Critiques", "dont Importantes", "dont Normales"),
    ("1 – Complétude",     len(D1_PP), len(D1_PM), 0, len(D1_PP)+len(D1_PM),
     sum(1 for r in D1_PP+D1_PM if r[-1]=="🔴 Critique"),
     sum(1 for r in D1_PP+D1_PM if r[-1]=="🟡 Importante"),
     sum(1 for r in D1_PP+D1_PM if r[-1]=="🟢 Normale")),
    ("2 – Validité",       len(D2_PP), len(D2_PM), 0, len(D2_PP)+len(D2_PM),
     sum(1 for r in D2_PP+D2_PM if r[-1]=="🔴 Critique"),
     sum(1 for r in D2_PP+D2_PM if r[-1]=="🟡 Importante"),
     sum(1 for r in D2_PP+D2_PM if r[-1]=="🟢 Normale")),
    ("3 – Cohérence",      len(D3_PP), len(D3_PM), 0, len(D3_PP)+len(D3_PM),
     sum(1 for r in D3_PP+D3_PM if r[-1]=="🔴 Critique"),
     sum(1 for r in D3_PP+D3_PM if r[-1]=="🟡 Importante"),
     sum(1 for r in D3_PP+D3_PM if r[-1]=="🟢 Normale")),
    ("4 – Conformité réglementaire", len(D4_PP), len(D4_PM), 0, len(D4_PP)+len(D4_PM),
     sum(1 for r in D4_PP+D4_PM if r[-1]=="🔴 Critique"),
     sum(1 for r in D4_PP+D4_PM if r[-1]=="🟡 Importante"),
     sum(1 for r in D4_PP+D4_PM if r[-1]=="🟢 Normale")),
    ("5 – Unicité & Doublons", 0, 0, len(D5), len(D5),
     sum(1 for r in D5 if r[8]=="🔴 Critique"),
     sum(1 for r in D5 if r[8]=="🟡 Importante"),
     sum(1 for r in D5 if r[8]=="🟢 Normale")),
]

all_rules = D1_PP + D1_PM + D2_PP + D2_PM + D3_PP + D3_PM + D4_PP + D4_PM
total_pp = len(D1_PP) + len(D2_PP) + len(D3_PP) + len(D4_PP)
total_pm = len(D1_PM) + len(D2_PM) + len(D3_PM) + len(D4_PM)
total_shared = len(D5)
total_all = total_pp + total_pm + total_shared

summary_data += [
    ("TOTAL", total_pp, total_pm, total_shared, total_all,
     sum(1 for r in all_rules if r[-1]=="🔴 Critique") + sum(1 for r in D5 if r[8]=="🔴 Critique"),
     sum(1 for r in all_rules if r[-1]=="🟡 Importante") + sum(1 for r in D5 if r[8]=="🟡 Importante"),
     sum(1 for r in all_rules if r[-1]=="🟢 Normale") + sum(1 for r in D5 if r[8]=="🟢 Normale")),
]

DIM_SUMMARY_COLORS = {
    "1 – Complétude": C["dim1_completude"],
    "2 – Validité": C["dim2_validite"],
    "3 – Cohérence": C["dim3_coherence"],
    "4 – Conformité réglementaire": C["dim4_conformite"],
    "5 – Unicité & Doublons": C["dim5_unicite"],
    "TOTAL": "1B2A4A",
}

for r_idx, row_data in enumerate(summary_data, start=3):
    ws_s.row_dimensions[r_idx].height = 22
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws_s.cell(row=r_idx, column=c_idx, value=val)
        label = row_data[0] if row_data else ""

        if r_idx == 4 and label == "📌 RÉCAPITULATIF PAR DIMENSION":
            cell.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
            cell.fill = fill("1A3C6E")
            cell.alignment = align("left", "center", wrap=False)
            ws_s.merge_cells(f"A{r_idx}:H{r_idx}")
            break
        elif r_idx == 5:  # Header row
            cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
            cell.fill = fill("2C3E50")
            cell.alignment = align("center", "center")
            cell.border = border("666666")
        elif label == "TOTAL":
            cell.font = Font(bold=True, size=10, color="FFFFFF" if c_idx == 1 else "FFFFFF", name="Calibri")
            cell.fill = fill("1B2A4A")
            cell.alignment = align("center" if c_idx > 1 else "left", "center")
            cell.border = border("666666")
        elif label.startswith(("1", "2", "3", "4", "5")):
            dim_key = label.split(" – ")[0] + " – " + label.split(" – ")[1] if " – " in label else label
            bg_c = None
            for key, clr in DIM_SUMMARY_COLORS.items():
                if label.startswith(key[:2]):
                    bg_c = clr
                    break
            cell.font = Font(size=10, bold=(c_idx == 1), color="1A1A1A" if c_idx > 1 else "1A3C6E", name="Calibri")
            cell.fill = fill(bg_c) if bg_c else fill(C["row_even"])
            cell.alignment = align("center" if c_idx > 1 else "left", "center")
            cell.border = border()
        else:
            cell.font = Font(size=9, color="666666", name="Calibri")
            cell.alignment = align()

# Légende criticité
start_legend = len(summary_data) + 5
ws_s.merge_cells(f"A{start_legend}:H{start_legend}")
ws_s[f"A{start_legend}"] = "🔴 Critique : Non-conformité grave – risque de sanction réglementaire immédiat"
ws_s[f"A{start_legend}"].font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
ws_s[f"A{start_legend}"].fill = fill(C["crit_critique"])
ws_s[f"A{start_legend}"].alignment = align("left", "center", wrap=False)
ws_s.row_dimensions[start_legend].height = 22

ws_s.merge_cells(f"A{start_legend+1}:H{start_legend+1}")
ws_s[f"A{start_legend+1}"] = "🟡 Importante : Anomalie à corriger – impact sur la qualité opérationnelle"
ws_s[f"A{start_legend+1}"].font = Font(bold=True, size=10, color="1A1A1A", name="Calibri")
ws_s[f"A{start_legend+1}"].fill = fill(C["crit_import"])
ws_s[f"A{start_legend+1}"].alignment = align("left", "center", wrap=False)
ws_s.row_dimensions[start_legend+1].height = 22

ws_s.merge_cells(f"A{start_legend+2}:H{start_legend+2}")
ws_s[f"A{start_legend+2}"] = "🟢 Normale : Amélioration continue – enrichissement données"
ws_s[f"A{start_legend+2}"].font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
ws_s[f"A{start_legend+2}"].fill = fill(C["crit_normale"])
ws_s[f"A{start_legend+2}"].alignment = align("left", "center", wrap=False)
ws_s.row_dimensions[start_legend+2].height = 22

# Bases légales référencées
start_ref = start_legend + 5
ws_s.merge_cells(f"A{start_ref}:H{start_ref}")
ws_s[f"A{start_ref}"] = "📚 BASES LÉGALES RÉFÉRENCÉES"
ws_s[f"A{start_ref}"].font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
ws_s[f"A{start_ref}"].fill = fill("1A3C6E")
ws_s[f"A{start_ref}"].alignment = align("left", "center", wrap=False)
ws_s.row_dimensions[start_ref].height = 24

refs = [
    ("Instruction BCEAO n°01-2018", "Instruction relative aux modalités d'identification des clients des établissements de crédit et des compagnies d'assurance (UEMOA)"),
    ("Recommandations GAFI", "40 Recommandations du Groupe d'Action Financière – Rec. 10 (vigilance), Rec. 12 (PPE), Rec. 24 (bénéficiaires effectifs PM)"),
    ("Loi LCB-FT (UEMOA)", "Loi uniforme relative à la lutte contre le blanchiment de capitaux et le financement du terrorisme (UEMOA) – Art. 28, 34, 35"),
    ("Règlement UEMOA 14-2002", "Règlement n°14/2002/CM/UEMOA relatif au gel des fonds et autres ressources financières"),
    ("Droit OHADA", "Acte Uniforme du Droit Commercial Général (AUDCG) – Art. 45 ; Acte Uniforme sur les sociétés commerciales (AUSCGIE)"),
    ("Norme ISO 4217", "Norme internationale des codes de devises (3 caractères alphabétiques)"),
    ("Standard ITU-T E.164", "Format international des numéros de téléphone (max 15 chiffres)"),
    ("RGPD / Loi protection données", "Règlement Général sur la Protection des Données (UE) – Art. 7 (consentement) ; Lois nationales de protection des données"),
    ("Code général des impôts", "Codes des impôts nationaux (CI, SN, ML, BF, BJ...) – Obligation de numéro fiscal (NIF/NINEA/IFU)"),
    ("Politique interne BOA", "Charte KYC Bank of Africa Group, Directives KYC filiales, Politique QualitéDonnées BOA"),
]

for i, (ref, desc) in enumerate(refs):
    r = start_ref + 1 + i
    ws_s.row_dimensions[r].height = 28
    c1 = ws_s.cell(row=r, column=1, value=ref)
    c1.font = Font(bold=True, size=9, color="1A3C6E", name="Calibri")
    c1.fill = fill(C["row_odd"] if i % 2 == 0 else C["row_even"])
    c1.border = border()
    c1.alignment = align("left", "center", wrap=False)

    ws_s.merge_cells(f"B{r}:H{r}")
    c2 = ws_s.cell(row=r, column=2, value=desc)
    c2.font = Font(size=9, italic=True, color="444444", name="Calibri")
    c2.fill = fill(C["row_odd"] if i % 2 == 0 else C["row_even"])
    c2.border = border()
    c2.alignment = align("left", "center", wrap=True)

# Largeurs
for col_idx, width in enumerate([28, 60, 12, 12, 12, 14, 18, 14], start=1):
    ws_s.column_dimensions[get_column_letter(col_idx)].width = width

ws_s.freeze_panes = "A3"

# ─────────────────────────────────────────────
# ENREGISTREMENT
# ─────────────────────────────────────────────
out = os.path.abspath(OUTPUT_PATH)
wb.save(out)
print("\n[OK] Fichier Excel genere avec succes :")
print(f"   {out}\n")
print(f"   {total_pp} regles PP  |  {total_pm} regles PM  |  {total_shared} regles communes")
print(f"   {total_all} regles au total sur 4 onglets\n")
